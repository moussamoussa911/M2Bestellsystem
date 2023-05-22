import math
import os
import sqlite3
import subprocess
import tempfile
import time
import tkinter
import tkinter.font
import tkinter.messagebox as MessageBox
import emoji
from threading import Timer
from tkinter import *
from tkinter import messagebox
from tkinter import ttk, Listbox
import matplotlib.pyplot as plt
import tkintermapview
import win32api
import win32print
import wmi
from PIL import Image, ImageTk
from keyboard import press
from ttkwidgets.autocomplete import AutocompleteEntry
import sys, os
import code
import hashlib
import queue
import sys
import threading
import tkinter as tk
import traceback
from tkinter.scrolledtext import ScrolledText
import docx
from idlelib.tooltip import Hovertip
import pandas as pd
import win32ui
import win32print
import win32con
import csv
import uuid
import re
import textwrap
import pyodbc
import datetime



def resource_path(relative_path):
    try:
        base_path=sys._MEIPASS2
    except Exception:
        base_path=os.path.abspath(".")
        return os.path.join(base_path,relative_path)
# Open the PLZ and Ort text files
with open(resource_path('Data\plz.txt'), 'r', encoding='utf-8') as plz_file:
    plz_text = plz_file.read()

with open(resource_path('Data\ort.txt'), 'r', encoding='utf-8') as ort_file:
    ort_text = ort_file.read()

# Extract the PLZ and Ort data using regular expressions
plz_list = re.findall(r'\d{5}', plz_text)
ort_list = re.findall(r'[a-zA-ZäöüÄÖÜß()-]+', ort_text)

# Combine the PLZ and Ort data into a dictionary
plz_ort_dict = dict(zip(plz_list, ort_list))
conn = sqlite3.connect(resource_path('Data\Pending.db'))
cur = conn.cursor()

cur.execute('CREATE TABLE IF NOT EXISTS pending (Anzahl INTEGER, Datum INTEGER, Gesamt INTEGER, Fahrer TEXT)')
cur.execute('CREATE TABLE IF NOT EXISTS Rechnung (Nr INTEGER, Datum TEXT, Preis INTEGER, Fahrer TEXT)')

conn.commit()
conn.close()

# Create a connection to the database
conn = sqlite3.connect(resource_path("Data\OrderData.db"))
cursor = conn.cursor()

# Check if the table exists
cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='lieferpara'")
result = cursor.fetchone()

if result:
    # Table exists, do nothing
    print("Table already exists")
else:
    # Table does not exist, create it
    cursor.execute("CREATE TABLE lieferpara(geld text, zeit Integer)")

# Check if the table exists
cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Rabattpara'")
result = cursor.fetchone()

if result:
    # Table exists, do nothing
    print("Table already exists")
else:
    # Table does not exist, create it
    cursor.execute("CREATE TABLE Rabattpara(Rabatt text, zeit Integer)")
# Commit changes and close the connection
conn.commit()
conn.close()


#
# df=pd.read_excel(r"C:\\Users\\Admin\\PycharmProjects\\Orignal version\\Kunden.xlsx", usecols=["Telefonnummer"
# , "Name", "Strasse", "Nummer", "Plz", "Ort","KdNr"])
# df.to_csv(r"C:\\Users\\Admin\\PycharmProjects\\Orignal version\\temp_file.csv", index=False)
#
# # Read the data back into a new DataFrame, without the index
# df = pd.read_csv(r"C:\\Users\\Admin\\PycharmProjects\\Orignal version\\temp_file.csv")
# df = df[df['KdNr'].notna()]
# conn=sqlite3.connect('Kundendaten.db')
# cursor=conn.cursor()
#
#
# add=0
# for index, row in df.iterrows():
#     values = row[['Telefonnummer', 'Name', 'Strasse', 'Nummer', 'Plz', 'Ort']].tolist()
#     cursor.execute(
#         "INSERT INTO kundendaten (Telefon, Name, Addresse , Nr, PLZ, ORT ) VALUES (?,?,?,?,?,?)", values)
#     add+=1
#
# conn.commit()

# import requests
#
# # TODO: replace with the URL of your Flask server
# verify_license_url = 'http://127.0.0.1:5000/verify_license'
#
#
# # TODO: replace with the license key or unique identifier to be verified
# license_key = '1'
#
# # send a POST request to the license verification endpoint with the license key
# response = requests.post(verify_license_url, data={'license_key': license_key})
#
# # parse the response data as JSON and retrieve the validation code or error message
# response_data = response.json()
#


# assuming last_check_time_string is 'YYYY-MM-DD HH:MM:SS'
# last_check_time_string = '2023-03-10 15:30:00'
#
# # convert the string to a datetime object
# last_check_time = datetime.datetime.strptime(last_check_time_string, '%Y-%m-%d %H:%M:%S')

# get the current time
# current_time = datetime.datetime.now()
#
# # calculate the time difference between the current time and the last check time
# time_since_last_check = current_time - last_check_time

# check if the last check was made more than 30 days ago


# invoice_data = {
#     'field1': 'Sample Value 1',
#     'field2': 'Sample Value 2',
#     # Add more fields as required by your database schema
# }
#
#
# import requests
#
# def send_invoice_to_api(invoice_data):
#     url = 'https://manage.wix.com/dashboard/6329b66d-c4c4-4a28-8613-de6ff782dcae/restaurants/orders?referralInfo=sidebar'
#     response = requests.post(url, json=invoice_data)
#
#     if response.status_code == 201:
#         "Invoice sent successfully")
#     else:
#         print("Error sending invoice:", response.text)
#
#
# send_invoice_to_api(invoice_data)
#
#
conn=sqlite3.connect(resource_path('Data\Einstellung.db'))
cur = conn.cursor()
cur.execute('select* from fast')
nado = cur.fetchall()
frei = 'Orignal style'
if nado:
    for t in nado:
        frei = t[1]
print(frei)
if frei=='90s style':
    colour0='lemon chiffon'
    colour1='red'
    colour2='gold'
    colour3='green'
    colour4='darkslategray4'
    colour5='yellow3'
    theme='alt'
else:
    colour0='white smoke'
    colour1='#2d3f4e'
    colour2='#c2b59f'
    colour3='grey'
    colour4='dark slate gray'
    colour5='dark slate gray'
    theme='clam'

# Connect to the database and execute a select statement
license_key = '1'
conn = sqlite3.connect(resource_path('Data\licenses101.db'))
cur = conn.cursor()

cur.execute('SELECT validation_code, expiration_date FROM licenses WHERE license_key = ?', (license_key,))
row = cur.fetchone()

# Convert the expiration date string to a date object
expiration_date = datetime.datetime.strptime(row[1], '%Y-%m-%d').date()

# Calculate the number of days left until the expiration date
daysleft = datetime.date.today() - expiration_date

# Check if the license is still valid
if datetime.date.today() <= expiration_date:
    from datetime import datetime

    root = None
#kiko

    def create_root():
        global root,Bestellliste
        if root is None:
            # Define the root window
            root = tkinter.Tk()
            '''''''''''''''Resturant Daten'''''''''''''''''''''''
            conn = sqlite3.connect(resource_path('Data\Einstellung.db'))
            cur = conn.cursor()
            cur.execute('select * from Drucker_Addresse ')
            restaddresse = cur.fetchall()
            for restaddes in restaddresse:
                restName = restaddes[0]
                restStr = restaddes[1]
                restTele = restaddes[2]
            '''''''''''''''Resturant Daten'''''''''''''''''''''''
            # Define the Exit function
            def Exit():
                sure = messagebox.askyesno("Exit", "Are you sure you want to exit?", parent=root)
                if sure:
                    root.destroy()
                    sys.exit()


            # Set the size and properties of the root window
            titlespace = " "
            window_height = 800
            window_width = 1200
            screen_width = root.winfo_screenwidth()
            screen_height = root.winfo_screenheight()
            show = screen_height + screen_width
            infosybol = (f'{emoji.emojize(":left_speech_bubble:")}')

            if show < 3500:
                root.eval('tk::PlaceWindow . center')

            x_cordinate = int((screen_width / 2) - (window_width / 2))
            y_cordinate = int((screen_height / 2) - (window_height / 2))
            root.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
            root.title(210 * titlespace + "M2 Bestellsystem")
            root['bg'] = colour4
            root.resizable(False, False)
            import turtle
            # Create a turtle canvas within the root window
            canvas = tk.Canvas(root, width=1100, height=120, bd=0, highlightthickness=0)
            canvas.place(x=40,y=20)
            canvas.configure(bg=colour4)
            font = ('Arial', 26, 'bold')
            t = turtle.RawTurtle(canvas)

            # Set the background color to white
            canvas.configure(bg=colour4)
            t.screen.bgcolor(colour4)

            # Set the coordinate system of the turtle canvas
            t.screen.setworldcoordinates(-150, -150, 150, 150)

            # Move the turtle to the bottom center of the canvas
            t.penup()
            t.goto(-130, -90)
            colors=[colour1,colour2,colour3]
            t.pendown()
            text=restName
            # Draw the text
            # Loop over each character in the text
            for i in range(len(text)):
                # Change the color of the turtle for every three characters
                color = colors[(i // 3) % 3]
                t.color(color)

                # Write the character and move the turtle forward
                t.write(text[i], font=font)
                t.penup()
                t.forward(15)
                t.pendown()

            # Hide the turtle and start the Tkinter main loop
            t.hideturtle()
            # Define global variables
            global counter, font_size, passo, openchef
            openchef=0
            passo = ['']
            counter = 1

            # Connect to the local database and execute a select statement
            conn = sqlite3.connect(resource_path('Data\Local.db'))
            cur = conn.cursor()
            cur.execute('SELECT * FROM local')
            get = cur.fetchall()

            # Get values from the local database and assign them to global variables
            for gib in get:
                Aktive2 = gib[0]
                Einstellung2 = gib[1]
                Kundendaten2 = gib[2]
                speisekarte12 = gib[3]
                ZutatenListe2 = gib[4]
                ZutatenPreise2 = gib[5]
                OrderData2 = gib[6]
                DataAnalysis=gib[7]
            ################################################# Show Passwort ########################################################

            '''''''''''''''''''''CONN EINSTELLUNG'''''''''''''''
            if Aktive2 == 'AKTIVE' and Einstellung2 != 'lokal':
                try:
                    connE = sqlite3.connect(Einstellung2)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu Einstellung {Einstellung2} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connE = sqlite3.connect(resource_path('Data\Einstellung.db'))
            else:
                connE = sqlite3.connect(resource_path('Data\Einstellung.db'))
            '''''''''''''''''''''CONN EINSTELLUNG'''''''''''''''

            '''''''''''''''''''''CONN KUNDENDATEN'''''''''''''''

            if Aktive2 == 'AKTIVE' and Kundendaten2 != 'lokal':
                try:
                    connK = sqlite3.connect(Kundendaten2)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {Kundendaten2} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connK = sqlite3.connect(resource_path('Data\Kundendaten.db'))
            else:
                connK = sqlite3.connect(resource_path('Data\Kundendaten.db'))
            '''''''''''''''''''''CONN KUNDENDATEN'''''''''''''''

            '''''''''''''''''''''CONN speisekarte1'''''''''''''''

            if Aktive2 == 'AKTIVE' and Kundendaten2 != 'lokal':
                try:
                    connS = sqlite3.connect(speisekarte12)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {speisekarte12} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connS = sqlite3.connect(resource_path('Data\speisekarte1.db'))
            else:
                connS = sqlite3.connect(resource_path('Data\speisekarte1.db'))

            '''''''''''''''''''''CONN speisekarte1'''''''''''''''

            '''''''''''''''''''''CONN ZutatenListe'''''''''''''''

            if Aktive2 == 'AKTIVE' and ZutatenListe2 != 'lokal':
                try:
                    connZ = sqlite3.connect(ZutatenListe2)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {ZutatenListe2} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connZ = sqlite3.connect(resource_path('Data\ZutatenListe.db'))
            else:
                connZ = sqlite3.connect(resource_path('Data\ZutatenListe.db'))

            '''''''''''''''''''''CONN ZutatenListe'''''''''''''''

            '''''''''''''''''''''CONN ZutatenPreise'''''''''''''''

            if Aktive2 == 'AKTIVE' and ZutatenPreise2 != 'lokal':
                try:
                    connZu = sqlite3.connect(ZutatenPreise2)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {ZutatenPreise2} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connZu = sqlite3.connect(resource_path('Data\ZutatenPreise.db'))
            else:
                connZu = sqlite3.connect(resource_path('Data\ZutatenPreise.db'))

            '''''''''''''''''''''CONN ZutatenPreise'''''''''''''''

            '''''''''''''''''''''CONN ZutatenPreise'''''''''''''''

            if Aktive2 == 'AKTIVE' and ZutatenPreise2 != 'lokal':
                try:
                    connZu = sqlite3.connect(ZutatenPreise2)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {ZutatenPreise2} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connZu = sqlite3.connect(resource_path('Data\ZutatenPreise.db'))
            else:
                connZu = sqlite3.connect(resource_path('Data\ZutatenPreise.db'))

            '''''''''''''''''''''CONN ZutatenPreise'''''''''''''''

            '''''''''''''''''''''CONN OrderData'''''''''''''''

            if Aktive2 == 'AKTIVE' and OrderData2 != 'lokal':
                try:
                    connO = sqlite3.connect(OrderData2)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {OrderData2} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connO = sqlite3.connect(resource_path('Data\OrderData.db'))
            else:
                connO = sqlite3.connect(resource_path('Data\OrderData.db'))

            '''''''''''''''''''''CONN OrderData'''''''''''''''

            '''''''''''''''''''''CONN DataAnalysis'''''''''''''''

            if Aktive2 == 'AKTIVE' and DataAnalysis != 'lokal':
                try:
                    connD = sqlite3.connect(DataAnalysis)
                except:
                    messagebox.showwarning(f'ERROR',
                                           f'der pfad zu KUNDENDATEN {DataAnalysis} IST NICHT RICHTIG ODER NICHT ERREICHBAR')
                    connD = sqlite3.connect(resource_path('Data\DataAnalysis.db'))
            else:
                connD = sqlite3.connect(resource_path('Data\DataAnalysis.db'))

            '''''''''''''''''''''CONN DataAnalysis'''''''''''''''

            cursor = connE.cursor()

            # Check if the table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='fast'")
            result = cursor.fetchone()

            if result:
                # Table exists, do nothing
                print("Table already exists")
            else:
                # Table does not exist, create it
                cursor.execute("CREATE TABLE fast (id INTEGER PRIMARY KEY,name TEXT)")

            # Commit changes and close the connection
            connE.commit()



            cur = connE.cursor()
            cur.execute('select* from Passwort ')
            connE = sqlite3.connect(resource_path('Data\Einstellung.db'))
            Passwort = cur.fetchall()
            Passwort1 = str(Passwort)
            Passwort2 = Passwort1.replace("(", "").replace(")", "").replace(",", "").replace("'", "").replace("[", "").replace(
                "]",
                "")
            connE.commit()
            ''''''''''''''''''''''''"dayesleft"''''''''''''''''''''''''''''''


            def makevisible():
                # Function to display the current password (Passwort2)
                Passwort23 = Label(root, text=MessageBox.showinfo('Aktueles Passwort', Passwort2), font=('Helvetica bold', 12),
                                   width=5)
                Passwort23.pack()
                return


            # Create button to call makevisible function
            btnpass = Button(root, font=('Helvetica bold', 10), bg='black', command=makevisible, bd=0)
            btnpass.place(x=0, y=1)


            # Define login function
            def login():
                style = ttk.Style(root)
                global username_login_entry, password_login_entry, login_screen
                login_screen = Toplevel(root)
                login_screen.title("Login")
                login_screen.geometry("300x250")
                login_screen.config(bg=colour4)
                window_height = 250
                window_width = 300
                screen_width = login_screen.winfo_screenwidth()
                screen_height = login_screen.winfo_screenheight()
                x_cordinate = int((screen_width / 2) - (window_width / 2))
                y_cordinate = int((screen_height / 2) - (window_height / 2))
                login_screen.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                Label(login_screen, text="Name und Passwort eingeben", bg=colour4, font=('Helvetica bold', 14)).pack()
                Label(login_screen, text="", bg=colour4).pack()
                global username_verify, password_verify
                username_verify = StringVar()
                password_verify = StringVar()
                Label(login_screen, text="Username  ", bg=colour4, font=('Helvetica bold', 12)).pack()
                username_login_entry = Entry(login_screen, textvariable=username_verify)
                username_login_entry.pack()
                username_login_entry.focus_force()


                Label(login_screen, text="", bg=colour4).pack()

                Label(login_screen, text="Password * ", bg=colour4, font=('Helvetica bold', 12)).pack()
                password_login_entry = Entry(login_screen, textvariable=password_verify, show='*')
                password_login_entry.pack()

                Label(login_screen, text="", bg=colour4).pack()
                Button(login_screen, text="Login", width=10, height=1, command=login_verify, bg='#4AB19D',
                       font=('Helvetica bold', 13)).pack()
                login_screen.bind('<Return>', login_verify)


            def login_verify(e=NONE):
                global username_login_entry, password_login_entry, login_screen, blob, passo
                username1 = username_verify.get()
                password1 = password_verify.get()
                username_login_entry.delete(0, END)
                password_login_entry.delete(0, END)

                cus = connE.cursor()
                cus.execute('select*  from Username where Name =(?)', (username1,))
                login_screen.destroy()
                blob = cus.fetchall()
                if blob:
                    for passo in blob:
                        if passo[2] != password1:
                            messagebox.showerror('Passwort', 'falsches passwort')
                            login_screen.destroy()
                        elif username1 != passo[1]:
                            messagebox.showerror('username', 'username exestiert nicht ')
                            login_screen.destroy()
                        else:
                            login_screen.destroy()
                            bestellung()

                # login_screen.bind('<Return>', login_verify)


            def bestellung():
                global counter,new_window2, print_Button,Bestellung_frame

                if counter < 3:
                    new_window2 = Toplevel(root)
                    root.iconify()
                    if show < 3200:
                        new_window2.state('zoomed')
                    new_window2.state(newstate=None)
                    new_window2.config(bd=2)
                    global font_size
                    font_size = ("ARIEL", 12, "bold")
                    new_window2.focus_set()
                    counter += 2
                    titlespace = " "
                    import sqlite3
                    import tkinter as tk

                    # Create a connection to the database




                    # Create a menu object
                    menu_bar = tk.Menu(root)

                    # Add a "File" menu
                    file_menu = tk.Menu(menu_bar, tearoff=0)

                    file_menu.add_separator()
                    file_menu.add_command(label="Exit", command=root.quit)
                    menu_bar.add_cascade(label="File", menu=file_menu)

                    # Add a "Options" menu
                    options_menu = tk.Menu(menu_bar, tearoff=0)


                    menu_bar.add_cascade(label="Options", menu=options_menu)

                    # Add the menu bar to the root window
                    new_window2.config(menu=menu_bar)
                    vardo = tk.IntVar()
                    vardo1 = tk.IntVar()
                    fast_colour=vardo1.get()

                    # Add a checkbutton to the menu
                    options_menu.add_checkbutton(label='Schnell Drucken', variable=vardo)
                    options_menu.add_checkbutton(label='90er style', variable=vardo1)

                    # Get the value of the checkbutton


                    new_window2.geometry("1920x1080")
                    new_window2.config(bg=colour4)
                    new_window2.resizable(width=False, height=False)
                    def on_closing():
                        if messagebox.askokcancel("Scliessen", "Sind Sie sicher?"):
                            global counter
                            counter -= 2
                            new_window2.destroy()

                    new_window2.protocol("WM_DELETE_WINDOW", on_closing)

                    Mainframe1 = Frame(new_window2, bd=4, width=300, height=900, relief=RIDGE, bg=colour4)
                    Mainframe1.grid()
                    Mainframe32 = Frame(new_window2, bd=4, width=290, height=100, relief=RIDGE, bg=colour4)
                    Mainframe32.place(x=5, y=400)
                    Mainframe2 = Frame(new_window2, bd=4, width=1250, height=350, relief=RIDGE, bg=colour4)
                    Mainframe2.grid(row=0, column=2, sticky='n')
                    Mainframe3 = Frame(new_window2, bd=4, width=1400, height=1550, bg=colour4)
                    Mainframe4 = Frame(Mainframe2, bd=5, height=85, width=190, relief=RIDGE, bg=colour4)
                    Mainframe4.place(x=252, y=115)
                    Mainframe5 = Frame(Mainframe2, bd=5, height=85, width=190, relief=RIDGE, bg=colour4)
                    Mainframe5.place(x=252, y=235)
                    Mainframe6 = Frame(Mainframe2, bd=5, height=55, width=1160, relief=RIDGE, bg=colour4)
                    Mainframe6.place(x=3, y=5)
                    Mainframe7 = Frame(Mainframe2, bd=5, height=220, width=200, relief=RIDGE, bg=colour4)
                    Mainframe7.place(x=960, y=75)
                    Mainframe10 = Frame(Mainframe2, bd=5, height=65, width=176, relief=RIDGE, bg=colour4)
                    Mainframe10.place(x=640, y=240)
                    Mainframe9 = Frame(Mainframe1, bd=5, height=220, width=390, relief=RIDGE, bg=colour4)
                    Mainframe9.place(x=0, y=500)
                    Mainframe11 = Frame(Mainframe2, bd=5, height=40, width=176, relief=RIDGE, bg=colour4)
                    Mainframe11.place(x=640, y=298)
                    abholframe = Frame(Mainframe3, height=40, width=210, bg=colour4, bd=4, relief=RIDGE)

                    # ------------------------------------------- seite schliessen-------------------------------------------------#
                    def close():
                        global counter
                        siko = messagebox.askyesno('Schliessen', 'Sind sie sicher')
                        if siko == 1:
                            new_window2.destroy()
                            counter -= 2
                        else:
                            pass

                    # closebutton = Button(new_window2, text='X', bd=2, bg='red', command=close)
                    # closebutton.place(x=1480, y=5)
                    ######################################speisen Labels and boxes##########################################################
                    global kl_gr
                    cur = connS.cursor()
                    kl_gr = ['Klein', 'Gross', 'Standard']
                    global clicked
                    clicked = StringVar()
                    Numbers = StringVar()
                    global drop
                    cur = connE.cursor()
                    cur.execute('select* from Gross ')
                    grosse_price = 0
                    grosse_name = ''
                    grosse = cur.fetchall()
                    grosse = cur.fetchall()
                    if len(grosse) >= 2:
                        grosse_price = grosse[1]
                        grosse_name = grosse[0]
                    else:
                        pass

                    def gross_pizza(event=None):
                        global downname
                        gorss = drop.get()
                        # function to handle gross pizza selection from the option menu
                        listbox1.selection_clear(ANCHOR)
                        zuti = []
                        zutatens = listbox2.get(0, END)
                        for zutat in zutatens:
                            if zutat[0] == '+':
                                zuti.append(zutat)
                        if len(zuti) > 0:
                            name = entrybox1.get()
                            price = ''
                            gorss = drop.get()

                            if gorss == 'Klein':
                                drop.set(gorss)
                                drop.config(state=DISABLED)
                                entrybox6.delete(0, END)
                                kat = entrybox5.get()
                                curs = connS.cursor()
                                curs.execute('SELECT * FROM Speisen WHERE Name=(?) ', (name,))
                                besonder = curs.fetchall()

                                # check if there is a price for the selected pizza

                                for beson in besonder:
                                    result=float(beson[3])
                                entrybox6.delete(0, END)

                                formatted_result = '{:.2f}'.format(result)
                                entrybox6.insert(0, formatted_result)
                                # clear the items with a plus sign from the listbox
                                cards = listbox2.get(0, END)
                                to_delete = []
                                for i, card in enumerate(cards):
                                    if card[0] == '+':
                                        to_delete.append(i)
                                for i in reversed(to_delete):
                                    listbox2.delete(i)
                                #enter the items from zuti in entrybox2
                                for i in range(len(zuti)):
                                    entrybox2.focus_force()
                                    entrybox2.insert(0, zuti[i][1:])

                                    addzutat()
                            elif gorss=='Gross':
                                drop.set(gorss)
                                drop.config(state=DISABLED)

                                drop.set(gorss)
                                kat = entrybox5.get()
                                curs = connS.cursor()
                                curs.execute('SELECT * FROM Speisen WHERE Name=(?) ', (name,))
                                besonder = curs.fetchall()
                                for info in besonder:
                                    price = info[5]
                                if price is None:
                                    # show warning message if there is no price for the selected pizza
                                    messagebox.showwarning('Gross zuschlag',
                                                           'Für diese Speise ist kein Zuschlag eingegeben. Bitte anderen auswählen oder Preis manuell eingeben.')
                                    return
                                # check if there is a price for the selected pizza

                                for beson in besonder:
                                    sacko=float(beson[3])
                                entrybox6.delete(0, END)
                                result = float(sacko) + float(price)
                                formatted_result = '{:.2f}'.format(result)
                                entrybox6.insert(0, formatted_result)
                                # clear the items with a plus sign from the listbox
                                cards = listbox2.get(0, END)
                                to_delete = []
                                for i, card in enumerate(cards):
                                    if card[0] == '+':
                                        to_delete.append(i)
                                for i in reversed(to_delete):
                                    listbox2.delete(i)
                                # enter the items from zuti in entrybox2
                                for i in range(len(zuti)):
                                    entrybox2.focus_force()
                                    entrybox2.insert(0, zuti[i][1:])

                                    addzutat()


                        else:
                            name = entrybox1.get()
                            price = ''
                            gorss = drop.get()




                            if gorss == 'Klein':

                                entrybox4.focus()
                                entrybox4.selection_range(0, END)
                                # drop.set(gorss)

                            else:
                                drop.set(gorss)
                                kat = entrybox5.get()
                                curs = connS.cursor()
                                curs.execute('SELECT * FROM Speisen WHERE Name=(?) ', (name,))
                                besonder = curs.fetchall()
                                for info in besonder:
                                    price = info[5]
                                if price is None:
                                    # show warning message if there is no price for the selected pizza
                                    messagebox.showwarning('Gross zuschlag',
                                                           'Für diese Speise ist kein Zuschlag eingegeben. Bitte anderen auswählen oder Preis manuell eingeben.')
                                    return
                                # check if there is a price for the selected pizza

                                sacko = float(entrybox6.get())
                                entrybox6.delete(0, END)
                                result = float(sacko) + float(price)
                                formatted_result = '{:.2f}'.format(result)
                                entrybox6.insert(0, formatted_result)
                                
                                entrybox4.focus()
                                entrybox4.select_range(0, END)
                            entrybox1.delete(0, END)
                            entrybox1.insert(0, name)
                    drop =ttk.Combobox (Mainframe2, values=kl_gr, font=('Helvetica bold', 16), width=8)
                    drop.place(x=255, y=70)
                    drop.bind("<<ComboboxSelected>>", gross_pizza)

                    global liso
                    liso = []

                    #############################################################Nummer einfugen#######################################
                    def resetall():
                        global abholframe
                        drop.config(state=NORMAL)
                        entrybox1.focus_set()
                        entrybox1.config(state=NORMAL)


                        zutatennummer_E.delete(0, END)
                        entrybox4.delete(0, END)
                        entrybox4.insert(0, '1')
                        entrybox2.delete(0, END)
                        entrybox1.delete(0, END)
                        listbox3.delete(0, END)
                        listbox2.delete(0, END)
                        entrybox3.delete(0, END)
                        entrybox5.delete(0, END)
                        entrybox6.delete(0, END)


                        LieferE.delete(0, END)
                        drop.set('')
                        entrybox12.delete(0, END)
                        drop.configure(font=('Helvetica bold', 16),   width=8)
                        fort = c1a.state()
                        fift = c2a.state()
                        sixt = c3a.state()
                        secent = c4a.state()
                        if 'selected' in fort:
                            c1a.invoke()
                        if 'selected' in fift:
                            c2a.invoke()
                        if 'selected' in sixt:
                            c3a.invoke()
                        if 'selected' in secent:
                            c4a.invoke()
                        liso.clear()

                    global Reset

                    Reset = Button(Mainframe2, text="Reset", bg=colour1, bd=2, command=resetall, height=3, width=7)
                    Reset.place(x=880, y=70)

                    def update():
                        Uhr.config(text="new text")

                    def uhr():
                        stunde = time.strftime("%H")
                        minute = time.strftime("%M")
                        seconde = time.strftime("%S")
                        utc = ' Uhr'
                        tag = time.strftime("%d")
                        monat = time.strftime("%m")
                        jahr = time.strftime("%y")
                        name = time.strftime("%a")
                        Uhr.config(text=stunde + ':' + minute + ':' + seconde + utc)
                        Datum.config(text=name + '.' + tag + '/' + monat + '/' + jahr)
                        Uhr.after(100, uhr)

                    ##########################################Bestellung /Labels and Buttons###############################################
                    Label1 = Label(Mainframe6, text='Speise ', font=("Helvetica", 18, 'bold'), fg="black", bg=colour4,
                                   bd=8)
                    Label1.place(x=10, y=0)
                    # Label2 = Label(Mainframe6, text='Kl/Gr ', font=("Helvetica", 18, 'bold'), fg="black", bg="dark slate gray", bd=8)
                    # Label2.place(x=270, y=0)
                    Label3 = Label(Mainframe6, text='Mit Zutaten ', font=("Helvetica", 18, 'bold'), fg="black",
                                   bg=colour4, bd=8)
                    Label3.place(x=450, y=0)

                    Label4 = Label(Mainframe6, text='Anzahl ', font=("Helvetica", 18, 'bold'), fg="black", bg=colour4,
                                   bd=8)
                    Label4.place(x=325, y=0)
                    Label6 = Label(Mainframe6, text='Nr: ', font=("Helvetica", 18, 'bold'), fg="black", bg=colour4,
                                   bd=8)
                    Label6.place(x=185, y=0)
                    Label7 = Label(Mainframe4, text='Kategorie:', font=("Helvetica", 12, 'bold'), fg='black',
                                   bg=colour4,
                                   bd=4)
                    Label7.place(x=0, y=0)
                    Label8 = Label(Mainframe4, text='Preis:', font=("Helvetica", 12, 'bold'), fg='black', bg=colour4,
                                   bd=4)
                    Label8.place(x=0, y=40)
                    Label9 = Label(Mainframe4, text='€', font=("Helvetica", 12, 'bold'), fg='black', bg=colour4,
                                   bd=2)
                    Label9.place(x=150, y=40)
                    Label10 = Label(Mainframe6, text='Gesamt Preise', font=("Helvetica", 18, 'bold'), fg='black',
                                    bg=colour4, bd=8)
                    Label10.place(x=950, y=0)
                    Label11 = Label(Mainframe7, text='Brutto:                           '
                                                     '€', font=("Helvetica", 12, 'bold'), fg='black', bg=colour4,
                                    bd=2)
                    Label11.place(x=5, y=5)
                    Label12 = Label(Mainframe7, text='Liefergeld:                      '
                                                     '€', font=("Helvetica", 12, 'bold'), fg='black',
                                    bg=colour4,
                                    bd=2)
                    Label12.place(x=5, y=45)
                    Label13 = Label(Mainframe7, text='Gutschrift:                    '
                                                     '€', font=("Helvetica", 12, 'bold'), fg='black',
                                    bg=colour4,
                                    bd=2)
                    Label13.place(x=5, y=85)
                    Label14 = Label(Mainframe7, text='Rabatt:                         '
                                                     '% ', font=("Helvetica", 12, 'bold'), fg='black',
                                    bg=colour4,
                                    bd=2)
                    Label14.place(x=5, y=125)
                    Label15 = Label(Mainframe7, text='Endpreis:                      '
                                                     '€', font=("Helvetica", 12, 'bold'), fg='black',
                                    bg=colour4,
                                    bd=2)
                    Label15.place(x=5, y=165)
                    Label16 = Label(Mainframe6, text='Ohne ', font=("Helvetica", 18, 'bold'), fg="black",
                                    bg=colour4, bd=8)
                    Label16.place(x=680, y=0)
                    Label17 = Label(Mainframe9, text='Bediener: ', font=("Helvetica", 12, 'bold'), fg="black",
                                    bg=colour4, bd=8)
                    Label17.place(x=0, y=130)
                    Label18 = Label(Mainframe9, text='ID nummer: ', font=("Helvetica", 12, 'bold'), fg="black",
                                    bg=colour4, bd=8)
                    Label18.place(x=0, y=162)
                    if passo != ['']:
                        Label19 = Label(Mainframe9, text=passo[0], font=("Helvetica", 12, 'bold'), fg=colour4,
                                        bg="Black", bd=0)
                        Label19.place(x=100, y=170)
                        Label20 = Label(Mainframe9, text=passo[1], font=("Helvetica", 12, 'bold'), fg=colour4,
                                        bg="black", bd=0)
                        Label20.place(x=90, y=136)
                    Label21 = Label(Mainframe2, text='Kommentar ', font=("Helvetica", 14, 'bold'), fg="black",
                                    bg=colour4, bd=8)
                    Label21.place(x=459, y=260)
                    # -----------------------------------------------------------------------------------------------------------------#
                    entrybox1 = Entry(Mainframe2, font=("Helvetica", 12, 'bold'), width=22, bd=5,bg=colour0)
                    entrybox1.place(x=10, y=70)
                    entrybox3 = Entry(Mainframe2, font=("Helvetica", 12, 'bold'), width=4, bd=5,bg=colour0)
                    entrybox3.place(x=200, y=70)
                    entrybox3.configure(font=font_size)
                    entrybox4 = Entry(Mainframe2, font=("Helvetica", 12, 'bold'), width=4, bd=5, textvariable=Numbers,bg=colour0)
                    entrybox4.place(x=385, y=70)
                    entrybox4.configure(font=font_size)
                    entrybox4.insert(0, "1")
                    entrybox5 = Entry(Mainframe4, font=("Helvetica", 12, 'bold'), width=9, bd=3, state=NORMAL,bg=colour0)
                    entrybox5.place(x=85, y=5)
                    entrybox6 = Entry(Mainframe4, font=("Helvetica", 12, 'bold'), width=6, bd=3 ,bg=colour0)
                    entrybox6.place(x=85, y=40)
                    entrybox7 = Entry(Mainframe7, font=("Helvetica", 12, 'bold'), width=6, bd=3,bg=colour0 )
                    entrybox7.place(x=95, y=5)
                    entrybox7.insert(0, float(0.0))
                    entrybox8 = Entry(Mainframe7, font=("Helvetica", 12, 'bold'), width=6, bd=3 ,bg=colour0)
                    entrybox8.place(x=95, y=45)
                    entrybox8.insert(0, float(0.0))
                    entrybox9 = Entry(Mainframe7, font=("Helvetica", 12, 'bold'), width=6, bd=3,bg=colour0 )
                    entrybox9.place(x=95, y=85)
                    entrybox9.insert(0, float(0.0))
                    entrybox10 = Entry(Mainframe7, font=("Helvetica", 12, 'bold'), width=6, bd=3 ,bg=colour0)
                    entrybox10.place(x=95, y=125)
                    entrybox10.insert(0, float(0.0))
                    entrybox11 = Entry(Mainframe7, font=("Helvetica", 12, 'bold'), width=6, bd=3 ,bg=colour0)
                    entrybox11.place(x=95, y=165)
                    entrybox11.insert(0, float(0.0))
                    entrybox12 = Entry(Mainframe2, font=("Helvetica", 12, 'bold'), width=18, bd=5 ,bg=colour0)
                    entrybox12.place(x=450, y=297)
                    global abholE
                    abholE = Entry(abholframe, width=5, bd=4)


                    # ----------------------------------------------------------------------------------------------------------------#
                    listbox1 = Listbox(Mainframe2, width=25, bd=5,bg=colour0)
                    listbox1.place(x=10, y=115)
                    listbox1.configure(font=font_size)
                    normal_height=7
                    listbox2 = Listbox(Mainframe2, width=18, bd=5, height=normal_height,bg=colour0)
                    listbox2.place(x=450, y=120)
                    listbox2.configure(font=("Helvetica", 11, 'bold'),bg=colour0)
                    listbox3 = Listbox(Mainframe2, width=18, bd=5, fg='red', height=8,bg=colour0)
                    listbox3.place(x=640, y=70)
                    listbox3.configure(font=("Helvetica", 10, 'bold'),bg=colour0)

                    def on_entry_focus_in(event):
                        # adjust the height of the listbox when the entry widget gets focus
                        listbox2.config(height=normal_height + 5)

                    def on_entry_focus_out(event):
                        # reset the height of the listbox when the entry widget loses focus
                        listbox2.config(height=normal_height)


                    # -----------------------------------------------------------------------------------------------------------------#
                    ########################################################speisen liste###################################################
                    # ------------------------------------------------------Alle Straßen----------------------------------------------------#
                    with open(resource_path('Data\streets.txt'), 'r', encoding='utf-8') as street:
                        Straßen = []
                        for stre in street:
                            if not stre.isspace():
                                streline = stre.strip()
                                Straßen.append(str(streline))
                    # -------------------------------------------------------Enterys--------------------------------------------------------#
                    KundenidE = Entry(Mainframe1, width=6, bd=5, bg=colour0)
                    KundenidE.place(x=0, y=28)
                    KundenidE.configure(font=font_size)
                    NameE = Entry(Mainframe1, bd=5, width=25, bg=colour0)
                    NameE.place(x=0, y=94)
                    NameE.configure(font=font_size)
                    AdresseE = AutocompleteEntry(Mainframe1, width=23, completevalues=Straßen, font=("Helvetica", 12, 'bold'))
                    AdresseE.place(x=0, y=162)
                    HauesnrE = Entry(Mainframe1, bd=1, width=4, bg=colour0)
                    HauesnrE.place(x=220, y=162)
                    HauesnrE.configure(font=font_size)
                    PLZE = Entry(Mainframe1, bd=5, width=7, bg=colour0)
                    PLZE.place(x=0, y=230)
                    PLZE.configure(font=font_size)
                    ORTE = Entry(Mainframe1, bd=5, width=15, bg=colour0)
                    ORTE.place(x=140, y=230)
                    ORTE.configure(font=font_size)
                    TelefonnummerE = Entry(Mainframe1, bd=5, width=16, bg=colour0)
                    TelefonnummerE.place(x=120, y=28)
                    TelefonnummerE.configure(font=font_size)
                    EmailE = Entry(Mainframe1, bd=5, width=25, bg=colour0)
                    EmailE.place(x=0, y=300)
                    EmailE.configure(font=font_size)
                    LieferE = Entry(Mainframe11, bd=5, width=5, bg=colour0)
                    LieferE.place(x=107, y=0)
                    LieferE.configure(font=font_size)
                    in_comment = Text(Mainframe9, width=30, height=1, bd=5)
                    in_comment.place(x=0, y=30)
                    in_comment.configure(font=('14'))
                    ext_comment = Text(Mainframe9, width=30, height=1, bd=5)
                    ext_comment.place(x=0, y=89)
                    ext_comment.configure(font=('14'))
                    Note_text = Text(Mainframe2, width=14, height=3, bd=5,bg=colour0)
                    Note_text.place(x=258, y=246)
                    Note_text.configure(font=('14'), fg='red')
                    edy = 'Note'
                    with open(resource_path('Data/Note.txt'), 'r') as file14:
                        for e in file14:
                            if e:
                                edy = e.strip()
                    note_label = Label(Note_text, text=edy, width=16, height=3, bg=colour0, fg='black', wraplengt=120,
                                       font=("Helvetica", 12, 'bold'))
                    note_label.pack(side=TOP)

                    def update_orte(event=None):
                        # Get the PLZ value from the PLZE Entry
                        plz = PLZE.get()

                        # Remove any non-digit characters from the PLZ value using regular expressions
                        plz = re.sub(r'\D', '', plz)

                        # Get the corresponding Ort value from the dictionary
                        ort = plz_ort_dict.get(plz, '')

                        # Convert the Ort value to a string if it is not already a string
                        if not isinstance(ort, str):
                            ort = str(ort)
                        # Set the value of the ORTE Entry to the Ort value
                        ORTE.delete(0, tk.END)
                        ORTE.insert(0, ort)

                    PLZE.bind('<Return>', update_orte)

                    ##################################Kunden Liste / Labels#############################################################
                    Kundenid = Label(Mainframe1, bd=0, text='ID', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=0, y=0)
                    Name = Label(Mainframe1, bd=0, text='Name:', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=0, y=65)
                    Adresse = Label(Mainframe1, bd=0, text='Straße:', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=0, y=130)
                    Hausnr = Label(Mainframe1, bd=0, text='Nr:', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=230, y=130)
                    PLZ = Label(Mainframe1, bd=0, text='PLZ:', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=0, y=200)
                    Ort = Label(Mainframe1, bd=0, text='ORT:', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=140, y=200)
                    Telefonnummer = Label(Mainframe1, bd=0, text='Telefonnummer:', font=('arial', 14, 'bold'),
                                          bg=colour4) \
                        .place(x=120, y=0)

                    Email = Label(Mainframe1, bd=0, text='Email:', font=('arial', 14, 'bold'), bg=colour4) \
                        .place(x=0, y=267)
                    comment_in = Label(Mainframe9, bd=0, text='Intern-Info:', font=('arial', 16, 'bold'), bg=colour4) \
                        .place(x=0, y=1)
                    comment_ex = Label(Mainframe9, bd=0, text='Extern-Info:(wird gedruckt)', font=('arial', 16, 'bold'),
                                       bg=colour4) \
                        .place(x=0, y=60)
                    Note = Label(Mainframe2, bd=0, text='Note von Chef', font=('arial', 16, 'bold'),
                                 bg=colour4) \
                        .place(x=270, y=220)
                    Uhr = Label(Mainframe10, bd=0, text='', font=('arial', 12, 'bold'), fg='black',
                                bg=colour4)
                    Uhr.place(x=65, y=0)
                    Datum = Label(Mainframe10, bd=0, text='', font=('arial', 12, 'bold'), fg='black',
                                  bg=colour4)
                    Datum.place(x=65, y=27)

                    ################################### checkboxes
                    so = ttk.Style()
                    so.theme_use('clam')
                    so.configure('chop.TCheckbutton', background=colour4)

                    so.configure('chop.TCheckbutton', font=('Helvetica', 14))
                    var = StringVar()
                    var1 = StringVar()
                    var2 = StringVar()
                    var3 = StringVar()
                    c1a = ttk.Checkbutton(Mainframe2, text='Schneiden', variable=var, cursor="cross", style='chop.TCheckbutton')
                    c1a.place(x=820, y=160)
                    c2a = ttk.Checkbutton(Mainframe2, text='Knusbrig', variable=var1, cursor="cross", style='chop.TCheckbutton')
                    c2a.place(x=820, y=200)
                    c3a = ttk.Checkbutton(Mainframe2, text='Hell Backen', variable=var2, cursor="cross",
                                          style='chop.TCheckbutton')
                    c3a.place(x=820, y=240)
                    c4a = ttk.Checkbutton(Mainframe2, text='Wenig Käse', variable=var3, cursor="cross",
                                          style='chop.TCheckbutton')
                    c4a.place(x=820, y=280)
                    TelefonnummerE.focus_force()
                    # --------------------------------------------------------UHR-----------------------------------------------------------#
                    imagen = Image.open(resource_path('Bilder\clockpng.jpg'))
                    imagser = imagen.resize((45, 45))
                    imagser.save(resource_path('Bilder\clock.jpg'))
                    clock = ImageTk.PhotoImage(Image.open(resource_path('Bilder\clock.jpg')))

                    def opencalen():
                        os.system(
                            'start explorer shell:appsfolder\microsoft.windowscommunicationsapps_8wekyb3d8bbwe!microsoft.windowslive.calendar')

                    labelclock = Button(Mainframe10, image=clock, bg='dark slate gray', command=opencalen)
                    labelclock.place(x=5, y=1)
                    Bestell = Label(Mainframe11, bd=0, text='Lieferzeit:', font=('arial', 16, 'bold'), bg=colour4) \
                        .place(x=0, y=0)
                    uhr()

                    # -------------------------------------------------------Karte----------------------------------------------------------#
                    def maps():
                        hob = str(AdresseE.get())
                        nop = str(HauesnrE.get())
                        zob = str(hob + ' ' + nop)
                        mapes = Toplevel(root)
                        root.wm_state('iconic')
                        mapes.state(newstate=None)
                        mapes.config(bd=2)
                        mapes.focus_set()
                        map = tkintermapview.TkinterMapView(mapes, width=400, height=250)
                        map.set_zoom(15)
                        map.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
                        window_height = 200
                        window_width = 260
                        screen_width = mapes.winfo_screenwidth()
                        screen_height = mapes.winfo_screenheight()
                        x_cordinate = int((screen_width / 2) - (window_width / 2))
                        y_cordinate = int((screen_height / 2) - (window_height / 2))
                        ad2 = map.set_address(zob + ',' + PLZE.get() + ',' + ORTE.get(), marker=True)
                        map.pack()

                    but = Button(Mainframe1, bg='#c2b59f', width=6, text='Karte', command=maps, ).place(x=80, y=130)
                    ###############################SQL TO TEXT'########################################################################
                    # con = sqlite3.connect('speisekarte1.db', isolation_level=None)
                    custor = connS.cursor()
                    custor.fetchall()
                    custor.execute('select Name from Speisen')
                    file = open(resource_path('Data\speisentabel.txt'), 'w')
                    for sack in custor:
                        sacko = "'".join(sack)
                        file.write(str(sacko) + '\n')
                    file.close()
                    with open(resource_path('Data\speisentabel.txt'), 'r') as file1:
                        toppings = []
                        for line in file1:
                            if line != '':
                                stripped_line = line.strip()
                                toppings.append(stripped_line)

                        ########################################################################################################################

                        cur = connZu.cursor()
                        cur.execute("select Zutat from zutatenpreise ")
                        maro = cur.fetchall()
                        dressing_zutaten = []
                        for izo in maro:
                            zutat = izo[0]
                            if zutat.endswith('Dressing'):
                                dressing_zutaten.append(zutat)


                        with open(resource_path('Data\zutatenpreise.txt'), 'w') as file2:
                            for zuz in maro:
                                ziko = ",".join(zuz)
                                file2.write(str(ziko) + '\n')
                        zutatenpreise = []

                        with open(resource_path('Data\zutatenpreise.txt'), 'r') as file3:
                            for zaro in file3:
                                zutatenpreise.append(str(zaro) + '\n')
                        connZu.commit()
                        #######################################################################################################################
                        # had to be moved here becouse of the AutocompleteEntry
                        global entrybox2
                        entrybox2 = AutocompleteEntry(Mainframe2, font=("Helvetica", 12, 'bold'), width=15,
                                                      completevalues=zutatenpreise,background='#FFFFE0')
                        entrybox2.place(x=450, y=70)
                        entrybox2.configure(font=font_size,background='#FFFFE0')
                        entrybox2.bind('<FocusIn>', on_entry_focus_in)
                        entrybox2.bind('<FocusOut>', on_entry_focus_out)

                        def combo_zutat(event):
                            # Your code here
                            selct = combozutat.get()
                            entrybox2.delete(0, END)
                            entrybox2.insert(0, selct)
                            entrybox2.focus_set()
                            addzutat()
                            combozutat.set('')


                        combozutat = ttk.Combobox(Mainframe2, values=dressing_zutaten, font=("Helvetica",9, 'bold'))
                        combozutat.place(x=450, y=97)
                        combozutat.bind("<<ComboboxSelected>>", combo_zutat)

                        def zutaten_nummer(event=None):
                            nummber = zutatennummer_E.get()

                            cur = connZu.cursor()
                            cur.execute("select Zutat from zutatenpreise where Nummer=(?) ", (nummber,))
                            maro = cur.fetchall()
                            entrybox2.insert(0, maro)

                            entrybox2.focus_set()
                            addzutat()
                            zutatennummer_E.delete(0, END)

                        zutatennummer_E = Entry(Mainframe2, font=("Helvetica", 12, 'bold'), width=2)
                        # zutatennummer_E.place(x=600, y=70)
                        zutatennummer_E.bind('<Return>', zutaten_nummer)

                        def komment():
                            siko = entrybox12.get()
                            listbox2.insert(0, '*' + siko)
                            entrybox12.delete(0, END)

                        kommb = Button(Mainframe2, text='ADD', bg=colour2, height=0, command=komment,font=("Helvetica", 8, 'bold'),)
                        kommb.place(x=608, y=270)

                        ################################################Familien Pizza Functions###############################################
                        def Familie_window():
                            global Counter
                            global Familienpizza
                            Familienpizza = Toplevel(root)
                            root.wm_state('iconic')
                            Familienpizza.state(newstate=None)
                            Familienpizza.config(bd=2)
                            global font_size
                            font_size = ("ARIEL", 12, "bold")
                            Familienpizza.focus_set()
                            titlespace = " "
                            window_height = 200
                            window_width = 260
                            screen_width = Familienpizza.winfo_screenwidth()
                            screen_height = Familienpizza.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizza.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizza.config(bg='dark slate gray')
                            Familienpizza.resizable(width=False, height=False)
                            voll = Button(Familienpizza, text='100%', bd=4, bg='white', width=10, command=Familiewindow1)
                            voll.place(x=15, y=20)
                            halb = Button(Familienpizza, text='50/50%', bd=4, bg='white', width=10, command=Familiewindow2)
                            halb.place(x=160, y=20)
                            drei = Button(Familienpizza, text='3x33%', bd=4, bg='white', width=10, command=Familiewindow3)
                            drei.place(x=15, y=100)
                            vier = Button(Familienpizza, text='4x25%', bd=4, bg='white', width=10, command=Familiewindow4)
                            vier.place(x=160, y=100)
                            funf = Button(Familienpizza, text='75/25%', bd=4, bg='white', width=10, command=Familiewindow5)
                            funf.place(x=90, y=60)
                            six = Button(Familienpizza, text='50/2x25%', bd=4, bg='white', width=10, command=Familiewindow6)
                            six.place(x=90, y=140)

                        def delete_selected_item(listbox):
                            # Get the index of the currently selected item
                            selection = listbox.curselection()
                            if selection:
                                index = selection[0]
                                # Delete the item at the selected index
                                listbox.delete(index)

                        ###############################################Familinepizza Voll#######################################################
                        def Familiewindow1():
                            global Familiewindow1
                            Familienpizza.destroy()
                            global Familienpizzavoll
                            Familienpizzavoll = Toplevel(root)
                            root.wm_state('iconic')
                            Familienpizzavoll.state(newstate=None)
                            Familienpizzavoll.config(bd=2)
                            Familienpizzavoll.focus_set()
                            titlespace = " "
                            window_height = 290
                            window_width = 260
                            screen_width = Familienpizzavoll.winfo_screenwidth()
                            screen_height = Familienpizzavoll.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizzavoll.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizzavoll.config(bg='dark slate gray')
                            Familienpizzavoll.resizable(width=False, height=False)

                            # ---------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to listbox
                            def addzutatvoll(event=None):
                                sado = str(entrybox2.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entrybox2.get()[:1])


                                    except ValueError:
                                        input_number = None


                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entrybox2.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()

                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()


                                        for rows in row[3:]:
                                            if rows!='' :
                                                listboxvoll.insert(END,rows)
                                                entrybox2.delete(0, END)
                                else:

                                    sado1 = sado.strip()
                                    listboxvoll.insert(END, sado1)
                                    entrybox2.delete(0, END)

                            # function to search and display ingredients for a specific pizza
                            def nummerfamilie():
                                cub = connZ.cursor()
                                naldo = entrybox3.get()
                                cub.execute("select * from zutaten where SpeiseName  =(?)", (naldo,))
                                foro = cub.fetchall()
                                for taha in foro:
                                    entrybox1.delete(0, END)
                                    entrybox1.insert(END, taha[0])

                            # function to finalize family pizza order
                            def fertig():
                                cur = connE.cursor()
                                cur.execute('select* from einstellung where Name = ?', ('FamilienPizza',))
                                nado = cur.fetchall()
                                Frei = ''
                                Preis = ''
                                for t in nado:
                                    Frei = str(t[1])

                                    Preis = str(t[2])
                                conn.commit()
                                sicko = str(listboxvoll.get(0, END))

                                saro = sicko.strip()
                                selko = saro.replace('(', '').replace(')', '').replace(',', '').replace("'", "").replace(' ',
                                                                                                                         '/')

                                # insert family pizza to order listbox with 100% additional cost
                                listbox2.insert(END, '+100%' + str(selko))
                                sado1 = len(sicko)

                                anzahl_list = []
                                for zahl in listboxvoll.get(0, END):
                                    anzahl_list.insert(0, zahl)
                                naldo = len(anzahl_list) - int(Frei)

                                anzahl = 0
                                if naldo > 0:
                                    anzahl = naldo
                                siko = float(entrybox6.get())
                                entrybox6.delete(0, END)
                                entrybox6.insert(END, siko + float(anzahl * float(Preis)))
                                Familienpizzavoll.destroy()

                                # naldo = len(loko[int(Frei[0]):])
                                # siko = float(entrybox6.get())
                                # entrybox6.delete(0, END)
                                # entrybox6.insert(END, siko + float(naldo * Preis))

                            # create entrybox for adding ingredients
                            entrybox2 = AutocompleteEntry(Familienpizzavoll, font=("Helvetica", 12, 'bold'), width=10,
                                                          completevalues=zutatenpreise)
                            entrybox2.place(x=30, y=20)
                            # bind the focus events of the entry widget to the functions that adjust the listbox height

                            # create button to add ingredient to listbox
                            einfbutton = Button(Familienpizzavoll, text='einfugen', width=6, bg='green', command=addzutatvoll)
                            einfbutton.place(x=190, y=20)

                            # create listbox for displaying added ingredients
                            listboxvoll = Listbox(Familienpizzavoll, width=25, bd=5)
                            listboxvoll.place(x=30, y=60)

                            # create button to finalize family pizza order
                            Fertigbutton = Button(Familienpizzavoll, text='fertig', width=10, bg='grey', command=fertig)
                            Fertigbutton.place(x=180, y=260)
                            entrybox2.bind('<Return>',addzutatvoll)
                            listboxvoll.bind('<Button-1>', lambda event: delete_selected_item(listboxvoll))

                        ###############################################Familinepizza Halb #######################################################
                        def Familiewindow3():
                            global Familiewindow3
                            Familienpizza.destroy()
                            global Familienpizzadrei
                            Familienpizzadrei = Toplevel(root)
                            root.wm_state('iconic')
                            Familienpizzadrei.state(newstate=None)
                            Familienpizzadrei.config(bd=2)
                            Familienpizzadrei.focus_set()
                            titlespace = " "
                            window_height = 550
                            window_width = 400
                            screen_width = Familienpizzadrei.winfo_screenwidth()
                            screen_height = Familienpizzadrei.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizzadrei.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizzadrei.config(bg='dark slate gray')
                            Familienpizzadrei.resizable(width=False, height=False)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to the first listbox in the third page of the order process
                            def addzutatdrei(event=None):
                                sado = str(entryboxdrei.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxdrei.get()[:1])


                                    except ValueError:
                                        input_number = None


                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxdrei.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows!='' :
                                                listboxdrei.insert(END,rows)
                                                entryboxdrei.delete(0, END)
                                else:

                                    sado1 = sado.strip()
                                    listboxdrei.insert(END, sado1)
                                    entryboxdrei.delete(0, END)


                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to the second listbox in the third page of the order process
                            def addzutatdrei1(event=None):
                                sado = str(entryboxdrei1.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxdrei1.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxdrei1.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()

                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()


                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxdrei1.insert(END, rows)
                                                entryboxdrei1.delete(0, END)
                                else:

                                    sado1 = sado.strip()
                                    listboxdrei1.insert(END, sado1)
                                    entryboxdrei1.delete(0, END)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to the third listbox in the third page of the order process
                            def addzutatdrei2(event=None):
                                sado = str(entryboxdrei2.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxdrei2.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxdrei2.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()

                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()


                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxdrei2.insert(END, rows)
                                                entryboxdrei2.delete(0, END)
                                else:

                                    sado1 = sado.strip()
                                    listboxdrei2.insert(END, sado1)
                                    entryboxdrei2.delete(0, END)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to finalize the order of the third page of the order process
                            def fertigdrei():
                                # get the ingredients from the three listboxes and concatenate them
                                sicko = listboxdrei.get(0, END)
                                soko = listboxdrei1.get(0, END)
                                soka = listboxdrei2.get(0, END)
                                saro = str(sicko).replace('(', '').replace(')', '').replace("'", "").replace(
                                    ',', '').replace(' ', '/')
                                saro1 = str(soko).replace('(', '').replace(')', '').replace("'", "").replace(
                                    ',', '').replace(' ', '/')
                                saro2 = str(soka).replace('(', '').replace(')', '').replace("'", "").replace(
                                    ',', '').replace(' ', '/')
                                if not saro:
                                    saro = 'Margherita'
                                if not saro1:
                                    saro1 = 'Margherita'
                                if not saro2:
                                    saro2 = 'Margherita'

                                listbox2.insert(END, f'+33%{saro}')
                                listbox2.insert(END, f'+33%{saro1}')
                                listbox2.insert(END, f'+33%{saro2}')

                                # calculate the price of the final order and display it in the corresponding entrybox
                                sado1 = sicko + soko + soka  # concatenate the contents of the three listboxes
                                soso = []  # initialize an empty list
                                for nano in sado1:  # iterate through the concatenated string
                                    if nano not in soso:  # if the character is not already in the list
                                        soso.append(nano)  # append it to the list
                                loko = []  # initialize another empty list
                                for t in soso:  # iterate through the list of unique characters
                                    loko.append(t)  # append each character to the new list
                                cur = connE.cursor()  # create a cursor object
                                cur.execute('select* from einstellung where Name = ?',
                                            ('FamilienPizza',))  # execute a select statement
                                nado = cur.fetchall()  # fetch all the results
                                Frei = ''  # initialize a variable
                                Preis = ''  # initialize another variable
                                for t in nado:  # iterate through the results
                                    Frei = str(t[1])  # assign the second element to the variable Frei
                                    Preis = t[2]  # assign the third element to the variable Preis
                                naldo = len(
                                    loko[int(Frei[0]):])  # calculate the number of ingredients beyond the first Frei elements
                                siko = float(entrybox6.get())  # get the value from entrybox6 as a float
                                entrybox6.delete(0, END)  # delete the current value in entrybox6
                                entrybox6.insert(END, siko + float(naldo * Preis))  # insert the new value in entrybox6

                                Familienpizzadrei.destroy()  # close the third page of the order process

                            # ------------------------------------------------------------------------------------------------------------------#
                            entryboxdrei = AutocompleteEntry(Familienpizzadrei, font=("Helvetica", 12, 'bold'), width=14,
                                                             completevalues=zutatenpreise)
                            entryboxdrei.place(x=30, y=20)
                            entryboxdrei1 = AutocompleteEntry(Familienpizzadrei, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxdrei1.place(x=220, y=20)
                            entryboxdrei2 = AutocompleteEntry(Familienpizzadrei, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxdrei2.place(x=120, y=280)
                            einfbutton = Button(Familienpizzadrei, text='einfugen', width=6, bg='green', command=addzutatdrei)
                            einfbutton.place(x=140, y=20)
                            einfbutton1 = Button(Familienpizzadrei, text='einfugen', width=6, bg='green', command=addzutatdrei1)
                            einfbutton1.place(x=325, y=20)
                            einfbutton2 = Button(Familienpizzadrei, text='einfugen', width=6, bg='green', command=addzutatdrei2)
                            einfbutton2.place(x=230, y=280)
                            listboxdrei = Listbox(Familienpizzadrei, width=20, bd=5)
                            listboxdrei.place(x=30, y=60)
                            listboxdrei1 = Listbox(Familienpizzadrei, width=20, bd=5)
                            listboxdrei1.place(x=220, y=60)
                            listboxdrei2 = Listbox(Familienpizzadrei, width=20, bd=5)
                            listboxdrei2.place(x=120, y=320)
                            Fertigbutton = Button(Familienpizzadrei, text='fertig', width=10, bg='grey', command=fertigdrei)
                            Fertigbutton.place(x=320, y=500)
                            entryboxdrei.bind('<Return>',addzutatdrei)
                            entryboxdrei1.bind('<Return>', addzutatdrei1)
                            entryboxdrei2.bind('<Return>', addzutatdrei2)
                            listboxdrei.bind('<Button-1>', lambda event: delete_selected_item(listboxdrei))
                            listboxdrei1.bind('<Button-1>', lambda event: delete_selected_item(listboxdrei1))
                            listboxdrei2.bind('<Button-1>', lambda event: delete_selected_item(listboxdrei2))

                        # ------------------------------------------------------------------------------------------------------------------#
                        def Familiewindow2():
                            # Declare global variables
                            global Familiewindow2
                            global Familienpizzahalb

                            # Close the current window and open a new one
                            Familienpizza.destroy()
                            Familienpizzahalb = Toplevel(root)

                            # Set the size and position of the new window
                            root.wm_state('iconic')
                            Familienpizzahalb.state(newstate=None)
                            Familienpizzahalb.config(bd=2)
                            Familienpizzahalb.focus_set()
                            titlespace = " "
                            window_height = 290
                            window_width = 400
                            screen_width = Familienpizzahalb.winfo_screenwidth()
                            screen_height = Familienpizzahalb.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizzahalb.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizzahalb.config(bg='dark slate gray')
                            Familienpizzahalb.resizable(width=False, height=False)

                            # Define functions to add ingredients to lists
                            def addzutathalb(event=None):
                                sado = str(entryboxhalb.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxhalb.get()[:1])


                                    except ValueError:
                                        input_number = None


                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxhalb.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows!='' :
                                                listboxhalb.insert(END,rows)
                                                entryboxhalb.delete(0, END)
                                else:


                                    sado = str(entryboxhalb.get())
                                    sado1 = sado.strip().replace(' ', '')
                                    listboxhalb.insert(END, sado1)
                                    entryboxhalb.delete(0, END)

                            def addzutathalb1(event=None):
                                sado = str(entryboxhalb1.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxhalb1.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxhalb1.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxhalb1.insert(END, rows)
                                                entryboxhalb1.delete(0, END)
                                else:

                                    sado = str(entryboxhalb1.get())
                                    sado1 = sado.strip().replace(' ', '')
                                    listboxhalb1.insert(END, sado1)
                                    entryboxhalb1.delete(0, END)

                            # Define a function to calculate the total price of the pizza and add it to the entry box
                            def fertighalb(event=None):
                                # Get the ingredients from listboxhalb and listboxhalb1
                                sicko = (listboxhalb.get(0, END))
                                soko = (listboxhalb1.get(0, END))
                                saro = str(sicko).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(
                                    ' ', '/')
                                saro1 = str(soko).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(
                                    ' ', '/')
                                if not saro:
                                    saro = 'Margherita'
                                if not saro1:
                                    saro1 = 'Margherita'

                                # Add the ingredients to listbox2 with a +50% mark-up
                                # Insert the values into listbox2
                                listbox2.insert(END, f'+50%{saro}')
                                listbox2.insert(END, f'+50%{saro1}')

                                # Combine the ingredients and remove duplicates
                                sado1 = sicko + soko
                                soso = []
                                for nano in sado1:
                                    if nano not in soso:
                                        soso.append(nano)
                                loko = []
                                for t in soso:
                                    loko.append(t)

                                # Get the current price and free ingredients for "FamilienPizza" from the database
                                cur = connE.cursor()
                                cur.execute('select* from einstellung where Name = ?', ('FamilienPizza',))
                                nado = cur.fetchall()
                                Frei = ''
                                Preis = ''
                                for t in nado:
                                    Frei = str(t[1])
                                    Preis = t[2]

                                # Calculate the number of non-free ingredients and add the cost to entrybox6
                                naldo = len(loko[int(Frei[0]):])
                                siko = float(entrybox6.get())
                                entrybox6.delete(0, END)
                                entrybox6.insert(END, siko + float(naldo * Preis))

                                # Clear entryboxhalb1 and close Familienpizzahalb window
                                entryboxhalb1.delete(0, END)
                                Familienpizzahalb.destroy()

                            # ----------------------------------------------------------------------------------------------------------------------#
                            entryboxhalb = AutocompleteEntry(Familienpizzahalb, font=("Helvetica", 12, 'bold'), width=14,
                                                             completevalues=zutatenpreise)
                            entryboxhalb.place(x=30, y=20)
                            einfbutton = Button(Familienpizzahalb, text='einfugen', width=6, bg='green',
                                                command=addzutathalb)
                            einfbutton.place(x=140, y=20)
                            listboxhalb = Listbox(Familienpizzahalb, width=20, bd=5)
                            listboxhalb.place(x=30, y=60)
                            Fertigbutton = Button(Familienpizzahalb, text='fertig', width=10, bg='grey', command=fertighalb)
                            Fertigbutton.place(x=325, y=260)
                            entryboxhalb1 = AutocompleteEntry(Familienpizzahalb, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxhalb1.place(x=220, y=20)
                            listboxhalb1 = Listbox(Familienpizzahalb, width=20, bd=5)
                            listboxhalb1.place(x=220, y=60)
                            einfbutton1 = Button(Familienpizzahalb, text='einfugen', width=6, bg='green',
                                                 command=addzutathalb1)
                            einfbutton1.place(x=325, y=20)
                            entryboxhalb.bind('<Return>',addzutathalb)
                            entryboxhalb1.bind('<Return>',addzutathalb1)
                            listboxhalb1.bind('<Button-1>', lambda event: delete_selected_item(listboxhalb1))
                            listboxhalb.bind('<Button-1>', lambda event: delete_selected_item(listboxhalb))

                        #####################################FamilienPizza vier###########################################################
                        def Familiewindow4():
                            global Familiewindow4
                            Familienpizza.destroy()
                            global Familienpizzavier

                            Familienpizzavier = Toplevel(root)
                            root.wm_state('iconic')
                            Familienpizzavier.state(newstate=None)
                            Familienpizzavier.config(bd=2)
                            Familienpizzavier.focus_set()
                            titlespace = " "
                            window_height = 550
                            window_width = 400
                            screen_width = Familienpizzavier.winfo_screenwidth()
                            screen_height = Familienpizzavier.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizzavier.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizzavier.config(bg='dark slate gray')
                            Familienpizzavier.resizable(width=False, height=False)

                            # ------------------------------------------------------------------------------------------------------------------#
                            def addzutatvier(event=None):
                                sado = str(entryboxvier.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxvier.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxvier.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxvier.insert(END, rows)
                                                entryboxvier.delete(0, END)
                                else:


                                    sado = str(entryboxvier.get())
                                    sado1 = sado.strip()
                                    listboxvier.insert(END, sado1)
                                    entryboxvier.delete(0, END)

                            # This function adds the value of the entryboxvier1 to the listboxvier1 and clears the entryboxvier1
                            def addzutatvier1(event=None):
                                sado = str(entryboxvier1.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxvier1.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxvier1.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxvier1.insert(END, rows)
                                                entryboxvier1.delete(0, END)
                                else:


                                    sado = str(entryboxvier1.get())
                                    sado1 = sado.strip()
                                    listboxvier1.insert(END, sado1)
                                    entryboxvier1.delete(0, END)

                            # This function adds the value of the entryboxvier2 to the listboxvier2 and clears the entryboxvier2
                            def addzutatvier3(event=None):
                                sado = str(entryboxvier2.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxvier2.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxvier2.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxvier2.insert(END, rows)
                                                entryboxvier2.delete(0, END)
                                else:


                                    sado = str(entryboxvier2.get())
                                    sado1 = sado.strip()
                                    listboxvier2.insert(END, sado1)
                                    entryboxvier2.delete(0, END)

                            # This function adds the value of the entryboxvier3 to the listboxvier3 and clears the entryboxvier3
                            def addzutatvier4(event=None):
                                sado = str(entryboxvier3.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxvier3.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxvier3.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxvier3.insert(END, rows)
                                                entryboxvier3.delete(0, END)
                                else:


                                    sado = str(entryboxvier3.get())
                                    sado1 = sado.strip()
                                    listboxvier3.insert(END, sado1)
                                    entryboxvier3.delete(0, END)

                            # This function calculates the total price for a pizza made from the ingredients in the listboxes,
                            # adds it to the entrybox6, and closes the window
                            def fertigvier():
                                # Get the values from the listboxes and remove any non-alphanumeric characters
                                sicko = listboxvier.get(0, END)
                                soko = listboxvier1.get(0, END)
                                soka = listboxvier2.get(0, END)
                                saka = listboxvier3.get(0, END)
                                sicko1 = str(sicko).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(' ','/').replace('[','').replace(']','')

                                soko1 =str(soko).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(' ','/').replace('[','').replace(']','')

                                soka1 =str(soka).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(' ','/').replace('[','').replace(']','')

                                saka1 = str(saka).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(' ','/').replace('[','').replace(']','')

                                if not sicko1:
                                    sicko1 = 'Margherita'
                                if not soko1:
                                    soko1 = 'Margherita'
                                if not soka1:
                                    soka1 = 'Margherita'
                                if not saka1:
                                    saka1 = 'Margherita'

                                # Add each ingredient to listbox2 with a 25% increase in price
                                listbox2.insert(END, f'+25%{sicko1}')
                                listbox2.insert(END, f'+25%{soko1}')
                                listbox2.insert(END, f'+25%{soka1}')
                                listbox2.insert(END, f'+25%{saka1}')

                                # Combine all the ingredients into one string and remove duplicates
                                sado1 = sicko + soko + soka + saka
                                soso = list(set(sado1))

                                loko = []
                                for t in soso:
                                    loko.append(t)

                                # Get the current price and free ingredients for "FamilienPizza" from the database
                                cur = connE.cursor()
                                cur.execute('select* from einstellung where Name = ?', ('FamilienPizza',))
                                nado = cur.fetchall()
                                Frei = ''
                                Preis = ''
                                for t in nado:
                                    Frei = str(t[1])
                                    Preis = t[2]

                                # Calculate the number of non-free ingredients and add the cost to entrybox6
                                naldo = len(loko[int(Frei[0]):])
                                siko = float(entrybox6.get())
                                entrybox6.delete(0, END)
                                entrybox6.insert(END, siko + float(naldo * Preis))
                                # Close the window
                                Familienpizzavier.destroy()

                            # ------------------------------------------------------------------------------------------------------------------#
                            entryboxvier = AutocompleteEntry(Familienpizzavier, font=("Helvetica", 12, 'bold'), width=14,
                                                             completevalues=zutatenpreise)
                            entryboxvier.place(x=30, y=20)
                            entryboxvier1 = AutocompleteEntry(Familienpizzavier, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxvier1.place(x=220, y=20)
                            entryboxvier2 = AutocompleteEntry(Familienpizzavier, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxvier2.place(x=220, y=280)
                            entryboxvier3 = AutocompleteEntry(Familienpizzavier, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxvier3.place(x=30, y=280)
                            einfbutton = Button(Familienpizzavier, text='einfugen', width=6, bg='green', command=addzutatvier)
                            einfbutton.place(x=140, y=20)
                            einfbutton1 = Button(Familienpizzavier, text='einfugen', width=6, bg='green', command=addzutatvier1)
                            einfbutton1.place(x=325, y=20)
                            einfbutton2 = Button(Familienpizzavier, text='einfugen', width=6, bg='green', command=addzutatvier3)
                            einfbutton2.place(x=325, y=280)
                            einfbutton3 = Button(Familienpizzavier, text='einfugen', width=6, bg='green', command=addzutatvier4)
                            einfbutton3.place(x=140, y=280)
                            listboxvier = Listbox(Familienpizzavier, width=20, bd=5)
                            listboxvier.place(x=30, y=60)
                            listboxvier1 = Listbox(Familienpizzavier, width=20, bd=5)
                            listboxvier1.place(x=220, y=60)
                            listboxvier3 = Listbox(Familienpizzavier, width=20, bd=5)
                            listboxvier3.place(x=30, y=320)
                            listboxvier2 = Listbox(Familienpizzavier, width=20, bd=5)
                            listboxvier2.place(x=220, y=320)
                            Fertigbutton = Button(Familienpizzavier, text='fertig', width=10, bg='grey', command=fertigvier)
                            Fertigbutton.place(x=320, y=500)
                            entryboxvier.bind('<Return>',addzutatvier)
                            entryboxvier1.bind('<Return>', addzutatvier1)
                            entryboxvier2.bind('<Return>', addzutatvier3)
                            entryboxvier3.bind('<Return>', addzutatvier4)
                            listboxvier.bind('<Button-1>', lambda event: delete_selected_item(listboxvier))
                            listboxvier1.bind('<Button-1>', lambda event: delete_selected_item(listboxvier1))
                            listboxvier2.bind('<Button-1>', lambda event: delete_selected_item(listboxvier2))
                            listboxvier3.bind('<Button-1>', lambda event: delete_selected_item(listboxvier3))

                        ##################################### FamilienPizza 75%+25% #####################################################
                        def Familiewindow5():
                            # Declare global variables
                            global Familiewindow2
                            global Familienpizzahalb

                            # Close the current window and open a new one
                            Familienpizza.destroy()
                            Familienpizzahalb = Toplevel(root)

                            # Set the size and position of the new window
                            root.wm_state('iconic')
                            Familienpizzahalb.state(newstate=None)
                            Familienpizzahalb.config(bd=2)
                            Familienpizzahalb.focus_set()
                            titlespace = " "
                            window_height = 290
                            window_width = 400
                            screen_width = Familienpizzahalb.winfo_screenwidth()
                            screen_height = Familienpizzahalb.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizzahalb.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizzahalb.config(bg='dark slate gray')
                            Familienpizzahalb.resizable(width=False, height=False)

                            # Define functions to add ingredients to lists
                            def addzutathalb(event=None):
                                sado = str(entryboxhalb.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxhalb.get()[:1])


                                    except ValueError:
                                        input_number = None


                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxhalb.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows!='' :
                                                listboxhalb.insert(END,rows)
                                                entryboxhalb.delete(0, END)
                                else:


                                    sado = str(entryboxhalb.get())
                                    sado1 = sado.strip().replace(' ', '')
                                    listboxhalb.insert(END, sado1)
                                    entryboxhalb.delete(0, END)

                            def addzutathalb1(event=None):
                                sado = str(entryboxhalb1.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxhalb1.get()[:1])


                                    except ValueError:
                                        input_number = None


                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxhalb1.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows!='' :
                                                listboxhalb1.insert(END,rows)
                                                entryboxhalb1.delete(0, END)
                                else:


                                    sado = str(entryboxhalb1.get())
                                    sado1 = sado.strip().replace(' ', '')
                                    listboxhalb1.insert(END, sado1)
                                    entryboxhalb1.delete(0, END)

                            # Define a function to calculate the total price of the pizza and add it to the entry box
                            def fertighalb(event=None):
                                # Get the ingredients from listboxhalb and listboxhalb1
                                sicko = (listboxhalb.get(0, END))
                                soko = (listboxhalb1.get(0, END))
                                saro = str(sicko).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(
                                    ' ', '/')
                                saro1 = str(soko).replace('(', '').replace(')', '').replace("'", "").replace(',', '').replace(
                                    ' ', '/')

                                # Add the ingredients to listbox2 with a +50% mark-up

                                if not saro:
                                    saro = 'Margherita'
                                if not saro1:
                                    saro1 = 'Margherita'
                                listbox2.insert(END, '+75%' + str(saro))
                                listbox2.insert(END, '+25%' + str(saro1))

                                # Combine the ingredients and remove duplicates
                                sado1 = sicko + soko
                                soso = []
                                for nano in sado1:
                                    if nano not in soso:
                                        soso.append(nano)
                                loko = []
                                for t in soso:
                                    loko.append(t)

                                # Get the current price and free ingredients for "FamilienPizza" from the database
                                cur = connE.cursor()
                                cur.execute('select* from einstellung where Name = ?', ('FamilienPizza',))
                                nado = cur.fetchall()
                                Frei = ''
                                Preis = ''
                                for t in nado:
                                    Frei = str(t[1])
                                    Preis = t[2]

                                # Calculate the number of non-free ingredients and add the cost to entrybox6
                                naldo = len(loko[int(Frei[0]):])
                                siko = float(entrybox6.get())
                                entrybox6.delete(0, END)
                                entrybox6.insert(END, siko + float(naldo * Preis))

                                # Clear entryboxhalb1 and close Familienpizzahalb window
                                entryboxhalb1.delete(0, END)
                                Familienpizzahalb.destroy()

                            # ----------------------------------------------------------------------------------------------------------------------#
                            entryboxhalb = AutocompleteEntry(Familienpizzahalb, font=("Helvetica", 12, 'bold'), width=14,
                                                             completevalues=zutatenpreise)
                            entryboxhalb.place(x=30, y=20)
                            label75 = Label(Familienpizzahalb, text='75%', bg='dark slate gray', font=("Helvetica", 12, 'bold'))
                            label75.place(x=30, y=240)
                            label25 = Label(Familienpizzahalb, text='25%', bg='dark slate gray', font=("Helvetica", 12, 'bold'))
                            label25.place(x=220, y=240)
                            einfbutton = Button(Familienpizzahalb, text='einfugen', width=6, bg='green',
                                                command=addzutathalb)
                            einfbutton.place(x=140, y=20)
                            listboxhalb = Listbox(Familienpizzahalb, width=20, bd=5)
                            listboxhalb.place(x=30, y=60)
                            Fertigbutton = Button(Familienpizzahalb, text='fertig', width=10, bg='grey', command=fertighalb)
                            Fertigbutton.place(x=325, y=260)
                            entryboxhalb1 = AutocompleteEntry(Familienpizzahalb, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxhalb1.place(x=220, y=20)
                            listboxhalb1 = Listbox(Familienpizzahalb, width=20, bd=5)
                            listboxhalb1.place(x=220, y=60)
                            einfbutton1 = Button(Familienpizzahalb, text='einfugen', width=6, bg='green',
                                                 command=addzutathalb1)
                            einfbutton1.place(x=325, y=20)
                            entryboxhalb1.bind('<Return>',addzutathalb1)
                            entryboxhalb.bind('<Return>', addzutathalb)
                            listboxhalb.bind('<Button-1>', lambda event: delete_selected_item(listboxhalb))
                            listboxhalb1.bind('<Button-1>', lambda event: delete_selected_item(listboxhalb1))

                        ##################################### FamilienPizza 1x50%+2x25% #################################################
                        def Familiewindow6():
                            global Familiewindow3
                            Familienpizza.destroy()
                            global Familienpizzadrei
                            Familienpizzadrei = Toplevel(root)
                            root.wm_state('iconic')
                            Familienpizzadrei.state(newstate=None)
                            Familienpizzadrei.config(bd=2)
                            Familienpizzadrei.focus_set()
                            titlespace = " "
                            window_height = 550
                            window_width = 400
                            screen_width = Familienpizzadrei.winfo_screenwidth()
                            screen_height = Familienpizzadrei.winfo_screenheight()
                            x_cordinate = int((screen_width / 2) - (window_width / 2))
                            y_cordinate = int((screen_height / 2) - (window_height / 2))
                            Familienpizzadrei.geometry(
                                "{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                            Familienpizzadrei.config(bg='dark slate gray')
                            Familienpizzadrei.resizable(width=False, height=False)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to the first listbox in the third page of the order process
                            def addzutatdrei(event=None):
                                sado = str(entryboxdrei.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxdrei.get()[:1])


                                    except ValueError:
                                        input_number = None


                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxdrei.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows!='' :
                                                listboxdrei.insert(END,rows)
                                                entryboxdrei.delete(0, END)
                                else:



                                    sado = str(entryboxdrei.get())
                                    sado1 = sado.strip()
                                    listboxdrei.insert(END, sado1)
                                    entryboxdrei.delete(0, END)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to the second listbox in the third page of the order process
                            def addzutatdrei1(event=None):
                                sado = str(entryboxdrei1.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxdrei1.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxdrei1.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxdrei1.insert(END, rows)
                                                entryboxdrei1.delete(0, END)
                                else:

                                    sado = str(entryboxdrei1.get())
                                    sado1 = sado.strip()
                                    listboxdrei1.insert(END, sado1)
                                    entryboxdrei1.delete(0, END)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to add ingredient to the third listbox in the third page of the order process
                            def addzutatdrei2(event=None):
                                sado = str(entryboxdrei2.get())
                                if sado[:1].isdigit():
                                    try:
                                        input_number = int(entryboxdrei2.get()[:1])


                                    except ValueError:
                                        input_number = None

                                    # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                                    if input_number is not None:
                                        input_numb_text = entryboxdrei2.get()

                                        cur1 = connS.cursor()
                                        cur1.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text,))
                                        row = cur1.fetchone()
                                        if row is not None:
                                            input_text = row[0]
                                        cur = connZ.cursor()
                                        cur.execute('SELECT * FROM zutaten WHERE SpeiseName = ?', (input_text,))
                                        row = cur.fetchone()
                                        for rows in row[3:]:
                                            if rows != '':
                                                listboxdrei2.insert(END, rows)
                                                entryboxdrei2.delete(0, END)
                                else:

                                    sado = str(entryboxdrei2.get())
                                    sado1 = sado.strip()
                                    listboxdrei2.insert(END, sado1)
                                    entryboxdrei2.delete(0, END)

                            # ------------------------------------------------------------------------------------------------------------------#
                            # function to finalize the order of the third page of the order process
                            def fertigdrei():
                                # get the ingredients from the three listboxes and concatenate them
                                sicko = listboxdrei.get(0, END)
                                soko = listboxdrei1.get(0, END)
                                soka = listboxdrei2.get(0, END)
                                saro = str(sicko).replace('(', '').replace(')', '').replace("'", "").replace(
                                    ',', '').replace(' ', '/')
                                saro1 = str(soko).replace('(', '').replace(')', '').replace("'", "").replace(
                                    ',', '').replace(' ', '/')
                                saro2 = str(soka).replace('(', '').replace(')', '').replace("'", "").replace(
                                    ',', '').replace(' ', '/')
                                if not saro:
                                    saro = 'Margherita'
                                if not saro1:
                                    saro1 = 'Margherita'
                                if not saro2:
                                    saro2 = 'Margherita'
                                listbox2.insert(END, f'+25%{saro}')
                                listbox2.insert(END, f'+25%{saro1}')
                                listbox2.insert(END, f'+50%{saro2}')

                                # calculate the price of the final order and display it in the corresponding entrybox
                                sado1 = sicko + soko + soka  # concatenate the contents of the three listboxes
                                soso = []  # initialize an empty list
                                for nano in sado1:  # iterate through the concatenated string
                                    if nano not in soso:  # if the character is not already in the list
                                        soso.append(nano)  # append it to the list
                                loko = []  # initialize another empty list
                                for t in soso:  # iterate through the list of unique characters
                                    loko.append(t)  # append each character to the new list
                                cur = connE.cursor()  # create a cursor object
                                cur.execute('select* from einstellung where Name = ?',
                                            ('FamilienPizza',))  # execute a select statement
                                nado = cur.fetchall()  # fetch all the results
                                Frei = ''  # initialize a variable
                                Preis = ''  # initialize another variable
                                for t in nado:  # iterate through the results
                                    Frei = str(t[1])  # assign the second element to the variable Frei
                                    Preis = t[2]  # assign the third element to the variable Preis
                                naldo = len(
                                    loko[int(Frei[0]):])  # calculate the number of ingredients beyond the first Frei elements
                                siko = float(entrybox6.get())  # get the value from entrybox6 as a float
                                entrybox6.delete(0, END)  # delete the current value in entrybox6
                                entrybox6.insert(END, siko + float(naldo * Preis))  # insert the new value in entrybox6

                                Familienpizzadrei.destroy()  # close the third page of the order process

                            # ------------------------------------------------------------------------------------------------------------------#
                            entryboxdrei = AutocompleteEntry(Familienpizzadrei, font=("Helvetica", 12, 'bold'), width=14,
                                                             completevalues=zutatenpreise)
                            entryboxdrei.place(x=30, y=20)
                            label25 = Label(Familienpizzadrei, text='25%', bg='dark slate gray', font=("Helvetica", 12, 'bold'))
                            label25.place(x=30, y=240)
                            label26 = Label(Familienpizzadrei, text='25%', bg='dark slate gray', font=("Helvetica", 12, 'bold'))
                            label26.place(x=220, y=240)
                            label26 = Label(Familienpizzadrei, text='50%', bg='dark slate gray', font=("Helvetica", 12, 'bold'))
                            label26.place(x=155, y=495)
                            entryboxdrei1 = AutocompleteEntry(Familienpizzadrei, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxdrei1.place(x=220, y=20)
                            entryboxdrei2 = AutocompleteEntry(Familienpizzadrei, font=("Helvetica", 12, 'bold'), width=14,
                                                              completevalues=zutatenpreise)
                            entryboxdrei2.place(x=120, y=280)
                            einfbutton = Button(Familienpizzadrei, text='einfugen', width=6, bg='green', command=addzutatdrei)
                            einfbutton.place(x=140, y=20)
                            einfbutton1 = Button(Familienpizzadrei, text='einfugen', width=6, bg='green', command=addzutatdrei1)
                            einfbutton1.place(x=325, y=20)
                            einfbutton2 = Button(Familienpizzadrei, text='einfugen', width=6, bg='green', command=addzutatdrei2)
                            einfbutton2.place(x=230, y=280)
                            listboxdrei = Listbox(Familienpizzadrei, width=20, bd=5)
                            listboxdrei.place(x=30, y=60)
                            listboxdrei1 = Listbox(Familienpizzadrei, width=20, bd=5)
                            listboxdrei1.place(x=220, y=60)
                            listboxdrei2 = Listbox(Familienpizzadrei, width=20, bd=5)
                            listboxdrei2.place(x=120, y=320)
                            Fertigbutton = Button(Familienpizzadrei, text='fertig', width=10, bg='grey', command=fertigdrei)
                            Fertigbutton.place(x=320, y=500)
                            entryboxdrei.bind('<Return>',addzutatdrei)
                            entryboxdrei1.bind('<Return>', addzutatdrei1)
                            entryboxdrei2.bind('<Return>', addzutatdrei2)
                            listboxdrei.bind('<Button-1>', lambda event: delete_selected_item(listboxdrei))
                            listboxdrei1.bind('<Button-1>', lambda event: delete_selected_item(listboxdrei1))
                            listboxdrei2.bind('<Button-1>', lambda event: delete_selected_item(listboxdrei2))

                    liso1 = []


                    #####################################zutaten funktion##############################################################
                    def zutati(event=None):
                        # Declare global variables

                        global drop, liso, kl_gr, liso1
                        # Define a constant variable
                        FamilienPizza = 'Familien Pizza'
                        # Get the input from entrybox1
                        input_text = entrybox1.get()
                        # Check if input_text is a number
                        try:
                            input_number = int(input_text[:1])
                        except ValueError:
                            input_number = None
                        # If input_text is a number, perform a select statement on the Speisen table to get the corresponding Speisename
                        if input_number is not None:
                            input_numb_text = entrybox1.get()
                            cur = connS.cursor()
                            cur.execute('SELECT name FROM Speisen WHERE nummer = ?', (input_numb_text ,))
                            row = cur.fetchone()
                            if row is not None:
                                input_text = row[0]

                        # Check if input_text is not empty
                        if input_text != '':
                            liko1 = input_text
                            entrybox1.delete(0,END)
                            entrybox1.insert(0,liko1)

                            # Connect to the database and execute a select statement
                            cur1 = connE.cursor()
                            cur1.execute('SELECT Anzahl FROM Freiezutaten WHERE Speisename = (?)', (liko1,))
                            mor = cur1.fetchall()
                            # If the result of the select statement is not empty, add the values to a list
                            if mor != '' or []:
                                for izo in mor:
                                    izo1 = str(izo).replace(',', '').replace('(', '').replace(')', '')
                                    silok = int(izo1)
                                    mous = 'one'
                                    while silok > 0:
                                        liso.insert(0, mous)
                                        silok -= 1

                            # Get the category of the selected dish from the Speisen table
                            nado = input_text
                            cu = connS.cursor()
                            cu.execute('SELECT K_G FROM Speisen WHERE name = (?)', (nado,))
                            opl = cu.fetchall()
                            kat = str(opl)
                            kat1 = kat.replace(',', '').replace('(', '').replace(')', '').replace("'", "").replace('[',
                                                                                                                   '').replace(
                                ']', '')
                            # Get the details of the selected dish from the Speisen table
                            vur = connS.cursor()
                            vur.execute("SELECT * FROM Speisen WHERE Name = (?)", (nado,))
                            # Define constant variables
                            Klein = 'Klein'
                            Gross = 'Gross'
                            Standard = 'Standard'
                            zob = []

                            # If the dish has a size option, create a new option menu widget
                            for nano in vur.fetchall():
                                ziko = str(nano[2])
                                if Klein in nano[4] and ziko==('Salattheke') or nado[-13:] =='Pizzabrötchen'  :

                                    kl_gr =  Gross,Klein
                                    drop.destroy()
                                    drop =ttk.Combobox (Mainframe2, values= kl_gr)
                                    drop.configure(font=('Helvetica bold', 16), width=8)
                                    drop.place(x=255, y=70)
                                    drop.set(Klein)
                                    combozutat.focus()
                                    combozutat.event_generate("<Down>")
                                elif  Klein not in nano[4] and ziko==('Salattheke'):
                                    drop = ttk.Combobox(Mainframe2, values=Gross)
                                    drop.configure(font=('Helvetica bold', 16), width=8)
                                    drop.place(x=255, y=70)
                                    drop.set(Gross)
                                    drop.bind("<<ComboboxSelected>>", gross_pizza)
                                    drop.set(nano[4])
                                    drop.config(state=DISABLED)



                                    combozutat.focus()
                                    combozutat.event_generate("<Down>")
                                if Klein in nano[4] and ziko !=('Salattheke'):
                                    kl_gr =  Gross,Klein
                                    drop.destroy()
                                    drop =ttk.Combobox (Mainframe2, values= kl_gr)
                                    drop.configure(font=('Helvetica bold', 16), width=8)
                                    drop.place(x=255, y=70)
                                    drop.focus()
                                    def open_dropdown(event):

                                        press('Down')
                                    drop.bind("<FocusIn>", open_dropdown)


                                # If the dish has no size option, disable the option menu widget
                                elif Standard in kat1 or Gross in kat1:
                                    kl_gr = Standard
                                    drop.destroy()
                                    drop = ttk.Combobox(Mainframe2, values=kl_gr)
                                    drop.configure(font=('Helvetica bold', 16), width=8)
                                    drop.place(x=255, y=70)
                                    drop.bind("<<ComboboxSelected>>", gross_pizza)
                                    drop.set(nano[4])
                                    drop.config(state=DISABLED)
                                    entrybox4.focus()
                                    entrybox4.select_range(0,END)
                                # Remove unnecessary characters from the string
                                ziko.replace("'", "").replace(",", "").replace("{", "").replace("(", "").replace(")",
                                                                                                                 "").replace(
                                    "}", "")
                                # Format the result and insert the values into the entrybox widgets
                                result = math.floor(float(nano[3]) * 100) / 100
                                formatted_result = '{:.2f}'.format(result)
                                entrybox3.delete(0, END)
                                entrybox5.delete(0, END)
                                entrybox6.delete(0, END)
                                entrybox3.insert(END, nano[1])
                                entrybox5.insert(END, ziko)
                                entrybox6.insert(END, formatted_result)
                            listbox2.delete(0, END)
                            listbox3.delete(0, END)
                            cur = connZ.cursor()
                            cur.execute("select* from zutaten where SpeiseName=(?)", (nado,))
                            for ziko in cur.fetchall():
                                listbox2.insert(END, ziko[1]),
                                [listbox2.insert(END, element) for element in ziko[2:] if element != '']
                            connZ.commit()

                            drop.bind("<<ComboboxSelected>>", gross_pizza)
                            if nado == FamilienPizza:
                                Familie_window()
                        else:
                            pass
                        entrybox1.config(state=DISABLED)



                    connS.commit()

                    def jump_to_entrybox5(event):
                        entrybox2.focus()

                    def listbox_to_enterybox(event):

                        speise = listbox1.get(ACTIVE)
                        entrybox1.delete(0, END)
                        entrybox1.insert(0, speise)
                        entrybox1.focus()
                        press('Enter')
                        entrybox1.config(state=DISABLED)








                    entrybox1.bind('<Return>', zutati)
                    listbox1.bind('<Double-Button-1>', zutati)
                    listbox1.bind('<Return>',listbox_to_enterybox)
                    entrybox1.bind('<Down>',lambda event:listbox1.focus_set())


                    # ------------------------------------------------------------------------------------------------------------------#
                    # Had to be moved here to be after the function
                    # ----------------------------------------------------------------------------------------------------------------------#
                    def Zutatinr(event=None):
                        cub = connS.cursor()
                        naldo = entrybox3.get()
                        cub.execute("select * from Speisen where Nummer =(?)", (naldo,))
                        foro = cub.fetchall()
                        for taha in foro:
                            entrybox1.delete(0, END)
                            entrybox1.insert(END, taha[0])

                    # s_button=Button(Mainframe2,text='S',width=2,bg='blue',command=Zutatinr)
                    # s_button.place(x=230,y=70)
                    entrybox3.bind('<Return>', Zutatinr)
                    entrybox3.bind('<Return>', zutati, add="+")

                    # ----------------------------------------------------------------------------------------------------------------------#
                    def deletezutaten(event=None):
                        global deletezutaten, liso

                        siko = listbox2.get(ACTIVE)
                        salo = siko[0]
                        salo1 = str(siko[1:])
                        if salo != '+':
                            liso.insert(0, siko)
                        sok = len(liso)

                        is_gross = False
                        cur = connE.cursor()
                        cur.execute("select* from Gross ")
                        gross = cur.fetchall()
                        if salo == '+':
                            for row in gross:
                                if salo1 in row:
                                    cur = connE.cursor()
                                    cur.execute("select* from Gross where Name1=(?) ", (salo1,))
                                    gross_price = cur.fetchone()
                                    float_gross_price = float(gross_price[1])
                                    is_gross = True

                                    sicko = float(entrybox6.get())
                                    nalo = (sicko - float_gross_price)
                                    entrybox6.delete(0, END)
                                    result = math.floor(float(nalo) * 100) / 100
                                    formatted_result = '{:.2f}'.format(result)

                                    entrybox6.insert(END, formatted_result)
                                    listbox2.delete(ANCHOR)
                                    break
                            if not is_gross:
                                big = drop.get()
                                sari = 0
                                if big == 'Gross':
                                    sari = 1
                                elif big == 'Klein':
                                    sari = 0.5
                                else:
                                    sari = 1

                                sicko = float(entrybox6.get())
                                entrybox6.delete(0, END)

                                nalo = (sicko - sari)
                                entrybox6.delete(0, END)
                                result = math.floor(float(nalo) * 100) / 100
                                formatted_result = '{:.2f}'.format(result)

                                entrybox6.insert(END, formatted_result)
                                listbox2.delete(ANCHOR)
                        elif salo == '*':
                            listbox2.delete(ANCHOR)

                        else:
                            listbox3.insert(END, '-' + listbox2.get(ACTIVE))
                            listbox2.delete(ANCHOR)

                    listbox2.bind("<Delete>", deletezutaten)
                    listbox2.bind('<Double-Button-1>', deletezutaten)

                    # deletezutat = Button(Mainframe2, text='Löschen', bd=2, width=6, bg='red', command=deletezutaten)
                    # deletezutat.place(x=630, y=100)
                    # ----------------------------------------------------------------------------------------------------------------------#
                    def undelete(event=None):
                        listbox2.insert(END, listbox3.get(ACTIVE)[1:])
                        listbox3.delete(ANCHOR)

                    listbox3.bind('<Double-Button-1>', undelete)

                    def addzutat(event=None):
                        # Make the following variables global so they can be accessed outside of the function
                        global addzutat, liso, liso1
                        sado = str(entrybox2.get())

                        sado1 = sado.strip()
                        # Get the value of the entry box and strip whitespace
                        cur = connZu.cursor()
                        zerolist = []
                        if sado1[:1].isdigit():
                            num=int(sado1)

                            cur = connZu.cursor()
                            cur.execute('select Preis€ from zutatenpreise where Nummer =(?)', (sado1,))
                            mako = cur.fetchall()
                        else:
                            cur.execute('select Preis€ from zutatenpreise where Zutat =(?)', (sado1,))
                            mako = cur.fetchall()

                        for sar_zero in mako:

                            zerolist.append(sar_zero[0])

                        liko = drop.get()
                        liko1 = liko.strip()
                        zutat_name=entrybox2.get()
                        zutat_name_s=zutat_name.strip()
                        cur = connZu.cursor()



                        cur=connZu.cursor()
                        cur.execute('select Zutat from zutatenpreise ')
                        mido=cur.fetchall()
                        zutaten_list = [m[0] for m in mido]
                        if zutat_name_s in zutaten_list or zutat_name_s .isdigit():


                            # Get the length of two empty lists
                            liso1 = []
                            modes = len(liso)
                            num = int(modes)
                            countk = 0
                            modes1 = len(liso1)
                            num1 = int(modes1)
                            # Check if the selected pizza is "Familien Pizza"
                            if liko1 == 'Standard'  :

                                if zerolist[0] == 0:
                                    entrybox2.delete(0, END)
                                    listbox2.insert(END, ('*') + sado1)
                                else:
                                    # Get the value of the second entry box and strip whitespace
                                    sado = str(entrybox2.get())

                                    sado1 = sado.strip()
                                    if sado1[:1].isdigit():
                                        cur = connZu.cursor()
                                        cur.execute('select Zutat from zutatenpreise where Nummer =(?)', (sado1,))
                                        mako = cur.fetchall()
                                        if mako is not None:
                                            input_num = str(mako[0]).replace('(','').replace(')','').replace(',','').replace("'","")
                                            sado1=input_num
                                            entrybox2.delete(0,END)

                                    # Insert the value of the second entry box into the second listbox with a "+" sign in front
                                    listbox2.insert(END, ('+') + sado1)
                                    # Execute a SQL query to get the price of the selected pizza ingredient
                                    cur = connZu.cursor()
                                    cur.execute('select* from zutatenpreise where Zutat =(?)', (sado1,))
                                    mako = cur.fetchall()
                                    haupt_list = ['Dönerfleisch', 'Hähn-Döner', 'Scampi', 'Hänchenfillet', 'Lammkotelett',
                                                  'Chicken-Nuggets', 'Chicken-Wings']
                                    # Iterate through the results and add the price to the value in the sixth entry box
                                    for saro in mako:

                                        if num > 0 and sado1 not in haupt_list:
                                            liso.pop(0)
                                        elif entrybox1.get()=='Gefüllt Pizzabrötchen'or entrybox1.get()=='Chickenburger':

                                            sari = 0.5
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + float(sari))
                                            result1 = math.floor(float(summ) * 100) / 100
                                            formatted_result1 = '{:.2f}'.format(result1)
                                            entrybox6.insert(0, formatted_result1)


                                        else:
                                            new_price_float = float(saro[2])
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)
                                    # Clear the second entry box
                                    entrybox2.delete(0, END)
                            # If the selected pizza is not "Familien Pizza"
                            if liko1 == 'Klein' :
                                if zerolist[0] ==0:

                                    entrybox2.delete(0, END)
                                    listbox2.insert(END, ('*') + sado1)
                                else:
                                    # Get the value of the second entry box and strip whitespace
                                    sado = str(entrybox2.get())
                                    # Get the values of the dropdown menu and fifth entry box
                                    big = drop.get()
                                    kat = entrybox5.get()
                                    sado1 = sado.strip()
                                    if sado1[:1].isdigit():
                                        cur = connZu.cursor()
                                        cur.execute('select Zutat from zutatenpreise where Nummer =(?)', (sado1,))
                                        mako = cur.fetchall()
                                        if mako is not None:
                                            input_num = str(mako[0]).replace('(', '').replace(')', '').replace(',', '').replace("'", "")
                                            sado1 = input_num
                                            entrybox2.delete(0, END)

                                    # Execute a SQL query to get all the pizza ingredient prices
                                    cur = connZu.cursor()
                                    cur.execute("select * from zutatenpreise")
                                    mako = cur.fetchall()
                                    # Get the length of two empty lists
                                    conn = connE
                                    cur = conn.cursor()
                                    cur.execute('select* from Gross ')
                                    gross = cur.fetchall()

                                    is_gross = False

                                    entrybox2.delete(0, END)
                                    listbox2.insert(END, ('+') + sado1)
                                    # If the length of the first empty list is greater than 0, remove the first item
                                    if num > 0:
                                        sari = 0
                                        if sado1 == 'Scampi':
                                            sari += 0.5
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + float(sari))
                                            result1 = math.floor(float(summ) * 100) / 100
                                            formatted_result1 = '{:.2f}'.format(result1)
                                            entrybox6.insert(0, formatted_result1)
                                        for row in gross:
                                            if sado1 in row and row[0] not in liso:
                                                is_gross = True
                                                cur.execute('select Price1 from Gross where Name1=(?)', (sado1,))
                                                new_price = cur.fetchone()
                                                new_price_float = float(new_price[0])

                                                sari = new_price_float - 0.5
                                                siko = float(entrybox6.get())
                                                entrybox6.delete(0, END)
                                                summ = (float(sari) + siko)
                                                result1 = math.floor(float(summ) * 100) / 100
                                                formatted_result1 = '{:.2f}'.format(result1)
                                                entrybox6.insert(0, formatted_result1)
                                        liso.pop(0)

                                    else:
                                        listos = []
                                        for row in gross:
                                            listos.insert(0, row[0])
                                        if sado1 in listos:

                                            cur.execute('select Price1 from Gross where Name1=(?)', (sado1,))
                                            new_price = cur.fetchone()
                                            new_price_float = float(new_price[0])
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)

                                        elif sado1 != 'Scampi':

                                            new_price_float = 0.5
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)
                                        else:
                                            new_price_float = 1
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)
                            if liko1 == 'Gross':
                                if zerolist[0] ==0:

                                    entrybox2.delete(0, END)
                                    listbox2.insert(END, ('*') + sado1)
                                else:
                                    # Get the value of the second entry box and strip whitespace
                                    sado = str(entrybox2.get())
                                    sado1 = sado.strip()
                                    if sado1[:1].isdigit():
                                        cur = connZu.cursor()
                                        cur.execute('select Zutat from zutatenpreise where Nummer =(?)', (sado1,))
                                        mako = cur.fetchall()
                                        if mako is not None:
                                            input_num = str(mako[0]).replace('(', '').replace(')', '').replace(',', '').replace("'", "")
                                            sado1 = input_num
                                            entrybox2.delete(0, END)

                                    # Get the values of the dropdown menu and fifth entry box
                                    big = drop.get()
                                    kat = entrybox5.get()


                                    # Execute a SQL query to get all the pizza ingredient prices
                                    cur = connZu.cursor()
                                    cur.execute("select * from zutatenpreise")
                                    mako = cur.fetchall()
                                    # Get the length of two empty lists
                                    conn = connE
                                    cur = conn.cursor()
                                    cur.execute('select* from Gross ')
                                    gross = cur.fetchall()
                                    is_gross = False
                                    entrybox2.delete(0, END)
                                    listbox2.insert(END, ('+') + sado1)
                                    # If the length of the first empty list is greater than 0, remove the first item
                                    if num > 0:
                                        sari = 0
                                        if sado1 == 'Scampi':
                                            sari += 1
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + float(sari))
                                            result1 = math.floor(float(summ) * 100) / 100
                                            formatted_result1 = '{:.2f}'.format(result1)
                                            entrybox6.insert(0, formatted_result1)

                                        liso.pop(0)

                                    else:
                                        listos = []
                                        for row in gross:
                                            listos.insert(0, row[0])

                                        if sado1 in listos:

                                            cur.execute('select Price1 from Gross where Name1=(?)', (sado1,))
                                            new_price = cur.fetchone()
                                            new_price_float = float(new_price[0])
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)

                                        elif sado1 != 'Scampi':

                                            new_price_float = 1
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)
                                        else:
                                            new_price_float = 1.5
                                            siko = float(entrybox6.get())
                                            entrybox6.delete(0, END)
                                            summ = (siko + new_price_float)
                                            result = math.floor(float(summ) * 100) / 100
                                            formatted_result = '{:.2f}'.format(result)
                                            entrybox6.insert(0, formatted_result)


                            else:
                                pass

                                # Commit changes to the database
                                connZu.commit()
                        else:
                            pass

                    ############################################      Lieferzuschlag          #############################################
                    global ako
                    cur12 = connE.cursor()
                    cur12.execute(' select * from Lieferzuschlg  ')
                    cob = cur12.fetchall()
                    liefako = '0'
                    for ido in cob:
                        if cob:
                            for ako in ido:
                                liefako = float(ako)
                    liefer = Label(Mainframe3, text='Mindestpreise:', bg=colour4, font=font_size, )
                    liefer.place(x=50, y=395)  # für die lieferzuschlag oder Mindensten wert
                    liefer1 = Label(Mainframe3, text=str(liefako) + '€', bg=colour4, font=font_size, fg='red')
                    liefer1.place(x=165, y=395)

                    ####################################Search and Autofill Functions#######################################################
                    def update(data):
                        # Clear the existing items in the listbox
                        listbox1.delete(0, END)
                        # Insert each item in the new data into the listbox
                        for tata in data:
                            listbox1.insert(END, tata)

                    def fillout(e):
                        # Clear the text in the entry box and insert the selected item into it
                        entrybox1.delete(0, END)
                        entrybox1.insert(0, listbox1.get(ANCHOR))

                    def check(e):
                        # Get the text in the first entry box
                        typed = entrybox1.get()
                        # Get the text in the third entry box
                        nuym = entrybox3.get()
                        if typed == "":
                            # If the first entry box is empty, show all items in the list
                            data = toppings[:]  # copy all items in the toppings list

                        else:
                            # If the first entry box is not empty, filter the items based on the text entered
                            data = []
                            for item in toppings:
                                if typed.lower() in item.lower():
                                    data.append(item)
                        # Update the listbox with the new data
                        update(data)


                    # Initialize the listbox with all the items in the toppings list
                    update(toppings)
                    # Bind the 'KeyRelease' event to the first entry box to filter the items in the listbox as the user types
                    entrybox1.bind("<KeyRelease>", check)
                    # Bind the 'ListboxSelect' event to the listbox to fill the first entry box with the selected item
                    listbox1.bind("<<ListboxSelect>>", fillout)

                    def Numbers_einfugen(event=None):
                        # Get the current delivery surcharge
                        entrybox1.config(state=NORMAL)
                        first=entrybox1.get()
                        if first!='':
                            cur13 = connE.cursor()
                            cur13.execute('select * from Lieferzuschlg ')
                            llieferzuschalg = cur13.fetchall()
                            if llieferzuschalg:
                                for izo1 in llieferzuschalg:
                                    lieferzuschalg = str(izo1).replace("'", "").replace(',', '').replace(')', '').replace('(', '')
                                    lieferzuschalg = float(lieferzuschalg)

                            # Get values from the input fields
                            solo = entrybox2.get()
                            item1 = (entrybox4.get())
                            komm = entrybox12.get()

                            # Check if a topping has been added
                            if solo != '':
                                sicko = messagebox.showinfo('Mit Zutat', 'Bitte den Zutat mit Enter Bestätigen')
                            else:
                                # Check if the quantity is a number and add data to the table
                                if (item1.isdigit()):
                                    drop.config(state=NORMAL)
                                    insertdatatree()

                                    # Clear and reset input fields
                                    entrybox1.config(state=NORMAL)
                                    entrybox4.delete(0, END)
                                    entrybox4.insert(0, '1')
                                    entrybox2.delete(0, END)
                                    entrybox1.delete(0, END)
                                    listbox3.delete(0, END)
                                    entrybox3.delete(0, END)
                                    entrybox5.delete(0, END)
                                    entrybox6.delete(0, END)



                                    entrybox10.delete(0, END)
                                    entrybox10.insert(0, float(0.0))

                                    entrybox12.delete(0, END)
                                    drop.configure(font=('Helvetica bold', 16),  width=8)
                                    drop.set('')
                                    entrybox1.focus_force()
                                    update(toppings)
                                    listbox2.delete(0, END)
                                    sopr = float(entrybox11.get())
                                    liso.clear()

                                    # Check if the delivery surcharge should be charged
                                    if AdresseE.get() != 'ABHOLUNG':
                                        if float(sopr) > float(lieferzuschalg):
                                            liefer1.config(fg='green')
                                        else:
                                            liefer1.config(fg='red')
                                else:
                                    tkinter.messagebox.showinfo("Error", "bei (Anzahl) ein nummer eingeben!!!")

                                    entrybox4.delete(0, END)
                        else:
                            messagebox.showwarning('leer','Speise Auswählen')
                    global einfugen

                    einfugen = Button(Mainframe2, text="Nächste", bg=colour3, bd=2, command=Numbers_einfugen, height=3)
                    einfugen.place(x=820, y=70)

                    ##################################Kunden liste Bestellung (functions)###################################################
                    def popup():
                        response = MessageBox.showinfo("!", 'kunde existiert nicht, Bitte neu speichern ')
                        Label(new_window2, text=response)

                    # ---------------------------------------------------------------------------------------------------------------------#
                    def reset():
                        KundenidE.delete(0, END)
                        NameE.delete(0, END)
                        AdresseE.delete(0, END)
                        TelefonnummerE.delete(0, END)
                        in_comment.delete('1.0', END)
                        ext_comment.delete('1.0', END)
                        EmailE.delete(0, END)
                        ORTE.delete(0, END)
                        HauesnrE.delete(0, END)
                        PLZE.delete(0, END)

                    # ------------------------------------------------------------------------------------------------------------------#

                    # Define a function to search for a customer's information using their phone number or ID

                    def search(event):
                        cur = connK.cursor()

                        ziko = TelefonnummerE.get()
                        tiko = KundenidE.get()
                        # If phone number is given, search by phone number
                        if ziko != '' and tiko == '':
                            cur.execute("select * from  kundendaten where Telefon=(?) ", (ziko,))
                            ar = cur.fetchall()

                            # Insert the customer information into the entry widgets
                            for sag in ar:
                                TelefonnummerE.delete(0, END)
                                KundenidE.delete(0, END)
                                NameE.delete(0, END)
                                AdresseE.delete(0, END)
                                in_comment.delete(1.0, END)
                                ext_comment.delete('1.0', END)
                                EmailE.delete(0, END)
                                ORTE.delete(0, END)
                                HauesnrE.delete(0, END)
                                PLZE.delete(0, END)
                                columns = [("kundenid", KundenidE), ("Name", NameE), ("Addresse", AdresseE),
                                           ("Telefon", TelefonnummerE), ("Nr", HauesnrE), ("PLZ", PLZE), ("ORT", ORTE),
                                           ("Email", EmailE), ("in_comment", in_comment)]

                                for i, (col_name, entry_widget) in enumerate(columns):
                                    if sag[i] is not None:
                                        entry_widget.insert(END, sag[i])

                        # If customer ID is given, search by customer ID
                        elif tiko != '':
                            cur.execute("select * from  kundendaten where Id=(?) ", (tiko,))
                            ar = cur.fetchall()

                            # Insert the customer information into the entry widgets
                            for sag in ar:
                                TelefonnummerE.delete(0, END)
                                KundenidE.delete(0, END)
                                NameE.delete(0, END)
                                AdresseE.delete(0, END)
                                in_comment.delete(1.0, END)
                                ext_comment.delete('1.0', END)
                                EmailE.delete(0, END)
                                ORTE.delete(0, END)
                                HauesnrE.delete(0, END)
                                PLZE.delete(0, END)
                                columns = [("kundenid", KundenidE), ("Name", NameE), ("Addresse", AdresseE),
                                           ("Telefon", TelefonnummerE), ("Nr", HauesnrE), ("PLZ", PLZE), ("ORT", ORTE),
                                           ("Email", EmailE), ("in_comment", in_comment)]

                                for i, (col_name, entry_widget) in enumerate(columns):
                                    if sag[i] is not None:
                                        entry_widget.insert(END, sag[i])

                            NameE.focus_force()

                        connK.commit()

                    def bind_function(event):
                        if einfugen["state"] == "normal":
                            Numbers_einfugen()
                    # Bind the search function to the return key in the phone number and customer ID entry widgets
                    TelefonnummerE.bind('<Return>', search)
                    KundenidE.bind('<Return>', search)
                    # Bind the numbers_einfugen function to the insert key
                    new_window2.bind('<KeyPress-Insert>', bind_function)
                    # Bind the "Return" key to the addzutat function
                    entrybox2.bind('<Return>', addzutat)
                    def jump_to_entrybox5(event):
                        entrybox2.focus()

                    entrybox4.bind('<Return>',jump_to_entrybox5)


                    ########################################################################################################################
                    def addDATA1():  # Define the function to add a new customer to the database
                        # Get the customer's phone number
                        maro = TelefonnummerE.get()
                        # Create a cursor for the kundendaten table and execute a query to get all phone numbers
                        custor = connK.cursor()
                        custor.execute('select Telefon from kundendaten')
                        # Fetch all phone numbers and store them in a list
                        nummber = [n[0] for n in custor.fetchall()]

                        # Check if any of the required fields are empty
                        if NameE.get() == "" or AdresseE.get() == "" or TelefonnummerE.get() == "" or HauesnrE.get() == "" or PLZE.get() == "" or ORTE.get() == "":
                            tkinter.messagebox.showerror('Error', 'alle felder sind pflichtfelder ')

                        # Check if the phone number already exists in the database
                        elif maro in nummber:
                            messagebox.showwarning('kunde exestiert', 'Kunder exestiert schon bitte auf update drucken')

                        # If all checks pass, insert the new customer data into the kundendaten table and commit the transaction
                        else:
                            custor.execute(
                                """insert into kundendaten
                                        (Name,Addresse,Telefon,Nr,PLZ,ORT,Email,int_comment)
                                         values(?,?,?,?,?,?,?,?)""",
                                (NameE.get(),
                                 AdresseE.get(),
                                 TelefonnummerE.get(),
                                 HauesnrE.get(),
                                 PLZE.get(),
                                 ORTE.get(),
                                 EmailE.get(),
                                 in_comment.get("1.0", END)
                                 ))
                            # Show a success message and commit the transaction
                            MessageBox.showinfo("!", 'kunde gespeichert')
                            custor.execute("commit")

                    ########################################################################################################################
                    def search1(event=None):
                        # create a cursor object
                        cur = connK.cursor()

                        # get the customer ID from the KundenidE entry widget
                        tiko = KundenidE.get()

                        # execute a SELECT query to find the customer with the given ID
                        if cur.execute("select * from  kundendaten where Id=(?) ", (tiko,)):
                            # if a customer with the given ID is found, retrieve their data and insert it into the entry widgets
                            ar = cur.fetchall()
                            for sag in ar:
                                KundenidE.delete(0, END)
                                NameE.delete(0, END)
                                AdresseE.delete(0, END)
                                TelefonnummerE.delete(0, END)
                                in_comment.delete('1.0', END)
                                ext_comment.delete('1.0', END)
                                EmailE.delete(0, END)
                                ORTE.delete(0, END)
                                HauesnrE.delete(0, END)
                                PLZE.delete(0, END)
                                columns = [("kundenid", KundenidE), ("Name", NameE), ("Addresse", AdresseE),
                                           ("Telefon", TelefonnummerE), ("Nr", HauesnrE), ("PLZ", PLZE), ("ORT", ORTE),
                                           ("Email", EmailE), ("in_comment", in_comment)]
                                for i, (col_name, entry_widget) in enumerate(columns):
                                    if sag[i] is not None:
                                        entry_widget.insert(END, sag[i])
                        else:
                            # if a customer with the given ID is not found, show an error popup
                            popup()
                        # commit the changes to the database
                        connK.commit()

                    # ---------------------------------------------------------------------------------------------------------------------#
                    def search3(event):
                        # create a cursor object
                        cur = connK.cursor()
                        # get the values of the name and customer id fields
                        ziko = NameE.get()
                        tiko = KundenidE.get()

                        # if the name field is not empty and customer id field is empty
                        if ziko != '' and tiko == '':
                            # execute a select query to retrieve the customer data with matching name
                            cur.execute("select * from kundendaten where Name=(?) ", (ziko,))
                            ar = cur.fetchall()
                            # insert the retrieved customer data into the corresponding entry widgets
                            for sag in ar:
                                TelefonnummerE.delete(0, END)
                                KundenidE.delete(0, END)
                                NameE.delete(0, END)
                                AdresseE.delete(0, END)
                                in_comment.delete(1.0, END)
                                ext_comment.delete('1.0', END)
                                EmailE.delete(0, END)
                                ORTE.delete(0, END)
                                HauesnrE.delete(0, END)
                                PLZE.delete(0, END)
                                columns = [("kundenid", KundenidE), ("Name", NameE), ("Addresse", AdresseE),
                                           ("Telefon", TelefonnummerE), ("Nr", HauesnrE), ("PLZ", PLZE), ("ORT", ORTE),
                                           ("Email", EmailE), ("in_comment", in_comment)]
                                for i, (col_name, entry_widget) in enumerate(columns):
                                    if sag[i] is not None:
                                        entry_widget.insert(END, sag[i])

                        # if the customer id field is not empty
                        elif tiko != '':
                            # execute a select query to retrieve the customer data with matching id
                            cur.execute("select * from  kundendaten where ID=(?) ", (tiko,))
                            ar = cur.fetchall()
                            # insert the retrieved customer data into the corresponding entry widgets
                            for sag in ar:
                                TelefonnummerE.delete(0, END)
                                KundenidE.delete(0, END)
                                NameE.delete(0, END)
                                AdresseE.delete(0, END)
                                in_comment.delete(1.0, END)
                                ext_comment.delete('1.0', END)
                                EmailE.delete(0, END)
                                ORTE.delete(0, END)
                                HauesnrE.delete(0, END)
                                PLZE.delete(0, END)
                                columns = [("kundenid", KundenidE), ("Name", NameE), ("Addresse", AdresseE),
                                           ("Telefon", TelefonnummerE), ("Nr", HauesnrE), ("PLZ", PLZE), ("ORT", ORTE),
                                           ("Email", EmailE), ("in_comment", in_comment)]
                                for i, (col_name, entry_widget) in enumerate(columns):
                                    if sag[i] is not None:
                                        entry_widget.insert(END, sag[i])

                        # commit the changes to the database
                        connK.commit()

                    ########################################################################################################################

                    def openwindow7():
                        # creates a new window
                        global new_window

                        font_size = ("ARIEL", 12, "bold")
                        new_window6 = Toplevel(root)
                        screen_width = new_window6.winfo_screenwidth()
                        screen_height = new_window6.winfo_screenheight()
                        x_cordinate = int((screen_width / 2) - (window_width / 2))
                        y_cordinate = int((screen_height / 2) - (window_height / 2))
                        new_window6.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
                        new_window6.config(bg='dark slate gray')
                        new_window6.resizable(width=False, height=False)
                        new_window6.configure(bg='SlateGray4')
                        new_window6.title(220 * titlespace + "Liste")
                        new_window6.resizable(width=False, height=False)
                        Frame1 = Frame(new_window6, width=550, bg='white', bd=4, height=1200, relief=RIDGE)
                        Frame1.place(x=0, y=0)

                    def update1():
                        # checks if all required fields are filled in
                        if NameE.get() == "" or AdresseE.get() == "" or TelefonnummerE.get() == "" or HauesnrE.get() == "" or PLZE.get() == "" or ORTE.get() == "":
                            tkinter.messagebox.showerror('Error', 'alle felder sind pflichtfelder ')
                        else:
                            miko = TelefonnummerE.get()
                            cur = connK.cursor()
                            # updates the customer data in the database
                            cur.execute(
                                """update  kundendaten  set Name=(?),Addresse=(?),int_comment=(?),Nr=(?),PLZ=(?),ORT=(?),Email=(?) where Telefon=(?)""",
                                (NameE.get(),
                                 AdresseE.get(),
                                 in_comment.get("1.0", END),
                                 HauesnrE.get(),
                                 PLZE.get(),
                                 ORTE.get(),
                                 EmailE.get(),
                                 TelefonnummerE.get()
                                 ))
                            connK.commit()
                            # displays a message box informing that the data has been updated
                            MessageBox.showinfo("!", 'Daten geandert')

                    ########################################################################################################################
                    def delete():
                        # delete customer info
                        cur = connK.cursor()
                        cur.execute("delete from  kundendaten where kundenid=%s", KundenidE.get())
                        connK.commit()
                        reset()

                    def printexcel1():
                        import win32com.client
                        actualpath = resource_path('Data\printer23.xlsx')
                        _path = os.path.abspath(actualpath)
                        excel = win32com.client.Dispatch('Excel.Application')
                        excel.Visible = False  # hide the Excel window
                        wb = excel.Workbooks.Open(_path)
                        wb.PrintOut()  # print the workbook
                        wb.Close(SaveChanges=False)  # close the workbook without saving changes
                        excel.Quit()

                    ####################################Kunden listeBestellung(Buttons)#####################################################
                    # btnsearch = Button(Mainframe1, font=('arial', 7, 'bold'), text='suche', bd=4, bg='red', pady=1,
                    #                          padx=24,
                    #                          width=4, height=1, command=search)
                    # btnsearch.place(x=175,y=340)
                    global text1
                    printframe = Frame(new_window2, width=60, height=60, bg='white')
                    scrolbar = Scrollbar(printframe)
                    scrolbar.pack(side=RIGHT, fill=Y)
                    btnspeicher = Button(Mainframe1, font=('arial', 10, 'bold'), text='Speichern', bd=4, bg=colour3, pady=1,
                                         padx=24,
                                         width=3, height=2, command=addDATA1)
                    btnspeicher.place(x=0, y=340)
                    text1 = Text(new_window2, bg='white', bd=0, font=('arial', 20,), wrap=WORD, width=250)
                    text2 = Text(new_window2, bg='white', bd=0, font=('arial', 20,), wrap=WORD, width=250)
                    btnsuche1 = Button(Mainframe1, font=('arial', 7, 'bold'), text='suche', bd=4, bg='red',
                                       pady=1,
                                       padx=24,
                                       width=4, height=1, command=search3)
                    KundenidE.bind('<Return>', search1)
                    NameE.bind('<Return>', search3)
                    btnsuche1.bind('<Return>', search3)
                    btnreset = Button(Mainframe1, font=('arial', 10, 'bold'), text='Reset', bd=4, bg=colour1, pady=1,
                                      padx=24,
                                      width=3, height=2, command=reset)
                    btnreset.place(x=90, y=340)

                    btndelet = Button(Mainframe1, font=('arial', 10, 'bold'), text='delete', bd=4, bg='red', pady=1,
                                      padx=24,
                                      width=3, height=2, command=delete)
                    btnupdate = Button(Mainframe1, font=('arial', 10, 'bold'), text='Update', bd=4, bg=colour2, pady=1,
                                       padx=24,
                                       width=3, height=2, command=update1)
                    btnupdate.place(x=180, y=340)
                    btnliste = Button(Mainframe1, font=('arial', 18), text="\u2399", bd=0, bg=colour4, pady=1,
                                      padx=24,
                                      width=1, command=printexcel1)
                    btnliste.place(x=50, y=187)
                    def lockin():

                        btnbearbeiten_show.config(state=DISABLED)
                        btnbearbeiten.config(state=DISABLED)
                        btnbearbeiten_unshow.config(state=DISABLED)
                    def unlockin():
                        Reset.config(state=NORMAL)
                        einfugen.config(state=NORMAL)
                        btnspeicher.config(state=NORMAL)
                        btnreset.config(state=NORMAL)
                        btnupdate.config(state=NORMAL)
                        btnLiefer.config(state=NORMAL)
                        btnAbholung.config(state=NORMAL)
                        btnImhaus.config(state=NORMAL)
                        btnbearbeiten_show.config(state=NORMAL)
                        btnbearbeiten.config(state=NORMAL)
                        btnbearbeiten_unshow.config(state=NORMAL)

                    # ---------------------------------------------------------- HIDE FUNCTION --------------------------------------------#
                    def hide():
                        global counts, text1, abholframe,print_Button,Bestellung_frame
                        if ido > 0:
                            entrybox8.delete(0, END)
                            entrybox8.insert(0, float(ido) * 100 / 100)
                        # Get customer information
                        tek = TelefonnummerE.get()
                        tik = NameE.get()
                        einfugen.config(state=NORMAL)
                        try:
                            print_Button.destroy()
                        except:
                            pass
                        try:
                            abholframe.destroy()
                        except:
                            pass
                        Bestellung_frame.destroy()
                        # Check if customer is in blacklist
                        lockin()
                        custor = connK.cursor()
                        custor1 = connK.cursor()
                        custor1.execute("select * from kundendatenB")
                        blacky = custor1.fetchall()
                        for tizi in blacky:
                            if tik in tizi:
                                messagebox.showwarning('Achtung', 'DIESE KUNDE STEHT AUF DIE BLACK LISTE')

                        # Check if customer exists, and prompt user to add if not
                        custor.execute('select Telefon from kundendaten')
                        conlo = custor.fetchall()
                        telefon = []
                        for holo in conlo:
                            salo = str(holo)
                            mako = salo.replace(',', '').replace("'", "").replace('(', '').replace(')', '')
                            telefon.append(mako)
                        if tek not in telefon and TelefonnummerE.get() != '' and AdresseE.get() != '':
                            saro = messagebox.askyesno('M2 Bestellsystem', 'Kunde exestiert nicht wollen Sie speichern? ')
                            if saro == 1:
                                addDATA1()

                        # Show address entry form
                        mou = AdresseE.get()
                        if mou != '':
                            counts = 1
                            Mainframe2.grid(row=0, column=2, sticky='n')
                            Mainframe3.place(x=310, y=350)
                            btnAbholung.config(state=DISABLED)
                            btnImhaus.config(state=DISABLED)
                            entrybox1.focus_set()
                        # Show message if address is not entered
                        else:
                            messagebox.showinfo('Adresse', 'Bitte Adresse eingeben!!')

                    def hide1():
                        global counts, text1, abholE, abholframe, abholz,print_Button,Bestellung_frame
                        try:
                            print_Button.destroy()
                        except:
                            pass
                        tek = TelefonnummerE.get()
                        tik = NameE.get()


                        # Check if the customer is in the black list
                        custor = connK.cursor()
                        custor1 = connK.cursor()
                        custor1.execute("select * from kundendatenB")
                        blacky = custor1.fetchall()
                        for tizi in blacky:
                            if tik in tizi:
                                messagebox.showwarning('Achtung', 'DIESE KUNDE STEHT AUF DIE BLACK LISTE')
                        Bestellung_frame.destroy()
                        # Check if the customer exists in the database, if not prompt to add
                        custor.execute('select Telefon from kundendaten')
                        conlo = custor.fetchall()
                        telefon = []

                        for holo in conlo:
                            salo = str(holo)
                            mako = salo.replace(',', '').replace("'", "").replace('(', '').replace(')', '')
                            telefon.append(mako)
                        if tek not in telefon and TelefonnummerE.get() != '' and AdresseE.get() != '':
                            saro = messagebox.askyesno('M2 Bestellsystem', 'Kunde exestiert nicht wollen Sie speichern? ')
                            if saro == 1:
                                addDATA1()
                        Mainframe2.grid(row=0, column=2, sticky='n')
                        Mainframe3.place(x=310, y=350)

                        # Configure widgets to disable customer details entry and enable pickup option

                        abholframe = Frame(Mainframe3, height=40, width=210, bg=colour4, bd=4, relief=RIDGE)
                        abholframe.place(x=630, y=0)
                        abholzeit = Label(abholframe, font=("Helvetica", 14, 'bold'), text='Abhol Zeit:', bg=colour4)
                        abholzeit.place(x=2, y=0)
                        abholE = Entry(abholframe, width=5, bd=4)
                        abholE.place(x=110, y=0)
                        abholzeitm = Label(abholframe, font=("Helvetica", 10, 'bold'), text='min', bg=colour4)
                        abholzeitm.place(x=160, y=3)

                        AdresseE.delete(0, END)
                        AdresseE.insert(0, 'ABHOLUNG')

                        HauesnrE.delete(0, END)

                        PLZE.delete(0, END)

                        ORTE.delete(0, END)

                        btnLiefer.config(state=DISABLED)
                        btnImhaus.config(state=DISABLED)

                        lockin()
                        entrybox1.focus_set()


                    def hide2():
                        # Get customer phone number and name
                        global counts, text1,abholframe,print_Button,Bestellung_frame

                        tek = TelefonnummerE.get()
                        tik = NameE.get()
                        try:
                            abholframe.destroy()

                        except:
                            pass
                        try:

                            print_Button.destroy()
                        except:
                            pass
                        # Check if customer is on black list
                        custor = connK.cursor()
                        custor1 = connK.cursor()
                        custor1.execute("select * from kundendatenB")
                        blacky = custor1.fetchall()
                        for tizi in blacky:
                            if tik in tizi:
                                messagebox.showwarning('Achtung', 'DIESE KUNDE STEHT AUF DIE BLACK LISTE')
                        Bestellung_frame.destroy()
                        # Get all phone numbers from customer data table
                        custor.execute('select Telefon from kundendaten')
                        conlo = custor.fetchall()
                        telefon = []
                        for holo in conlo:
                            salo = str(holo)
                            mako = salo.replace(',', '').replace("'", "").replace('(', '').replace(')', '')
                            telefon.append(mako)

                        # Disable customer info input fields and set default values for delivery
                        Mainframe2.grid(row=0, column=2, sticky='n')
                        Mainframe3.place(x=310, y=350)



                        AdresseE.delete(0, END)
                        AdresseE.insert(0, 'IM HAUS')

                        HauesnrE.delete(0, END)

                        PLZE.delete(0, END)

                        ORTE.delete(0, END)

                        btnLiefer.config(state=DISABLED)
                        btnAbholung.config(state=DISABLED)
                        liefer.config(state=DISABLED)
                        liefer1.config(state=DISABLED)
                        entrybox1.focus_set()
                        lockin()

                    # Buttons for  LIEFERUNG ABHLOUNG IMHAUS
                    btnLiefer = Button(Mainframe32, font=('arial', 10, 'bold'), text='Lieferung', bd=5, bg=colour3, pady=1,
                                     padx=24,
                                     width=3, height=4, highlightcolor="red", command=hide)
                    btnLiefer.place(x=1, y=10)
                    btnAbholung = Button(Mainframe32, font=('arial', 10, 'bold'), text='Abholung', bd=5, bg=colour1, pady=1,
                                      padx=24,
                                      width=3, height=4, command=hide1)
                    btnAbholung.place(x=96, y=10)
                    btnImhaus = Button(Mainframe32, font=('arial', 10, 'bold'), text='Im Haus', bd=5, bg=colour2, pady=1,
                                      padx=24,
                                      width=3, height=4, command=hide2)
                    btnImhaus.place(x=190, y=10)
                    def show_bestellung():
                        global count, pos,newtree,Bestellung_frame,y_scrollbar
                        global counts
                        Bestellung_frame = Frame(new_window2, bd=4, width=1200, height=900, relief=RIDGE,
                                                 bg=colour4)
                        Bestellung_frame.place(x=300, y=0)
                        count = 0
                        counts = 1
                        pos = 0

                        newtree = ttk.Treeview(Bestellung_frame, height=15)
                        # Create a vertical scrollbar
                        y_scrollbar = ttk.Scrollbar(Bestellung_frame, orient='vertical', command=newtree.yview)


                        # Set the scrollbar to control the Treeview's vertical scroll
                        newtree.configure(yscrollcommand=y_scrollbar.set)
                        newtree['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                        newtree.column("#0", width=0, stretch=NO)
                        newtree.column("Nr", anchor=CENTER, width=120, stretch=TRUE, )
                        newtree.column("Name", anchor=CENTER, width=120)
                        newtree.column("Datum/Uhrzeit", anchor=W, width=160)
                        newtree.column("Straße", anchor=W, width=120)
                        newtree.column("Haus/nr", anchor=W, width=120)
                        newtree.column("Preis", anchor=W, width=120)

                        newtree.heading("#0", text="", anchor=W)
                        newtree.heading("Nr", text="Nr", anchor=W)
                        newtree.heading("Name", text="Name", anchor=W)
                        newtree.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                        newtree.heading("Straße", text="Straße", anchor=W)
                        newtree.heading("Haus/nr", text="nr", anchor=W)
                        newtree.heading("Preis", text="Preis", anchor=W)
                        newtree.tag_configure('even', background='#E8F6F3')  # Set background color for even rows
                        newtree.tag_configure('odd', background='#D2E6E3')
                    def show_hidden():
                        global newtree,y_scrollbar,Bestellung_frame
                        for widget in Bestellung_frame.winfo_children():
                            widget.destroy()
                        FIRSTLABEL = Label(Bestellung_frame, text='auf Lieferung oder Abholung drücken',
                                           font=("ARIEL", 26, "bold"),
                                           bg=colour4)
                        FIRSTLABEL.place(x=265, y=400)
                        newtree = ttk.Treeview(Bestellung_frame, height=15)
                        # Create a vertical scrollbar
                        y_scrollbar = ttk.Scrollbar(Bestellung_frame, orient='vertical', command=newtree.yview)


                        # Set the scrollbar to control the Treeview's vertical scroll
                        newtree.configure(yscrollcommand=y_scrollbar.set)
                        newtree['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                        newtree.column("#0", width=0, stretch=NO)
                        newtree.column("Nr", anchor=CENTER, width=120, stretch=TRUE, )
                        newtree.column("Name", anchor=CENTER, width=120)
                        newtree.column("Datum/Uhrzeit", anchor=W, width=160)
                        newtree.column("Straße", anchor=W, width=120)
                        newtree.column("Haus/nr", anchor=W, width=120)
                        newtree.column("Preis", anchor=W, width=120)

                        newtree.heading("#0", text="", anchor=W)
                        newtree.heading("Nr", text="Nr", anchor=W)
                        newtree.heading("Name", text="Name", anchor=W)
                        newtree.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                        newtree.heading("Straße", text="Straße", anchor=W)
                        newtree.heading("Haus/nr", text="nr", anchor=W)
                        newtree.heading("Preis", text="Preis", anchor=W)
                        newtree.tag_configure('even', background='#E8F6F3')  # Set background color for even rows
                        newtree.tag_configure('odd', background='#D2E6E3')
                        newtree.place(x=0,y=450)
                        y_scrollbar.place(x=750, y=450, height=330)



                    #################################### change the order
                    # Get data from SQLite database
                        pos=0
                        county=0
                        curd = connO.cursor()
                        curd.execute('SELECT * FROM kundeinfo  ORDER BY ID DESC')
                        rows = curd.fetchall()

                        for row in rows:


                        # Create a list to store selected names
                            if pos %2 == 0:
                                newtree.insert(parent='', open=False, index='end', iid=county, text='',
                                                   values=(row[0], row[3], row[10],
                                                           row[4], row[5],row[13]
                                                           ), tags=('odd',))
                                county+=1


                            else:
                                newtree.insert(parent='', open=False, index='end', iid=county, text='',
                                                   values=(row[0], row[3], row[10],
                                                           row[4], row[5],row[13]
                                                           ), tags=('even',))

                                county += 1
                                pos += 1
                    show_bestellung()
                    def save_name(event=None):
                        global counts, pos, count, print_Button,newtree,bearbeiten

                        KundenidE.delete(0, END)
                        NameE.delete(0, END)
                        AdresseE.delete(0, END)
                        HauesnrE.delete(0, END)
                        PLZE.delete(0, END)
                        ORTE.delete(0, END)
                        TelefonnummerE.delete(0, END)
                        EmailE.delete(0, END)
                        zeit=[]
                        bestellnumber=[]

                        curItem = newtree.focus()

                        values=newtree.item(curItem) ['values']





                        name=values[1]
                        time_str = values[2]

                        # Convert the time string to a datetime object (optional)
                        curd = connO.cursor()

                        curd.execute('select * from kundeinfo where kname =(?) and zeit=(?)', (name, time_str))
                        all_info = curd.fetchall()
                        for info in all_info:
                            bestellnumber.insert(0,info[0])
                            KundenidE.insert(0, info[1])
                            NameE.insert(0, info[3])
                            AdresseE.insert(0, info[4])
                            HauesnrE.insert(0, info[5])
                            PLZE.insert(0, info[6])
                            ORTE.insert(0, info[7])
                            TelefonnummerE.insert(0, '0' + str(info[2]))
                            EmailE.insert(0, info[8])
                            zeit.insert(0,info[10])
                            LieferE.insert(0, info[11])
                            ext_comment.insert(1.0,info[12])


                            if info[4]=='ABHOLUNG':
                                hide1()
                            elif info[4]=='IM HAUS':
                                hide2()
                            else:
                                hide()
                            counts = len(BestellTree.get_children()) + 1
                            pos = 0
                            count = 0
                            curd.execute('select * from speiseinfo where zeit =(?) ', (time_str,))

                            speisen = curd.fetchall()
                            curd.execute('select geld from lieferpara where zeit=(?)',(time_str,))
                            liefer_speisen=curd.fetchall()
                            for speise in speisen:

                                BestellTree.insert(parent='', open=False, index='end', iid=counts, text='',
                                                   values=(speise[1], speise[2], speise[3],
                                                           speise[4], speise[5],
                                                           speise[6], speise[7], speise[8], speise[9]))
                                counts += 1
                                pos += 1
                                count += 1
                            if liefer_speisen:
                                lieferspeise=(str(liefer_speisen[0]).replace('(','').replace(')','').replace("'","").replace(',',''))
                                entrybox8.delete(0,END)
                                entrybox8.insert(0,lieferspeise)
                                entrybox11.delete(0, END)
                                result1 = (float(info[13]-float(lieferspeise)))
                                entrybox11.insert(0, result1)
                            else:
                                entrybox11.delete(0, END)
                                result1 = (float(info[13]))
                                entrybox8.delete(0,END)
                                entrybox8.insert(0,'0.0')
                                entrybox11.insert(0,result1)
                            curds=connO.cursor()
                            curds.execute('select * from Rabattpara where zeit =(?) ',(time_str,))
                            rabatt=curds.fetchone()
                            if rabatt:
                                rab=rabatt[0]
                                entrybox10.delete(0,END)
                                entrybox10.insert(0,rab)


                        def Änderendrucken():
                            curd = connO.cursor()
                            curd.execute('delete from speiseinfo where zeit =(?)', (zeit[0],))
                            connO.commit()

                            # Loop through the items in the treeview and insert them into the speiseinfo table

                            # Get values from all items in the treeview
                            values = [[BestellTree.item(item)['values'][i] for i in range(9)]
                                      for item in BestellTree.get_children()]

                            # Create a cursor for the database connection

                            # Do something with the values, for example:
                            for sick100, sick101, sick102, sick103, sick104, sick105, sick106, sick107, sick108 in values:
                                cur1 = connO.cursor()
                                # Insert the item into the speiseinfo table
                                cur1.execute("""INSERT INTO speiseinfo (zeit, pos, grosse, anzahl, nr, speise, mit, ohne, katagorie, preis, name) 
                                                                   VALUES (:zeit, :pos, :grosse, :anzahl, :nr, :speise, :mit, :ohne, :katagorie, :preis, :name)""",
                                             {'zeit': zeit[0],
                                              'pos': sick100,
                                              'grosse': sick101,
                                              'anzahl': sick102,
                                              'nr': sick103,
                                              'speise': sick104,
                                              'mit': sick105 + ' ',
                                              'ohne': sick106,
                                              'katagorie': sick107,
                                              'preis': sick108,
                                              'name': NameE.get()
                                              })
                                connO.commit()
                            sick15 = entrybox11.get()

                            sicked = BestellTree.get_children()
                            mab = [BestellTree.item(tom)['values'][7] for tom in sicked]
                            cur = connE.cursor()
                            cur.execute('select* from Katagorie ')
                            fur = cur.fetchall()

                            # Create a set of the categories
                            categories = {cat[0] for cat in fur}
                            alle = []
                            # Get all the categories
                            for kat in fur:
                                top = str(kat)
                                replace = {'[': '',
                                           ']': '',
                                           '(': '',
                                           ')': '',
                                           ',': '',
                                           "'": ''}
                                tops = top.translate(str.maketrans(replace))
                                alle.insert(0, tops)
                            sick1 = KundenidE.get()
                            sick2 = TelefonnummerE.get()
                            sick3 = NameE.get()
                            sick4 = AdresseE.get()
                            sick5 = HauesnrE.get()
                            sick6 = PLZE.get()
                            sick7 = ORTE.get()
                            sick8 = EmailE.get()
                            sick9 = passo[1]
                            comment = ext_comment.get(1.0, END)

                            # Clear the text widget
                            # text1.delete(1.0, END)

                            # Check if there is a comment
                            comment1 = len(comment)

                            from openpyxl import Workbook, load_workbook
                            from openpyxl.styles import Font
                            from openpyxl.styles import PatternFill
                            from openpyxl.utils import get_column_letter

                            # Use the `os` module to delete the workbook
                            try:
                                file_path = resource_path('Data\printer23.xlsx')
                                os.remove(file_path)
                            except:
                                pass
                            wa = load_workbook(resource_path('Data\printer2.xlsx'))

                            # Save the workbook as a new file
                            wa.save(resource_path('Data\printer23.xlsx'))

                            # Create an instance of the workbook
                            wb = load_workbook(resource_path('Data\printer23.xlsx'))
                            ws = wb.active

                            # ws.delete_cols(1,5)
                            # ws.delete_rows(1, 100)
                            ws.column_dimensions['C'].width = 12
                            ws.column_dimensions['D'].width = 5
                            row_num = 2
                            # Define the new height (in this example, 30)
                            new_height = 15
                            # Set the row height for the specified row number
                            wb.save(resource_path('Data\printer23.xlsx'))
                            # Select the active sheet
                            # Save the changes to the workbook
                            # cell1=ws['A1']
                            # cell2=ws['B2']
                            # cell3=ws['A3']
                            # cell4 = ws['A4']
                            # cell5 = ws['A5']
                            # cell6 = ws['A6']
                            # cell7 = ws['A7']
                            # cell8 = ws['B4']
                            # cell9 = ws['B7']
                            # cell10 = ws['A8']
                            if sick4 == 'ABHOLUNG':
                                if abholE.get() == '':
                                    abhol_z = 'Sofort'
                                else:
                                    abhol_z = abholE.get()
                            # # Define the new height (in this example, 30)
                            new_height = 5
                            # # Set the row height for the specified row number
                            ws.row_dimensions[row_num].height = new_height
                            # # Get the row object for the specified row number

                            # Define the color to be used (in this example, yellow)
                            fill_color = PatternFill(start_color="FF0F00", end_color="FF0F00", fill_type="solid")
                            light_fill_color = PatternFill(start_color="FF9F60", end_color="FF9F60", fill_type="solid")
                            # Iterate through each cell in the row
                            smallfont = Font(name='Calibri', size=11, italic=True)
                            nfont = Font(name='Calibri', size=12, bold=True, italic=True)
                            font = Font(name='Calibri', size=18, bold=True, italic=True)
                            add_font = Font(name='Calibri', size=16, bold=True, italic=True)
                            font1 = Font(name='Calibri', size=25, bold=True, italic=True)
                            ufont = Font(underline='single')
                            # Write the first row with the headers
                            if sick4 == 'ABHOLUNG':
                                import datetime
                                now = datetime.datetime.now()

                                # Add 20 minutes
                                if abhol_z.isdigit():
                                    future_time = now + datetime.timedelta(minutes=int(abhol_z))

                                    # Convert the future time to a string in a specific format
                                    formatted_time = future_time.strftime("%H:%M")
                                else:
                                    formatted_time = abhol_z

                                ws.oddHeader.center.text = (f'{sick4}    ZU {formatted_time}')
                                ws.oddHeader.center.size = 25
                                ws.oddHeader.center.font = "Tahoma,Bold"
                                ws.oddHeader.center.color = "FF0F00"
                            else:
                                ws.oddHeader.center.text = (f'LIEFERUNG\nA{bestellnumber[0]}')
                                ws.oddHeader.center.size = 25
                                ws.oddHeader.center.font = "Tahoma,Bold"
                                ws.oddHeader.center.color = "FF0F00"
                            next_row = ws.max_row + 3
                            # ws['B{}'.format(next_row)] = f'      A{hub}'
                            # NUM_row = ws[f'B{next_row}']
                            # next_row = ws.max_row +1
                            #
                            # NUM_row.font = add_font

                            ws['A{}'.format(next_row)] = f'\U0001F550  Nach Druck{zeit[0]}    '
                            next_row = ws.max_row + 1
                            ws['A{}'.format(next_row)] = f'kd:{sick1}'
                            ws['B{}'.format(next_row)] = f' \u260E:{sick2}'
                            next_row = ws.max_row + 1
                            ws['A{}'.format(next_row)] = f'Name:  {sick3}'
                            add = f'{sick4} {sick5}'
                            next_row = ws.max_row + 1

                            wrapped_add = textwrap.wrap(add, width=25)
                            for line in wrapped_add:
                                ws['A{}'.format(next_row)] = line
                                add_row = ws[f'A{next_row}']
                                add_row.font = add_font
                                next_row = ws.max_row + 1
                            next_row = ws.max_row + 1
                            ws['A{}'.format(next_row)] = f'{sick6}   {sick7}'
                            add_row_1 = ws[f'A{next_row}']
                            add_row_1.font = add_font
                            next_row = ws.max_row + 1
                            sick12 = LieferE.get()
                            if sick12 != '':

                                next_row = ws.max_row + 1
                                ws['B{}'.format(next_row)] = 'VORBESTELLUNG'
                                vorbestellen_row = ws[f'A{next_row}']
                                vorbestellen_row1 = ws[f'B{next_row}']
                                vorbestellen_row2 = ws[f'C{next_row}']
                                vorbestellen_row3 = ws[f'D{next_row}']
                                vorbestellen_row.fill = light_fill_color
                                vorbestellen_row1.fill = light_fill_color
                                vorbestellen_row2.fill = light_fill_color
                                vorbestellen_row3.fill = light_fill_color

                                vorbestellen_row.font = font
                                next_row = ws.max_row + 1

                                ws['B{}'.format(next_row)] = (f'ZU  {sick12}  UHR LIEFERN ')
                                liefern_row = ws[f'A{next_row}']
                                liefern_row1 = ws[f'B{next_row}']
                                liefern_row2 = ws[f'C{next_row}']
                                liefern_row3 = ws[f'D{next_row}']
                                liefern_row.fill = light_fill_color
                                liefern_row1.fill = light_fill_color
                                liefern_row2.fill = light_fill_color
                                liefern_row3.fill = light_fill_color

                                next_row = ws.max_row + 2
                            elif sick12 != '' and sick4 == 'ABHOLUNG' :

                                next_row = ws.max_row + 1
                                ws['B{}'.format(next_row)] = 'VORBESTELLUNG'
                                vorbestellen_row = ws[f'A{next_row}']
                                vorbestellen_row1 = ws[f'B{next_row}']
                                vorbestellen_row2 = ws[f'C{next_row}']
                                vorbestellen_row3 = ws[f'D{next_row}']
                                vorbestellen_row.fill = light_fill_color
                                vorbestellen_row1.fill = light_fill_color
                                vorbestellen_row2.fill = light_fill_color
                                vorbestellen_row3.fill = light_fill_color

                                vorbestellen_row.font = font
                                next_row = ws.max_row + 1

                                ws['B{}'.format(next_row)] = (f'ZU  {sick12}  UHR ABHOLEN')
                                liefern_row = ws[f'A{next_row}']
                                liefern_row1 = ws[f'B{next_row}']
                                liefern_row2 = ws[f'C{next_row}']
                                liefern_row3 = ws[f'D{next_row}']
                                liefern_row.fill = light_fill_color
                                liefern_row1.fill = light_fill_color
                                liefern_row2.fill = light_fill_color
                                liefern_row3.fill = light_fill_color

                                next_row = ws.max_row + 2
                            if comment1 > 1:
                                # Write the title of the comment section to the next row

                                next_row = ws.max_row + 1
                                ws['A{}'.format(next_row)] = 'Kunden Komentare'

                                komment_row1 = ws[f'A{next_row}']
                                komment_row1.font = font

                                next_row = ws.max_row + 1
                                # Split the comment into multiple lines
                                wrapped_comment = textwrap.wrap(comment, width=35)

                                # Write each line of the comment to a separate row
                                for line in wrapped_comment:
                                    ws['A{}'.format(next_row)] = line
                                    next_row = ws.max_row + 2

                            # Update the row number for the next record
                            row_num = ws.max_row + 2

                            row_num = next_row

                            ws.row_dimensions[row_num].height = new_height
                            EIGHTROW = ws[f'A{next_row}']
                            EIGHTROW1 = ws[f'B{next_row}']
                            EIGHTROW2 = ws[f'C{next_row}']
                            EIGHTROW3 = ws[f'D{next_row}']

                            mit = []
                            ohne = []
                            next_row = ws.max_row + 1
                            lenalle = len(alle)

                            # loop through the speisen and add every katagorie
                            if lenalle > 0:
                                for i in range(lenalle):
                                    if alle[i] in mab:
                                        sob = mab.count(alle[i])

                                        ws['B{}'.format(next_row)] = (f'--------------------------')
                                        next_row += 1
                                        katrow1 = ws[next_row]
                                        katrow1[0].font = nfont
                                        ws['A{}'.format(next_row)] = (f'------{alle[i]}----{sob}--')
                                        next_row += 1
                                    sicked = BestellTree.get_children()
                                    for tom in sicked:
                                        values = BestellTree.item(tom)['values']
                                        sick100 = values[0]
                                        sick101 = values[1]
                                        sick102 = values[2]
                                        sick103 = values[3]
                                        sick104 = values[4]
                                        sick105 = values[5]

                                        sick106 = values[6]
                                        sick107 = values[7]
                                        sick108 = values[8]


                                        if sick107 == alle[i]:

                                            def truncate_text(text, max_length):
                                                if len(text) > max_length:
                                                    return text[:max_length]
                                                else:
                                                    return text

                                            max_length = 38  # Define your maximum length here
                                            max_length1 = 5

                                            if sick101 != 'Standard':
                                                katrow1 = ws[next_row]
                                                katrow1[0].font = nfont
                                                cell_content = f'{sick102}x({sick103} {sick101}){sick104} '
                                                ws['A{}'.format(next_row)] = truncate_text(cell_content, max_length)
                                                cell_content1 = f'{sick108}€'
                                                ws['D{}'.format(next_row)] = truncate_text(cell_content1, max_length1)

                                                katrow101 = ws[next_row]
                                                katrow101[1].font = smallfont
                                                next_row += 1





                                            elif float(sick108) > 0:
                                                katrow1 = ws[next_row]
                                                katrow1[0].font = nfont
                                                ws['A{}'.format(next_row)] = (f' {sick102}x ({sick103}) {sick104}')
                                                cell_content1 = f'{sick108}€'
                                                ws['D{}'.format(next_row)] = truncate_text(cell_content1, max_length1)

                                                next_row += 1


                                            elif float(sick108) == 0.00:
                                                    katrow101 = ws[next_row]
                                                    katrow101[1].font = smallfont
                                                    ws['A{}'.format(next_row)] = (
                                                        f'              {sick104}     ')

                                            if len(sick106) > 0:
                                                for without in [sick106]:
                                                    within = str(without).split(' ')
                                                ohne.append(within)
                                                for tim in ohne:
                                                    for taza in tim:
                                                        ws['B{}'.format(next_row)] = (f'  {taza}')
                                                        next_row += 1
                                                        ohne.clear()

                                            if sick105 != "":
                                                hozo = [s for s in sick105.split(' ') if s.strip()]
                                                mit.append(hozo)
                                                for azo in mit:
                                                    for tizo in azo:
                                                        print(tizo)
                                                        if sick104 == 'Familien Pizza':
                                                            unwraped_Familie = tizo
                                                            wrapped_Familie = textwrap.wrap(unwraped_Familie, width=38)
                                                            for line in wrapped_Familie:
                                                                ws['A{}'.format(next_row)] = (f'{line}')
                                                                next_row += 1
                                                                mit.clear()
                                                        else:
                                                            for taza in [tizo]:


                                                                ws['B{}'.format(next_row)] = (f'{taza}')
                                                                next_row += 1
                                                                mit.clear()
                            row_num = next_row
                            row = ws[row_num]
                            # Define the color to be used (in this example, yellow)
                            ws.row_dimensions[row_num].height = new_height
                            # Iterate through each cell in the row
                            for cell in row:
                                cell.fill = fill_color
                            next_row += 1
                            katrow5 = ws[next_row]
                            katrow5[1].font = nfont
                            if entrybox8.get()!=0.0:
                                formatted_gesamt = '{:.2f}'.format(float(entrybox11.get())+float(entrybox8.get()))
                            else:
                                formatted_gesamt = '{:.2f}'.format(float(entrybox11.get()))
                            ws['B{}'.format(next_row)] = (f"\n\t Gesamtpreis: {formatted_gesamt}€")
                            next_row += 1
                            dold = AdresseE.get()
                            lofro=entrybox8.get()
                            if lofro=='':
                                lofro=0.0

                            if dold.strip() != 'ABHOLUNG' and dold.strip() != 'IM HAUS':

                                try:
                                    if dif:
                                        formatted_dif = '{:.2f}'.format(dif)
                                        ws['A{}'.format(next_row)] = (f"\n\tLieferzuschlag: {formatted_dif}€")
                                        next_row += 1
                                except:
                                    pass
                                print(lofro)
                                if float(lofro) > 0 :
                                    ws['A{}'.format(next_row)] = (
                                        f"\n\tincl Liefergeld: {lofro} €")
                                    next_row += 1
                                if entrybox10.get()!='0.0':
                                    print(entrybox10.get())
                                    ws['A{}'.format(next_row)] = (f"\n incl Rabatt \t{entrybox10.get()}%")
                                    next_row += 1
                                    row_num = next_row
                                    row = ws[row_num]

                                # if float(sazo12) > 0 and 'selected' not in first:
                                #     # katrow3 = ws[next_row]
                                #     # katrow3[0].font = nfont
                                #     formatted_dif = '{:.2f}'.format(float(sazo12))
                                #     ws['A{}'.format(next_row)] = (
                                #         f"\n\tincl Liefergeld: {str(sazo12).replace('[', '').replace(']', '')}€")
                                #     next_row += 1
                            mwst = (7 * float(formatted_gesamt)) / 100.0
                            mwst1 = (math.floor(float(mwst) * 100) / 100)
                            mwst_haus = (19 * float(formatted_gesamt)) / 100.0
                            mwst1_haus = (math.floor(float(mwst) * 100) / 100)
                            # katrow4 = ws[next_row]
                            # katrow4[0].font = nfont
                            if sick4 != 'IM HAUS':
                                ws['A{}'.format(next_row)] = (f"\n incl 7%Umsatzsteuer :\t{mwst1}€")
                                next_row += 1
                                row_num = next_row
                                row = ws[row_num]
                            else:
                                ws['A{}'.format(next_row)] = (f"\n incl 19%Umsatzsteuer :\t{mwst1_haus}€")
                                next_row += 1
                                row_num = next_row
                                row = ws[row_num]
                            # Define the color to be used (in this example, yellow)
                            ws.row_dimensions[row_num].height = new_height
                            # Iterate through each cell in the row
                            for cell in row:
                                cell.fill = fill_color
                            next_row += 1

                            # Write each line of the comment to a separate row

                            # Write each line of the comment to a separate row

                            ws['A{}'.format(next_row)] = (f'--------------------------------------------')
                            next_row = ws.max_row + 1
                            ws['B{}'.format(next_row)] = (f'{restName}')
                            next_row = ws.max_row + 1
                            ws['B{}'.format(next_row)] = (f'{restStr}')
                            next_row = ws.max_row + 1
                            ws['B{}'.format(next_row)] = (f'{restTele}')
                            next_row = ws.max_row + 1
                            ws['A{}'.format(next_row)] = (f'--------------------------------------------')
                            next_row = ws.max_row + 1
                            katrow1 = ws[next_row]
                            katrow1[1].font = add_font
                            ws['B{}'.format(next_row)] = ('Lieferschein')

                            EIGHTROW.fill = fill_color
                            EIGHTROW1.fill = fill_color
                            EIGHTROW2.fill = fill_color
                            EIGHTROW3.fill = fill_color
                            curd = connO.cursor()
                            curd.execute('UPDATE kundeinfo  set  gesamtepreis =(?) where zeit=(?)', (formatted_gesamt, zeit[0]))
                            rows = curd.fetchall()
                            wb.save(resource_path('Data\printer23.xlsx'))
                            print_Button.destroy()
                            # os.remove("output.txt")
                            remove_all()


                            show_bestellung()
                            new_window2.update()

                            def printexcel():
                                default_printer = win32print.GetDefaultPrinter()

                                time.sleep(0.5)
                                # win32print.SetDefaultPrinter(default_printer)
                                import win32com.client
                                actualpath = resource_path('Data\printer23.xlsx')
                                _path = os.path.abspath(actualpath)
                                excel = win32com.client.Dispatch('Excel.Application')
                                excel.Visible = False  # hide the Excel window
                                wb = excel.Workbooks.Open(_path)
                                wb.PrintOut()  # print the workbook
                                wb.Close(SaveChanges=False)  # close the workbook without saving changes
                                excel.Quit()

                            printexcel()
                            entrybox10.delete(0,END)
                            entrybox10.insert(0,'0.0')
                            unlockin()
                            FIRSTLABEL = Label(Bestellung_frame, text='auf Lieferung oder Abholung drücken',
                                               font=("ARIEL", 26, "bold"),
                                               bg=colour4)
                            FIRSTLABEL.place(x=265, y=400)

                        print_Button = Button(endframe, text=' Änderen', width=12, bg=colour3, height=2,
                                                  command=Änderendrucken)
                        print_Button.pack()

                        # insert price with voucher into price entry box
                    def unshow_bestellung():
                        global newtree,county
                        try:
                            newtree.delete(*newtree.get_children())
                            newtree.place_forget()
                            y_scrollbar.place_forget()

                            new_window2.update()

                        except:
                            pass

                    btnbearbeiten=Button(Mainframe1,text='Bearbeiten',bg=colour3,width=10,command=save_name)
                    btnbearbeiten.place(x=205, y=730)
                    btnbearbeiten_löschen = Button(Mainframe1, text='Löschen', bg=colour1, width=10, state=DISABLED)
                    btnbearbeiten_löschen.place(x=205, y=770)
                    btnbearbeiten_Drucken = Button(Mainframe1, text='Drucken', bg=colour2, width=10, state=DISABLED)
                    btnbearbeiten_Drucken.place(x=205, y=810)
                    btnbearbeiten_show = Button(Mainframe1, text='Liste', bg=colour0, width=10,command=show_hidden,font=('arial',15,'bold') )
                    btnbearbeiten_show.place(x=5, y=730)
                    btnbearbeiten_unshow = Button(Mainframe1, text='Schliessen', bg=colour0, width=10, command=unshow_bestellung,
                                             font=('arial', 15, 'bold'))
                    btnbearbeiten_unshow.place(x=5, y=780)
                    FIRSTLABEL = Label(Bestellung_frame, text='auf Lieferung oder Abholung drücken',
                                       font=("ARIEL", 26, "bold"),
                                       bg=colour4)
                    FIRSTLABEL.place(x=265, y=400)
                    def lockall():
                        KundenidE.delete(0, END)
                        TelefonnummerE.delete(0, END)
                        NameE.delete(0, END)
                        AdresseE.delete(0, END)
                        HauesnrE.delete(0, END)
                        PLZE.delete(0, END)
                        ORTE.delete(0, END)
                        EmailE.delete(0, END)
                        in_comment.delete('1.0', END)
                        ext_comment.delete('1.0', END)
                        KundenidE.config(state=DISABLED)
                        TelefonnummerE.config(state=DISABLED)
                        NameE.config(state=DISABLED)
                        AdresseE.config(state=DISABLED)
                        HauesnrE.config(state=DISABLED)
                        PLZE.config(state=DISABLED)
                        ORTE.config(state=DISABLED)
                        EmailE.config(state=DISABLED)
                        in_comment.config(state=DISABLED)
                        ext_comment.config(state=DISABLED)

                        btnspeicher.config(state=DISABLED)
                        btnreset.config(state=DISABLED)
                        btnupdate.config(state=DISABLED)
                        btnLiefer.config(state=DISABLED)
                        btnAbholung.config(state=DISABLED)
                        btnImhaus.config(state=DISABLED)
                        btnbearbeiten_show.config(state=DISABLED)
                        btnbearbeiten.config(state=DISABLED)
                        btnbearbeiten_unshow.config(state=DISABLED)
                    def unlockall():
                        for widget in Bestellung_frame.winfo_children():
                            widget.destroy()
                        FIRSTLABEL = Label(Bestellung_frame, text='auf Lieferung oder Abholung drücken',
                                           font=("ARIEL", 26, "bold"),
                                           bg=colour4)
                        FIRSTLABEL.place(x=265, y=400)
                        KundenidE.config(state=NORMAL)
                        TelefonnummerE.config(state=NORMAL)
                        NameE.config(state=NORMAL)
                        AdresseE.config(state=NORMAL)
                        HauesnrE.config(state=NORMAL)
                        PLZE.config(state=NORMAL)
                        ORTE.config(state=NORMAL)
                        EmailE.config(state=NORMAL)
                        in_comment.config(state=NORMAL)
                        ext_comment.config(state=NORMAL)

                        btnspeicher.config(state=NORMAL)
                        btnreset.config(state=NORMAL)
                        btnupdate.config(state=NORMAL)
                        btnLiefer.config(state=NORMAL)
                        btnAbholung.config(state=NORMAL)
                        btnImhaus.config(state=NORMAL)
                        btnbearbeiten_show.config(state=NORMAL)
                        btnbearbeiten.config(state=NORMAL)
                        btnbearbeiten_unshow.config(state=NORMAL)

                    def closepopup():
                        top.destroy()

                    # ------------------------------------------------------------------------------------------------------------------#
                    def passwortget(event=None):

                        cur = connE.cursor()
                        cur.execute('select* from Passwort ')
                        Passwort = cur.fetchall()
                        Passwort1 = str(Passwort)
                        Passwort2 = Passwort1.replace("(", "").replace(")", "").replace(",", "").replace("'",
                                                                                                         "").replace(
                            "[",
                            "").replace(
                            "]", "")
                        connE.commit()
                        Password = Passwort2
                        passwordE = entery.get()
                        if passwordE == Password:
                            Fieramt()
                            closepopup()


                        else:
                            closepopup()

                    # ----------------------------------------------------------------------------------------------------------------------#
                    def popupchef(event=None):
                        global top
                        top = Toplevel(new_window2)
                        w13 = 350
                        h13 = 100
                        # Get the width and height of the screen
                        screen_width = root.winfo_screenwidth()
                        screen_height = root.winfo_screenheight()
                        # Calculate the x and y coordinates for the top-left corner of the Toplevel window
                        x = (screen_width - w13) // 2
                        y = (screen_height - h13) // 2
                        top.geometry(f"{w13}x{h13}+{x}+{y}")
                        top.config(bg=colour4)
                        top.title('CHEF Passwort')
                        global entery
                        entery = Entry(top, show="*", width=25)
                        label1 = Label(top, text="Passwort eingeben", font=("Ariel", 12), bg=colour4)
                        entery.focus_force()
                        label1.pack()
                        entery.pack()

                        def password_enter(e):
                            passwortget()

                        # ----------------------------------------------------------------------------------------------------------------------#

                        Button1 = Button(top, text="Ok", width=10, command=passwortget)
                        Button2 = Button(top, text="Quit", width=10, command=closepopup)
                        Button1.place(x=40, y=60)
                        Button2.place(x=185, y=60)
                        top.bind("<Return>", passwortget)


                    def Fieramt():
                        global new_window, new_window2, counter,Bestellung_frame
                        global kundenframe2

                        for widget in Bestellung_frame.winfo_children():
                            widget.destroy()
                        lockall()
                        # new_window6.attributes('-fullscreen', True)
                        Frame6 = Frame(Bestellung_frame , width=300, bg=colour4, bd=4, height=100, relief=RIDGE)
                        Frame6.place(x=0, y=684)
                        Frame2 = Frame(Bestellung_frame , width=450, bg='grey', bd=4, height=400, relief=RIDGE)
                        Frame2.place(x=0, y=60)
                        Frame3 = Frame(Bestellung_frame , width=560, bg='grey', bd=4, height=350, relief=RIDGE)
                        Frame3.place(x=600, y=60)
                        Frame4 = Frame(Bestellung_frame , width=450, bg='grey', bd=4, height=400, relief=RIDGE)
                        Frame4.place(x=0, y=500)
                        Frame7 = Frame(Bestellung_frame, width=450, bg='grey', bd=4, height=400, relief=RIDGE)
                        Frame7.place(x=600, y=500)
                        # Frame6 = Frame(new_window6, width=600, bg='white', bd=4, height=260, relief=RIDGE)
                        # Frame6.place(x=0, y=650)
                        Frame5 = Frame(Bestellung_frame , width=600, bg=colour4, bd=4, height=140, relief=RIDGE)
                        # Frame5.place(x=5, y=650)
                        btnschliessen=Button(Bestellung_frame,text='X',bg=colour4,font=('arial', 16, 'bold'),command=unlockall)
                        btnschliessen.place(x=1100,y=5)
                        # conn=sqlite3.connect('Einstellung.db')
                        # cur=conn.cursor()
                        # cur.execute("""Create Table Lieferanten ( ID integer primary key AUTOINCREMENT, Name TEXT ) """)
                        # conn.commit()
                        # ------------------------------------------------ functions--------------------------------------------------------#


                        # ------------------------------------------------Labels------------------------------------------------------------#


                        namef = Label(Bestellung_frame, text='Fahrer auswählen:', bg=colour4, bd=0, relief=RIDGE,
                                      font=('arial', 18, 'bold'))
                        namef.place(x=0, y=644)
                        nummer = Label(Bestellung_frame , text=' BestellungsNr:', bg=colour4, bd=0,
                                       font=('arial', 16, 'bold'))
                        nummer.place(x=0, y=25)
                        ofennebes = Label(Bestellung_frame , text=' Offene Bestellung:', bg=colour4, bd=0,
                                          font=('arial', 16, 'bold'))
                        ofennebes.place(x=600, y=25)
                        pendingl = Label(Bestellung_frame , text=' Pending:', bg=colour4, bd=0,
                                         font=('arial', 16, 'bold'))
                        pendingl.place(x=0, y=470)
                        Fahrerl = Label(Bestellung_frame, text=' Fahrer Liste:', bg=colour4, bd=0,
                                         font=('arial', 16, 'bold'))
                        Fahrerl.place(x=600, y=470)
                        Anzahl = Label(Frame5, text=' Anzahl/B', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                        Anzahl.place(x=175, y=10)
                        Datum = Label(Frame5, text=' Datum', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                        Datum.place(x=290, y=10)
                        GPreis = Label(Frame5, text=' Gesamt preis', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                        GPreis.place(x=390, y=10)
                        ffahrer = Label(Bestellung_frame, text=' Fahrer auswählen', bg=colour4, bd=0, font=('arial', 18, 'bold'))
                        ffahrer.place(x=0, y=900)
                        # -----------------------------------------------------------Enetry_Liste----------------------------------------------#


                        nummerL = Listbox(Bestellung_frame , width=3, height=0, font=font_size)
                        nummerL.place(x=810, y=25)
                        nummerE = Entry(Bestellung_frame , width=3, font=font_size, bd=3)
                        nummerE.place(x=165, y=24)
                        AnzE = Entry(Frame5, width=3, font=font_size, bd=3)
                        AnzE.place(x=195, y=40)
                        DatE = Entry(Frame5, width=9, font=font_size, bd=3)
                        DatE.place(x=290, y=40)
                        GesE = Entry(Frame5, width=9, font=font_size, bd=3)
                        GesE.place(x=410, y=40)
                        # -----------------------------------------------------------DropBox----------------------------------------------------#
                        options = []
                        options2 = []
                        cur = connE.cursor()
                        cur.execute('select Name from Lieferanten ')
                        siko = cur.fetchall()
                        for ido in siko:
                            name=ido[0]
                            options.insert(0, name)
                        conn12 = sqlite3.connect(resource_path('Data\Pending.db'))
                        cur12 = conn12.cursor()
                        cur12.execute('select Fahrer from pending ')
                        open_fahrer = cur12.fetchall()
                        if open_fahrer:
                            for fahrers in open_fahrer:
                                options2.insert(0, fahrers)
                        else:
                            options2.insert(0, ' ')
                        clicked = StringVar()

                        clicked1 = StringVar()
                        clicked2 = StringVar()

                        clicked.set(ido[0])

                        drop1 = OptionMenu(Frame6, clicked1, *options)
                        drop1.configure(width=10, bd=0, bg='grey', font=('arial', 16, 'bold'))
                        drop1.place(x=0, y=30)

                        # ------------------------------------------------------------------------------------------------------------------#
                        def name_lösch():

                            cur = connE.cursor()
                            siko = clicked.get()
                            soko = siko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"',
                                                                                                                    '')
                            cur.execute('delete from Lieferanten where name =(?)', (soko,))
                            connE.commit()
                            new_window6.destroy()
                            openwindow7()

                        # ----------------------------------------------------------#

                        cur = connE.cursor()
                        cur.execute('select Name from Lieferanten ')
                        connE.commit()
                        miko = cur.fetchall()

                        # ---------------------------------------------------------------Buttons------------------------------------------------#

                        # conn = sqlite3.connect('Pending.db')
                        # cur = conn.cursor()
                        # cur.execute("""Create Table Rechnung ( Nr INTEGER, Name TEXT,Datum TEXT, Stasse TEXT, Hnr TEXT,Preis INTEGER ,Fahrer TEXT ) """)
                        # conn.commit()
                        # ------------------------------------------------------Tree1---------------------------------------------------------#
                        # style = ttk.Style()
                        # style.theme_use("clam")
                        # style.configure("Tree1", background='white', foreground="black", rowheight=20, font=("ARIEL", 9, 'bold'),
                        #                 fieldbackground="silver")
                        # style.map('Tree1', background=[('selected', 'red')])
                        # style.configure("Tree1.Heading", font=('bold', 12))
                        # style.configure("Tree1.column", font=('bold', 6))
                        # style.configure('Tree1', rowheight=20)
                        # style.configure('Tree1.row', font=('bold', 20))
                        besteltree = ttk.Treeview(Frame2, height=18)
                        besteltree['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                        besteltree.column("#0", width=0, stretch=NO)
                        besteltree.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                        besteltree.column("Name", anchor=CENTER, width=120)
                        besteltree.column("Datum/Uhrzeit", anchor=W, width=160)
                        besteltree.column("Straße", anchor=W, width=120)
                        besteltree.column("Haus/nr", anchor=W, width=40)
                        besteltree.column("Preis", anchor=W, width=60)
                        besteltree.tag_configure('pos', background='white')
                        besteltree.heading("#0", text="", anchor=W)
                        besteltree.heading("Nr", text="Nr", anchor=W)
                        besteltree.heading("Name", text="Name", anchor=W)
                        besteltree.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                        besteltree.heading("Straße", text="Straße", anchor=W)
                        besteltree.heading("Haus/nr", text="nr", anchor=W)
                        besteltree.heading("Preis", text="Preis", anchor=W)
                        besteltree.pack()
                        # ------------------------------------------------------- Tree2---------------------------------------------------------#
                        besteltree1 = ttk.Treeview(Frame3, height=18)
                        besteltree1['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                        besteltree1.column("#0", width=0, stretch=NO)
                        besteltree1.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                        besteltree1.column("Name", anchor=CENTER, width=120)
                        besteltree1.column("Datum/Uhrzeit", anchor=W, width=160)
                        besteltree1.column("Straße", anchor=W, width=120)
                        besteltree1.column("Haus/nr", anchor=W, width=40)
                        besteltree1.column("Preis", anchor=W, width=60)
                        besteltree1.tag_configure('pos', background='white')
                        besteltree1.heading("#0", text="", anchor=W)
                        besteltree1.heading("Nr", text="Nr", anchor=W)
                        besteltree1.heading("Name", text="Name", anchor=W)
                        besteltree1.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                        besteltree1.heading("Straße", text="Straße", anchor=W)
                        besteltree1.heading("Haus/nr", text="nr", anchor=W)
                        besteltree1.heading("Preis", text="Preis", anchor=W)
                        besteltree1.pack()
                        # ----------------------------------------------------- Tree 3 ---------------------------------------------------------#
                        besteltree2 = ttk.Treeview(Frame4, height=5)
                        besteltree2['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                        besteltree2.column("#0", width=0, stretch=NO)
                        besteltree2.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                        besteltree2.column("Name", anchor=CENTER, width=120)
                        besteltree2.column("Datum/Uhrzeit", anchor=W, width=160)
                        besteltree2.column("Straße", anchor=W, width=120)
                        besteltree2.column("Haus/nr", anchor=W, width=40)
                        besteltree2.column("Preis", anchor=W, width=60)
                        besteltree2.tag_configure('pos', background='white')
                        besteltree2.heading("#0", text="", anchor=W)
                        besteltree2.heading("Nr", text="Nr", anchor=W)
                        besteltree2.heading("Name", text="Name", anchor=W)
                        besteltree2.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                        besteltree2.heading("Straße", text="Straße", anchor=W)
                        besteltree2.heading("Haus/nr", text="nr", anchor=W)
                        besteltree2.heading("Preis", text="Preis", anchor=W)
                        besteltree2.pack()
                        #---------------------------------------------Tree4---------------------------------------------------------------------#
                        besteltree3 = ttk.Treeview(Frame7, height=12)
                        besteltree3['columns'] = ("Fahrer", "Anzahl", "Datum", "Gesamte Preis")
                        besteltree3.column("#0", width=0, stretch=NO)
                        besteltree3.column("Fahrer", anchor=CENTER, width=140, stretch=TRUE, )
                        besteltree3.column("Anzahl", anchor=CENTER, width=130)
                        besteltree3.column("Datum", anchor=W, width=160)
                        besteltree3.column("Gesamte Preis", anchor=W, width=120)

                        besteltree3.tag_configure('pos', background='white')
                        besteltree3.heading("#0", text="", anchor=W)
                        besteltree3.heading("Fahrer", text="Fahrer", anchor=W)
                        besteltree3.heading("Anzahl", text="Anzahl", anchor=W)
                        besteltree3.heading("Datum", text="Datum", anchor=W)
                        besteltree3.heading("Gesamte Preis", text="Gesamte Preis", anchor=W)

                        besteltree3.pack()
                        # -------------------------------------------- Tress Functions ---------------------------------------------------------#
                        def besteliste():
                            for record in besteltree1.get_children():
                                besteltree1.delete(record)
                            count = 0
                            conn1 = sqlite3.connect(resource_path('Data\OrderData.db'))
                            cur1 = conn1.cursor()
                            cur1.execute('select * from kundeinfo ')
                            firo = cur1.fetchall()
                            for ziko in firo:
                                besteltree1.insert(parent='', open=True, index='0', iid=count, text='',
                                                   values=(ziko[0], ziko[3], ziko[10], ziko[4], ziko[5], ziko[13]))
                                count += 1
                            zuzu = besteltree1.get_children()
                            moso = len(zuzu)
                            nummerL.delete(0, END)
                            nummerL.insert(0, moso)
                        besteliste()
                        def fahreliste():
                            for record in besteltree3.get_children():
                                besteltree3.delete(record)
                            count = 0
                            conn1 = sqlite3.connect(resource_path('Data\Pending.db'))
                            cur1 = conn1.cursor()
                            cur1.execute('select * from Rechnung ')
                            firo = cur1.fetchall()
                            for ziko in firo:
                                besteltree3.insert(parent='', open=True, index='0', iid=count, text='',
                                                   values=(ziko[3], ziko[0], ziko[1], ziko[2]))
                                count += 1
                            zuzu = besteltree1.get_children()
                            moso = len(zuzu)
                            nummerL.delete(0, END)
                            nummerL.insert(0, moso)

                        fahreliste()
                        global zaro
                        zaro = 0

                        # ----------------------------------------------------------------------------------------------------------------------#
                        def add_fun(event=None):
                            global zaro
                            zopl = nummerE.get()
                            cur = connO.cursor()
                            cur.execute('select* from kundeinfo where ID=(?)', (zopl,))
                            bol = cur.fetchall()
                            ipo=''
                            for opo in bol:
                                besteltree.insert(parent='', open=True, index='end', iid=zaro, text='',
                                                  values=(opo[0], opo[3], opo[10], opo[4], opo[5], opo[13]))
                                zaro += 1
                                ipo=opo[10]

                            cur.execute('select* from speiseinfo where zeit=(?)', (ipo,))
                            samo = cur.fetchall()

                            for sam in samo:

                                curd=connD.cursor()

                                curd.execute("""INSERT INTO speiseinfo (zeit, pos, grosse, anzahl, nr, speise, mit, ohne, katagorie, preis, name)
                                                VALUES (:zeit, :pos, :grosse, :anzahl, :nr, :speise, :mit, :ohne, :katagorie, :preis, :name)""",
                                             {'zeit': sam[0],
                                              'pos': sam[1],
                                              'grosse': sam[2],
                                              'anzahl': sam[3],
                                              'nr': sam[4],
                                              'speise': sam[5],
                                              'mit': sam[6] + ' ',
                                              'ohne': sam[7],
                                              'katagorie': sam[8],
                                              'preis': sam[9],
                                              'name': sam[10]
                                              })

                            for opo in bol:
                                curd.execute("""INSERT  into kundeinfo (ID,kid,ktelefonnummer,kname,kstrasse,khausnr,
                                                kplz,kort,kemail,bediener,zeit,bestellzeit,externinfo,gesamtepreis)
                                             VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                            (opo[0],
                                             opo[1],
                                             opo[2],
                                             opo[3],
                                             opo[4],
                                             opo[5],
                                             opo[6],
                                             opo[7],
                                             opo[8],
                                             opo[9],
                                             opo[10],
                                             opo[11],
                                             opo[12]
                                             ))
                            connD.commit()
                            connO.commit()

                            cur.execute('delete from kundeinfo where ID=(?)', (zopl,))
                            connO.commit()
                            nummerE.delete(0, END)
                            besteliste()

                        nummerE.bind('<Return>',add_fun)

                        # ----------------------------------------------------------------------------------------------------------------------#
                        def add_abhol():
                            global zaro
                            zopl = 'ABHOLUNG'
                            cur = connO.cursor()
                            cur.execute('select* from kundeinfo where kstrasse =(?)', (zopl,))
                            bol = cur.fetchall()
                            for opo in bol:
                                besteltree.insert(parent='', open=True, index='end', iid=zaro, text='',
                                                  values=(opo[0], opo[3], opo[10], opo[4], opo[5], opo[13]))
                                zaro += 1
                            cur.execute('delete from kundeinfo WHERE kstrasse =(?)', (zopl,))
                            connO.commit()
                            nummerE.delete(0, END)
                            besteliste()

                        # ----------------------------------------------------------------------------------------------------------------------#
                        def pending():
                            global zazo
                            selected = besteltree.selection()[0]
                            values = besteltree.item(selected, 'values')
                            besteltree.delete(selected)
                            countss = 0
                            besteltree2.insert(parent='', open=True, index='end', text='',
                                               values=(
                                               values[0], values[1], values[2], values[3], values[4], values[5]))
                            countss += 1

                        # ------------------------------------------------- Tree Buttons -------------------------------------------------------#
                        add = Button(Bestellung_frame , text='Add', bg='green', width=4, bd=4, command=add_fun,
                                     state=DISABLED)
                        add.place(x=210, y=24)
                        pend = Button(Bestellung_frame , text='Pend', bg='orange', width=7, bd=4, command=pending,
                                      state=DISABLED)
                        pend.place(x=265, y=25)
                        addabh = Button(Bestellung_frame , text='Alle Abholung', bg='green', width=11, bd=4,
                                        command=add_abhol,
                                        state=DISABLED)
                        addabh.place(x=870, y=23)

                        # ----------------------------------------------------------------------------------------------------------------------#
                        def Speichern():
                            siko = str(clicked1.get())
                            soko = siko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"',
                                                                                                                    '')

                            conn = sqlite3.connect(resource_path('Data\Pending.db'))
                            cur = conn.cursor()
                            cur.execute("SELECT * FROM Rechnung WHERE Fahrer = ?", (soko,))
                            row = cur.fetchone()
                            sick106 = len(besteltree.get_children())

                            total_price = 0  # Initialize total_price variable to hold the sum of sick105

                            for line in besteltree.get_children():
                                sick105 = float(besteltree.item(line)['values'][5])  # convert sick105 to float
                                total_price += sick105  # update total_price with sick105 sum
                                sick102 = (besteltree.item(line)['values'][2])

                            if row is not None:
                                cur.execute("UPDATE Rechnung SET Preis = Preis + ?, Nr= Nr + ?  WHERE Fahrer = ?",
                                            (round(total_price,2),sick106, soko))  # use total_price in update statement
                            else:
                                conn = sqlite3.connect(resource_path('Data\Pending.db'))
                                cur = conn.cursor()
                                cur.execute(
                                    """ insert into Rechnung (Nr,Datum,Preis,Fahrer)values(:Nr,:Datum,:Preis,:Fahrer)""",
                                    {'Nr': sick106,
                                     'Datum': sick102,
                                     'Preis': total_price,  # use total_price for the Preis value
                                     'Fahrer': soko
                                     })
                                conn.commit()

                            conn.commit()


                            for record in besteltree.get_children():
                                besteltree.delete(record)

                            drop1.config(state=ACTIVE)
                            clicked1.set('')
                            Strat.config(state=ACTIVE)
                            add.config(state=DISABLED)
                            pend.config(state=DISABLED)
                            addabh.config(state=DISABLED)
                            Done.config(state=DISABLED)
                            fahreliste()

                        Done = Button(Bestellung_frame , text='Fertig', font=font_size, bg='red', width=5, height=1, bd=6,
                                      command=Speichern, state=DISABLED)
                        Done.place(x=400, y=18)

                        # ---------------------------------------- Straten Abrechnung ------------------------------------------------------#
                        def abrechnung():
                            sok = clicked1.get()
                            sol = len(sok)
                            if sol < 3:
                                messagebox.showerror('Error', 'Fahrer aussuchen')
                            else:
                                drop1.config(state=DISABLED)
                                add.config(state=ACTIVE)
                                pend.config(state=ACTIVE)
                                addabh.config(state=ACTIVE)

                                Strat.config(state=DISABLED)
                                Done.config(state=ACTIVE)

                        Strat = Button(Frame6, text='Abrechnung', font=font_size, bg=colour1, bd=6, height=2,command=abrechnung)
                        Strat.place(x=170, y=10)

                        def reset():
                            AnzE.config(state=NORMAL)
                            DatE.config(state=NORMAL)
                            GesE.config(state=NORMAL)
                            AnzE.delete(0, END)
                            DatE.delete(0, END)
                            GesE.delete(0, END)

                            clicked2.set(' ')

                        rest = Button(Frame5, text='Reset', bg='red', font=font_size, bd=2, command=reset)
                        rest.place(x=20, y=90)

                        # --------------------------------- jeder Fahrer Rechnung(summe) --------------------------------------------------#
                        def fahrer(e=None):
                            count = 0
                            fahrer_list = []
                            AnzE.config(state=NORMAL)
                            DatE.config(state=NORMAL)
                            GesE.config(state=NORMAL)

                            soko = str(clicked2.get())

                            siko = soko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"',
                                                                                                                    '')
                            conn = sqlite3.connect(resource_path('Data\Pending.db'))
                            cur = conn.cursor()
                            cur1 = conn.cursor()
                            cur2 = conn.cursor()
                            cur3 = conn.cursor()
                            cur.execute('select * from Rechnung ')
                            cur3.execute('select Fahrer from Rechnung')
                            solo = cur3.fetchall()
                            for sol in solo:
                                fahrer_list.append(
                                    str(sol).replace('(', '').replace(')', '').replace(',', '').replace("'", ""))

                            if siko in fahrer_list:
                                cur2.execute('select ROUND (SUM(Preis),0) from Rechnung Where Fahrer =(?)', (siko,))
                                cur1.execute('select count(*) from Rechnung Where Fahrer =(?)', (siko,))
                                tol = cur1.fetchall()
                                zoro = cur2.fetchall()
                                fob = cur.fetchall()
                                connE.commit()
                                datum = datetime.now().strftime("%d.%m.%Y")
                                for ziko in fob:
                                    anzahl = ziko[0]
                                    AnzE.delete(0, END)
                                    DatE.delete(0, END)
                                    GesE.delete(0, END)

                                    AnzE.insert(0, tol)
                                    GesE.insert(0, zoro)
                                    DatE.insert(0, datum)
                                AnzE.config(state=DISABLED)
                                DatE.config(state=DISABLED)
                                GesE.config(state=DISABLED)

                            else:
                                AnzE.delete(0, END)
                                DatE.delete(0, END)
                                GesE.delete(0, END)


                        drop2 = OptionMenu(Frame5, clicked2, *options, command=fahrer)
                        drop2.configure(width=10, bd=0, bg='grey', font=('arial', 16, 'bold'))
                        drop2.place(x=0, y=45)

                        # ------------------------------------------------------------------------------------------------------------------#
                        def unpending():
                            selected = besteltree2.selection()[0]
                            values = besteltree2.item(selected, 'values')
                            besteltree2.delete(selected)
                            zazo = 0
                            besteltree.insert(parent='', open=True, index='end', text='',
                                              values=(values[0], values[1], values[2], values[3], values[4], values[5]))

                        unpend = Button(Bestellung_frame , text='Unpend', font=('arial', 10, 'bold'), bg='green', bd=4, command=unpending)
                        unpend.place(x=105, y=465)
                        conn = sqlite3.connect(resource_path('Data\Pending.db'))
                        cur = conn.cursor()
                        conn.commit()


                        ##################Rsest

                        ############################################# Alle Rechnung von Alle Fahrer##########################################
                        def fahreren():

                            selected_item = besteltree3.selection()[0]  # get the item ID of the selected row
                            values = besteltree3.item(selected_item, 'values')
                            siko=values[0]
                            an=values[1]
                            da=values[2]
                            gm=values[3]
                            text2 = Text(Bestellung_frame, height=40, bg='white', bd=0, font=('arial', 16, 'bold'))
                            text2.insert(END, f'\t\t\t\n{restName}   \n')
                            text2.insert(END, f'\t\n=======================')
                            text2.insert(END, f'\t\t\t\n Fahrer:{siko}')
                            text2.insert(END, f'\t\n=======================')
                            text2.insert(END, f'\t\t\t\t\n Anzahl:{an}')
                            text2.insert(END, f'\t\n=======================')
                            text2.insert(END, f'\t\t\t\t\n Datum:{da}')
                            text2.insert(END, f'\t\n=======================')
                            text2.insert(END, f'\t\t\t\t\n Preis:{gm} €')
                            text2.insert(END, f'\t\n=======================')
                            text2.insert(END, f'\t\nFür Ihre Buchhaltung wird automatisch eine Rechnung erstellt \n und ein Duplikat wird am Ende des Tages für Ihren Chef ausgedruckt')
                            text2.insert(END, f'\t\n')
                            text2.insert(END, f'\t\n')
                            lp = text2.get("1.0", "end-1c")

                            printfile = tempfile.mktemp(".txt")
                            open(printfile, 'w').write(lp)
                            os.startfile(printfile, "print")
                            conn = sqlite3.connect(resource_path('Data\Pending.db'))
                            cur = conn.cursor()
                            cur.execute('Insert into pending (Anzahl,Datum,Gesamt,Fahrer) values(?,?,?,?)',
                                        (an, da, gm, siko))
                            conn.commit()

                            conn1 = sqlite3.connect(resource_path('Data\Pending.db'))
                            cur1 = conn1.cursor()
                            cur1.execute('delete  from Rechnung where fahrer =(?)', (siko,))
                            conn1.commit()

                            # new_window6.destroy()
                            # openwindow7()

                            fahreliste()








                        dru = Button(Bestellung_frame, text='Drucken', bg='green', font=('arial', 10, 'bold'), bd=4, command=fahreren)
                        dru.place(x=750, y=465)

                        # liefer = Label(Frame5, text='Mindestpreise :', bg='white', font=font_size, bd=2)
                        # liefer.place(x=50, y=195)
                        # liefer = Label(Frame5, text='Mindestpreise :', bg='white', font=font_size, bd=2)
                        # liefer.place(x=50, y=195)

                        ####--------------------------------------     FIERAMT      -------------------------------------------------------#
                        def fieramt():
                            global new_window2



                            storno='storno'
                            them=len(besteltree3.get_children())
                            if them > 0:
                                messagebox.showwarning('Fahrer', 'BITTE ALLE FAHRER RECHNUNG AUSDRUCKEN')
                                # register the cleanup function to be called when the program is closing


                            else:

                                zuzu = besteltree1.get_children()
                                moso = int(nummerL.get(0))
                                if moso > 0:
                                    messagebox.showerror('Warning', 'da sind noch offene Bestellung bitte zu ordnen')
                                else:
                                    ask = messagebox.askyesno('Feierabend', 'Sind Sie sicher')
                                    if ask == 1:

                                        conn = sqlite3.connect(resource_path('Data\Pending.db'))
                                        cur = conn.cursor()
                                        cur.execute('select * from pending where Fahrer !=(?) ', (storno,))
                                        all = cur.fetchall()
                                        conn1 = sqlite3.connect(resource_path('Data\Pending.db'))
                                        cur1 = conn.cursor()
                                        cur1.execute('select * FROM pending where Fahrer =(?) ', (storno,))
                                        them = cur1.fetchall()
                                        text3 = Text(Bestellung_frame, height=40, bg='white', bd=0,
                                                     font=('arial', 10, 'bold'))
                                        text3.insert(END, f'\n{restName}  \n')
                                        text3.insert(END, f'\n=============')
                                        for alles in all:
                                            text3.insert(END, f'\n Fahrer:{alles[3]}')
                                            text3.insert(END, f'\n Anzahl:{alles[0]}')
                                            text3.insert(END, f'\n Datum:{alles[1]}')
                                            text3.insert(END, f'\n Preis:{alles[2]} €')
                                            text3.insert(END, f'\n')
                                            text3.insert(END, f'\n')
                                        text3.insert(END, f'\n STORNO:')
                                        for the in them:
                                            text3.insert(END, f'\n Datum:{the[1]}')
                                            text3.insert(END, f'\n Bestellung Nummer:{the[2]} ')
                                            # Select the columns "Preis" and "Anzahl" from the "pending" table
                                        cur.execute('SELECT Gesamt, Anzahl FROM pending')

                                        # Fetch all rows and calculate the sum of "Preis" and "Anzahl"
                                        rows = cur.fetchall()
                                        preis_sum = sum(float(row[0]) for row in rows)
                                        anzahl_sum = sum(row[1] for row in rows)

                                        text3.insert(END, f'\n=============')
                                        text3.insert(END, f'\n Gesamt')
                                        text3.insert(END, f'\n Anzahl:{anzahl_sum}')
                                        text3.insert(END, f'\n Preis:{preis_sum} €')
                                        lp = text3.get("1.0", "end-1c")
                                        printfile = tempfile.mktemp(".txt")
                                        open(printfile, 'w').write(lp)
                                        os.startfile(printfile, "print")
                                        text3.delete("1.0", "end-1c")


                                        time.sleep(3)

                                        messagebox.showinfo('FIERAMT',
                                                            'DAS PROGRAM BITTE NICHT SCHLIESEN DER SCHLIEST VON ALLEINE EIN SCHÖNES FIERABEND')

                                        connK.close()
                                        connZu.close()
                                        connO.close()
                                        connS.close()
                                        connZ.close()
                                        conn1.close()
                                        conn.close()



                                        try:
                                            new_window2.destroy()
                                        except:
                                            pass
                                        import subprocess



                                        root.destroy()
                                        os.remove(resource_path('Data\Pending.db'))
                                        os.remove(resource_path('Data\OrderData.db'))
                                        time.sleep(3)

                                        conn = sqlite3.connect(resource_path('Data\Pending.db'))
                                        cur = conn.cursor()
                                        cur.execute(
                                            'Create Table pending (Anzahl INTEGER , Datum INTEGER, Gesamt INTEGER ,Fahrer TEXT)')

                                        cur.execute(
                                            """Create Table Rechnung ( Nr INTEGER, Datum TEXT, Preis INTEGER ,Fahrer TEXT ) """)
                                        conn.commit()
                                        conn1 = sqlite3.connect(resource_path('Data\OrderData.db'))
                                        cur1 = conn1.cursor()
                                        cur1.execute("""Create Table kundeinfo(

                                                ID integer primary key AUTOINCREMENT UNIQUE,
                                                kid INTEGER,
                                                ktelefonnummer INTEGER,
                                                kname TEXT,
                                                kstrasse TEXT,
                                                khausnr INTEGER,
                                                kplz INTEGER,
                                                kort TEXT,
                                                kemail TEXT,
                                                bediener TEXT,
                                                zeit INTEGER,
                                                bestellzeit INTEGER,
                                                externinfo TEXT,
                                                gesamtepreis INTEGER)""")
                                        cur1.execute("""Create Table speiseinfo(
                                                    zeit INTEGER,
                                                    pos INTEGER,
                                                    grosse TEXT,
                                                    anzahl INTEGER,
                                                    nr INTEGER,
                                                    speise TEXT,
                                                    mit TEXT,
                                                    ohne TEXT,
                                                    katagorie TEXT,
                                                    preis INTEGER,
                                                    name TEXT )""")
                                        cur1.execute("CREATE TABLE lieferpara(geld text, zeit Integer)")
                                        cur1.execute("CREATE TABLE Rabattpara(Rabatt text, zeit Integer)")
                                        conn1.commit()
                                        conn.commit()
                                        conn1.close()
                                        conn.close()


                                        try:
                                            # terminate the program
                                            process_app = "MindMeshLab.exe"

                                            # call the taskkill command to terminate the process
                                            subprocess.call(["taskkill", "/F", "/IM", process_app])
                                        except:
                                            pass




                        fierB = Button(Bestellung_frame , text='Feierabend', bd=4, font=font_size, height=3, bg='red',
                                       command=fieramt)
                        fierB.place(x=1080, y=810)
                    Abrechnung = Button(Mainframe1, text='Chef', bg=colour1, width=8,
                                               command=popupchef,
                                               font=('arial', 15, 'bold'))
                    Abrechnung .place(x=80, y=840)


                    newtree.bind("<Double-Button-1>", save_name)


                    endframe = Frame(Mainframe3, bd=4, width=80, height=80, bg=colour4)
                    endframe.place(x=1050, y=385)


                    global call, call1
                    # create empty lists to check if the a new nummber is calling
                    call = []
                    call1 = []

                    def last_nummber():
                        TelefonnummerE.focus_force()
                        TelefonnummerE.delete(0,END)
                        TelefonnummerE.insert(0, '0' + call1[0])

                    Telefonnummer_sign = Button(Mainframe1, bd=0, text='\u260E', font=('arial', 14, 'bold'),
                                                bg=colour4, command=last_nummber, anchor='n') \
                        .place(x=82, y=0)

                    ##################################################### Fritzcaller #####################################################
                    def fritz():
                        global call, call1, counter,cono  # global variables used in the function
                        if counter < 3:  # if the counter is less than 3, return and do nothing
                            return

                        cono = sqlite3.connect(resource_path('Data\journal.sqlite'))  # connect to SQLite database
                        curo = cono.cursor()  # create a cursor object to interact with the database
                        curo.execute(
                            'select areacode,number  from calls ORDER BY uuid  DESC  ')  # execute a query to get the latest call data
                        cico = curo.fetchone()  # fetch one row of data from the query result
                        call1.clear()  # clear the call1 list
                        sarmi = str(cico).replace('(', '').replace("'", "").replace(' ', '').replace(')', '').replace(',',
                                                                                                                      '')  # format the data and save it in a string variable
                        call1.append(sarmi)  # append the formatted string to the call1 list
                        def press_enetr():
                            press('Enter')



                        status = TelefonnummerE.get()
                        # get the current value of the TelefonnummerE widget

                        if not status:  # if the widget is empty
                            if call != call1:  # if the current call data is different from the previous call data
                                TelefonnummerE.focus_force()  # set the focus on the TelefonnummerE widget
                                for saro in call1:  # for each call data in call1 list
                                    TelefonnummerE.insert(0, '0' + saro)  # insert the call data into the widget
                                    press_enetr()
                                    press_enetr() # simulate pressing the enter key
                                    call1.clear()  # clear the call1 list
                                    call.clear()  # clear the call list
                                    call1.append(saro)  # append the call data to the call1 list
                                    call.append(saro)  # append the call data to the call list
                                    Timer(2, fritz).start()  # wait for 5 seconds and call the fritz() function again
                            elif call == call1:  # if the current call data is the same as the previous call data
                                Timer(2, fritz).start()  # wait for 5 seconds and call the fritz() function again
                        else:  # if the widget is not empty
                            Timer(2, fritz).start()  # wait for 5 seconds and call the fritz() function again

                    fritz()
                    # isdn = Button(Mainframe1, text='ISDN', bg='blue', font=("ARIEL", 8, 'bold'), command=fritz)
                    # isdn.place(x=120, y=10)
                    ##################################################### Tree view Bestellung ############################################
                    style = ttk.Style()
                    style.theme_use("clam")
                    style.configure("Custom1.Treeview", background="light blue", foreground="black", fieldbackground="white",
                                    font=('Calibri', 13, 'bold'), rowheight=30)
                    style.map('Custom1.Treeview', background=[('selected', '#6AC7C8')], foreground=[('selected', 'white')])
                    style.configure("Custom1.Treeview.Heading", font=('Calibri', 13))
                    BestellTree = ttk.Treeview(Mainframe3, height=10, style="Custom1.Treeview")

                    BestellTree['columns'] = ("pos", "Größe", "Anzahl", "Nr", "Speise", "Mit", "Ohne", "Kategorie", "Preis",)
                    BestellTree.column("#0", width=0, stretch=NO)
                    BestellTree.column("pos", anchor=CENTER, width=40, stretch=TRUE, )
                    BestellTree.column("Größe", anchor=CENTER, width=80)
                    BestellTree.column("Anzahl", anchor=CENTER, width=35)
                    BestellTree.column("Nr", anchor=CENTER, width=35)
                    BestellTree.column("Speise", anchor=W, width=180)
                    BestellTree.column("Mit", anchor=W, width=400, stretch=TRUE)
                    BestellTree.column("Ohne", anchor=W, width=180, stretch=TRUE)
                    BestellTree.column("Kategorie", anchor=CENTER, width=140)
                    BestellTree.column("Preis", anchor=CENTER, width=60)
                    BestellTree.tag_configure('pos', background='gray')
                    BestellTree.heading("#0", text="", anchor=W)
                    BestellTree.heading("pos", text="pos", anchor=W)
                    BestellTree.heading("Größe", text="Grosse", anchor=W)
                    BestellTree.heading("Anzahl", text="Anzahl", anchor=W)
                    BestellTree.heading("Nr", text="Nr", anchor=W)
                    BestellTree.heading("Speise", text="Speise", anchor=CENTER)
                    BestellTree.heading("Mit", text="Mit", anchor=CENTER)
                    BestellTree.heading("Ohne", text="Ohne", anchor=CENTER)
                    BestellTree.heading("Kategorie", text="Katagorie", anchor=CENTER)
                    BestellTree.heading("Preis", text="Preis", anchor=W)
                    BestellTree.place(x=0, y=5)
                    BestellTree.tag_configure('even', background='light blue')
                    BestellTree.tag_configure('odd', background='light grey')

                    #  ################################    Delete all Data  ########################################################
                    def remove_all1():
                        global count,print_Button

                        if entrybox1.get()!='':
                            messagebox.showwarning('warnung','Speise bitte bestätigen')
                        else:


                            entrybox1.config(state=NORMAL)
                            update(toppings)
                            mos = 0
                            miko = messagebox.askyesno('Abbrechen', 'wollen sie wirklich abbrechen')
                            for i in BestellTree.get_children():
                                if not BestellTree.item(i, "values"):
                                    mos = 0
                                else:
                                    mos = 1
                            if miko == 1:
                                show_bestellung()
                                try:
                                    abholframe.destroy()

                                except:
                                    pass
                                try:
                                    bestätigen.destroy()
                                except:
                                    pass
                                if mos == 0:
                                    resetall()
                                    try:
                                        print_Button.destroy()
                                    except:
                                        pass
                                    unlockin()
                                    entrybox1.delete(0, END)
                                    entrybox11.delete(0, END)
                                    entrybox11.insert(0, 0.0)
                                    entrybox8.delete(0, END)
                                    entrybox8.insert(0, 0.0)
                                    entrybox10.delete(0, END)
                                    entrybox10.insert(0, 0.0)
                                    count = 0
                                    Mainframe2.grid_forget()
                                    Mainframe3.place_forget()
                                    KundenidE.configure(state=NORMAL)
                                    TelefonnummerE.configure(state=NORMAL)
                                    NameE.configure(state=NORMAL)
                                    AdresseE.configure(state=NORMAL)
                                    HauesnrE.configure(state=NORMAL)
                                    PLZE.configure(state=NORMAL)
                                    ORTE.configure(state=NORMAL)
                                    EmailE.configure(state=NORMAL)
                                    liefer.config(state=NORMAL)
                                    liefer1.config(state=NORMAL)
                                    reset()
                                elif mos != 0:
                                    for record in BestellTree.get_children():
                                        BestellTree.delete(record)
                                        resetall()
                                        unlockin()
                                        entrybox1.delete(0, END)
                                        entrybox11.delete(0, END)
                                        entrybox11.insert(0, 0.0)
                                        entrybox8.delete(0, END)
                                        entrybox8.insert(0, 0.0)
                                        entrybox10.delete(0, END)
                                        entrybox10.insert(0, 0.0)
                                        count = 0
                                        Mainframe2.grid_forget()
                                        Mainframe3.place_forget()
                                        KundenidE.configure(state=NORMAL)
                                        TelefonnummerE.configure(state=NORMAL)
                                        NameE.configure(state=NORMAL)
                                        AdresseE.configure(state=NORMAL)
                                        HauesnrE.configure(state=NORMAL)
                                        PLZE.configure(state=NORMAL)
                                        ORTE.configure(state=NORMAL)
                                        EmailE.configure(state=NORMAL)
                                        reset()

                                FIRSTLABEL = Label(Bestellung_frame, text='auf Lieferung oder Abholung drücken',
                                                   font=("ARIEL", 26, "bold"),
                                                   bg=colour4)
                                FIRSTLABEL.place(x=265, y=400)
                            else:
                                pass

                    # ----------------------------------------------------------------------------------------------------------------#
                    def remove_all():
                        entrybox1.config(state=NORMAL)
                        global count, mos, print_Button
                        try:

                            print_Button.destroy()
                        except:
                            pass

                        update(toppings)
                        mos = 0
                        for i in BestellTree.get_children():
                            if not BestellTree.item(i, "values"):
                                mos = 0
                            else:
                                mos = 1
                        if mos == 0:
                            unlockall()
                            resetall()
                            entrybox1.delete(0, END)
                            entrybox11.delete(0, END)
                            entrybox8.delete(0,END)
                            entrybox8.insert(0, 0.0)
                            entrybox11.insert(0, 0.0)
                            entrybox10.delete(0, END)
                            entrybox10.insert(0, 0.0)
                            count = 0
                            Mainframe2.grid_forget()
                            Mainframe3.place_forget()
                            KundenidE.configure(state=NORMAL)
                            TelefonnummerE.configure(state=NORMAL)
                            NameE.configure(state=NORMAL)
                            AdresseE.configure(state=NORMAL)
                            HauesnrE.configure(state=NORMAL)
                            PLZE.configure(state=NORMAL)
                            ORTE.configure(state=NORMAL)
                            EmailE.configure(state=NORMAL)
                            reset()
                        elif mos != 0:
                            for record in BestellTree.get_children():
                                BestellTree.delete(record)
                                resetall()
                                entrybox1.delete(0, END)
                                entrybox11.delete(0, END)
                                entrybox11.insert(0, 0.0)
                                entrybox10.delete(0, END)
                                entrybox10.insert(0, 0.0)
                                entrybox8.delete(0, END)
                                entrybox8.insert(0, 0.0)
                                count = 0
                                Mainframe2.grid_forget()
                                Mainframe3.place_forget()
                                KundenidE.configure(state=NORMAL)
                                TelefonnummerE.configure(state=NORMAL)
                                NameE.configure(state=NORMAL)
                                AdresseE.configure(state=NORMAL)
                                HauesnrE.configure(state=NORMAL)
                                PLZE.configure(state=NORMAL)
                                ORTE.configure(state=NORMAL)
                                EmailE.configure(state=NORMAL)
                                reset()
                        else:
                            pass

                    Bestellungabb = Button(Mainframe3, font=("Helvetica", 11
                                                             , 'bold'), text='Bestellung abbrechen', bd=4, bg=colour1,
                                           command=remove_all1)
                    Bestellungabb.place(x=20, y=0)

                    #   ################################  eine Speise HInzüfugen  ####################################################

                    def insertdatatree():
                        global counts, clicked, count, pos, sop, labo, tako, lop, sip

                        # Increase the count variable
                        count += 1

                        # Check the state of the checkboxes and insert items into listbox2 accordingly
                        first = c1a.state()
                        if 'selected' in first:
                            listbox2.insert(0, '*Schneiden')
                        second = c2a.state()
                        if 'selected' in second:
                            listbox2.insert(0, '*Knusprig')
                        third = c3a.state()
                        if 'selected' in third:
                            listbox2.insert(0, '*Hell Backen'),
                        fourth = c4a.state()
                        if 'selected' in fourth:
                            listbox2.insert(0, '*Wenig Käse')

                        # Get the selected items from listbox2 and process them
                        sip = (listbox2.get(0, END))
                        ton = drop.get()
                        fibo = entrybox1.get()
                        komm = entrybox12.get()
                        sop = []
                        lop = []
                        kom = []
                        for i in sip:
                            o = str(i)
                            o.replace("'", ' ')
                            try:
                                if o[0] == '+':
                                    sop.append(o)
                                elif o[0] == '*':
                                    kom.append(o)
                                else:
                                    lop.append(o)
                            except:
                                pass

                        # Get the selected items from listbox3 and process them
                        tip = listbox3.get(0, END)
                        nop = []
                        for he in tip:
                            hes = str(he)
                            hes.strip()
                            hes.split()
                            hos = hes.replace('{', '').replace('}', '').replace('/', '').replace('(', '').replace(')',
                                                                                                                  '').replace(
                                "'", '').replace("'", '')
                            if hos != '':
                                nop.append(hos)
                        nip = ' '.join(nop)
                        ohne = str(nip)

                        # Get the values from the entry boxes and process them
                        labo = (float(entrybox6.get()))
                        tako = (float(entrybox11.get()))
                        entrybox11.delete(0, END)
                        mado = int(entrybox4.get())
                        kat = str(entrybox5.get())
                        kat.strip()
                        mido = labo * mado
                        result = math.floor(float(mido) * 100) / 100
                        tiko = '{:.2f}'.format(result)
                        count += 1




                        # Update the ID of the item you want to insert


                        # Insert the values into the treeview widget
                        if pos % 2 == 0:
                            BestellTree.insert(parent='', open=False, index='end', text=f'{counts + 1}',
                                               values=(counts, ton, entrybox4.get(),
                                                       entrybox3.get(), entrybox1.get(),
                                                       sop + kom, ohne, kat, tiko), tags=('odd',))

                            if komm != '':
                                BestellTree.insert(counts, index=counts,
                                                   values=('', '', '', '', '', komm))

                        else:
                            BestellTree.insert(parent='', open=False, index='end', text=f'{counts + 1}',
                                               values=(counts, ton, entrybox4.get(),
                                                       entrybox3.get(), entrybox1.get(),
                                                       sop + kom, ohne, kat, tiko), tags=('even',))

                            if komm != '':
                                BestellTree.insert(counts, index=counts,
                                                   values=('', '', '', '', '', komm))
                        counts += 1

                        pos += 1
                        # Clear the state of the checkboxes
                        forts = c1a.state()
                        fifts = c2a.state()
                        sixts = c3a.state()
                        secents = c4a.state()
                        if 'selected' in forts:
                            c1a.invoke()
                        if 'selected' in fifts:
                            c2a.invoke()
                        if 'selected' in sixts:
                            c3a.invoke()
                        if 'selected' in secents:
                            c4a.invoke()

                        # Increase the counts variable and update the value in entrybox11

                        result = math.floor(float(mido + tako) * 100) / 100
                        formatted_result = '{:.2f}'.format(result)
                        entrybox11.insert(0, formatted_result)
                        entrybox12.delete(0, END)

                    # ----------------------------------------------------------------------------------------------------------------------#
                    def bestätigens():
                        global counts, posent, bestätigen, clicked, count, pos, labo, tako, lop, sip, sop, nop

                        # enable entry boxes
                        entrybox1.config(state=NORMAL)
                        entrybox3.config(state=NORMAL)

                        # get selected item from treeview
                        selected = BestellTree.focus()
                        values = BestellTree.item(selected, 'values')

                        # get position from combobox
                        one = posent.get()

                        # clear lists
                        nop, sop, lop, hob, kommk = [], [], [], [], []

                        # get toppings from checkbox and add to list
                        komm = entrybox12.get()
                        first = c1a.state()
                        if 'selected' in first:
                            listbox2.insert(0, '*Schneiden')
                        second = c2a.state()
                        if 'selected' in second:
                            listbox2.insert(0, '*Knusbrig')
                        third = c3a.state()
                        if 'selected' in third:
                            listbox2.insert(0, '*Hellbacken'),
                        fourth = c4a.state()
                        if 'selected' in fourth:
                            listbox2.insert(0, '*wenigkäse')
                        sip = (listbox2.get(0, END))
                        ton = drop.get()
                        for i in sip:
                            o = str(i)
                            o.replace("'", '')
                            try:
                                if o[0] == '+':
                                    sop.append(o)
                                elif o[0] == '*':
                                    kommk.append(o)
                                else:
                                    lop.append(o)
                            except:
                                pass

                        # get extras from listbox and add to list
                        tip = listbox3.get(0, END)
                        for he in tip:
                            hes = str(he)
                            hes.strip()
                            if hes != '':
                                nop.insert(0, hes)
                        for tiz in nop:
                            tiz.strip()

                        # get pizza price and quantity
                        labo = (float(entrybox6.get()))
                        tako = (float(entrybox11.get()))
                        entrybox11.delete(0, END)
                        mado = int(entrybox4.get())
                        mido = labo * mado
                        tiko = (math.floor(float(mido) * 100) / 100)
                        formatted_tiko = '{:.2f}'.format(tiko)

                        # update selected item in treeview
                        BestellTree.item(selected, text='', values=(one, ton, entrybox4.get(),
                                                                    entrybox3.get(), entrybox1.get(),
                                                                    sop + kommk, nop, entrybox5.get(), formatted_tiko))
                        if komm != '':
                            BestellTree.insert(counts, index=counts,
                                               values=('', '', '', '', '', komm))

                        # calculate total price and update entry box
                        result = math.floor(float(mido + tako) * 100) / 100
                        formatted_result = '{:.2f}'.format(result)
                        entrybox11.insert(0, formatted_result)
                        # Clear and Enter new data
                        counts += 1
                        einfugen['state'] = ACTIVE
                        Reset['state'] = ACTIVE
                        bestätigen.destroy()
                        posent.destroy()
                        entrybox4.delete(0, END)
                        entrybox4.insert(0, '1')
                        entrybox2.delete(0, END)
                        entrybox1.delete(0, END)
                        listbox3.delete(0, END)
                        listbox2.delete(0, END)
                        entrybox3.delete(0, END)
                        entrybox5.delete(0, END)
                        entrybox6.delete(0, END)
                        entrybox12.delete(0, END)
                        drop.set('')
                        for i in BestellTree.selection():
                            BestellTree.selection_remove(i)
                        forts = c1a.state()
                        fifts = c2a.state()
                        sixts = c3a.state()
                        secents = c4a.state()
                        if 'selected' in forts:
                            c1a.invoke()
                        if 'selected' in fifts:
                            c2a.invoke()
                        if 'selected' in sixts:
                            c3a.invoke()
                        if 'selected' in secents:
                            c4a.invoke()

                    #   ################################# eine speise bearbeiten     ###############################################
                    def speisebearbeiten():
                        global counts, lop, nop, bestätigen, sop, einfugen, Reset, posent,drop
                        # Get all the items from the treeview
                        siko = BestellTree.get_children()
                        zablo = len(siko)

                        # If there are no items, show an error message
                        if zablo == 0:
                            messagebox.showerror('Error', 'diese action ist nicht gültig')

                        # If there are items, allow editing
                        else:
                            nop = []

                            # Decrease the count if there are items
                            if counts > 0:
                                counts -= 1

                            # Disable the 'Einfügen' and 'Reset' buttons
                            einfugen['state'] = DISABLED
                            Reset['state'] = DISABLED

                            # Create the 'Bestätigen' button and position it
                            bestätigen = Button(Mainframe3, text='Bestätigen', font=font_size, bd=4, bg='green',
                                                command=bestätigens)
                            bestätigen.place(x=450, y=5)

                            # Create the 'posent' entry box and position it
                            posent = Entry(Mainframe3, width=3, bd=4, font=font_size)
                            posent.place(x=580, y=5)

                            # Clear all the fields
                            resetall()

                            # Get the selected item
                            selected = BestellTree.focus()
                            values = BestellTree.item(selected, 'values')
                            alo = BestellTree.get_children(selected)
                            child = BestellTree.item(alo, 'values')

                            # Delete the text from the 'Anzahl' and 'Preis' entry boxes
                            entrybox12.delete(0, END)
                            entrybox4.delete(0, END)



                            # Fill in the 'Anzahl' and 'Name' entry boxes
                            entrybox4.insert(0, values[2])
                            entrybox3.insert(0, values[3])
                            entrybox1.insert(0, values[4])
                            entrybox1.config(state=DISABLED)
                            entrybox3.config(state=DISABLED)

                            # Get the 'Zutaten' list for the selected dish
                            vur = connZ.cursor()
                            nado = values[4]
                            solo = [values[6]]

                            # Add the 'Zutaten' to the 'Zutaten' listbox
                            for i in solo:
                                nop.insert(0, i)
                            for to in nop:
                                sopo = ''.join(to).split()
                                for sipo in sopo:
                                    listbox3.insert(0, sipo)
                            soso = solo[1:]
                            vur.execute("select* from zutaten where SpeiseName=(?)", (nado,))
                            farto = vur.fetchall()
                            for ziko in farto:
                                if (ziko[1] != '') and (ziko[1] not in sopo):
                                    listbox2.insert(END, ziko[1]),
                                if (ziko[2] != '') and (ziko[2] not in sopo):
                                    listbox2.insert(END, ziko[2]),
                                if (ziko[3] != '') and (ziko[3] not in sopo):
                                    listbox2.insert(END, ziko[3]),
                                if (ziko[4] != '') and (ziko[4] not in sopo):
                                    listbox2.insert(END, ziko[4]),
                                if (ziko[5] != '') and (ziko[5] not in sopo):
                                    listbox2.insert(END, ziko[5]),
                                if (ziko[6] != '') and (ziko[6] not in sopo):
                                    listbox2.insert(END, ziko[6]),
                                if (ziko[7] != '') and (ziko[7] not in sopo):
                                    listbox2.insert(END, ziko[7]),
                                if (ziko[8] != '') and (ziko[8] not in sopo):
                                    listbox2.insert(END, ziko[8]),
                                if (ziko[9] != '') and (ziko[9] not in sopo):
                                    listbox2.insert(END, ziko[9]),
                                if (ziko[10] != '') and (ziko[10] not in sopo):
                                    listbox2.insert(END, ziko[10])
                            if child != '':
                                # if the selected item has a child, insert the child's comment into entrybox12
                                entrybox12.insert(0, child[5])
                                # insert the position number from the selected item into posent
                            posent.insert(0, values[0])
                            salo = [values[5]]
                            for izo in salo:
                                # iterate through each ingredient in the list of ingredients for the selected item
                                izoz = ''.join(izo).split()
                                for oko in izoz:
                                    # insert each ingredient into listbox2
                                    listbox2.insert(END, oko)
                            # insert the category from the selected item into entrybox5
                            entrybox5.insert(0, values[7])
                            # calculate the price per item
                            selko = float(values[8])
                            felko = float(values[2])
                            niko = selko / felko
                            # insert the calculated price per item into entrybox6
                            formatted_niko = '{:.2f}'.format(niko)
                            entrybox6.insert(0, formatted_niko)

                            # update the total price
                            num = float(entrybox11.get())
                            neu = float(num) - float(values[8])
                            entrybox11.delete(0, END)
                            result = math.floor(float(neu) * 100) / 100
                            formatted_result = '{:.2f}'.format(result)
                            entrybox11.insert(0, formatted_result)
                            vur = connS.cursor()
                            vur.execute("SELECT * FROM Speisen WHERE Name = (?)", (nado,))
                            # Define constant variables
                            kl_gr=''
                            Klein = 'Klein'
                            Gross = 'Gross'
                            Standard = 'Standard'
                            zob = []
                            drop.destroy()

                            # If the dish has a size option, create a new option menu widget
                            for nano in vur.fetchall():
                                ziko = str(nano[2])
                            if Klein in nano[4]:
                                kl_gr = Gross, Klein


                            elif Standard in nano[4] :
                                kl_gr = Standard

                            elif Klein not in nano[4] and Standard not in nano[4]  :
                                kl_gr = Gross
                            drop = ttk.Combobox(Mainframe2, values=kl_gr)
                            drop.configure(font=('Helvetica bold', 16), width=8)
                            drop.place(x=255, y=70)
                            drop.set(values[1])
                            drop.bind("<<ComboboxSelected>>", gross_pizza)








                    bearbeiten = Button(Mainframe3, text='Bearbeiten', font=("Helvetica", 11, 'bold'), bd=4, bg=colour3,
                                        command=speisebearbeiten)
                    # switch back to normal frames
                    bearbeiten.place(x=350, y=0)
                    BestellTree.pack(pady=50)
                    Mainframe2.grid_forget()
                    Mainframe3.place_forget()

                    # -------------------------------------    eine Speise Löschen ---------------------------------------------#
                    def speiselöschen():
                        global counts
                        global count
                        # Get all the children nodes of the BestellTree
                        siko = BestellTree.get_children()
                        # Get the length of siko
                        zablo = len(siko)
                        # If there are no children, show an error message
                        if zablo == 0:
                            messagebox.showerror('Error', 'diese action ist nicht gültig')
                        else:
                            # Get the selected item in the BestellTree
                            selected = BestellTree.focus()
                            # Get the values of the selected item
                            values = BestellTree.item(selected, 'values')
                            # Get the current total value in the entrybox11
                            bako = float(entrybox11.get())
                            # Get the value of the selected item
                            nako = float(values[8])
                            # Subtract the selected item value from the current total value and update entrybox11
                            entrybox11.delete(0, END)
                            result = math.floor(float(bako - nako) * 100) / 100
                            formatted_result = '{:.2f}'.format(result)
                            entrybox11.insert(0, formatted_result)
                            # Delete the selected item from the BestellTree
                            one = BestellTree.selection()[0]
                            BestellTree.delete(one)
                            counts-=1

                    # Create a button to call the speiselöschen function
                    speiselösch = Button(Mainframe3, font=("Helvetica", 11, 'bold'), text=' Speise Löschen', bd=4,
                                         bg=colour2,
                                         command=speiselöschen)
                    speiselösch.place(x=205, y=0)

                    # ----------------------------------------------- Berchnen ABBRECHEN ---------------------------------------------------#
                    def berechabbrech():
                        # enable entry boxes
                        entrybox10.config(state=NORMAL)
                        entrybox9.config(state=NORMAL)
                        entrybox8.config(state=NORMAL)

                        # delete the values in the entry boxes and set them to 0
                        entrybox10.delete(0, END)
                        entrybox9.delete(0, END)

                        entrybox10.insert(0, float(0.0))
                        entrybox9.insert(0, float(0.0))
                        entrybox8.delete(0, END)
                        entrybox8.insert(0, float(0.0))

                        # create a list of floats for the prices of all the items in the order
                        floats = []
                        for line in BestellTree.get_children():
                            sick108 = (BestellTree.item(line[0])['values'][8])
                            floats.insert(0, float(sick108))

                        # calculate the total price and format it to 2 decimal places
                        total = math.fsum(floats)
                        result = math.floor(float(total) * 100) / 100
                        formatted_result = '{:.2f}'.format(result)

                        # update the total price in the entry box
                        entrybox11.delete(0, END)
                        entrybox11.insert(0, formatted_result)

                    # create a button to cancel the order calculation and reset the entry boxes
                    # btnrechab = Button(Mainframe2, font=('arial', 11, 'bold'), text='Abbrechen', bd=4, bg=colour1, pady=1,
                    #                    padx=24,
                    #                    width=3, height=1, command=berechabbrech)
                    # btnrechab.place(x=1065, y=295)

                    # get the value of the last row in the Liefergeld table in the database
                    cur12 = connE.cursor()
                    cur12.execute('select * from Liefergeld')
                    sol12 = cur12.fetchall()
                    for idolo in sol12:
                        for ido in idolo:
                            ido = ido

                    # if the value is greater than 0, create a checkbox to deactivate the delivery fee once




                    #####################################   Bestellung Data und Drücken  ###########################################
                    def addbestellungdata(event=None):
                        global text1, coto, count, counts, pos, abholz, abholE,endframe,Bestellung_frame
                        mos = 0
                        counts = 0
                        count = 0
                        pos = 0
                        #make sure that all speisen are added else show error
                        empty=entrybox1.get()
                        if empty !='':
                            messagebox.showwarning('warnung','Speise Bitte bestätigen ')
                        else:

                            end = float(entrybox11.get())
                            rabb = float(entrybox10.get())
                            liefermoney=entrybox8.get()
                            if liefermoney=='':
                                liefermoney=0.0


                            # calculate price with discount
                            if rabb != 0.0:
                                # disable discount entry box and clear price entry box
                                entrybox11.delete(0, END)


                                # apply discount percentage to each item in the tree
                                updated_records = []
                                for item in BestellTree.get_children():
                                    values = BestellTree.item(item)['values']
                                    old_price = float(values[8])
                                    new_price = old_price * (100.0 - rabb) / 100.0  # calculate new price with discount
                                    rop = math.floor(float(new_price) * 100) / 100
                                    formatted_result = '{:.2f}'.format(rop)
                                    values[8] = formatted_result
                                    updated_record = tuple(values)
                                    updated_records.append(updated_record)

                                # clear the tree and insert the updated records

                                BestellTree.delete(*BestellTree.get_children())  # delete all records in the tree

                                for record in updated_records:
                                    BestellTree.insert('', 'end', values=record)

                                # calculate discounted price

                            # for item in BestellTree.get_children():
                            #     values = BestellTree.item(item)['values']
                            # sumprice = float (sum(values[8]))

                            floats = []
                            for line in BestellTree.get_children():
                                sick108 = (BestellTree.item(line)['values'][8])
                                floats.insert(0, float(sick108))

                            # calculate the total price and format it to 2 decimal places
                            total = math.fsum(floats)
                            result = math.floor(float(total) * 100) / 100
                            formatted_result = '{:.2f}'.format(result)

                            # update the total price in the entry box
                            entrybox11.delete(0, END)
                            entrybox11.insert(0, formatted_result)
                            # Check if at least one item is selected in the treeview
                            for i in BestellTree.get_children():
                                if not BestellTree.item(i, "values"):
                                    mos = 0
                                else:
                                    mos = 1
                            if mos == 0:
                                # If no items are selected, display an error message
                                messagebox.showerror('ERROR', 'Bitte Mindesten eine Speise Aüswahlen')



                            else:

                                cur12 = connE.cursor()
                                cur12.execute('select * from Lieferzuschlg')
                                cob = cur12.fetchall()
                                # Get the delivery charge from the database
                                ako = 0
                                for ido in cob:
                                    if cob:
                                        for ako in ido:
                                            ako = float(ako)
                                floats = []
                                moks = []
                                # Get customer information from the GUI
                                sick1 = KundenidE.get()
                                sick2 = TelefonnummerE.get()
                                sick3 = NameE.get()
                                sick4 = AdresseE.get()
                                sick5 = HauesnrE.get()
                                sick6 = PLZE.get()
                                sick7 = ORTE.get()
                                sick8 = EmailE.get()
                                sick9 = passo[1]
                                comment = ext_comment.get(1.0, END)
                                timed = datetime.now().strftime("%d-%m-%y %H:%M:%S") + ' Uhr'
                                moss = (BestellTree.get_children())
                                long = (len(moss))
                                sol = []
                                entrybox10.config(state=NORMAL)
                                entrybox9.config(state=NORMAL)
                                entrybox8.config(state=NORMAL)
                                entrybox7.config(state=NORMAL)
                                if sick4 == 'ABHOLUNG':
                                    if abholE.get() == '':
                                        abhol_z = 'Sofort'
                                    else:
                                        abhol_z = abholE.get()

                                # Loop through the items in the treeview and insert them into the speiseinfo table
                                conn=connO
                                cur=conn.cursor()
                                if float(liefermoney)>0.0:
                                    cur.execute('insert into lieferpara values(?,?)',(liefermoney,timed))
                                    conn.commit()
                                if float(rabb)!=0.0:
                                    cur.execute('insert into Rabattpara values(?,?)', (rabb, timed))
                                    conn.commit()

                                if BestellTree.get_children():
                                    # Get values from all items in the treeview
                                    values = [[BestellTree.item(item)['values'][i] for i in range(9)]
                                              for item in BestellTree.get_children()]

                                    # Create a cursor for the database connection
                                    cur1 = connO.cursor()

                                    # Do something with the values:
                                    for sick100, sick101, sick102, sick103, sick104, sick105, sick106, sick107, sick108 in values:
                                        # Do something with each value


                                        # Insert the item into the speiseinfo table
                                        cur1.execute("""INSERT INTO speiseinfo (zeit, pos, grosse, anzahl, nr, speise, mit, ohne, katagorie, preis, name) 
                                                        VALUES (:zeit, :pos, :grosse, :anzahl, :nr, :speise, :mit, :ohne, :katagorie, :preis, :name)""",
                                                     {'zeit': timed,
                                                      'pos': sick100,
                                                      'grosse': sick101,
                                                      'anzahl': sick102,
                                                      'nr': sick103,
                                                      'speise': sick104,
                                                      'mit': sick105 + ' ',
                                                      'ohne': sick106,
                                                      'katagorie': sick107,
                                                      'preis': sick108,
                                                      'name': sick3
                                                      })
                                        connO.commit()
                                # Get the delivery charge and delivery surcharge from the database
                                cur12 = connE.cursor()
                                cur13 = connE.cursor()
                                cur12.execute('select * from Liefergeld')
                                liefergeld = cur12.fetchall()
                                cur13.execute('select * from Lieferzuschlg ')
                                llieferzuschalg = cur13.fetchall()

                                # Get the current state of the checkbox


                                # Get the total amount and check if the delivery surcharge needs to be added
                                sick15 = entrybox11.get()

                                # Check if there is a delivery surcharge and add it to the total amount
                                if llieferzuschalg:
                                    for izo1 in llieferzuschalg:
                                        lieferzuschalg = str(izo1).replace("'", "").replace(',', '').replace(')', '').replace('(',
                                                                                                                              '')
                                        lieferzuschalg = float(lieferzuschalg)
                                        if float(sick15) < lieferzuschalg and sick4 != 'ABHOLUNG' and sick4 != 'IM HAUS':
                                            mok = messagebox.askyesno("Lieferzuschlag", "soll lieferzuschlag gerechnet werden ?")
                                            if mok == 1:
                                                dif = lieferzuschalg - float(sick15)
                                                sick15 = float(sick15) + float(dif)


                                # Check if there is any discount and add it to the total amount


                                if float(liefermoney) > 0  and sick4 != 'ABHOLUNG' and sick4 != 'IM HAUS':
                                    sick15 = float(liefermoney) + float(sick15)


                                # Get the delivery time from the entry box
                                # get the user input


                                # get the user input
                                sick12 = LieferE.get()

                                # set delivery time to None if user input is empty
                                if not sick12:
                                    time_obj = None
                                else:
                                    # try to parse the time string
                                    try:
                                        # first try to parse as 24-hour format
                                        time_obj = datetime.strptime(sick12, '%H:%M')
                                    except ValueError:
                                        # if parsing fails, try to parse as plain integer (e.g. "1800")
                                        try:
                                            time_obj = datetime.strptime(sick12, '%H%M')
                                        except ValueError:
                                            # if parsing fails again, show an error message to the user
                                            messagebox.showerror('Invalid time',
                                                                 'Bitte die Lieferziet korrigieren (Z.b. "1800" oder "18:00").')

                                # convert the time object back to a string in the desired format
                                if time_obj is None:
                                    sick12 = ''
                                else:
                                    sick12 = datetime.strftime(time_obj, '%H:%M')

                                # Insert the customer information into the kundeinfo table
                                cur = connO.cursor()
                                cur.execute("""INSERT  into kundeinfo (ID,kid,ktelefonnummer,kname,kstrasse,khausnr,
                                                kplz,kort,kemail,bediener,zeit,bestellzeit,externinfo,gesamtepreis)
                                             VALUES(NULL,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                                            (sick1,
                                             sick2,
                                             sick3,
                                             sick4,
                                             sick5,
                                             sick6,
                                             sick7,
                                             sick8,
                                             sick9,
                                             timed,
                                             sick12,
                                             comment,
                                             sick15
                                             ))
                                connO.commit()

                                # Get the data from the kundeinfo and speiseinfo tables
                                cur.execute('select* from kundeinfo')
                                cur1.execute('select* from speiseinfo where zeit=(?) ', (timed,))
                                zob = cur1.fetchall()
                                hob = cur.fetchall()
                                hib = len(hob)
                                hub = hib
                                sor = (BestellTree.get_children())
                                coto = 0

                                # Configure the text box for printing the receipt
                                import win32ui
                                import win32print
                                import win32print
                                default_printer = win32print.GetDefaultPrinter()
                                show_bestellung()

                                def printing101():
                                    # Declare variables
                                    global logo1, coto, text1, mok, imag12, imag13, iamg1,print_fast
                                    print_fast = vardo.get()

                                    mit = []
                                    alle = []
                                    mab = []
                                    gesamt = []
                                    lon = []
                                    ohne = []

                                    # Get the categories from the database
                                    cur = connE.cursor()
                                    cur.execute('select* from Katagorie ')
                                    fur = cur.fetchall()

                                    # Get the state of the checkbox
                                    #
                                    # if float(liefermoney) <0.01:
                                    #
                                    #     # Calculate the total if the checkbox is unchecked
                                    #     if BestellTree.get_children():
                                    #         # Get values from all items in the treeview
                                    #         values = [[BestellTree.item(item)['values'][i] for i in range(9)]
                                    #                   for item in BestellTree.get_children()]
                                    #         # Sum the total of the values
                                    #         total = math.fsum([float(item[8]+float(liefermoney)) for item in values])

                                    # Get the delivery charge values from the database
                                    dold = AdresseE.get()
                                    cur12 = connE.cursor()
                                    cur12.execute('select * from Liefergeld')
                                    sol12 = cur12.fetchall()
                                    # for izo12 in sol12:
                                    #     sazo12 = str(izo12).replace("'", "").replace(',', '').replace(')', '').replace('(', '')

                                    # Get all the categories
                                    for kat in fur:
                                        top = str(kat)
                                        replace = {'[': '',
                                                   ']': '',
                                                   '(': '',
                                                   ')': '',
                                                   ',': '',
                                                   "'": ''}
                                        tops = top.translate(str.maketrans(replace))
                                        alle.insert(0, tops)
                                    lenalle = len(alle)

                                    # Get the information about the ordered items from the treeview
                                    sicked = BestellTree.get_children()
                                    for tom in sicked:
                                        values = BestellTree.item(tom)['values']
                                        sick100 = values[0]
                                        sick101 = values[1]
                                        sick102 = values[2]
                                        sick103 = values[3]
                                        sick104 = values[4]
                                        sick105 = values[5]
                                        sick106 = values[6]
                                        sick107 = values[7]
                                        sick108 = values[8]
                                        mab.insert(0, sick107)

                                    # Clear the text widget
                                    # text1.delete(1.0, END)

                                    # Check if there is a comment
                                    comment1 = len(comment)

                                    from openpyxl import Workbook, load_workbook
                                    from openpyxl.styles import Font
                                    from openpyxl.styles import PatternFill
                                    from openpyxl.utils import get_column_letter

                                    # Use the `os` module to delete the workbook
                                    try:
                                        file_path = resource_path('Data\printer23.xlsx')
                                        os.remove(file_path)
                                    except:
                                        pass
                                    wa = load_workbook(resource_path('Data\printer2.xlsx'))

                                    # Save the workbook as a new file
                                    wa.save(resource_path('Data\printer23.xlsx'))

                                    # Create an instance of the workbook
                                    wb = load_workbook(resource_path('Data\printer23.xlsx'))
                                    ws = wb.active
                                    ws.column_dimensions['C'].width = 12
                                    ws.column_dimensions['D'].width = 5
                                    row_num = 2
                                    # Define the new height (in this example, 30)
                                    new_height = 15
                                    # Set the row height for the specified row number
                                    wb.save(resource_path('Data\printer23.xlsx'))
                                    # Select the active sheet
                                    # Save the changes to the workbook
                                    # cell1=ws['A1']
                                    # cell2=ws['B2']
                                    # cell3=ws['A3']
                                    # cell4 = ws['A4']
                                    # cell5 = ws['A5']
                                    # cell6 = ws['A6']
                                    # cell7 = ws['A7']
                                    # cell8 = ws['B4']
                                    # cell9 = ws['B7']
                                    # cell10 = ws['A8']

                                    # # Define the new height (in this example, 30)
                                    new_height = 5
                                    # # Set the row height for the specified row number
                                    ws.row_dimensions[row_num].height = new_height
                                    # # Get the row object for the specified row number

                                    # Define the color to be used (in this example, yellow)
                                    fill_color = PatternFill(start_color="FF0F00", end_color="FF0F00", fill_type="solid")
                                    light_fill_color = PatternFill(start_color="FF9F60", end_color="FF9F60", fill_type="solid")
                                    # Iterate through each cell in the row
                                    smallfont = Font(name='Calibri', size=11, italic=True,bold=True)
                                    nfont = Font(name='Calibri', size=12, bold=True, italic=True)
                                    font = Font(name='Calibri', size=18, bold=True, italic=True)
                                    add_font = Font(name='Calibri', size=16, bold=True, italic=True)
                                    font1 = Font(name='Calibri', size=25, bold=True, italic=True)
                                    ufont = Font(underline='single')
                                    # Write the first row with the headers
                                    if sick4 == 'ABHOLUNG':
                                        import datetime
                                        now = datetime.datetime.now()

                                        # Add 20 minutes
                                        if abhol_z.isdigit():
                                            future_time = now + datetime.timedelta(minutes=int(abhol_z))

                                            # Convert the future time to a string in a specific format
                                            formatted_time = future_time.strftime("%H:%M")
                                        else:
                                            formatted_time = abhol_z

                                        ws.oddHeader.center.text = (f'{sick4}  ZU {formatted_time}')
                                        ws.oddHeader.center.size = 20
                                        ws.oddHeader.center.font = "Tahoma,Bold"
                                        ws.oddHeader.center.color = "FF0F00"
                                    elif sick4 == 'IM HAUS':
                                        ws.oddHeader.center.text = ('IM HAUS')
                                        ws.oddHeader.center.size = 20
                                        ws.oddHeader.center.font = "Tahoma,Bold"
                                        ws.oddHeader.center.color = "FF0F00"

                                    else:
                                        ws.oddHeader.center.text = (f'LIEFERUNG\nA{hub}')
                                        ws.oddHeader.center.size = 20
                                        ws.oddHeader.center.font = "Tahoma,Bold"
                                        ws.oddHeader.center.color = "FF0F00"
                                    next_row = ws.max_row + 3
                                    # ws['B{}'.format(next_row)] = f'      A{hub}'
                                    # NUM_row = ws[f'B{next_row}']
                                    # next_row = ws.max_row +1
                                    #
                                    # NUM_row.font = add_font

                                    ws['A{}'.format(next_row)] = f'\U0001F550 {timed[0:8]}  {timed[9:]}  '
                                    next_row = ws.max_row + 1
                                    ws['A{}'.format(next_row)] = f'kd:{sick1}'
                                    ws['B{}'.format(next_row)] = f' \u260E:{sick2}'
                                    next_row = ws.max_row + 1
                                    ws['A{}'.format(next_row)] = f'Name:  {sick3}'
                                    add = f'{sick4} {sick5}'
                                    next_row = ws.max_row + 1

                                    wrapped_add = textwrap.wrap(add, width=25)
                                    for line in wrapped_add:
                                        ws['A{}'.format(next_row)] = line
                                        add_row = ws[f'A{next_row}']
                                        add_row.font = add_font
                                        next_row = ws.max_row + 1
                                    next_row = ws.max_row + 1
                                    ws['A{}'.format(next_row)] = f'{sick6}   {sick7}'
                                    add_row_1 = ws[f'A{next_row}']
                                    add_row_1.font = add_font
                                    next_row = ws.max_row + 1

                                    if sick12 != '' and  sick4 != 'ABHOLUNG':
                                        # next_row = ws.max_row + 1
                                        # ws['B{}'.format(next_row)] = 'VORBESTELLUNG'
                                        next_row = ws.max_row + 1
                                        ws['B{}'.format(next_row)] = 'VORBESTELLUNG'
                                        vorbestellen_row = ws[f'A{next_row}']
                                        vorbestellen_row1 = ws[f'B{next_row}']
                                        vorbestellen_row2 = ws[f'C{next_row}']
                                        vorbestellen_row3 = ws[f'D{next_row}']
                                        vorbestellen_row.fill = light_fill_color
                                        vorbestellen_row1.fill = light_fill_color
                                        vorbestellen_row2.fill = light_fill_color
                                        vorbestellen_row3.fill = light_fill_color

                                        vorbestellen_row.font = font
                                        next_row = ws.max_row + 1

                                        ws['B{}'.format(next_row)] = (f'ZU  {sick12}  UHR LIEFERN ')
                                        liefern_row = ws[f'A{next_row}']
                                        liefern_row1 = ws[f'B{next_row}']
                                        liefern_row2 = ws[f'C{next_row}']
                                        liefern_row3 = ws[f'D{next_row}']
                                        liefern_row.fill = light_fill_color
                                        liefern_row1.fill = light_fill_color
                                        liefern_row2.fill = light_fill_color
                                        liefern_row3.fill = light_fill_color

                                        next_row = ws.max_row + 2
                                    elif sick12 != '' and  sick4 == 'ABHOLUNG':
                                        # next_row = ws.max_row + 1
                                        # ws['B{}'.format(next_row)] = 'VORBESTELLUNG'
                                        next_row = ws.max_row + 1
                                        ws['B{}'.format(next_row)] = 'VORBESTELLUNG'
                                        vorbestellen_row = ws[f'A{next_row}']
                                        vorbestellen_row1 = ws[f'B{next_row}']
                                        vorbestellen_row2 = ws[f'C{next_row}']
                                        vorbestellen_row3 = ws[f'D{next_row}']
                                        vorbestellen_row.fill = light_fill_color
                                        vorbestellen_row1.fill = light_fill_color
                                        vorbestellen_row2.fill = light_fill_color
                                        vorbestellen_row3.fill = light_fill_color

                                        vorbestellen_row.font = font
                                        next_row = ws.max_row + 1

                                        ws['B{}'.format(next_row)] = (f'ZU  {sick12}  UHR ABHOLUNG ')
                                        liefern_row = ws[f'A{next_row}']
                                        liefern_row1 = ws[f'B{next_row}']
                                        liefern_row2 = ws[f'C{next_row}']
                                        liefern_row3 = ws[f'D{next_row}']
                                        liefern_row.fill = light_fill_color
                                        liefern_row1.fill = light_fill_color
                                        liefern_row2.fill = light_fill_color
                                        liefern_row3.fill = light_fill_color

                                        next_row = ws.max_row + 2
                                    if comment1 > 1:
                                        # Write the title of the comment section to the next row

                                        next_row = ws.max_row + 1
                                        ws['A{}'.format(next_row)] = 'Kunden Komentare'

                                        komment_row1 = ws[f'A{next_row}']
                                        komment_row1.font = font

                                        next_row = ws.max_row + 1
                                        # Split the comment into multiple lines
                                        wrapped_comment = textwrap.wrap(comment, width=35)

                                        # Write each line of the comment to a separate row
                                        for line in wrapped_comment:
                                            ws['A{}'.format(next_row)] = line
                                            next_row = ws.max_row + 2

                                    # Update the row number for the next record
                                    row_num = ws.max_row + 2

                                    row_num = next_row

                                    ws.row_dimensions[row_num].height = new_height
                                    EIGHTROW = ws[f'A{next_row}']
                                    EIGHTROW1 = ws[f'B{next_row}']
                                    EIGHTROW2 = ws[f'C{next_row}']
                                    EIGHTROW3 = ws[f'D{next_row}']

                                    next_row = ws.max_row + 1
                                    # loop through the speisen and add every katagorie
                                    if lenalle > 0:
                                        for i in range(lenalle):
                                            if alle[i] in mab:
                                                sob = mab.count(alle[i])
                                                ws['B{}'.format(next_row)] = (f'--------------------------')
                                                next_row += 1
                                                katrow1 = ws[next_row]
                                                katrow1[1].font = nfont
                                                ws['B{}'.format(next_row)] = (f'------{alle[i]}----{sob}--')
                                                next_row += 1
                                                ws['B{}'.format(next_row)] = (f'--------------------------')
                                                next_row += 1
                                            sicked = BestellTree.get_children()
                                            for tom in sicked:
                                                values = BestellTree.item(tom)['values']
                                                sick100 = values[0]
                                                sick101 = values[1]
                                                sick102 = values[2]
                                                sick103 = values[3]
                                                sick104 = values[4]
                                                sick105 = values[5]
                                                sick106 = values[6]
                                                sick107 = values[7]
                                                sick108 = values[8]
                                                if sick107 == alle[i]:

                                                    def truncate_text(text, max_length):
                                                        if len(text) > max_length:
                                                            return text[:max_length]
                                                        else:
                                                            return text

                                                    max_length = 38 # Define your maximum length here
                                                    max_length1 = 5

                                                    if sick101 != 'Standard':
                                                        katrow1 = ws[next_row]
                                                        katrow1[0].font = nfont
                                                        cell_content = f'{sick102}x({sick103} {sick101}){sick104} '
                                                        ws['A{}'.format(next_row)] = truncate_text(cell_content, max_length)
                                                        cell_content1 =f'{sick108}€'
                                                        ws['D{}'.format(next_row)]=truncate_text(cell_content1, max_length1)


                                                        katrow101 = ws[next_row]
                                                        katrow101[1].font = smallfont
                                                        next_row+=1





                                                    elif float(sick108) > 0:
                                                        katrow1 = ws[next_row]
                                                        katrow1[0].font = nfont
                                                        ws['A{}'.format(next_row)] = (f' {sick102}x ({sick103}) {sick104}')
                                                        cell_content1 = f'{sick108}€'
                                                        ws['D{}'.format(next_row)] = truncate_text(cell_content1, max_length1)

                                                        next_row += 1


                                                    elif float(sick108) == 0.00:
                                                        ws['A{}'.format(next_row)] = (
                                                            f'              {sick104}     ')

                                                    if sick106 != "":
                                                        for without in [sick106]:
                                                            within = str(without).split(' ')
                                                        ohne.append(within)
                                                        for tim in ohne:
                                                            for taza in tim:
                                                                ws['B{}'.format(next_row)] = (f'  {taza}')
                                                                next_row += 1
                                                                ohne.clear()

                                                    if sick105 != "":
                                                        for izo in [sick105]:
                                                            hozo = str(izo).split(' ')
                                                        mit.append(hozo)
                                                        for azo in mit:
                                                            for tizo in azo:
                                                                if sick104 == 'Familien Pizza':
                                                                    unwraped_Familie = tizo
                                                                    wrapped_Familie = textwrap.wrap(unwraped_Familie, width=36)
                                                                    for line in wrapped_Familie:
                                                                        ws['A{}'.format(next_row)] = (f'{line}')
                                                                        next_row += 1
                                                                        mit.clear()
                                                                else:

                                                                    ws['B{}'.format(next_row)] = (f'{tizo}')
                                                                    next_row += 1
                                                                    mit.clear()

                                    row_num = next_row
                                    row = ws[row_num]
                                    # Define the color to be used (in this example, yellow)
                                    ws.row_dimensions[row_num].height = new_height
                                    # Iterate through each cell in the row
                                    for cell in row:
                                        cell.fill = fill_color
                                    next_row += 1
                                    katrow5 = ws[next_row]
                                    katrow5[1].font = nfont
                                    formatted_gesamt = '{:.2f}'.format(float(sick15))
                                    ws['B{}'.format(next_row)] = (f"\n\t Gesamtpreis: {formatted_gesamt}€")
                                    next_row += 1
                                    if dold.strip() != 'ABHOLUNG' and dold.strip() != 'IM HAUS':

                                        try:
                                            if dif:

                                                formatted_dif = '{:.2f}'.format(dif)
                                                ws['A{}'.format(next_row)] = (f"\n\tLieferzuschlag: {formatted_dif}€")
                                                next_row += 1
                                        except:
                                            pass
                                        if float(liefermoney) > 0  :
                                            # katrow3 = ws[next_row]
                                            # katrow3[0].font = nfont
                                            formatted_dif = '{:.2f}'.format(float(liefermoney))
                                            ws['A{}'.format(next_row)] = (
                                                f"\n\tincl Liefergeld: {str(liefermoney).replace('[', '').replace(']', '')}€")
                                            next_row += 1
                                    mwst = (7 * float(sick15)) / 100.0
                                    mwst1 = (math.floor(float(mwst) * 100) / 100)
                                    mwst_haus = (19 * float(sick15)) / 100.0
                                    mwst1_haus = (math.floor(float(mwst) * 100) / 100)
                                    # katrow4 = ws[next_row]
                                    # katrow4[0].font = nfont
                                    if sick4 != 'IM HAUS':
                                        ws['A{}'.format(next_row)] = (f"\n incl 7%Umsatzsteuer :\t{mwst1}€")
                                        next_row += 1
                                        row_num = next_row
                                        row = ws[row_num]
                                    else:
                                        ws['A{}'.format(next_row)] = (f"\n incl 19%Umsatzsteuer :\t{mwst1_haus}€")
                                        next_row += 1
                                        row_num = next_row
                                        row = ws[row_num]
                                    if rabb>0:
                                        ws['A{}'.format(next_row)] = (f"\n incl Rabatt \t{rabb}%")
                                        next_row += 1
                                        row_num = next_row
                                        row = ws[row_num]
                                    # Define the color to be used (in this example, yellow)
                                    ws.row_dimensions[row_num].height = new_height
                                    # Iterate through each cell in the row
                                    for cell in row:
                                        cell.fill = fill_color
                                    next_row += 1


                                    # Write each line of the comment to a separate row

                                    ws['A{}'.format(next_row)] = (f'--------------------------------------------------')
                                    next_row = ws.max_row + 1
                                    ws['B{}'.format(next_row)] = (f'{restName}')
                                    next_row = ws.max_row + 1
                                    ws['B{}'.format(next_row)] = (f'{restStr}')
                                    next_row = ws.max_row + 1
                                    ws['B{}'.format(next_row)] = (f'{restTele}')
                                    next_row = ws.max_row + 1
                                    ws['A{}'.format(next_row)] = (f'--------------------------------------------------')
                                    next_row = ws.max_row + 1
                                    katrow1 = ws[next_row]
                                    katrow1[1].font = add_font
                                    ws['B{}'.format(next_row)] = ('Lieferschein')




                                    # cell1.font = font
                                    # cell2.font = font1
                                    # cell3.font = nfont
                                    # cell4.font = nfont
                                    # cell5.font = nfont
                                    # cell6.font = nfont
                                    # cell7.font = nfont
                                    # cell8.font = nfont
                                    # cell9.font = nfont
                                    # cell10.font = ufont
                                    EIGHTROW.fill = fill_color
                                    EIGHTROW1.fill = fill_color
                                    EIGHTROW2.fill = fill_color
                                    EIGHTROW3.fill = fill_color

                                    wb.save(resource_path('Data\printer23.xlsx'))
                                    liefer1.config(fg='red')
                                    # reset everything to start a new order
                                    update(toppings)
                                    remove_all()
                                    TelefonnummerE.focus_force()

                                def printer1():

                                    cur = connE.cursor()
                                    cur.execute('select * from Katagorie')
                                    mik = cur.fetchall()
                                    sazo12=entrybox8.get()
                                    if sazo12=='':
                                        sazo12=0.0

                                    listi = []
                                    listo = []

                                    for ray in mik:

                                        for element in ray:
                                            if element:
                                                listo.insert(0, element)
                                        for iop in listo:
                                            result = str(iop).replace('[', '').replace(']', '').replace('(', '').replace(')',
                                                                                                                         '').replace(
                                                ',',
                                                '').replace(
                                                "'", "").replace(' ', '')
                                            if result != '':
                                                listi.insert(0, result)

                                    all = []
                                    cur.execute('select * from Katagorie')
                                    fur = cur.fetchall()

                                    for kat in fur:

                                        top = str(kat)
                                        replace = {'[': '',
                                                   ']': '',
                                                   '(': '',
                                                   ')': '',
                                                   ',': '',
                                                   "'": ''}
                                        tops = top.translate(str.maketrans(replace))
                                        if tops in listi:
                                            all.insert(0, tops)

                                    tol = len(all)
                                    sicked = BestellTree.get_children()
                                    mab = [BestellTree.item(tom)['values'][7] for tom in sicked]
                                    for i in all:
                                        if i in mab:

                                            sicked = BestellTree.get_children()
                                            mab = [BestellTree.item(tom)['values'][7] for tom in sicked]

                                            text2.delete(1.0, END)
                                            if AdresseE.get()=='ABHOLUNG':

                                                import datetime
                                                now = datetime.datetime.now()

                                                # Add 20 minutes
                                                if abhol_z.isdigit():
                                                    future_time = now + datetime.timedelta(minutes=int(abhol_z))

                                                    # Convert the future time to a string in a specific format
                                                    formatted_time = future_time.strftime("%H:%M")
                                                else:
                                                    formatted_time = abhol_z
                                                text2.insert(END, f' \t  ABHOLUNG \n\n')
                                                text2.insert(END,f' \t ZU {formatted_time}\n\n')

                                            elif AdresseE.get()=='IM HAUS':
                                                text2.insert(END, f' \t  IM HAUS \n\n')
                                            else:
                                                text2.insert(END, f' \t  LIEFERUNG \n\n')

                                            comment1 = len(comment)
                                            if comment1 >1:
                                                text2.insert(END, f'======================\n')
                                                text2.insert(END, f' {comment}\n')
                                                text2.insert(END, f'======================\n')
                                            if sick12 != '':

                                                text2.insert(END, f'\n VORBESTELLUNG')
                                                text2.insert(END,f'\n ZU  {sick12}  UHR LIEFERN \n')
                                            text2.insert(END, f'\n======================')
                                            text2.insert(END, f'\n{timed}\tBes_nr:{hub}')
                                            text2.insert(END, f'\n kd:{sick1}\tTel:{sick2}')
                                            text2.insert(END, f'\n{sick3}')
                                            text2.insert(END, f'\n{sick4}  \t{sick5}')
                                            text2.insert(END, f'\n{sick6}  \t{sick7}')
                                            text2.insert(END, f'\n======================')


                                            for i in range(tol):
                                                if all[i] in mab:

                                                    sob = mab.count(all[i])
                                                    text2.insert(END, f'\n-------{all[i]}------{sob}--')
                                                    for tom in sicked:
                                                        values = BestellTree.item(tom)['values']
                                                        if values[7] == all[i]:
                                                            text2.insert(END,
                                                                         f'\n{values[2]}x) {values[4]} ({values[1]}) {values[8]}€')
                                                            if values[5] != "":
                                                                for izo in [values[5]]:
                                                                    mit = str(izo).split(' ')
                                                                    for azo in mit:
                                                                        text2.insert(END, f'\n\t{azo}')
                                                            text2.insert(END, f'\n\t{values[6]}\n')

                                            text2.insert(END, f"\n=======================\n")
                                            text2.insert(END, f"\n\t Gesamtpreis:  {sick15}€")
                                            for ito in moks:
                                                if int(ito) > 1:
                                                    text2.insert(END,
                                                                 f"\n\t\t Lieferzuschlag:\t{str(ito).replace('[', '').replace(']', '')}€")

                                            if AdresseE.get() != 'ABHOLUNG' or AdresseE.get() != 'IM HAUS':
                                                liefer1.config(fg='red')
                                                liefer1.config(state=NORMAL)
                                                liefer.config(state=NORMAL)
                                            mwst = (7 * float(sick15)) / 100.0
                                            mwst1 = (math.floor(float(mwst) * 100) / 100)
                                            mwst_haus = (19 * float(sick15)) / 100.0
                                            mwst1_haus = (math.floor(float(mwst) * 100) / 100)
                                            if sick4 != 'IM HAUS':
                                                text2.insert(END,f"\n incl 7%Umsatzsteuer :{mwst1}€")



                                            else:
                                                text2.insert(END,f"\n incl 19%Umsatzsteuer :{mwst1_haus}€")
                                            dold = AdresseE.get()
                                            if dold.strip() != 'ABHOLUNG' and dold.strip() != 'IM HAUS':

                                                try:
                                                    if dif:
                                                        formatted_dif = '{:.2f}'.format(dif)
                                                        text2.insert(END,f"\n\tLieferzuschlag: {formatted_dif}€")

                                                except:
                                                    pass
                                                if float(sazo12) > 0 :
                                                    # katrow3 = ws[next_row]
                                                    # katrow3[0].font = nfont
                                                    formatted_dif = '{:.2f}'.format(float(sazo12))
                                                    text2.insert(END,
                                                        f"\n\tincl Liefergeld: {str(sazo12).replace('[', '').replace(']', '')}€")
                                                text2.insert(END, f"\n=======================\n")
                                                text2.insert(END, f'   {restName}\n')
                                                text2.insert(END, f'   {restStr}\n')
                                                text2.insert(END, f'   {restTele}')

                                            cur = connE.cursor()
                                            cur.execute('select printer from printer1 ')
                                            printins = cur.fetchall()

                                            printernames1 = str(printins).replace(',', '').replace("'", "").replace('(',
                                                                                                                    '').replace(
                                                ')', '').replace('[', '').replace(']', '').replace('"', "'")
                                            printername1 = printernames1

                                            printer_list = []
                                            printer_name = (repr(printername1.encode().decode()))
                                            last_printer = printer_name.replace('"', '').replace("'", "")

                                    import tempfile
                                    import os
                                    import win32api
                                    import win32print

                                    def print_to_default_printer(text):
                                        # Get the name of the default printer
                                        default_printer_name = win32print.GetDefaultPrinter()

                                        filename = tempfile.mktemp(".txt")
                                        with open(filename, 'w') as f:
                                            f.write(text)

                                        win32api.ShellExecute(0, "print", filename, '/d:"%s"' % default_printer_name, ".",
                                                              0)

                                    q = text2.get("1.0", "end-1c")
                                    if q != '':
                                        try:
                                            print_to_default_printer(q)
                                        except:
                                            messagebox.showwarning('Drucker',
                                                                   'Bestellung Gespeichert aber Drucker nicht gefunden')
                                    update(toppings)
                                    remove_all()
                                    TelefonnummerE.focus_force()
                                def printer2():

                                    cur = connE.cursor()
                                    cur.execute('select * from printer1')
                                    mik = cur.fetchall()

                                    listi = []
                                    listo = []

                                    for ray in mik:

                                        for element in ray:
                                            if element:
                                                listo.insert(0, element)
                                        for iop in listo:
                                            result = str(iop).replace('[', '').replace(']', '').replace('(', '').replace(')',
                                                                                                                         '').replace(
                                                ',',
                                                '').replace(
                                                "'", "").replace(' ', '')
                                            if result != '':
                                                listi.insert(0, result)

                                    all = []
                                    cur.execute('select * from Katagorie')
                                    fur = cur.fetchall()

                                    for kat in fur:

                                        top = str(kat)
                                        replace = {'[': '',
                                                   ']': '',
                                                   '(': '',
                                                   ')': '',
                                                   ',': '',
                                                   "'": ''}
                                        tops = top.translate(str.maketrans(replace))
                                        if tops in listi:
                                            all.insert(0, tops)

                                    tol = len(all)
                                    sicked = BestellTree.get_children()
                                    mab = [BestellTree.item(tom)['values'][7] for tom in sicked]
                                    for i in all:
                                        if i in mab:

                                            sicked = BestellTree.get_children()
                                            mab = [BestellTree.item(tom)['values'][7] for tom in sicked]

                                            text2.delete(1.0, END)
                                            text2.insert(END, f'\n   {restName}\n')
                                            text2.insert(END, f'\n===================================')
                                            text2.insert(END, f'\n{timed}\tBes_nr:{hub}')
                                            text2.insert(END, f'\n\n\n kd:{sick1}\t\ttel:{sick2}')
                                            text2.insert(END, f'\n{sick3}')
                                            text2.insert(END, f'\n{sick4}\t\t{sick5}')
                                            text2.insert(END, f'\n{sick6}\t\t{sick7}')
                                            text2.insert(END, f'\n===================================')
                                            for i in range(tol):
                                                if all[i] in mab:

                                                    sob = mab.count(all[i])
                                                    text2.insert(END, f'\n\t------------{all[i]}------{sob}--')
                                                    for tom in sicked:
                                                        values = BestellTree.item(tom)['values']
                                                        if values[7] == all[i]:
                                                            text2.insert(END,
                                                                         f'\n{values[2]}x) {values[4]} ({values[1]}) {values[8]}€')
                                                            if values[5] != "":
                                                                for izo in [values[5]]:
                                                                    mit = str(izo).split(' ')
                                                                    for azo in mit:
                                                                        text2.insert(END, f'\n\t{azo}')
                                                            text2.insert(END, f'\n\t{values[6]}\n')

                                            text2.insert(END, f"\n==================")
                                            text2.insert(END, f"\n\t Gesamtpreis:  {sick15}€")
                                            for ito in moks:
                                                if int(ito) > 1:
                                                    text2.insert(END,
                                                                 f"\n\t\t Lieferzuschlag:\t{str(ito).replace('[', '').replace(']', '')}€")

                                            if AdresseE.get() != 'ABHOLUNG' or AdresseE.get() != 'IM HAUS':
                                                liefer1.config(fg='red')
                                                liefer1.config(state=NORMAL)
                                                liefer.config(state=NORMAL)

                                            cur = connE.cursor()
                                            cur.execute('select printer from printer1 ')
                                            printins = cur.fetchall()

                                            printernames1 = str(printins).replace(',', '').replace("'", "").replace('(',
                                                                                                                    '').replace(
                                                ')', '').replace('[', '').replace(']', '').replace('"', "'")
                                            printername1 = printernames1

                                            printer_list = []
                                            printer_name = (repr(printername1.encode().decode()))
                                            last_printer = printer_name.replace('"', '').replace("'", "")

                                            def print_to_printer(printer_name, text):
                                                import tempfile
                                                import os
                                                import win32api
                                                import win32print

                                                # Set the default printer
                                                win32print.SetDefaultPrinter(printer_name)

                                                filename = tempfile.mktemp(".txt")
                                                with open(filename, 'w') as f:
                                                    f.write(text)

                                                win32api.ShellExecute(0, "print", filename, '/d:"%s"' % printer_name, ".", 0)

                                            q = text2.get("1.0", "end-1c")

                                            if q != '':
                                                try:
                                                    print_to_printer(last_printer, q)
                                                except:
                                                    messagebox.showwarning('Drucker',
                                                                           'Bestellung Geschpeichert aber Drucker nicht gefunden')

                                def printer3():
                                    cur = connE.cursor()
                                    cur.execute('select * from printer2')
                                    mik = cur.fetchall()

                                    listi = []
                                    listo = []

                                    for ray in mik:

                                        for element in ray:
                                            if element:
                                                listo.insert(0, element)
                                        for iop in listo:
                                            result = str(iop).replace('[', '').replace(']', '').replace('(', '').replace(')',
                                                                                                                         '').replace(
                                                ',',
                                                '').replace(
                                                "'", "").replace(' ', '')
                                            if result != '':
                                                listi.insert(0, result)

                                    all = []
                                    cur.execute('select * from Katagorie')
                                    fur = cur.fetchall()

                                    for kat in fur:

                                        top = str(kat)
                                        replace = {'[': '',
                                                   ']': '',
                                                   '(': '',
                                                   ')': '',
                                                   ',': '',
                                                   "'": ''}
                                        tops = top.translate(str.maketrans(replace))
                                        if tops in listi:
                                            all.insert(0, tops)

                                    tol = len(all)
                                    sicked = BestellTree.get_children()
                                    mab = [BestellTree.item(tom)['values'][7] for tom in sicked]
                                    for i in all:
                                        if i in mab:

                                            sicked = BestellTree.get_children()
                                            mab = [BestellTree.item(tom)['values'][7] for tom in sicked]

                                            text2.delete(1.0, END)
                                            text2.insert(END, f'\n   {restName}\n')
                                            text2.insert(END, f'\n===================================')
                                            text2.insert(END, f'\n{timed}\tBes_nr:{hub}')
                                            text2.insert(END, f'\n\n\n kd:{sick1}\t\ttel:{sick2}')
                                            text2.insert(END, f'\n{sick3}')
                                            text2.insert(END, f'\n{sick4}\t\t{sick5}')
                                            text2.insert(END, f'\n{sick6}\t\t{sick7}')
                                            text2.insert(END, f'\n===================================')
                                            for i in range(tol):
                                                if all[i] in mab:

                                                    sob = mab.count(all[i])
                                                    text2.insert(END, f'\n\t------------{all[i]}------{sob}--')
                                                    for tom in sicked:
                                                        values = BestellTree.item(tom)['values']
                                                        if values[7] == all[i]:
                                                            text2.insert(END,
                                                                         f'\n{values[2]}x) {values[4]} ({values[1]}) {values[8]}€')
                                                            if values[5] != "":
                                                                for izo in [values[5]]:
                                                                    mit = str(izo).split(' ')
                                                                    for azo in mit:
                                                                        text2.insert(END, f'\n\t{azo}')
                                                            text2.insert(END, f'\n\t{values[6]}\n')

                                            text2.insert(END, f"\n==================")
                                            text2.insert(END, f"\n\t Gesamtpreis:  {sick15}€")
                                            for ito in moks:
                                                if int(ito) > 1:
                                                    text2.insert(END,
                                                                 f"\n\t\t Lieferzuschlag:\t{str(ito).replace('[', '').replace(']', '')}€")

                                            if AdresseE.get() != 'ABHOLUNG' or AdresseE.get() != 'IM HAUS':
                                                liefer1.config(fg='red')
                                                liefer1.config(state=NORMAL)
                                                liefer.config(state=NORMAL)

                                            cur = connE.cursor()
                                            cur.execute('select printer from printer2 ')
                                            printins = cur.fetchall()

                                            printernames1 = str(printins).replace(',', '').replace("'", "").replace('(',
                                                                                                                    '').replace(
                                                ')', '').replace('[', '').replace(']', '').replace('"', "'")
                                            printername1 = printernames1

                                            printer_list = []
                                            printer_name = (repr(printername1.encode().decode()))
                                            last_printer = printer_name.replace('"', '').replace("'", "")

                                            def print_to_printer(printer_name, text):
                                                import tempfile
                                                import os
                                                import win32api
                                                import win32print

                                                # Set the default printer
                                                win32print.SetDefaultPrinter(printer_name)

                                                filename = tempfile.mktemp(".txt")
                                                with open(filename, 'w') as f:
                                                    f.write(text)

                                                win32api.ShellExecute(0, "print", filename, '/d:"%s"' % printer_name, ".", 0)

                                            q = text2.get("1.0", "end-1c")

                                            if q != '':
                                                try:
                                                    print_to_printer(last_printer, q)
                                                except:
                                                    messagebox.showwarning('Drucker',
                                                                           'Bestellung Geschpeichert aber Drucker nicht gefunden')

                                def printexcel():
                                    try:
                                        time.sleep(0.5)
                                        win32print.SetDefaultPrinter(default_printer)
                                        import win32com.client
                                        actualpath = resource_path('Data\printer23.xlsx')
                                        _path = os.path.abspath(actualpath)
                                        excel = win32com.client.Dispatch('Excel.Application')
                                        excel.Visible = False  # hide the Excel window
                                        wb = excel.Workbooks.Open(_path)
                                        wb.PrintOut()  # print the workbook
                                        wb.Close(SaveChanges=False)  # close the workbook without saving changes
                                        excel.Quit()
                                    except:
                                        MessageBox.showinfo('Drucker', 'Bestellung Geschpeichert aber Drucker nicht gefunden')
                                print_fast=vardo.get()
                                cur2 = connE.cursor()
                                cur2.execute('select* from DruckAnzahl')
                                nado2 = cur2.fetchall()
                                anzahl = ''

                                for t2 in nado2:
                                    anzahl = t2[1]

                                printer2()
                                printer3()


                                if print_fast >0:
                                    if anzahl > 1:
                                        for i in range(anzahl):
                                            printer1()
                                    else:
                                        printer1()

                                else:
                                    printing101()
                                    new_window2.update()

                                    if anzahl > 1:
                                        for i in range(anzahl):
                                            printexcel()
                                    else:
                                        printexcel()
                                entrybox10.delete(0,END)
                                entrybox10.insert(0,'0.0')
                                unlockin()
                                unlockall()
                            FIRSTLABEL = Label(Bestellung_frame, text='auf Lieferung oder Abholung drücken',
                                               font=("ARIEL", 26, "bold"),
                                               bg=colour4)
                            FIRSTLABEL.place(x=265, y=400)



                    end = Button(endframe, text='Drucken', width=8, bg=colour3, height=1, command=addbestellungdata).place(
                        x=2, y=2)
                    entrybox1.bind("<KeyPress-End>", addbestellungdata)

                    liefer = Label(Mainframe3, text='Mindestpreis :', bg=colour0, font=font_size, bd=2)
                    newtree.bind("<Double-Button-1>", save_name)
                    new_window2.mainloop()


            ########################################################################################################################
            btn = Button(root, text="Bestellung🍕", activebackground="blue",  bg=colour4,  width=12,height=2,
                         font=('Helvetica', 36,
                               "italic bold"),  fg="black", command=login)
            btn.pack(padx=10, pady=10)
            btn.place(x=80, y=250)
            global newebi, kundenframe2


            def zutatenpreise():
                # Clear any existing widgets in the kundenframe2 frame
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()

                # Create three frames to lay out the GUI
                kundenframe2.config(width=830, bg='SlateGray4',height=780)
                Frame1 = Frame(kundenframe2, width=450, bg='SlateGray4', bd=4, height=760, relief=RIDGE)
                Frame1.place(x=0, y=0)
                Frame2 = Frame(kundenframe2, width=200, bg='SlateGray4', bd=4, height=760, relief=RIDGE, padx=2)
                Frame2.place(x=610, y=0)
                Frame3 = Frame(kundenframe2, width=450, bg='white', height=760, relief=RIDGE)
                Frame3.place(x=40, y=110)

                # Create labels and entry widgets for the user to enter new ingredient prices
                font_ = ("ARIEL", 13, "bold")
                Zutat = Label(Frame1, text='Zutat', width=15, font=("arial", 16, "bold"), bd=4, bg='SlateGray4')
                Zutat.grid(row=1, column=2)
                ZutatE = Entry(Frame1, width=20, bd=4, font=font_)
                ZutatE.grid(row=2, column=2)
                Preis = Label(Frame1, text='Preis/€', width=15, font=("arial", 16, "bold"), bd=4, bg='SlateGray4')
                Preis.grid(row=1, column=3)
                PreisE = Entry(Frame1, width=8, bd=4, font=font_)
                PreisE.grid(row=2, column=3)
                Nummer = Label(Frame1, text='Nummer', width=15, font=("arial", 16, "bold"), bd=4, bg='SlateGray4')
                Nummer.grid(row=1, column=1)
                NummerE = Entry(Frame1, width=4, bd=4, font=font_)
                NummerE.grid(row=2, column=1)
                def jump_zutat(event=None):
                    ZutatE.focus_force()
                jump_zutat()
                NummerE.bind('<Return>',jump_zutat)

                ########################################## zutatenPreise Functions  ####################################################
                kol = []

                # -------------------------------------------Speise Nummer rausholen------------------------------------------------#
                def numm():
                    cur = connZu.cursor()
                    cur.execute('select * from zutatenpreise')
                    fet = cur.fetchall()
                    # Clear the NummerE entry box
                    NummerE.delete(0, END)

                    # Add all fetched rows to the kol list and get the length
                    for izo in fet:
                        kol.insert(0, izo)
                    hob = len(kol)

                    # Insert the next available number in the NummerE entry box
                    NummerE.insert(0, hob + 1)

                    # Disable the NummerE entry box
                    NummerE.config(state=DISABLED)

                    # Clear the kol list
                    kol.clear()

                numm()

                # ----------------------------------------------- Speise Nummer Frei stellen ---------------------------------------#
                def frei():
                    # Allow the user to enter a new zutat number
                    NummerE.config(state=NORMAL)
                    NummerE.delete(0, END)
                    NummerE.focus_force()

                    # ------------------------------------------------------------------------------------------------------------------#

                def Speicher_zutat():  # Save a new zutat to the database
                    # Allow the user to enter a new zutat number
                    NummerE.config(state=NORMAL)
                    # Get the PreisE and ZutatE values
                    siko = PreisE.get()
                    soko = ZutatE.get()

                    # Check for empty values and invalid formats
                    if ' ' in soko:
                        messagebox.showerror('Error', 'kein leere tasten')
                    else:
                        if ',' in siko:
                            MessageBox.showerror('Error', 'es muss ein punkt sein und nicht ein komma')
                        else:
                            numb = NummerE.get()
                            saldo = str(ZutatE.get())
                            cur1 = connZu.cursor()
                            cur1.execute('select * from zutatenpreise')
                            ford = cur1.fetchall()

                            # Check if the zutat number already exists
                            if ford:
                                for ido in ford:
                                    sol = str(ido[0]).replace(')', '').replace('(', '').replace(',', '')
                                    sold = str(ido[1]).replace(')', '').replace('(', '').replace(',', '')

                                if numb in sol:
                                    messagebox.showerror('Nummer vergeben ', 'diese Nummer exestiert schon ')
                                elif saldo in sold:
                                    messagebox.showerror('Zutat ', 'diese Zutat exestiert schon ')
                                else:

                                    # Insert the new zutat into the database
                                    cur = connZu.cursor()
                                    saldo = str(ZutatE.get())
                                    cur.execute("insert into zutatenpreise values(?,?,?)",
                                                (NummerE.get(),
                                                 saldo.strip(),
                                                 PreisE.get()
                                                 ))

                            # Update the zutat display and reset the entry boxes
                            Displayzutat()
                            Reseten()
                            connZu.commit()

                            # Call the numm function to get the next available number
                            numm()

                    # --------------------------------------------------------------------------------------------------------------#

                # Reset all input fields
                def Reseten():
                    NummerE.delete(0, END)
                    ZutatE.delete(0, END)
                    PreisE.delete(0, END)

                # Search for a specific ingredient in the database
                def suchenzutat():
                    cur = connZu.cursor()
                    ziko = NummerE.get()
                    cur.execute("select* from zutatenpreise where Nummer=(?)", (ziko,))
                    arko = cur.fetchall()
                    for tiko in arko:
                        Reseten()
                        NummerE.insert(END, tiko[0])
                        ZutatE.insert(END, tiko[1])
                        PreisE.insert(END, tiko[2])
                    connZu.commit()

                # Delete a specific ingredient from the database
                def deletezutat():
                    cur = connZu.cursor()
                    nido = ZutatE.get()
                    cur.execute("delete from zutatenpreise where Zutat=(?)", (nido,))
                    Reseten()
                    Displayzutat()
                    connZu.commit()

                # Display all ingredients and their prices in the database
                def Displayzutat():
                    cur = connZu.cursor()
                    cur.execute("SELECT * FROM zutatenpreise ORDER BY Nummer")
                    result = cur.fetchall()

                    if "-" * len(result) != 0:
                        ZutatenPreise.delete(*ZutatenPreise.get_children())
                        for row in result:
                            ZutatenPreise.insert('', END, values=row)


                    connZu.commit()

                # Display information about a specific ingredient when selected in the GUI
                def ZutatenpreiseInfo(ev):
                    NummerE.config(state=NORMAL)
                    viewInfo = ZutatenPreise.focus()
                    lerandata = ZutatenPreise.item(viewInfo)
                    Reseten()
                    row = lerandata['values']
                    Nummer = NummerE.insert(0, row[0])
                    Zutat = ZutatE.insert(0, row[1])
                    Preis = PreisE.insert(0, row[2])

                    # --------------------------------------------------------------------------------------------------------------#

                def updatezutat():
                    # create cursor object
                    if ' ' in ZutatE.get():
                        messagebox.showwarning('error','Ohne leer tasten')
                    else:
                        cur = connZu.cursor()
                        cur.execute("""
                            update  zutatenpreise  set
                            Nummer = :Nummer,
                            Zutat = :Zutat ,
                            Preis€=:Preis€
                            where Nummer = :Nummer """,
                                    {'Nummer': NummerE.get(),
                                     'Zutat': ZutatE.get(),
                                     'Preis€': PreisE.get(),
                                     })
                        Reseten()
                        Displayzutat()
                        connZu.commit()
                    ########################################## zutatenPreise Buttons  ######################################################

                # create buttons with associated commands
                SpeicherButton = Button(kundenframe2, text='Speichern', font=font_, width=12, height=2, bd=2, bg='grey'
                                        , activebackground='red', command=Speicher_zutat)
                SpeicherButton.place(x=617, y=20)
                UpdateButton = Button(kundenframe2, text='Update', font=font_, width=12, height=2, bd=2, bg='grey'
                                      , activebackground='red', command=updatezutat)
                UpdateButton.place(x=617, y=90)
                LöschenButton = Button(kundenframe2, text='Löschen', font=font_, width=12, height=2, bd=2, bg='grey'
                                       , activebackground='red', command=deletezutat)
                LöschenButton.place(x=617, y=160)
                SucheButton = Button(kundenframe2, text='Suchen', font=font_, width=12, height=2, bd=2, bg='grey'
                                     , activebackground='red', command=suchenzutat)
                SucheButton.place(x=617, y=230)
                ResetButton = Button(kundenframe2, text='Reset', font=font_, width=12, height=2, bd=2, bg='grey'
                                     , activebackground='red', command=Reseten)
                ResetButton.place(x=617, y=300)
                DisplayButton = Button(kundenframe2, text='Display', font=font_, width=12, height=2, bd=2, bg='grey'
                                       , activebackground='red', command=Displayzutat)
                DisplayButton.place(x=617, y=370)
                frei_schalten = Button(Frame1, text='F', width=3, font=("arial", 8, "bold"), bd=4, bg='darkSlateGray4',
                                       command=frei)
                frei_schalten.place(x=5, y=5)
                ######################################### Tree view zuztatenpreise #####################################################
                scroll_y = Scrollbar(Frame3, orient=VERTICAL)
                ZutatenPreise = ttk.Treeview(Frame3, height=30,
                                             columns=("Nummer", "Zutat", "Preis/€",)
                                             , yscrollcommand=scroll_y.set)
                scroll_y.config(command=ZutatenPreise.yview)
                scroll_y.pack(side=RIGHT, fill=Y)

                ZutatenPreise.heading('Nummer', text='Nummer')
                ZutatenPreise.heading('Zutat', text='Zutat')
                ZutatenPreise.heading('Preis/€', text='Preis/€')
                ZutatenPreise['show'] = 'headings'
                ZutatenPreise.column('Nummer', width=160)
                ZutatenPreise.column('Zutat', width=160)
                ZutatenPreise.column('Preis/€', width=160)
                ZutatenPreise.pack(fill=BOTH, expand=1)
                ZutatenPreise.bind("<ButtonRelease>", ZutatenpreiseInfo)

                Displayzutat()


            def Bestellung():
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                kundenframe2.config(width=850, bg='SlateGray4',height=780)
                Mainframe = Frame(kundenframe2, bd=4, width=850, height=None, relief=RIDGE, bg='SlateGray4')
                Mainframe.place(x=6, y=0)
                Titlefram = Frame(Mainframe, bd=4, width=850, height=None, relief=RIDGE)
                Titlefram.grid(row=0, column=0)
                Topframe = Frame(Mainframe, bd=4, width=None, height=None, relief=RIDGE)
                Topframe.grid(row=1, column=0)
                Leftframe = Frame(Topframe, bd=4, width=None, height=None, padx=2, relief=RIDGE, bg='SlateGray4')
                Leftframe.pack(side=LEFT, expand=True, fill=BOTH)
                Leftframe1 = Frame(Leftframe, bd=4, width=None, height=None, relief=RIDGE, bg='SlateGray4')
                Leftframe1.pack(side=TOP, padx=0, pady=0, expand=True, fill=BOTH)

                Rightfram1 = Frame(Topframe, bd=4, width=None, height=None, bg='SlateGray4', )
                Rightfram1.pack(side=RIGHT, expand=True, fill=BOTH)
                Name = StringVar()
                Nummer = StringVar()
                Kategorie = StringVar()
                Preis = StringVar()
                k_g = StringVar()
                Zuschlag = StringVar()
                cur = connS.cursor()

                ##################################speise liste chef(Funktions)######################################################
                def addspeisen():
                    # Get the user inputs from the entry fields and the drop-down menu
                    siko = str(PreisE.get())
                    if NameE.get() == ('') or Nummer1.get() == ('') or clicked.get() == ('') or PreisE.get() == (
                            '') or K_G2.get() == (''):
                        # Check if any of the required fields are empty
                        tkinter.messagebox.showerror('Erorr', 'Alle felder sind pflicht')
                    elif ',' in siko:
                        # Check if the user entered a comma instead of a period for the price
                        messagebox.showerror('Error', 'beim Preis muss ein Punkt sein und nicht ein Komma')
                        PreisE.delete(0, END)
                    else:
                        # If all fields are filled in and the price is properly formatted, add the dish to the database
                        cur = connS.cursor()
                        saldo = str(NameE.get())
                        saldo1 = saldo.strip()  # Remove leading/trailing whitespace from the dish name
                        sako = clicked.get()  # Get the selected category from the drop-down menu

                        cur.execute("insert into Speisen values(?,?,?,?,?,?)",  # Add the dish to the database
                                    (saldo1, Nummer1.get(), sako, siko, K_G2.get(), ZuschlagE.get()))
                        reset1()  # Reset the entry fields to their default values
                        connE.commit()  # Save changes to the database

                ########################################################################################################################
                # Remove any rows where the 'Name' column is empty or whitespace
                cur = connS.cursor()
                cur.execute("DELETE FROM Speisen WHERE Name IS NULL OR trim(Name) = ''")
                connS.commit()

                def displaydata1():
                    # Select all rows from 'Speisen' table and display in 'speiseliste' treeview
                    custor = connS.cursor()
                    custor.execute("SELECT * FROM Speisen")
                    result = custor.fetchall()
                    if "-" * len(result) != 0:
                        speiseliste.delete(*speiseliste.get_children())
                        for row in result:
                            speiseliste.insert('', END, values=row)
                    connS.commit()

                # Initialize 'Nameold' variable to empty string
                global Nameold
                Nameold = ''

                ########################################################################################################################
                def suche1():
                    # Create a cursor object to interact with the database
                    cur = connS.cursor()
                    # Get the value of Nummer1 entry box
                    nada = Nummer1.get()
                    # Execute a SQL query to select all fields from the Speisen table where Nummer matches the entered value
                    cur.execute("select * from Speisen where Nummer = " + nada)
                    # Fetch all matching rows
                    ars = cur.fetchall()
                    # Loop through the results and insert the values into their respective entry boxes
                    for sags in ars:
                        Name = NameE.insert(0, sags[0])
                        KategorieE.insert(0, sags[2])
                        PreisE.insert(0, sags[3])
                        K_G2.set(sags[4])
                    # Commit the changes to the database
                    connS.commit()

                ########################################################################################################################

                def update1():
                    global Nameold
                    # Create a cursor object to interact with the database
                    cur = connS.cursor()
                    # Execute a SQL query to update the values in the Speisen table
                    cur.execute(
                        "UPDATE Speisen SET Name = ?, Nummer = ?, Kategorie = ?, Preis = ?, K_G = ?, Zuschlag = ? WHERE Name = ?",
                        (
                            NameE.get(),
                            Nummer1.get(),
                            clicked.get(),
                            PreisE.get(),
                            K_G2.get(),
                            ZuschlagE.get(),
                            Nameold
                        )
                    )
                    # Commit the changes to the database
                    connS.commit()
                    # Reset the entry boxes to their default values
                    reset1()

                ####################################################################################################################

                def reset1():
                    # Delete the values in all entry boxes
                    NameE.delete(0, END),
                    Nummer1.delete(0, END),
                    clicked.set('Auswählen')
                    PreisE.delete(0, END),
                    K_G2.set("")
                    ZuschlagE.delete(0, END)

                ########################################################################################################################

                def delet1():
                    # Create a cursor object to interact with the database
                    cur = connS.cursor()
                    # Get the value of Nummer1 entry box and remove any quotation marks
                    nada1 = str(Nummer1.get())
                    nada2 = nada1.replace('"', '')
                    # Execute a SQL query to delete the row from the Speisen table where Nummer matches the entered value
                    cur.execute("delete from Speisen where Nummer =(?)", (nada2,))
                    # Commit the changes to the database
                    connS.commit()
                    # Call the displaydata1 function to update the GUI with the current data in the Speisen table
                    displaydata1()
                    # Reset the entry boxes to their default values
                    reset1()

                #######################################################################################################################
                def deletall():
                    pass

                ######################################################################################################################
                def speisenInfo(ev):
                    global Nameold

                    # get the selected item from the treeview
                    viewInfo = speiseliste.focus()
                    lerandata = speiseliste.item(viewInfo)
                    reset1()

                    # populate the entry fields with the data from the selected item
                    row = lerandata['values']
                    Name = NameE.insert(0, row[0])
                    Nameold = row[0]
                    Nummer = Nummer1.insert(0, row[1])
                    Kategorie = clicked.set(row[2])
                    Preis = PreisE.insert(0, row[3])
                    K_G = K_G2.set(row[4])
                    if row[5] != 'None':
                        Zuschlag = ZuschlagE.insert(0, row[5])
                    else:
                        Zuschlag = ZuschlagE.insert(0, '0')

                ########################################  Autocomplete for Zutaten  ###############################################

                # connect to the database and get the zutaten list
                cur = connZ.cursor()

                cur.execute('select SpeiseName from zutaten')
                baro = cur.fetchall()

                # write the zutaten list to a file
                with open(resource_path('Data\zutatenliste.txt'), 'w') as file1:
                    for naro in baro:
                        ziko = ",".join(naro)
                        file1.write(str(ziko) + '\n')

                # read the zutaten list from the file and store in a list
                zutatenliste = []
                with open(resource_path('Data\zutatenliste.txt'), 'r') as fsrg:
                    for ziz in fsrg:
                        zutatenliste.append(str(ziz))

                conn.commit()

                # ----------------------------------Speisen Liste Labels and Enterys----------------------------------------------------#
                lbtitle = Label(Titlefram, font=('arial', 35, 'bold'), text="Speise karte bearbeiten", bd=7)
                lbtitle.grid(row=0, column=0, padx=132)
                Name = Label(Leftframe1, font=('arial', 12, 'bold'), text="Name", bd=5, bg='SlateGray4')
                Name.grid(row=1, column=0, sticky='w', padx=5)
                NameE = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=44, justify='left',
                                          completevalues=zutatenliste, textvariable=Name)
                NameE.grid(row=1, column=1, sticky='w', padx=5)
                Nummer = Label(Leftframe1, font=('arial', 12, 'bold'), text="Nummer", bd=7, bg='SlateGray4')
                Nummer.grid(row=2, column=0, sticky='w', padx=5)
                Nummer1 = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left', textvariable=Nummer)
                Nummer1.grid(row=2, column=1, sticky='w', padx=5)
                Kategorie = Label(Leftframe1, font=('arial', 12, 'bold'), text="Kategorie", bd=7, bg='SlateGray4')
                Kategorie.grid(row=3, column=0, sticky='w', padx=5)
                # ------------------------------------------#

                # Get all categories from the database and add them to the drop-down menu
                cur = connE.cursor()
                cur.execute('select* from Katagorie ')
                nado = cur.fetchall()
                alle = []
                for t in nado:
                    top = str(t)

                    # Remove unwanted characters from the string
                    replace = {'[': '',
                               ']': '',
                               '(': '',
                               ')': '',
                               ',': '',
                               "'": ''}
                    tops = top.translate(str.maketrans(replace))

                    # Add the cleaned up string to the list
                    alle.insert(0, tops)

                connS.commit()

                # Create a dropdown menu for selecting categories
                clicked = StringVar()
                clicked.set("Auswählen")
                KategorieE = OptionMenu(Leftframe1, clicked, *alle, )
                KategorieE.place(x=108, y=67)
                KategorieE.configure(font=('arial', 12, 'bold'), bg="SlateGray4", bd=2, width=10)
                Preis = Label(Leftframe1, font=('arial', 12, 'bold'), text="Preis", bd=7, bg='SlateGray4')
                Preis.grid(row=5, column=0, sticky='w', padx=5)
                PreisE = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=10, justify='left', textvariable=Preis)
                PreisE.place(x=105, y=104)
                # Create entry field for surcharge
                ZuschlagE = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=5, justify='left', textvariable=Zuschlag)
                ZuschlagE.place(x=400, y=104)
                # Create labels for currency symbol and surcharge description
                Labeleuro = Label(Leftframe1, font=('arial', 12, 'bold'), text='€', bd=5, bg='SlateGray4')
                Labeleuro1 = Label(Leftframe1, font=('arial', 12, 'bold'), text='€', bd=5, bg='SlateGray4')

                Labeleuro.place(x=210, y=105)
                Labeleuro1.place(x=460, y=105)
                ZuschlagL = Label(Leftframe1, font=('arial', 12, 'bold'), text='Gross differenc', bd=5, bg='SlateGray4')
                ZuschlagL.place(x=260, y=105)
                # Create label and dropdown menu for portion size
                K_G1 = Label(Leftframe1, font=('arial', 12, 'bold'), text='K_G', bd=5, bg='SlateGray4')
                K_G1.grid(row=6, column=0, padx=5)
                K_G2 = ttk.Combobox(Leftframe1, font=('arial', 12, 'bold'), width=10, state='readonly')
                K_G2['values'] = ('', 'Klein', 'Gross', 'Standard')
                K_G2.current(0)
                K_G2.grid(row=6, column=1, sticky='W')
                #######################################  speiseliste tree view  ########################################################
                # Create a tree view for displaying the menu items
                scroll_y = Scrollbar(Leftframe, orient=VERTICAL)
                speiseliste = ttk.Treeview(Leftframe, height=22, columns=("Name", "Nummer", "Kategorie", "Preis", "Kl_Gr")
                                           )
                scroll_y.config(command=speiseliste.yview)
                scroll_y.pack(side=RIGHT, fill=Y)
                speiseliste.pack(fill=BOTH, expand=1)

                speiseliste.heading('Name', text='Name')
                speiseliste.heading('Nummer', text='Nummer')
                speiseliste.heading('Kategorie', text='Kategorie')
                speiseliste.heading('Preis', text='Preis')
                speiseliste.heading('Kl_Gr', text='Kl_Gr')
                speiseliste['show'] = 'headings'
                speiseliste.column('Name', width=80)
                speiseliste.column('Nummer', width=60)
                speiseliste.column('Kategorie', width=80)
                speiseliste.column('Preis', width=60)
                speiseliste.column('Kl_Gr', width=60)
                speiseliste.pack(fill=BOTH, expand=1)
                speiseliste.bind("<ButtonRelease>", speisenInfo)
                ################################  speise liste buttons  ################################################################
                btnspei = Button(Rightfram1, font=('arial', 12, 'bold'), text='Hinzufügen', bd=4, bg='grey', pady=10, padx=24,
                                 width=8, height=1, command=addspeisen)
                btnspei.grid(row=0, column=0, padx=1,pady=5)
                btndel = Button(Rightfram1, font=('arial', 12, 'bold'), text='Löschen', bd=4, bg='grey', pady=10,
                                padx=24,
                                width=8, height=1, command=delet1)
                btndel.grid(row=1, column=0, padx=1,pady=5)
                btnand = Button(Rightfram1, font=('arial', 12, 'bold'), text='update', bd=4, bg='grey', pady=10,
                                padx=24,
                                width=8, height=1, command=update1)
                btnand.grid(row=2, column=0, padx=1,pady=5)
                btnsear = Button(Rightfram1, font=('arial', 12, 'bold'), text='Suchen', bd=4, bg='grey', pady=10,
                                 padx=24,
                                 width=8, height=1, command=suche1)
                btnsear.grid(row=3, column=0, padx=1,pady=5)
                btnreset = Button(Rightfram1, font=('arial', 12, 'bold'), text='reset', bd=4, bg='grey', pady=10,
                                  padx=24,
                                  width=8, height=1, command=reset1)
                btnreset.grid(row=4, column=0, padx=1,pady=5)
                btnexit = Button(Rightfram1, font=('arial', 12, 'bold'), text='beenden', bd=4, bg='grey', pady=10,
                                 padx=24, width=8, height=1, command=deletall)
                btnexit.grid(row=5, column=0, padx=1,pady=5)
                btndishow = Button(Rightfram1, font=('arial', 12, 'bold'), text='Display', bd=4, bg='grey', pady=10,
                                   padx=24,
                                   width=8, height=1, command=displaydata1)
                btndishow.grid(row=6, column=0, padx=1,pady=5)

                ##################### Title


            def zutaten():
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                kundenframe2.config(width=850, bg='SlateGray4',height=780)
                Mainframe = Frame(kundenframe2, bd=4, width=900, height=700,  bg='SlateGray4')
                Mainframe.place(x=6, y=0)
                Titlefram = Frame(Mainframe, bd=4, width=900, height=700,relief=RIDGE)
                Titlefram.grid(row=0, column=0,pady=10)
                Topframe = Frame(Mainframe, bd=4, width=860, height=710, bg='SlateGray4',relief=RIDGE)
                Topframe.grid(row=1, column=0,pady=10)
                Leftframe = Frame(Mainframe, bd=4, width=700, height=700, padx=2,  bg='SlateGray4',relief=RIDGE)
                Leftframe.place(x=390,y=120)
                Leftframe1 = Frame(Mainframe, bd=4, width=480, height=700,  bg='SlateGray4',relief=RIDGE)
                Leftframe1.place(x=5,y=120)
                # Rightframe = Frame(Mainframe, bd=10, width=350, height=700, relief=RIDGE, bg='red')
                # Rightframe.place(x=100,y=100)
                Rightfram1 = Frame(Mainframe, bd=4, width=450, height=700,  bg='SlateGray4',relief=RIDGE)
                Rightfram1.place(x=620,y=120)
                global oldname
                oldname = ''

                ################################## Zutaten liste chef(Funktions) ###################################################
                def addZutaten():
                    name = SpeiseNameE.get()
                    cur = connZ.cursor()
                    cur.execute("select SpeiseName from zutaten ")
                    ars = cur.fetchall()
                    listofnames = []
                    for ido in ars:
                        listofnames.append(ido)
                    if name in [t[0] for t in listofnames]:
                        messagebox.showwarning('Existiert', 'Diese Speise existiert schon')
                    else:


                        cur = connZ.cursor()
                        karo = str(Zutaten1E.get(), )
                        naldo = str(Zutaten2E.get(), )
                        saldo = str(Zutaten3E.get(), )
                        ziko = str(Zutaten4E.get(), )
                        fiko = str(Zutaten5E.get(), )
                        tiko = str(Zutaten6E.get(), )
                        hiko = str(Zutaten7E.get(), )
                        aiko = str(Zutaten8E.get(), )
                        biko = str(Zutaten9E.get(), )
                        niko = str(Zutaten10E.get(), )
                        karo1 = karo.strip()
                        naldo1 = naldo.strip()
                        saldo1 = saldo.strip()
                        ziko1 = ziko.strip()
                        fiko1 = fiko.strip()
                        tiko1 = tiko.strip()
                        hiko1 = hiko.strip()
                        aiko1 = aiko.strip()
                        biko1 = biko.strip()
                        niko1 = niko.strip()
                        cur.execute(
                            "insert into zutaten values(?,?,?,?,?,?,?,?,?,?,?)",
                            (SpeiseNameE.get(),
                             karo1,
                             naldo1,
                             saldo1,
                             ziko1,
                             fiko1,
                             tiko1,
                             hiko1,
                             aiko1,
                             biko1,
                             niko1
                             ))
                        cur.execute("update zutaten set SpeiseName= LTRIM(RTRIM(SpeiseName))")
                        resetZutaten()
                        connZ.commit()

                ####################################################################################################################
                def displayZutaten():

                    custor = connZ.cursor()
                    custor.execute(
                        "SELECT * FROM zutaten")
                    result = custor.fetchall()
                    if "-" * len(result) != 0:
                        ZutatenListe.delete(*ZutatenListe.get_children())
                        for row in result:
                            ZutatenListe.insert('', END, values=row)
                    connZ.commit()

                ####################################################################################################################
                def sucheZutaten():
                    cur = connZ.cursor()
                    nado = SpeiseNameE.get()
                    cur.execute("select * from zutaten where SpeiseName = (?)", (nado,))
                    ars = cur.fetchall()
                    for sags in ars:
                        SpeiseNameE.delete(0, END)
                        SpeiseName = SpeiseNameE.insert(0, sags[0])
                        Zutaten1E.insert(0, sags[1]),
                        Zutaten2E.insert(0, sags[2]),
                        Zutaten3E.insert(0, sags[3]),
                        Zutaten4E.insert(0, sags[4]),
                        Zutaten5E.insert(0, sags[5]),
                        Zutaten6E.insert(0, sags[6]),
                        Zutaten7E.insert(0, sags[7]),
                        Zutaten8E.insert(0, sags[8]),
                        Zutaten9E.insert(0, sags[9]),
                        Zutaten10E.insert(0, sags[10]),
                    connZ.commit()

                ####################################################################################################################
                def updateZutaten():
                    oldspeise = SpeiseNameE.get()
                    cur = connZ.cursor()
                    sql = "UPDATE zutaten SET SpeiseName = ?, Zutaten1 = ?, Zutaten2 = ?, Zutaten3 = ?, Zutaten4 = ?, Zutaten5 = ?, Zutaten6 = ?, Zutaten7 = ?, Zutaten8 = ?, Zutaten9 = ?, Zutaten10 = ? WHERE SpeiseName = ?"
                    values = (
                        SpeiseNameE.get(),
                        Zutaten1E.get(),
                        Zutaten2E.get(),
                        Zutaten3E.get(),
                        Zutaten4E.get(),
                        Zutaten5E.get(),
                        Zutaten6E.get(),
                        Zutaten7E.get(),
                        Zutaten8E.get(),
                        Zutaten9E.get(),
                        Zutaten10E.get(),
                        oldname
                    )
                    cur.execute(sql, values)
                    connZ.commit()
                    resetZutaten()
                    displayZutaten()

                ####################################################################################################################
                def resetZutaten():
                    SpeiseNameE.delete(0, END),
                    Zutaten1E.delete(0, END),
                    Zutaten2E.delete(0, END),
                    Zutaten3E.delete(0, END),
                    Zutaten4E.delete(0, END),
                    Zutaten5E.delete(0, END),
                    Zutaten6E.delete(0, END),
                    Zutaten7E.delete(0, END),
                    Zutaten8E.delete(0, END),
                    Zutaten9E.delete(0, END),
                    Zutaten10E.delete(0, END)

                ####################################################################################################################
                def deletZutaten():
                    cur = connZ.cursor()
                    nada3 = SpeiseNameE.get()
                    cur.execute("delete from zutaten where SpeiseName =?",
                                (nada3,))
                    connZ.commit()
                    displayZutaten()

                    resetZutaten()

                # ##################################################################################################################
                def deletallZutaten():

                    pass

                ####################################################################################################################
                def ZutatenInfo(ev):
                    global oldname
                    viewInfo = ZutatenListe.focus()
                    lerandata = ZutatenListe.item(viewInfo)
                    resetZutaten()
                    row = lerandata['values']
                    SpeiseName = SpeiseNameE.insert(0, row[0])
                    oldname = row[0]
                    Zutaten1 = Zutaten1E.insert(0, row[1])
                    Zutaten2 = Zutaten2E.insert(0, row[2])
                    Zutaten3 = Zutaten3E.insert(0, row[3])
                    Zutaten4 = Zutaten4E.insert(0, row[4])
                    Zutaten5 = Zutaten5E.insert(0, row[5])
                    Zutaten6 = Zutaten6E.insert(0, row[6])
                    Zutaten7 = Zutaten7E.insert(0, row[7])
                    Zutaten8 = Zutaten8E.insert(0, row[8])
                    Zutaten9 = Zutaten9E.insert(0, row[9])
                    Zutaten10 = Zutaten10E.insert(0, row[10])

                ######################################## Autocomplete function for Zutaten #########################################
                cur = connZu.cursor()
                cur.execute("select trim(Zutat) from zutatenpreise ")
                maro = cur.fetchall()
                with open(resource_path('Data\zutatenpreise.txt'), 'w') as file2:
                    for zuz in maro:
                        ziko = ",".join(zuz)
                        ziko.strip()
                        ziko.split()
                        file2.write(str(ziko) + '\n')
                zutatenpreise = []
                with open(resource_path('Data\zutatenpreise.txt'), 'r') as file3:
                    for zaro in file3:
                        zaro = zaro.rstrip()
                        zutatenpreise.append(str(zaro) + '\n')
                ######################################## Zutaten liste chef (Buttons\Labells) ######################################
                lbtitle = Label(Titlefram, font=('arial', 30, 'bold'), text="Zutaten Liste bearbeiten", bd=7)
                lbtitle.grid(row=0, column=0, padx=132)
                SpeiseName = Label(Leftframe1, font=('arial', 12, 'bold'), text="SpeiseName", bd=7, bg='SlateGray4')
                SpeiseName.grid(row=1, column=0, sticky='w', padx=5,pady=5)
                SpeiseNameE = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=25, justify='left',
                                    textvariable=SpeiseName)
                SpeiseNameE.grid(row=1, column=1, sticky='w', padx=5,pady=5)
                Zutaten1 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten1", bd=7, bg='SlateGray4')
                Zutaten1.grid(row=2, column=0, sticky='w', padx=5,pady=5)
                Zutaten1E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten1, completevalues=zutatenpreise)
                Zutaten1E.grid(row=2, column=1, sticky='w', padx=5,pady=5)
                Zutaten2 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten2", bd=7, bg='SlateGray4')
                Zutaten2.grid(row=3, column=0, sticky='w', padx=5,pady=5)
                Zutaten2E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten2, completevalues=zutatenpreise)
                Zutaten2E.grid(row=3, column=1, sticky='w', padx=5,pady=5)
                Zutaten3 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten3", bd=7, bg='SlateGray4')
                Zutaten3.grid(row=4, column=0, sticky='w', padx=5,pady=5)
                Zutaten3E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten3, completevalues=zutatenpreise)
                Zutaten3E.grid(row=4, column=1, sticky='w', padx=5,pady=5)
                Zutaten4 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten4", bd=7, bg='SlateGray4')
                Zutaten4.grid(row=5, column=0, sticky='w', padx=5,pady=5)
                Zutaten4E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten4, completevalues=zutatenpreise)
                Zutaten4E.grid(row=5, column=1, sticky='w', padx=5,pady=5)
                Zutaten5 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten5", bd=7, bg='SlateGray4')
                Zutaten5.grid(row=6, column=0, sticky='w', padx=5,pady=5)
                Zutaten5E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten5, completevalues=zutatenpreise)
                Zutaten5E.grid(row=6, column=1, sticky='w', padx=5,pady=5)
                Zutaten6 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten6", bd=7, bg='SlateGray4')
                Zutaten6.grid(row=7, column=0, sticky='w', padx=5,pady=5)
                Zutaten6E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten6, completevalues=zutatenpreise)
                Zutaten6E.grid(row=7, column=1, sticky='w', padx=5,pady=5)
                Zutaten7 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten7", bd=7, bg='SlateGray4')
                Zutaten7.grid(row=8, column=0, sticky='w', padx=5,pady=5)
                Zutaten7E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten7, completevalues=zutatenpreise)
                Zutaten7E.grid(row=8, column=1, sticky='w', padx=5,pady=5)
                Zutaten8 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten8", bd=7, bg='SlateGray4')
                Zutaten8.grid(row=9, column=0, sticky='w', padx=5,pady=5)
                Zutaten8E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten8, completevalues=zutatenpreise)
                Zutaten8E.grid(row=9, column=1, sticky='w', padx=5,pady=5)
                Zutaten9 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten9", bd=7, bg='SlateGray4')
                Zutaten9.grid(row=10, column=0, sticky='w', padx=5,pady=5)
                Zutaten9E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                              textvariable=Zutaten9, completevalues=zutatenpreise)
                Zutaten9E.grid(row=10, column=1, sticky='w', padx=5,pady=5)
                Zutaten10 = Label(Leftframe1, font=('arial', 12, 'bold'), text="Zutaten10", bd=7, bg='SlateGray4')
                Zutaten10.grid(row=11, column=0, sticky='w', padx=5,pady=5)
                Zutaten10E = AutocompleteEntry(Leftframe1, font=('arial', 12, 'bold'), width=25, justify='left',
                                               textvariable=Zutaten10, completevalues=zutatenpreise)
                Zutaten10E.grid(row=11, column=1, sticky='w', padx=5,pady=5)

                ####################################### ZUTATENLISTE tree view #####################################################
                scroll_y = Scrollbar(Leftframe, orient=VERTICAL)
                ZutatenListe = ttk.Treeview(Leftframe, height=28,
                                            columns=("SpeiseName")
                                            , yscrollcommand=scroll_y.set)
                scroll_y.pack(side=RIGHT, expand=True, fill=BOTH)
                ZutatenListe.heading('SpeiseName', text='SpeiseName')
                scroll_y.config(command=ZutatenListe.yview)
                # ZutatenListe.heading('Zutaten1', text='Zutaten1')
                # ZutatenListe.heading('Zutaten2', text='Zutaten2')
                # ZutatenListe.heading('Zutaten3', text='Zutaten3')
                # ZutatenListe.heading('Zutaten4', text='Zutaten4')
                # ZutatenListe.heading('Zutaten5', text='Zutaten5')
                # ZutatenListe.heading('Zutaten6', text='Zutaten6')
                # ZutatenListe.heading('Zutaten7', text='Zutaten7')
                # ZutatenListe.heading('Zutaten8', text='Zutaten8')
                # ZutatenListe.heading('Zutaten9', text='Zutaten9')
                # ZutatenListe.heading('Zutaten10', text='Zutaten10')
                ZutatenListe['show'] = 'headings'
                ZutatenListe.column('SpeiseName', width=160)
                # ZutatenListe.column('Zutaten1', width=80)
                # ZutatenListe.column('Zutaten2', width=80)
                # ZutatenListe.column('Zutaten3', width=80)
                # ZutatenListe.column('Zutaten4', width=80)
                # ZutatenListe.column('Zutaten5', width=80)
                # ZutatenListe.column('Zutaten6', width=80)
                # ZutatenListe.column('Zutaten7', width=80)
                # ZutatenListe.column('Zutaten8', width=80)
                # ZutatenListe.column('Zutaten9', width=80)
                # ZutatenListe.column('Zutaten10', width=80)
                ZutatenListe.pack(expand=True, fill=BOTH)
                ZutatenListe.bind("<ButtonRelease>", ZutatenInfo)
                ############################### speise liste buttons ###############################################################
                btnspei = Button(Rightfram1, font=('arial', 12, 'bold'), text='Hinzufügen', bd=4, bg='grey', pady=5,
                                 padx=24,
                                 width=8, height=2, command=addZutaten)
                btnspei.grid(row=0, column=0, padx=1,pady=10)

                btndel = Button(Rightfram1, font=('arial', 12, 'bold'), text='Löschen', bd=4, bg='grey', pady=5,
                                padx=24,
                                width=8, height=2, command=deletZutaten)
                btndel.grid(row=1, column=0, padx=1,pady=10)

                btnand = Button(Rightfram1, font=('arial', 12, 'bold'), text='Update', bd=4, bg='grey', pady=5,
                                padx=24,
                                width=8, height=2, command=updateZutaten)
                btnand.grid(row=2, column=0, padx=1,pady=10)

                btnsear = Button(Rightfram1, font=('arial', 12, 'bold'), text='Suchen', bd=4, bg='grey', pady=5,
                                 padx=24,
                                 width=8, height=2, command=sucheZutaten)
                btnsear.grid(row=3, column=0, padx=1,pady=10)

                btnreset = Button(Rightfram1, font=('arial', 12, 'bold'), text='reset', bd=4, bg='grey', pady=5,
                                  padx=24,
                                  width=8, height=2, command=resetZutaten)
                btnreset.grid(row=4, column=0, padx=1,pady=10)

                btnexit = Button(Rightfram1, font=('arial', 12, 'bold'), text='Alle Löschen', bd=4, bg='grey', pady=5,
                                 padx=24, width=8, height=2, command=deletallZutaten)
                btnexit.grid(row=5, column=0, padx=1,pady=10)
                btndishow = Button(Rightfram1, font=('arial', 12, 'bold'), text='Display', bd=4, bg='grey', pady=5,
                                   padx=24,
                                   width=8, height=2, command=displayZutaten)
                btndishow.grid(row=6, column=0, padx=1,pady=10)


            def kunden():
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                kundenframe2.config(width=980, bg='SlateGray4',height=780)
                font = ('arial', 12, 'bold')
                Mainfram = Frame(kundenframe2, bd=10, width=1800, height=1800, bg='cadet blue')
                Mainfram.place(x=0, y=0)
                Titlefram = Frame(Mainfram, bd=10, width=1200, height=1200, bg='cadet blue')
                Titlefram.place(x=0, y=0)
                Topframe = Frame(Titlefram, bd=10, width=1200, height=1200, bg='dark slate grey')
                Topframe.place(x=0, y=0)
                Leftframe = Frame(Titlefram, bd=10, width=1200, height=1200, relief=RIDGE, bg='grey')
                Leftframe.place(x=8, y=340)
                Leftframe1 = Frame(Mainfram, bd=10, width=1200, height=1200, relief=RIDGE, bg='cadet blue')
                Leftframe1.place(x=15, y=15)
                Rightframe = Frame(Mainfram, bd=10, width=220, height=280, relief=RIDGE, bg='cadet blue')
                Rightframe.place(x=730, y=60)
                Rightfram1 = Frame(Mainfram, bd=10, width=1200, height=1200, bg='cadet blue', relief=RIDGE)
                Rightfram1.place(x=610, y=60)
                ###############################################Labels###################################################################
                lbtitle = Label(Titlefram, font=('arial', 20, 'bold'), text="Kunden Liste", bd=7, bg='cadet blue')
                lbtitle.place(x=620, y=0)
                kundenid = Label(Leftframe1, font=('arial', 12, 'bold'), text="Kunden Nummer", bd=7, bg='cadet blue')
                kundenid.grid(row=1, column=0, sticky='w', padx=5)
                kundenEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                                  textvariable=kundenid, state='normal')
                kundenEnt.grid(row=1, column=1, sticky='w', padx=5)
                Name = Label(Leftframe1, font=('arial', 12, 'bold'), text="Vor/Nachname", bd=7, bg='cadet blue')
                Name.grid(row=2, column=0, sticky='w', padx=5)
                NameEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left', textvariable=Name)
                NameEnt.grid(row=2, column=1, sticky='w', padx=5)
                Addresse = Label(Leftframe1, font=('arial', 12, 'bold'), text="Addresse", bd=7, bg='cadet blue')
                Addresse.grid(row=3, column=0, sticky='w', padx=5)
                AddresseEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                                    textvariable=Addresse)
                AddresseEnt.grid(row=3, column=1, sticky='w', padx=5)
                Telefon = Label(Leftframe1, font=('arial', 12, 'bold'), text="Telefonnummer", bd=7, bg='cadet blue')
                Telefon.grid(row=4, column=0, sticky='w', padx=5)
                TelefonEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                                   textvariable=Telefon)
                TelefonEnt.grid(row=4, column=1, sticky='w', padx=5)
                Hausnr = Label(Leftframe1, font=('arial', 12, 'bold'), text="Nr", bd=7, bg='cadet blue')
                Hausnr.grid(row=5, column=0, sticky='w', padx=5)
                HausnrEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                                  textvariable=Hausnr)
                HausnrEnt.grid(row=5, column=1, sticky='w', padx=5)
                PLZ = Label(Leftframe1, font=('arial', 12, 'bold'), text="PLZ", bd=7, bg='cadet blue')
                PLZ.grid(row=6, column=0, sticky='w', padx=5)
                PLZEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                               textvariable=PLZ)
                PLZEnt.grid(row=6, column=1, sticky='w', padx=5)
                ORT = Label(Leftframe1, font=('arial', 12, 'bold'), text="ORT", bd=7, bg='cadet blue')
                ORT.grid(row=7, column=0, sticky='w', padx=5)
                ORTEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                               textvariable=ORT)
                ORTEnt.grid(row=7, column=1, sticky='w', padx=5)
                Email = Label(Leftframe1, font=('arial', 12, 'bold'), text="Email", bd=7, bg='cadet blue')
                Email.grid(row=8, column=0, sticky='w', padx=5)
                comment = Label(Leftframe1, font=('arial', 12, 'bold'), text="comment", bd=7, bg='cadet blue')
                comment.grid(row=9, column=0, sticky='w', padx=5)
                EmailEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                                 textvariable=Email)
                EmailEnt.grid(row=8, column=1, sticky='w', padx=5)
                CommentEnt = Entry(Leftframe1, font=('arial', 12, 'bold'), bd=5, width=44, justify='left',
                                   textvariable=comment)
                CommentEnt.grid(row=9, column=1, sticky='w', padx=5)
                schwarzliste = Label(Rightframe, font=('arial', 12, 'bold'), text="SchwarzeListe", bd=2, bg='cadet blue',
                                     width=15)
                schwarzliste.place(x=25, y=0)
                schwarzname = Label(Rightframe, font=('arial', 16, 'bold'), text="Name", bd=7,
                                    bg='cadet blue', )
                schwarzname.place(x=25, y=55)
                # schwarztel = Label(Rightframe, font=('arial', 16, 'bold'), text="Tele", bd=7,
                #                    bg='dark slate grey', )
                # schwarztel.place(x=240, y=55)
                schwarzlisteboxn = Listbox(Rightframe, width=15, height=6, bg='#232323', font=('arial', 16, 'bold'), bd=4,
                                           fg='white')
                schwarzlisteboxn.place(x=5, y=90)
                # schwarzlisteboxt = Listbox(Rightframe, width=15, height=30, bg='#232323', font=('arial', 16, 'bold'), bd=4,
                #                            fg='white')
                # schwarzlisteboxt.place(x=200, y=90)
                #########################################################Variabels######################################################
                kundenid = StringVar()
                Name = StringVar()
                Addresse = StringVar()
                Telefon = StringVar()

                ###############################################kunden liste chef(functions)#############################################
                def iExit():
                    iExit = tkinter.messagebox.askyesno("KundenListe", "Sind SIE SICHER")
                    if iExit > 0:
                        root.destroy()
                        return

                # ------------------------------------------------------------------------------------------------------------------#
                def Reset():
                    kundenEnt.delete(0, END)
                    NameEnt.delete(0, END)
                    AddresseEnt.delete(0, END)
                    TelefonEnt.delete(0, END)
                    HausnrEnt.delete(0, END)
                    PLZEnt.delete(0, END)
                    ORTEnt.delete(0, END)
                    EmailEnt.delete(0, END)
                    CommentEnt.delete(0, END)

                ########################################################################################################################
                def addDATA():
                    custor = connK.cursor()
                    telefon = TelefonEnt.get()
                    if NameEnt.get() == "" or AddresseEnt.get() == "" or TelefonEnt.get() == "":
                        tkinter.messagebox.showerror('Error', 'alle felder sind pflichtfelder ')
                    elif NameEnt.get() == "" or AddresseEnt.get() == "" or TelefonEnt.get() == "" in custor:
                        tkinter.messagebox.showerror('kunde exestiert ')
                    else:
                        custor.execute(
                            "insert into kundendaten(Name,Addresse,Telefon,Nr,PLZ,ORT,Email,int_comment) values(?,?,?,?,?,?,?,?)",
                            (NameEnt.get(),
                             AddresseEnt.get(),
                             telefon,
                             HausnrEnt.get(),
                             PLZEnt.get(),
                             ORTEnt.get(),
                             EmailEnt.get(),
                             CommentEnt.get()
                             ))
                        custor.execute("commit")

                        MessageBox.showinfo("!", 'Kunde geschpeichert')
                        Reset()

                ########################################################################################################################
                def displaydata():
                    custor = connK.cursor()
                    custor1 = connK.cursor()
                    custor1.execute("select * from kundendatenB")
                    custor.execute("select * from kundendaten")
                    black = custor1.fetchall()
                    schwarzlisteboxn.delete(0, END)
                    # schwarzlisteboxt.delete(0, END)
                    for blacks in black:
                        schwarzlisteboxn.insert(0, blacks[1])
                        # schwarzlisteboxt.insert(0, blacks[3])
                    result = custor.fetchall()
                    if "-" * len(result) != 0:
                        KUNDENliste.delete(*KUNDENliste.get_children())
                        for row in result:
                            KUNDENliste.insert('', END, values=row)
                    connK.commit()

                ########################################################################################################################
                def TraineeInfo(ev):
                    viewInfo = KUNDENliste.focus()
                    lerandata = KUNDENliste.item(viewInfo)
                    Reset()
                    row = lerandata['values']
                    kundenid = kundenEnt.insert(0, row[0])
                    Name = NameEnt.insert(0, row[1])
                    Addresse = AddresseEnt.insert(0, row[2])
                    Telefon = TelefonEnt.insert(0, row[3])
                    Nr = HausnrEnt.insert(0, row[4])
                    PLZ = PLZEnt.insert(0, row[5])
                    ORT = ORTEnt.insert(0, row[6])
                    Email = EmailEnt.insert(0, row[7])
                    cmnt = CommentEnt.insert(0, row[8])

                ########################################################################################################################
                def update():
                    cur = connK.cursor()
                    cur.execute(
                        "update  kundendaten  set Name=(?),Addresse=(?),Telefon=(?),Nr=(?),PLZ=(?),ORT=(?),Email=(?),int_comment=(?) where Id=(?) ",
                        (NameEnt.get(),
                         AddresseEnt.get(),
                         TelefonEnt.get(),
                         HausnrEnt.get(),
                         PLZEnt.get(),
                         ORTEnt.get(),
                         EmailEnt.get(),
                         CommentEnt.get(),
                         kundenEnt.get()
                         ))
                    connK.commit()

                    MessageBox.showinfo("!", 'Daten geandert')

                ########################################################################################################################
                def delete():
                    cur = connK.cursor()
                    cur.execute("delete from  kundendaten where Id=(?)", (kundenEnt.get(),))
                    connK.commit()
                    displaydata()

                    Reset()

                ########################################################################################################################
                def suche(event=None):
                    sqlCon = connK.connect('Kundendaten.db')
                    cur = sqlCon.cursor()

                    cur.execute("select * from  kundendaten where Id=(?) ", (kundenEnt.get(),))
                    kundenEnt.delete(0, END)
                    NameEnt.delete(0, END)
                    AddresseEnt.delete(0, END)
                    TelefonEnt.delete(0, END)
                    HausnrEnt.delete(0, END)
                    PLZEnt.delete(0, END)
                    ORTEnt.delete(0, END)
                    EmailEnt.delete(0, END)
                    CommentEnt.delete(0, END)
                    ar = cur.fetchall()
                    for sag in ar:
                        Kundenid = kundenEnt.insert(0, sag[0])
                        Name = NameEnt.insert(0, sag[1])
                        Addresse = AddresseEnt.insert(0, sag[2])
                        Telefon = TelefonEnt.insert(0, sag[3])
                        Nr = HausnrEnt.insert(0, sag[4])
                        PLZ = PLZEnt.insert(0, sag[5])
                        ORT = ORTEnt.insert(0, sag[6])
                        Email = EmailEnt.insert(0, sag[7])
                        CommentEnt.insert(0, sag[8])
                        kundenEnt.configure(state=NORMAL)
                        connK.commit()
                        connK.close()

                ########################################################################################################################
                def AddBlackList():
                    kun = kundenEnt.get()
                    if kun == '':
                        messagebox.showerror('Error', 'Auf den Namen doppelt clicken ')
                    else:
                        custor = connK.cursor()
                        custor.execute(
                            "insert into kundendatenB(Name,Addresse,Telefon) values(?,?,?)",
                            (NameEnt.get(),
                             AddresseEnt.get(),
                             TelefonEnt.get(),
                             ))
                        custor.execute("commit")
                        custor.close()
                        MessageBox.showinfo("!", 'Kunde zu BlackList Hinzufügt')
                        Reset()

                addblacklistB = Button(Titlefram, text='Blacklisten', bg='#232323', fg='white', bd=4, command=AddBlackList)
                addblacklistB.place(x=600, y=270)

                ########################################################################################################################
                def Unlisten():
                    custor = connK.cursor()
                    arsa = str(schwarzlisteboxn.get(ACTIVE))

                    custor.execute(
                        "delete from  kundendatenB where Name =(?) ",
                        (arsa,))
                    schwarzlisteboxn.delete(0, END)
                    # schwarzlisteboxt.delete(0, END)
                    custor.execute("commit")
                    custor.close()
                    displaydata()

                unblacklistB = Button(Rightframe, text='Auslisten', bg='WHITE', fg='black', bd=4, width=8, command=Unlisten)
                unblacklistB.place(x=105, y=60)
                ########################################################################################################################
                # -------------------------------------------Treeview-----------------------------------------------------#
                scroll_y = Scrollbar(Leftframe, orient=VERTICAL)
                KUNDENliste = ttk.Treeview(Leftframe, height=18, columns=(
                    "kundenid", "Name", "Addresse", "Telefon", "HausNr", "PLZ", "ORT", "Email", "cmnt")
                                           , yscrollcommand=scroll_y.set)
                scroll_y.pack(side=RIGHT, fill=Y)
                KUNDENliste.heading('kundenid', text='Kundenid')
                KUNDENliste.heading('Name', text='Name')
                KUNDENliste.heading('Addresse', text='Adresse')
                KUNDENliste.heading('Telefon', text='Telefon')
                KUNDENliste.heading('cmnt', text='cmnt')
                KUNDENliste.heading('HausNr', text='HausNr')
                KUNDENliste.heading('PLZ', text='PLZ')
                KUNDENliste.heading('ORT', text='ORT')
                KUNDENliste.heading('Email', text='Email')
                KUNDENliste['show'] = 'headings'
                KUNDENliste.column('kundenid', width=40)
                KUNDENliste.column('Name', width=140)
                KUNDENliste.column('Addresse', width=140)
                KUNDENliste.column('Telefon', width=140)
                KUNDENliste.column('cmnt', width=0)
                KUNDENliste.column('HausNr', width=50)
                KUNDENliste.column('PLZ', width=60)
                KUNDENliste.column('ORT', width=80)
                KUNDENliste.column('Email', width=160)
                KUNDENliste.pack(fill=BOTH, expand=1)
                KUNDENliste.bind("<Double-Button-1>", TraineeInfo)
                ######################################kunden liste chef (Buttons#######################################################
                btnHinz = Button(Rightfram1, font=('arial', 10, 'bold'), text='Hinzufügen', bd=2, bg='cadet blue', padx=24,
                                 pady=1,
                                 width=5, height=1, command=addDATA, )
                btnHinz.grid(row=0, column=0)
                btnlösch = Button(Rightfram1, font=('arial', 10, 'bold'), text='Löschen', bd=2, bg='cadet blue', padx=24,
                                  pady=1,
                                  width=5, height=1, command=delete)
                btnlösch.grid(row=1, column=0)
                btnupd = Button(Rightfram1, font=('arial', 10, 'bold'), text='update', bd=2, bg='cadet blue', padx=24, pady=1,

                                width=5, height=1, command=update)
                btnupd.grid(row=2, column=0)
                btnsuche = Button(Rightfram1, font=('arial', 10, 'bold'), text='Suchen', bd=2, bg='cadet blue', padx=24, pady=1,

                                  width=5, height=1, command=suche)
                btnsuche.grid(row=3, column=0)
                btnres = Button(Rightfram1, font=('arial', 10, 'bold'), text='reset', bd=2, bg='cadet blue', padx=24, pady=1,

                                width=5, height=1, command=Reset)
                btnres.grid(row=4, column=0)
                btnbeen = Button(Rightfram1, font=('arial', 10, 'bold'), text='beenden', bd=2, bg='cadet blue', width=5,
                                 height=1,
                                 padx=24, pady=1,
                                 command=iExit)
                btnbeen.grid(row=5, column=0)
                btndisplay = Button(Rightfram1, font=('arial', 10, 'bold'), text='Display', bd=2, bg='cadet blue', pady=1,
                                    padx=24,
                                    width=5, height=1, command=displaydata)
                btndisplay.grid(row=6, column=0)


            def Einstellung():
                global kundenframe2

                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                font_size = ('arial', 12, 'bold')
                kundenframe2.config(width=930, bg='DarkSlateGray4',height=780)
                Frame1 = Frame(kundenframe2, width=500, bg='DarkSlateGray4', bd=4, height=900, relief=RIDGE)
                Frame1.place(x=0, y=0)
                Frame2 = Frame(kundenframe2, width=400, bg='light grey', bd=4, height=900, relief=RIDGE)
                Frame2.place(x=500, y=0)
                Update = Label(kundenframe2, text='Update ', font=font_size, bd=4, bg='SlateGray4', height=1, width=48)
                Update.place(x=0, y=5)
                Aktuell = Label(kundenframe2, text='Aktuell ', font=font_size, bd=4, bg='dark grey', height=1, width=35)
                Aktuell.place(x=500, y=5)

                # --------------------------------------------Familien Pizza ----------------------------------------------------------##
                def Update1():
                    freit = Familien_FreiE.get()
                    preist = Familien_PreisE.get()

                    cur = connE.cursor()
                    Familie = 'FamilienPizza'
                    cur.execute("""
                                update  einstellung set
                                Name = :Name,
                                einstellung1 = :einstellung1 ,
                                einstellung2=:einstellung2
                                where Name = :Name """,
                                {'Name': Familie,
                                 'einstellung1': freit,
                                 'einstellung2': preist,
                                 })
                    conn.commit()
                    conn.close()
                    Familien_FreiE.delete(0, END)
                    Familien_PreisE.delete(0, END)
                    Einstellung()

                # --------------------------------------------------------------------------------------------------------#

                cur = connE.cursor()
                cur.execute('select* from einstellung where Name = ?', ('FamilienPizza',))
                nado = cur.fetchall()
                frei = ''
                preis = ''
                for t in nado:
                    frei = t[1]
                    preis = t[2]

                connE.commit()
                # -------------------------------------------------------------------------------------------------------#
                Familien_ = Label(Frame1, text='Familien Pizza ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                  width=48)
                Familien_.place(x=0, y=35)
                Familien_tip = Label(Frame1, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1,
                                     width=5)
                Familien_tip.place(x=310, y=35)
                Familien_Frei = Label(Frame1, text='Anzahl Freie zutaten :', font=font_size, bd=4, bg='DarkSlateGray4',
                                      height=1)
                Familien_Frei.place(x=0, y=65)
                Familien_FreiE = Entry(Frame1, width=3, bd=4, font=font_size)
                Familien_FreiE.place(x=180, y=65)
                Familien_Preis = Label(Frame1, text='Preis :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Familien_Preis.place(x=230, y=65)
                Familien_euro = Label(Frame1, text='€', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Familien_euro.place(x=340, y=65)
                Familien_PreisE = Entry(Frame1, width=3, bd=4, font=font_size)
                Familien_PreisE.place(x=300, y=65)
                Familien_Button = Button(Frame1, text='Update', width=8, height=1, command=Update1, bg='#1fc5a8')
                Familien_Button.place(x=380, y=65)
                Familien_A = Label(kundenframe2, text='Familien Pizza ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                   width=35)
                Familien_A.place(x=500, y=39)
                Familien_Frei = Label(kundenframe2, text='Anzahl Freie zutaten :', font=font_size, bd=4, bg='light grey',
                                      height=1)
                Familien_Frei.place(x=500, y=70)
                Familien_F = Label(kundenframe2, text=frei, font=font_size, bd=4, bg='light grey', height=1)
                Familien_F.place(x=680, y=70)
                Familien_Preis = Label(kundenframe2, text='Preis :', font=font_size, bd=4, bg='light grey', height=1)
                Familien_Preis.place(x=720, y=70)
                Familien_P = Label(kundenframe2, text=preis, font=font_size, bd=4, bg='light grey', height=1)
                Familien_P.place(x=780, y=70)
                Familien_euro = Label(kundenframe2, text='€', font=font_size, bd=4, bg='light grey', height=1)
                Familien_euro.place(x=795, y=70)

                ############################################# Katagorie ################################################################
                def addkatagorie():

                    cur = connE.cursor()
                    siko = str(Katagorie_NameE.get())
                    if siko == '':
                        messagebox.showerror('Katagorie', 'Entry darf nicht leer sein')
                    else:
                        siko.strip()
                        cur.execute('INSERT INTO Katagorie values (?)', (siko,))
                        connE.commit()
                        Katagorie_NameE.delete(0, END)
                        Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#
                def deletekatagorie():

                    cur = connE.cursor()
                    siko = clicked.get()
                    cur.execute('delete from Katagorie where Name=(?)', (siko,))
                    conn.commit()
                    Katagorie_NameE.delete(0, END)
                    Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#

                cur = connE.cursor()
                cur.execute('select* from Katagorie ')
                nado = cur.fetchall()
                alle = []

                for t in nado:
                    top = str(t)
                    replace = {'[': '',
                               ']': '',
                               '(': '',
                               ')': '',
                               ',': '',
                               "'": ''}
                    tops = top.translate(str.maketrans(replace))
                    alle.insert(0, tops)

                connE.commit()
                # ------------------------------------------------------------------------------------------------------------------#
                Katagorie_ = Label(Frame1, text='Katagorie ', font=font_size, bd=4, bg='dark slate grey', height=1, width=48)
                Katagorie_.place(x=0, y=100)
                Katagorie_tip = Label(Frame1, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1, width=5)
                Katagorie_tip.place(x=310, y=100)
                Katagorie_Name = Label(Frame1, text=' Katagorie Typ :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Katagorie_Name.place(x=0, y=135)
                Katagorie_NameE = Entry(Frame1, width=18, bd=4, font=font_size)
                Katagorie_NameE.place(x=150, y=135)
                Katagorie_Button = Button(Frame1, text='Speichern', width=8, height=1, command=addkatagorie, bg='#1fc5a8')
                Katagorie_Button.place(x=330, y=135)
                Katagorie_Button1 = Button(Frame1, text='Löschen', width=8, height=1, bg='red', command=deletekatagorie)
                Katagorie_Button1.place(x=410, y=135)
                Katagorie_ = Label(kundenframe2, text='Katagorie ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                   width=35)
                Katagorie_.place(x=500, y=103)
                Katagorie_Name = Label(kundenframe2, text=' Katagorie Typ :', font=font_size, bd=4, bg='light grey', height=1)
                Katagorie_Name.place(x=500, y=135)
                clicked = StringVar()
                clicked.set("Auswählen")
                drop = OptionMenu(kundenframe2, clicked, *alle, )
                drop.place(x=630, y=135)
                drop.configure(font=font_size, bg='dark grey', bd=2, width=10)

                ################################################ Lieferzuschlag #######################################################
                def addLieferzuschlag():

                    cur = connE.cursor()
                    sika = Lieferzuschlag_NameE.get()
                    if sika == '':
                        messagebox.showerror('Lieferzuschlag', 'Entry darf nicht leer sein')
                    else:
                        cur.execute('INSERT INTO Lieferzuschlg values (?)', (sika,))
                        connE.commit()
                        Lieferzuschlag_NameE.delete(0, END)
                        Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#
                def Update_Liefer():

                    cur = connE.cursor()
                    siko = Lieferzuschlag_NameE.get()
                    if siko == '':
                        messagebox.showerror('Lieferzuschlag', 'Entry darf nicht leer sein')
                    else:
                        cur.execute('Update Lieferzuschlg set Name =(?)', (siko,))
                        conn.commit()
                        conn.close()
                        Lieferzuschlag_NameE.delete(0, END)
                        Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#

                cur = connE.cursor()
                cur.execute('select* from Lieferzuschlg ')
                Lieferpreis = cur.fetchall()
                connE.commit()
                # -----------------------------------------------------------------------------------------------------------------#
                Lieferzuschlag_ = Label(Frame1, text='Lieferzuschlag ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                        width=48)
                Lieferzuschlag_.place(x=0, y=165)
                Lieferzuschlag_tip = Label(Frame1, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1,
                                           width=5)
                Lieferzuschlag_tip.place(x=310, y=165)
                Lieferzuschlag_Name = Label(Frame1, text=' Lieferzuschlag ab :', font=font_size, bd=4, bg='DarkSlateGray4',
                                            height=1)
                Lieferzuschlag_Name.place(x=0, y=200)
                Lieferzuschlag_Euro = Label(Frame1, text=' € ', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Lieferzuschlag_Euro.place(x=240, y=200)
                Lieferzuschlag_NameE = Entry(Frame1, width=8, bd=4, font=font_size)
                Lieferzuschlag_NameE.place(x=165, y=200)
                Lieferzuschlag_Button = Button(Frame1, text='Speichern', width=8, height=1, command=addLieferzuschlag,
                                               bg='#1fc5a8')
                Lieferzuschlag_Button.place(x=265, y=200)
                Lieferzuschlag_ = Label(kundenframe2, text='Lieferzuschlag  ', font=font_size, bd=4, bg='dark slate grey',
                                        height=1,
                                        width=35)
                Lieferzuschlag_.place(x=500, y=168)
                Lieferzuschlag_Name = Label(kundenframe2, text=' Lieferzuschlag ab:', font=font_size, bd=4, bg='light grey',
                                            height=1)
                Lieferzuschlag_Name.place(x=500, y=202)
                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Lieferzuschlg ( Name TEXT ) """)
                # conn.commit()
                Lieferzuschlag_P = Label(kundenframe2, text=Lieferpreis, font=font_size, bd=4, bg='light grey', height=1)
                Lieferzuschlag_P.place(x=655, y=202)
                Lieferzuschlag_Euro = Label(kundenframe2, text=' € ', font=font_size, bd=4, bg='light grey', height=1)
                Lieferzuschlag_Euro.place(x=679, y=202)

                ########################################## Passwort ####################################################################
                def addPasswort():

                    cur = connE.cursor()
                    sika = Passwort_NameE.get()
                    if sika == '':
                        messagebox.showerror('PASSWORT', 'Entry darf nicht leer sein')
                    else:
                        cur.execute('Update Passwort set Name =(?)', (sika,))
                        connE.commit()
                        Passwort_NameE.delete(0, END)
                        Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#

                cur = connE.cursor()
                cur.execute('select* from Passwort ')
                Passwort = cur.fetchall()
                connE.commit()
                # ------------------------------------------------------------------------------------------------------------------#
                Passwort_ = Label(Frame1, text='Passwort ', font=font_size, bd=4, bg='dark slate grey', height=1, width=48)
                Passwort_.place(x=0, y=230)
                Passwort_tip = Label(Frame1, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1, width=5)
                Passwort_tip.place(x=310, y=230)
                Passwort_Name = Label(Frame1, text=' Neue Passwort :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Passwort_Name.place(x=0, y=260)
                Passwort_NameE = Entry(Frame1, width=5, bd=4, font=font_size)
                Passwort_NameE.place(x=160, y=260)
                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Passwort( Name TEXT ) """)
                # conn.commit()
                Passwort_Button = Button(Frame1, text='Speichern', width=8, height=1, command=addPasswort, bg='#1fc5a8')
                Passwort_Button.place(x=265, y=260)
                Passwort_ = Label(kundenframe2, text='Passwort ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                  width=35)
                Passwort_.place(x=500, y=238)
                Passwort_Name = Label(kundenframe2, text=' Aktueles Passwort:', font=font_size, bd=4, bg='light grey',
                                      height=1)
                Passwort_Name.place(x=500, y=270)
                Passwort_P = Label(kundenframe2, text=Passwort, font=font_size, bd=4, bg='light grey', height=1)
                Passwort_P.place(x=655, y=270)
                conn = connE
                cur = conn.cursor()
                cur.execute('select* from Gross ')
                gross = cur.fetchall()
                cur = connZu.cursor()
                cur.execute('select Zutat from zutatenpreise ')
                nado = cur.fetchall()
                alle_zutaten = []

                for t in nado:
                    if nado:
                        top = str(t)
                        replace = {'[': '',
                                   ']': '',
                                   '(': '',
                                   ')': '',
                                   ',': '',
                                   "'": ''}
                        tops = top.translate(str.maketrans(replace))
                        alle_zutaten.insert(0, tops)

                connE.commit()

                # -----------------------------------------------------------------------------------------------------------------#
                def Gross_preise():
                    cur = connE.cursor()
                    cur.execute('select * from Gross')
                    soka = clicked12.get()
                    sika = Gross_NameE.get()
                    rows = cur.fetchall()
                    found = False
                    for row in rows:
                        if row[0] == soka:
                            found = True
                            cur.execute('update Gross set Price1 = ? where Name1 like ?', (sika, soka))
                            break
                    if not found:
                        cur.execute('insert into Gross  values (?,?)', (soka, sika))
                    connE.commit()
                    Gross_NameE.delete(0, END)
                    clicked12.set('Auswählen')
                    Einstellung()

                def delete_Gross():
                    cur = connE.cursor()
                    soka = clicked12.get()
                    cur.execute('delete from Gross where Name1=(?)', (soka,))
                    connE.commit()
                    Gross_NameE.delete(0, END)
                    clicked12.set('Auswählen')
                    Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#

                clicked12 = StringVar()
                clicked12.set("Auswählen")
                drop12 = OptionMenu(kundenframe2, clicked12, *alle_zutaten, )
                drop12.place(x=5, y=330)
                drop12.configure(font=font_size, bg="dark slate grey", bd=2, width=10)

                Gross_NameE = Entry(Frame1, width=5, bd=4, font=font_size)
                Gross_NameE.place(x=190, y=325)
                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Gross ( Name INTEGER ) """)
                # conn.commit()
                Gross_Button = Button(Frame1, text='Speichern', width=8, height=1, command=Gross_preise, bg='#1fc5a8')
                Gross_Button.place(x=265, y=325)
                löschen_Button = Button(Frame1, text='Löschen', width=8, height=1, command=delete_Gross, bg='red')
                löschen_Button.place(x=350, y=325)
                Gross__ = Label(kundenframe2, text='Extra Zutaten ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                width=35)
                Gross__.place(x=500, y=298)
                Gross_ = Label(kundenframe2, text='Extra Zutaten ', font=font_size, bd=4, bg='dark slate grey', height=1,
                               width=48)
                Gross_.place(x=5, y=298)
                Gross_tip = Label(Gross_, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1, width=5)
                Gross_tip.place(x=300, y=0)
                Gross_listbox = Listbox(kundenframe2, bd=2, width=20, height=5, font=font_size)
                Gross_listbox.place(x=500, y=330)
                Price_listbox = Listbox(kundenframe2, bd=2, width=5, height=5, font=font_size)
                Price_listbox.place(x=700, y=330)

                for list in gross:
                    Gross_listbox.insert(END, list[0])
                    Price_listbox.insert(END, (str(list[1]) + ' €'))
                conn.commit()

                ########################################################################################################################
                def Note():
                    cub = connE.cursor()
                    sikp = Note_NameE.get()
                    if sikp == '':
                        messagebox.showerror('Note', 'Note darf nicht leer sein')
                    else:
                        cub.execute('UPDATE Note SET Name = (?)', (sikp,))
                        connE.commit()
                        Note_NameE.delete(0, END)
                        Einstellung()

                # ------------------------------------------------------------------------------------------------------------------#
                mess_note = ''
                cub = connE.cursor()
                cub.execute('select* from Note ')
                noties = cub.fetchall()
                with open(resource_path('Data/Note.txt'), 'w') as file13:
                    for zuz in noties:
                        ziko = ",".join(zuz)
                        ziko.strip()
                        # ziko.split()
                        file13.write(str(ziko))
                with open(resource_path('Data/Note.txt'), 'r') as file14:
                    for i in file14:
                        if i:
                            mess_note = i.strip()
                connE.commit()
                Note_ = Label(Frame1, text='Note von Chef ', font=font_size, bd=4, bg='dark slate grey', height=1, width=48)
                Note_.place(x=0, y=360)
                Note_tip = Label(Frame1, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1, width=5)
                Note_tip.place(x=310, y=360)
                Note_Name = Label(Frame1, text=' Note  :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Note_Name.place(x=0, y=390)
                Note_NameE = Entry(Frame1, width=30, bd=4, font=font_size)
                Note_NameE.place(x=100, y=390)
                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Note( Name TEXT ) """)
                # conn.commit()
                Note_Button = Button(Frame1, text='Speichern', width=8, height=1, command=Note, bg='#1fc5a8')
                Note_Button.place(x=390, y=390)
                # Note__ = Label(kundenframe2, text='Note von chef ', font=font_size, bd=4, bg='dark slate grey', height=1,
                #                width=35)
                # Note__.place(x=500, y=365)
                # Note__Name = Label(kundenframe2, text=' Aktueles Note:', font=font_size, bd=4, bg='light grey',
                #                    height=1)
                # Note__Name.place(x=500, y=400)
                Note__P = Label(kundenframe2, text=mess_note, font=font_size, bd=0, bg='DarkSlateGray4', height=1)
                Note__P.place(x=5, y=430)

                ########################################################################################################################
                def add_name():

                    cub = connE.cursor()
                    sikol = Login_nameE.get()
                    sikola = Login_passwortE.get()
                    if sikol == '' or sikola == '':
                        messagebox.showerror('username', 'Entery darf nicht leer sein')
                    else:
                        cub.execute('insert into Username VALUES(Null,?,?)',
                                    (
                                        sikol,
                                        sikola))
                        connE.commit()
                        Login_nameE.delete(0, END)
                        Login_passwortE.delete(0, END)
                        Einstellung()

                # ----------------------------------------------------------------------------------------------------------------------#
                def delete_name():

                    cub = connE.cursor()
                    sikon = IDE.get()
                    if sikon == '':
                        messagebox.showerror('username', 'Entery darf nicht leer sein')
                    else:
                        cub.execute('delete from Username where ID=(?)', (sikon,))
                        connE.commit()
                        IDE.delete(0, END)
                        Einstellung()

                # -------------------------------------------------------------------------------------------------------------------#
                Login_ = Label(Frame1, text='Username ', font=font_size, bd=4, bg='dark slate grey', height=1, width=48)
                Login_.place(x=0, y=452)
                Login_tip = Label(Frame1, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1, width=5)
                Login_tip.place(x=310, y=452)
                Login_name = Label(Frame1, text='Name :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Login_name.place(x=0, y=490)
                ID_name = Label(Frame1, text='ID :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                ID_name.place(x=0, y=540)
                Login_nameE = Entry(Frame1, width=11, bd=4, font=font_size)
                Login_nameE.place(x=65, y=490)
                IDE = Entry(Frame1, width=6, bd=4, font=font_size)
                IDE.place(x=65, y=540)
                Login_passwort = Label(Frame1, text='Passwort :', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                Login_passwort.place(x=190, y=490)
                Login_passwortE = Entry(Frame1, width=10, bd=4, font=font_size)
                Login_passwortE.place(x=280, y=490)
                Add_Button = Button(Frame1, text='Add', width=8, height=1, command=add_name, bg='#1fc5a8')
                Add_Button.place(x=400, y=490)
                LOSCHEN_Button = Button(Frame1, text='löschen', width=8, height=1, bg='red', command=delete_name)
                LOSCHEN_Button.place(x=160, y=540)
                Login_A = Label(kundenframe2, text='Usernames ', font=font_size, bd=4, bg='dark slate grey', height=1, width=35)
                Login_A.place(x=500, y=460)
                Login_names = Label(kundenframe2, text='Username ', font=font_size, bd=4, bg='light grey', height=1)
                Login_names.place(x=515, y=500)
                Login_F = Label(kundenframe2, text=frei, font=font_size, bd=4, bg='light grey', height=1)
                Login_F.place(x=680, y=70)
                Login_Pass = Label(kundenframe2, text='ID ', font=font_size, bd=4, bg='light grey', height=1)
                Login_Pass.place(x=710, y=500)
                Login_P = Label(kundenframe2, text=preis, font=font_size, bd=4, bg='light grey', height=1)
                Login_P.place(x=780, y=70)
                Preis_er = Label(kundenframe2, text='Preise Erhöhen ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                 width=48)
                Preis_er.place(x=0, y=580)
                Preis_tip = Label(kundenframe2, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1, width=5)
                Preis_tip.place(x=310, y=580)
                um_ = Label(kundenframe2, text='Um ', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                um_.place(x=180, y=620)
                preosE = Entry(kundenframe2, width=4, bd=4, font=font_size)
                preosE.place(x=225, y=620)
                um_er = Label(kundenframe2, text='Erhöhen ', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                um_er.place(x=290, y=620)

                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Username(ID integer primary key AUTOINCREMENT, Name TEXT, Passwort TEXT ) """)
                # conn.commit()
                userlist = Listbox(kundenframe2, bd=0, width=13, height=6, bg='light grey')
                userlist.place(x=520, y=540)
                IDlist = Listbox(kundenframe2, bd=0, width=5, height=6, bg='light grey')
                IDlist.place(x=690, y=540)
                IDlist.config(font=font_size, bg='light grey')
                userlist.config(font=font_size, bg='light grey')

                cus = connE.cursor()
                cus.execute('select Name from Username ')
                blob = cus.fetchall()
                for user in blob:
                    userlist.insert(END, user)
                cus.execute('select ID from Username ')
                blob = cus.fetchall()
                for IDES in blob:
                    IDlist.insert(END, IDES)

                ###################################    Preise Erhohen    ###############################################################
                clicked1 = StringVar()
                clicked1.set("Auswählen")
                drop1 = OptionMenu(kundenframe2, clicked1, *alle, )
                drop1.place(x=10, y=620)
                drop1.configure(font=font_size, bg="dark slate grey", bd=2, width=10)

                # ------------------------------------------------------------------------------------------------------------------#
                def preise_erhöhen():

                    soso = clicked1.get()
                    sisi = preosE.get()
                    cur = connS.cursor()
                    primary_key_value = 1
                    cur.execute('INSERT OR REPLACE INTO Liefergeld (primary_key, Preis) VALUES (?, ?)',
                                (primary_key_value, sisi))
                    connS.commit()
                    Einstellung()

                er_Button = Button(kundenframe2, text='Speichern', width=8, height=1, command=preise_erhöhen, bg='#1fc5a8')
                er_Button.place(x=380, y=620)

                # --------------------------------------------- Liefergeld------------------------------------------------------#
                def lieferpreisf():

                    cur = connE.cursor()
                    prei = liefergeldE.get()

                    cur.execute('INSERT OR REPLACE INTO Liefergeld (Preis) VALUES (?)', (prei,))
                    connE.commit()
                    liefergeldE.delete(0, END)
                    Einstellung()

                liefergeld = Label(kundenframe2, text='Liefergeld', font=font_size, bd=4, bg='dark slate grey', height=1,
                                   width=48)
                liefergeld.place(x=0, y=680)
                liefergeld_tip = Label(kundenframe2, text=infosybol, font=font_size, bd=4, bg='dark slate grey', height=1,
                                       width=5)
                liefergeld_tip.place(x=310, y=680)
                liefergeld1 = Label(kundenframe2, text='Liefergeld', font=font_size, bd=4, bg='dark slate grey', height=1,
                                    width=33)
                liefergeld1.place(x=500, y=680)
                liefergeldE = Entry(kundenframe2, width=8, bd=4, font=font_size)
                liefergeldE.place(x=120, y=720)
                lieferpreis = Label(Frame1, text='Preis:', font=font_size, bd=4, bg='DarkSlateGray4', height=1)
                lieferpreis.place(x=2, y=720)
                liefer_Button = Button(kundenframe2, text='Speichern', width=8, height=1, command=lieferpreisf, bg='#1fc5a8')
                liefer_Button.place(x=320, y=720)

                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Liefergeld(  Preis int ) """)
                # conn.commit()

                cur12 = connE.cursor()
                cur12.execute('select * from Liefergeld')
                sol12 = cur12.fetchall()
                sazo12 = ''

                for izo12 in sol12:
                    if izo12:
                        sazo12 = str(izo12[0]).replace("'", "").replace(',', '').replace(')', '').replace('(', '')

                lieferln = Label(kundenframe2, text=sazo12 + '€', font=font_size, bd=4, bg='light grey', height=1)
                lieferln.place(x=620, y=720)
                tip1 = Hovertip(Familien_tip, 'eingabe, wie viele frei Zutaten der Kunde wählen darf'
                                              'Preis, je weitere Zutat')
                tip2 = Hovertip(Katagorie_tip, 'Kategorie einfügen z.b Pizza oder Getränke.\n die Folge der Kategorien '
                                               'kann bei der Kategorie Freiezutaten weiter bearbeitet werden')
                tip3 = Hovertip(Lieferzuschlag_tip, 'falls es geliefert werden muss\n'
                                                    'wie hoch, ist der mindestbetrag, um zu liefern')
                tip4 = Hovertip(Passwort_tip, 'passwort erstellen, für Admin um  auf die Einstellungs Seite zu greifen')
                tip5 = Hovertip(Gross_tip, 'die Difference, in Euro zwischen kleine und grosse Speisen \n'
                                           'z.b falls eine kleine Pizza 5€ kostet  und der wert hier auf 1€ ist\n'
                                           ' eingestellet wird eine grosse Pizza 6€ kosten  ')
                tip6 = Hovertip(Note_tip, 'Was hier geschrieben wird, erscheint auf die Bestellungsseite für die Mitarbeiter\n'
                                          'z.b "Nr 98 ausverkauft ","immmer freundlich mit dem kunden sein"')
                tip7 = Hovertip(Login_tip, 'für jeden Mitarbeiter ein Username erstellen \n'
                                           'um zu wissen, welcher Mitarbeiter die Bestellung aufgenomen hat')
                tip8 = Hovertip(Preis_tip, 'Preise aller Speisen  in der Kategorie erhöhen ')
                tip9 = Hovertip(liefergeld_tip,
                                'es wird für jede Bestellung Liefergeld berechnet, abgesehen davon, wie hoch der Betrag ist\n'
                                'es kann einmalig Deaktiviert werden, in der Bestellungsseite')


            def Drucker():
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                kundenframe2.config(
                    width=750, bg='DarkSlateGray4',height=780)
                global font_size

                def installed_printer():
                    printers = win32print.EnumPrinters(2)
                    for p in printers:
                        return (p)

                printerdef = ''

                # connec=sqlite3.connect('Einstellung.db')
                # cur=connec.cursor()
                # cur.execute("""create Table printer3 (printer TEXT,Kat1 TEXT,Kat2 TEXT,
                # Kat3 TEXT,Kat4 TEXT,Kat5 TEXT,Kat6 TEXT,Kat7 TEXT,Kat8 TEXT,Kat9 TEXT,Kat10 TEXT)""")
                # connec.commit()
                cure = connE.cursor()

                #####################################################Frames and Labels #############################################
                Frame1 = Frame(kundenframe2, width=620, height=500, bg='grey', bd=4)
                Frame1.place(x=10, y=400)
                Frame2 = Frame(kundenframe2, width=190, height=350, bg='azure2', bd=4)
                Frame2.place(x=10, y=5)
                Frame3 = Frame(kundenframe2, width=190, height=350, bg='light grey', bd=4)
                Frame3.place(x=200, y=5)
                Frame4 = Frame(kundenframe2, width=310, height=140, bg='light grey', bd=4)
                Frame4.place(x=400, y=5)
                Frame5 = Frame(kundenframe2, width=260, height=140, bg='light grey', bd=4)
                Frame5.place(x=400, y=140)
                listbox = Listbox(Frame5, width=40, height=8, bd=2)
                listbox.pack()

                Label1 = Label(kundenframe2, text='Aktuell', font=('arial', 20, 'bold'), bg='azure2')
                Label1.place(x=290, y=360)
                NameL = Label(Frame4, text='Name:', font=('arial', 12, 'bold'), bg='light grey')
                NameL.place(x=10, y=10)
                AddresseL = Label(Frame4, text='Addresse:', font=('arial', 12, 'bold'), bg='light grey')
                AddresseL.place(x=10, y=50)
                TelefonL = Label(Frame4, text='Telefon:', font=('arial', 12, 'bold'), bg='light grey')
                TelefonL.place(x=10, y=90)
                NameE = Entry(Frame4, width=30, bd=4)
                NameE.place(x=100, y=10)
                AdresseE = Entry(Frame4, width=30, bd=4)
                AdresseE.place(x=100, y=50)
                TelefonE = Entry(Frame4, width=30, bd=4)
                TelefonE.place(x=100, y=90)
                s = ttk.Style()
                s.theme_use('clam')
                cure.execute('select * from Drucker_Addresse ')
                actual = cure.fetchall()
                for addre in actual:
                    for addresep in addre:
                        listbox.insert(END, addresep)

                # cure=connE.cursor()
                # cure.execute('create table Drucker_Addresse ( Name Text, Addesse Text, Telefon Text ) ')
                # connE.commit()
                def add_Addresse():
                    name = NameE.get()
                    addresse = AdresseE.get()
                    telefon = TelefonE.get()

                    cure.execute('update Drucker_Addresse set Name=(?),Addesse=(?),Telefon=(?)', (name, addresse, telefon))
                    connE.commit()
                    NameE.delete(0, END)
                    AdresseE.delete(0, END)
                    TelefonE.delete(0, END)
                    Drucker()

                addresseButton = Button(kundenframe2, text='Speichern', bg='green', command=add_Addresse)
                addresseButton.place(x=650, y=145)

                # Configure the style of Heading in Treeview widget
                s.configure('printtree.Heading', background="green3")
                s.configure("printtree", background='red', foreground="black", rowheight=25, font=("ARIEL", 9, 'bold'),
                            fieldbackground="black")
                s.map('printtree', background=[('selected', 'red')])
                s.configure("printtreeHeading", font=('bold', 12))
                s.configure("printtree.column", font=('bold', 6))
                s.configure('printtree', rowheight=40)
                s.configure('printtree.row', font=('bold', 20))
                s.configure('Red.TCheckbutton', background='azure2')
                s.configure('yellow.TCheckbutton', background='light grey')

                printtree = ttk.Treeview(Frame1, height=10)
                printtree['columns'] = ("Drucker", "Kat1", "Kat2", "Kat3", "Kat4", "Kat5", "Kat6", "Kat7", "Kat8")
                printtree.column("#0", width=0, stretch=NO)
                printtree.column("Drucker", anchor=CENTER, width=120, stretch=TRUE, )
                printtree.column("Kat1", anchor=CENTER, width=60)
                printtree.column("Kat2", anchor=W, width=60)
                printtree.column("Kat3", anchor=W, width=60)
                printtree.column("Kat4", anchor=W, width=60)
                printtree.column("Kat5", anchor=W, width=60)
                printtree.column("Kat6", anchor=W, width=60)
                printtree.column("Kat7", anchor=W, width=60)
                printtree.column("Kat8", anchor=W, width=60)

                printtree.heading("#0", text="", anchor=W)
                printtree.heading("Drucker", text="Drucker", anchor=W)
                printtree.heading("Kat1", text="Kat1", anchor=W)
                printtree.heading("Kat2", text="Kat2", anchor=W)
                printtree.heading("Kat3", text="Kat3", anchor=W)
                printtree.heading("Kat4", text="Kat4", anchor=W)
                printtree.heading("Kat5", text="Kat5", anchor=W)
                printtree.heading("Kat6", text="Kat6", anchor=W)
                printtree.heading("Kat7", text="Kat7", anchor=W)
                printtree.heading("Kat8", text="Kat8", anchor=W)
                # printtree.place(x=0,y=60)
                printtree.pack()

                #################################################### SQLDATA #######################################################
                global lon, katlist
                katlist = []
                cur = connE.cursor()
                cur.execute('select* from Katagorie  ')
                sor = cur.fetchall()
                for izo in sor:
                    katlist.insert(0, izo)
                lon = len(katlist)

                var = StringVar()
                var5 = StringVar()
                var6 = StringVar()
                var7 = StringVar()
                var8 = StringVar()
                var9 = StringVar()
                var10 = StringVar()
                var11 = StringVar()
                var12 = StringVar()
                varb = StringVar()
                varb5 = StringVar()
                varb6 = StringVar()
                varb7 = StringVar()
                varb8 = StringVar()
                varb9 = StringVar()
                varb10 = StringVar()
                varb11 = StringVar()
                varb12 = StringVar()
                kol = ['']
                kol1 = ['']
                kol2 = ['']
                kol3 = ['']
                kol4 = ['']
                kol5 = ['']
                kol6 = ['']
                kol7 = ['']
                kol8 = ['']
                kol9 = ['']
                kol10 = ['']
                kolb = ['']
                kolb1 = ['']
                kolb2 = ['']
                kolb3 = ['']
                kolb4 = ['']
                kolb5 = ['']
                kolb6 = ['']
                kolb7 = ['']
                kolb8 = ['']
                kolb9 = ['']
                kolb10 = ['']
                #####################################################Drucker auswahlen #############################################

                var1 = StringVar()
                Combo1 = ttk.Combobox(Frame2, width=25, textvariable=var1)
                Combo1.set('Drucker 1')
                print_list = []
                printers = list(win32print.EnumPrinters(2))
                for i in printers:
                    print_list.append(i[2])

                # Put printers in combobox
                var2 = StringVar()
                Combo2 = ttk.Combobox(Frame3, width=25, textvariable=var2)

                Combo1['values'] = print_list
                Combo1.place(x=5, y=5)
                print_list1 = ['']
                printers1 = list(win32print.EnumPrinters(2))
                for ip in printers1:
                    print_list1.append(ip[2])
                Combo2['values'] = print_list1
                Combo2.set('Drucker 2')
                Combo2.place(x=5, y=5)

                listo = []
                listob = []

                def fargini1():
                    cur = connE.cursor()
                    cur.execute('select * from printer2')
                    mik = cur.fetchall()
                    for ray in mik:
                        for element in ray:
                            if element:
                                listo.insert(0, element)
                        for iop in listo:
                            result = str(iop).replace('[', '').replace(']', '').replace('(', '').replace(')', '').replace(',',
                                                                                                                          '').replace(
                                "'", "")
                            # if result!='':
                            #     listbox1.insert(0,result)

                def fargini():
                    cur = connE.cursor()
                    cur.execute('select * from printer1')
                    mik = cur.fetchall()
                    for ray in mik:

                        for element in ray:
                            if element:
                                listo.insert(0, element)
                        for iop in listo:
                            result = str(iop).replace('[', '').replace(']', '').replace('(', '').replace(')', '').replace(',',
                                                                                                                          '').replace(
                                "'", "")

                #################################################

                buttons = []
                kolbs = []
                c1 = ttk.Checkbutton(Frame2, text=katlist[0], variable=var, cursor="cross", style='Red.TCheckbutton')

                c1.place(x=5, y=40)
                if lon > 1:
                    c2 = ttk.Checkbutton(Frame2, text=katlist[1], variable=var5, style='Red.TCheckbutton')

                    c2.place(x=5, y=60)
                else:
                    pass
                if lon > 2:
                    c3 = ttk.Checkbutton(Frame2, text=katlist[2], variable=var6, style='Red.TCheckbutton')

                    c3.place(x=5, y=80)
                else:
                    pass
                if lon > 3:
                    c4 = ttk.Checkbutton(Frame2, text=katlist[3], variable=var7, style='Red.TCheckbutton')

                    c4.place(x=5, y=100)
                else:
                    pass
                if lon > 4:
                    c5 = ttk.Checkbutton(Frame2, text=katlist[4], variable=var8, style='Red.TCheckbutton')

                    c5.place(x=5, y=120)
                else:
                    pass
                if lon > 5:
                    c6 = ttk.Checkbutton(Frame2, text=katlist[5], variable=var9, style='Red.TCheckbutton')
                    c6.place(x=5, y=140)
                else:
                    pass
                if lon > 6:
                    c7 = ttk.Checkbutton(Frame2, text=katlist[6], variable=var10, style='Red.TCheckbutton')
                    c7.place(x=5, y=160)
                else:
                    pass
                if lon > 7:
                    c8 = ttk.Checkbutton(Frame2, text=katlist[7], variable=var11, style='Red.TCheckbutton')
                    c8.place(x=5, y=180)
                else:
                    pass
                if lon > 8:
                    c9 = ttk.Checkbutton(Frame2, text=katlist[8], variable=var11, style='Red.TCheckbutton')
                    c9.place(x=5, y=200)
                else:
                    pass
                if lon > 9:
                    c10 = ttk.Checkbutton(Frame2, text=katlist[9], variable=var11, style='Red.TCheckbutton')
                    c10.place(x=5, y=220)
                else:
                    pass
                b1 = ttk.Checkbutton(Frame3, text=katlist[0], variable=varb, cursor="cross", style='yellow.TCheckbutton')
                b1.place(x=5, y=40)
                if lon > 1:
                    b2 = ttk.Checkbutton(Frame3, text=katlist[1], variable=varb5, style='yellow.TCheckbutton')

                    b2.place(x=5, y=60)
                else:
                    pass
                if lon > 2:
                    b3 = ttk.Checkbutton(Frame3, text=katlist[2], variable=varb6, style='yellow.TCheckbutton')

                    b3.place(x=5, y=80)
                else:
                    pass
                if lon > 3:
                    b4 = ttk.Checkbutton(Frame3, text=katlist[3], variable=varb7, style='yellow.TCheckbutton')

                    b4.place(x=5, y=100)
                else:
                    pass
                if lon > 4:
                    b5 = ttk.Checkbutton(Frame3, text=katlist[4], variable=varb8, style='yellow.TCheckbutton')

                    b5.place(x=5, y=120)
                else:
                    pass
                if lon > 5:
                    b6 = ttk.Checkbutton(Frame3, text=katlist[5], variable=varb9, style='yellow.TCheckbutton')
                    b6.place(x=5, y=140)
                else:
                    pass
                if lon > 6:
                    b7 = ttk.Checkbutton(Frame3, text=katlist[6], variable=varb10, style='yellow.TCheckbutton')
                    b7.place(x=5, y=160)
                else:
                    pass
                if lon > 7:
                    b8 = ttk.Checkbutton(Frame3, text=katlist[7], variable=varb11, style='yellow.TCheckbutton')
                    b8.place(x=5, y=180)
                else:
                    pass
                if lon > 8:
                    b9 = ttk.Checkbutton(Frame3, text=katlist[8], variable=varb11, style='yellow.TCheckbutton')
                    b9.place(x=5, y=200)
                else:
                    pass
                if lon > 9:
                    b10 = ttk.Checkbutton(Frame3, text=katlist[8], variable=varb11, style='yellow.TCheckbutton')
                    b10.place(x=5, y=220)
                else:
                    pass

                def show():
                    listo.clear()
                    value1 = Combo1.get()
                    kol = [value1]

                    curo = connE.cursor()
                    curo.execute('delete from printer1')
                    connE.commit()
                    printername = var1.get()
                    if lon > 0:
                        first = c1.state()
                        if 'selected' in first:
                            kol1.insert(0, katlist[0])
                            c1.invoke()
                    else:
                        pass
                    if lon > 1:
                        second = c2.state()
                        if 'selected' in second:
                            kol2.insert(0, katlist[1])
                            c2.invoke()
                    if lon > 2:
                        third = c3.state()
                        if 'selected' in third:
                            kol3.insert(0, katlist[2])
                            c3.invoke()
                    else:
                        pass
                    if lon > 3:
                        fourth = c4.state()
                        if 'selected' in fourth:
                            kol4.insert(0, katlist[3])
                            c4.invoke()
                    else:
                        pass
                    if lon > 4:
                        fifth = c5.state()
                        if 'selected' in fifth:
                            kol5.insert(0, katlist[4])
                            c5.invoke()
                    else:
                        pass
                    if lon > 5:
                        six = c6.state()
                        if 'selected' in six:
                            kol6.insert(0, katlist[5])
                            c6.invoke()
                    else:
                        pass
                    if lon > 6:
                        seven = c7.state()
                        if 'selected' in seven:
                            kol7.insert(0, katlist[6])
                            c7.invoke()
                    else:
                        pass
                    if lon > 7:
                        eighth = c8.state()
                        if 'selected' in eighth:
                            kol8.insert(0, katlist[7])
                            c8.invoke()
                    else:
                        pass
                    if lon > 8:
                        neinth = c9.state()
                        if 'selected' in neinth:
                            kol9.insert(0, katlist[8])
                            c9.invoke()
                    else:
                        pass
                    if lon > 9:
                        tenth = c10.state()
                        if 'selected' in tenth:
                            kol10.insert(0, katlist[9])
                            c10.invoke()
                    else:
                        pass

                    cur = connE.cursor()
                    cur.execute("""insert into printer1 (printer,kat1,kat2,kat3,kat4,kat5,kat6,kat7,
                                        kat8,kat9,kat10) values (?,?,?,?,?,?,?,?,?,?,?)""",
                                (str(kol), str(kol1), str(kol2), str(kol3), str(kol4), str(kol5), str(kol6), str(kol7),
                                 str(kol8), str(kol9), str(kol10),))
                    connE.commit()
                    Combo1.set('Drucker 1')
                    Combo2.set('Drucker 2')
                    Drucker()

                def show1():
                    listo.clear()
                    value1 = Combo2.get()
                    kolb.insert(0, str(value1))

                    curo = connE.cursor()
                    curo.execute('delete from printer2')
                    connE.commit()

                    printername = var1.get()

                    if lon > 0:
                        first = b1.state()
                        if 'selected' in first:
                            kolb1.insert(0, katlist[0])
                            b1.invoke()
                    else:
                        pass
                    if lon > 1:
                        second = b2.state()
                        if 'selected' in second:
                            kolb2.insert(0, katlist[1])
                            b2.invoke()
                    if lon > 2:
                        third = b3.state()
                        if 'selected' in third:
                            kolb3.insert(0, katlist[2])
                            b3.invoke()
                    else:
                        pass
                    if lon > 3:
                        fourth = b4.state()
                        if 'selected' in fourth:
                            kolb4.insert(0, katlist[3])
                            b4.invoke()
                    else:
                        pass
                    if lon > 4:
                        fifth = b5.state()
                        if 'selected' in fifth:
                            kolb5.insert(0, katlist[4])
                            b5.invoke()
                    else:
                        pass
                    if lon > 5:
                        six = b6.state()
                        if 'selected' in six:
                            kolb6.insert(0, katlist[5])
                            b6.invoke()
                    else:
                        pass
                    if lon > 6:
                        seven = b7.state()
                        if 'selected' in seven:
                            kolb7.insert(0, katlist[6])
                            b7.invoke()
                    else:
                        pass
                    if lon > 7:
                        eighth = b8.state()
                        if 'selected' in eighth:
                            kolb8.insert(0, katlist[7])
                            b8.invoke()
                    else:
                        pass
                    if lon > 8:
                        neinth = b9.state()
                        if 'selected' in neinth:
                            kolb9.insert(0, katlist[8])
                            b9.invoke()
                    else:
                        pass
                    if lon > 9:
                        tenth = b10.state()
                        if 'selected' in tenth:
                            kolb10.insert(0, katlist[9])
                            b10.invoke()
                    else:
                        pass
                    Combo1.set('Drucker 1')
                    Combo2.set('Drucker 2')

                    cur = connE.cursor()
                    cur.execute("""insert into printer2 (printer,kat1,kat2,kat3,kat4,kat5,kat6,kat7,
                                    kat8,kat9,kat10) values (?,?,?,?,?,?,?,?,?,?,?)""",
                                (str(kolb), str(kolb1), str(kolb2), str(kolb3), str(kolb4), str(kolb5), str(kolb6), str(kolb7),
                                 str(kolb8), str(kolb9), str(kolb10),))
                    connE.commit()
                    Drucker()

                # Put printers in combobox

                # menubar = Menu(root)
                # root.config(menu=menubar)
                #
                # file_menu = Menu(menubar)
                # menubar.add_cascade(label="File", menu=file_menu)
                # file_menu.add_command(label="printer", command=locprinter)

                zaro = 0

                def shows():
                    global zaro

                    cur1 = connE.cursor()
                    cur1.execute('select* from  printer1 ')
                    bol1 = cur1.fetchall()

                    zaro = 0
                    for opo1 in bol1:
                        values1 = [
                            str(x).replace('[', '').replace(']', '').replace("'", "").replace(',', '').replace(')', '').replace(
                                '(',
                                '')
                            for x in opo1 if x is not None]
                        printtree.insert(parent='', open=True, index='end', iid=zaro, text='', values=values1)
                        zaro += 1

                def shows1():
                    global zaro

                    cur = connE.cursor()
                    cur.execute('select* from  printer2 ')
                    bol = cur.fetchall()

                    for opo in bol:
                        values = [
                            str(x).replace('[', '').replace(']', '').replace("'", "").replace(',', '').replace(')', '').replace(
                                '(',
                                '')
                            for x in opo if x is not None]
                        printtree.insert(parent='', open=True, index='end', iid=zaro, text='', values=values)
                        zaro += 1

                shows()
                shows1()

                LAB = Label(root, text="Comment")
                # T2 = Text(root, width=40, height=10, wrap=WORD)
                hpeichern = Button(Frame2, text='Speichern', command=show, width=10, bg='green')
                hpeichern.place(x=105, y=315)
                hpeichern1 = Button(Frame3, text='Speichern', command=show1, width=10, bg='green')
                hpeichern1.place(x=105, y=315)
                # zieg = Button(Frame1, text='zeigen', command=fargini,width=20)
                # zieg.place(x=250,y=250)
                # zieg1 = Button(Frame1, text='zeigen', command=fargini1, width=20)
                # zieg1.place(x=500, y=250)


            def Freizutaten():
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                kundenframe2.config(
                    width=720, bg='DarkSlateGray4',height=780)
                global font_size
                name = StringVar
                curE = connE.cursor()
                curE.execute('select * from Katagorie')
                katlist = curE.fetchall()

                cur = connZ.cursor()
                cur.execute('select SpeiseName from zutaten')
                baro = cur.fetchall()
                with open(resource_path('Data\zutatenliste.txt'), 'w') as file1:
                    for naro in baro:
                        ziko = ",".join(naro)
                        file1.write(str(ziko) + '\n')
                zutatenliste = []
                with open(resource_path('Data\zutatenliste.txt'), 'r') as fsrg:
                    for ziz in fsrg:
                        zutatenliste.append(str(ziz))
                conn.commit()
                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute('create table Freiezutaten(Speisename text, Anzahl int)')
                # conn.commit()

                cur = connE.cursor()
                cur.execute('select * from Freiezutaten')
                mor = cur.fetchall()

                ###################Frames
                frame1 = Frame(kundenframe2, bd=4, width=650, height=650, bg='light grey',relief=RIDGE)
                frame1.place(x=10, y=20)
                frame2 = Frame(kundenframe2, bd=4, width=650, height=850, bg='light grey',relief=RIDGE)
                frame2.place(x=735, y=20)
                frame3 = Frame(frame2, bd=4, width=350, height=400, bg='light grey',relief=RIDGE)
                frame3.place(x=10, y=20)
                frame4 = Frame(kundenframe2, bd=4, width=250, height=400, bg='light grey',relief=RIDGE)
                frame4.place(x=400, y=230)

                ###################Tree view
                class ReorderableListbox(tk.Listbox):
                    """ A Tkinter listbox with drag & drop reordering of lines """

                    def __init__(self, master, **kw):
                        kw['selectmode'] = tk.EXTENDED
                        tk.Listbox.__init__(self, master, kw)
                        self.bind('<Button-1>', self.setCurrent)
                        self.bind('<Control-1>', self.toggleSelection)
                        self.bind('<B1-Motion>', self.shiftSelection)
                        self.bind('<Leave>', self.onLeave)
                        self.bind('<Enter>', self.onEnter)
                        self.selectionClicked = False
                        self.left = False
                        self.unlockShifting()
                        self.ctrlClicked = False

                    def orderChangedEventHandler(self):
                        pass

                    def onLeave(self, event):
                        # prevents changing selection when dragging
                        # already selected items beyond the edge of the listbox
                        if self.selectionClicked:
                            self.left = True
                            return 'break'

                    def onEnter(self, event):
                        # TODO
                        self.left = False

                    def setCurrent(self, event):
                        self.ctrlClicked = False
                        i = self.nearest(event.y)
                        self.selectionClicked = self.selection_includes(i)
                        if (self.selectionClicked):
                            return 'break'

                    def toggleSelection(self, event):
                        self.ctrlClicked = True

                    def moveElement(self, source, target):
                        if not self.ctrlClicked:
                            element = self.get(source)
                            self.delete(source)
                            self.insert(target, element)

                    def unlockShifting(self):
                        self.shifting = False

                    def lockShifting(self):
                        # prevent moving processes from disturbing each other
                        # and prevent scrolling too fast
                        # when dragged to the top/bottom of visible area
                        self.shifting = True

                    def shiftSelection(self, event):
                        if self.ctrlClicked:
                            return
                        selection = self.curselection()
                        if not self.selectionClicked or len(selection) == 0:
                            return

                        selectionRange = range(min(selection), max(selection))
                        currentIndex = self.nearest(event.y)

                        if self.shifting:
                            return 'break'

                        lineHeight = 15
                        bottomY = self.winfo_height()
                        if event.y >= bottomY - lineHeight:
                            self.lockShifting()
                            self.see(self.nearest(bottomY - lineHeight) + 1)
                            self.master.after(500, self.unlockShifting)
                        if event.y <= lineHeight:
                            self.lockShifting()
                            self.see(self.nearest(lineHeight) - 1)
                            self.master.after(500, self.unlockShifting)

                        if currentIndex < min(selection):
                            self.lockShifting()
                            notInSelectionIndex = 0
                            for i in selectionRange[::-1]:
                                if not self.selection_includes(i):
                                    self.moveElement(i, max(selection) - notInSelectionIndex)
                                    notInSelectionIndex += 1
                            currentIndex = min(selection) - 1
                            self.moveElement(currentIndex, currentIndex + len(selection))
                            self.orderChangedEventHandler()
                        elif currentIndex > max(selection):
                            self.lockShifting()
                            notInSelectionIndex = 0
                            for i in selectionRange:
                                if not self.selection_includes(i):
                                    self.moveElement(i, min(selection) + notInSelectionIndex)
                                    notInSelectionIndex += 1
                            currentIndex = max(selection) + 1
                            self.moveElement(currentIndex, currentIndex - len(selection))
                            self.orderChangedEventHandler()
                        self.unlockShifting()
                        return 'break'

                listogbox = ReorderableListbox(frame4, width=21, height=12)
                listogbox.place(x=0, y=5)
                listogbox.config(font=('Calibri', 14,'bold'))
                for kat in katlist:
                    listogbox.insert(0, kat)
                style = ttk.Style()
                style.configure("Custom2.Treeview", background="#33737A", foreground="White", fieldbackground="white",
                                font=('Calibri', 13), rowheight=20)
                style.map('Custom2.Treeview', background=[('selected', '#6AC7C8')], foreground=[('selected', 'white')])
                style.configure("Custom2.Treeview.Heading", font=('Calibri', 13))

                zutattree = ttk.Treeview(frame1, height=15, style="Custom2.Treeview")
                zutattree['columns'] = ("Speisename", "Anzahl",)
                zutattree.column("#0", width=0, stretch=NO)
                zutattree.column("Speisename", anchor=CENTER, width=120, stretch=TRUE, )
                zutattree.column("Anzahl", anchor=CENTER, width=60)

                zutattree.heading("#0", text="", anchor=W)
                zutattree.heading("Speisename", text="Speisename", anchor=W)
                zutattree.heading("Anzahl", text="Anzahl", anchor=W)

                zutattree.place(x=10, y=200)

                ################### functions
                def show():
                    for item in zutattree.get_children():
                        zutattree.delete(item)

                    cur = connE.cursor()
                    cur.execute('select * from Freiezutaten')
                    mor = cur.fetchall()

                    for izo in mor:
                        zazo = 0
                        zutattree.insert(parent='', open=True, index='end', text='',
                                         values=(izo[0], izo[1]))

                def loschen():
                    name = zutattree.focus()
                    sol = (zutattree.item(name)['values'][0])

                    cur = connE.cursor()
                    cur.execute('delete from Freiezutaten where Speisename =(?) ', (sol,))
                    connE.commit()
                    show()



                def speichern():
                    name = entery1.get()
                    name1 = str(name).strip()
                    anzahl = entery2.get()

                    cur = connE.cursor()
                    cur.execute('select * from Freiezutaten')

                    mor = cur.fetchall()

                    cur.execute('insert or replace into Freiezutaten (Speisename,Anzahl) values (?,?)', (name1, anzahl))

                    # if name in izo :
                    #     cur.execute('UPDATE Freiezutaten SET Anzahl = (?) where  Speisename=(?)', (anzahl, name1))

                    connE.commit()
                    entery1.delete(0, END)
                    entery2.delete(0, END)
                    show()

                lost = []

                def katspeichern():
                    curE = connE.cursor()
                    curE.execute('Drop Table Katagorie')
                    curE.execute('create table Katagorie (name text)')
                    kats = listogbox.get(0, END)
                    for kato in kats:
                        lost.insert(0, kato)
                    for kat in lost:
                        kat = str(kat).replace(',', '').replace('(', '').replace(')', '').replace("'", "")
                        curE.execute('insert into Katagorie values(:name)', {'name': kat})
                    connE.commit()
                    Freizutaten()

                ################### Labels,Buttons,Entterys

                label1 = Label(frame1, text='Speisen Mit Freizutaten', bg='DarkSlateGray4', font=('arial', 20, 'bold'),bd=4,relief=RIDGE)
                label1.place(x=140, y=10)
                label2 = Label(frame1, text='Speisename:', bg='DarkSlateGray4', font=('arial', 16, 'bold'),relief=RIDGE)
                label2.place(x=10, y=80)
                label3 = Label(frame1, text='Anzahl:', bg='DarkSlateGray4', font=('arial', 16, 'bold'),relief=RIDGE)
                label3.place(x=210, y=80)
                label4 = Label(frame1, text=' Kategorien Sortiern', bg='DarkSlateGray4', font=('arial', 16, 'bold'),relief=RIDGE)
                label4.place(x=365, y=170)
                label5 = Label(frame1, text=infosybol, bg='DarkSlateGray4', font=('arial', 16, 'bold'),relief=RIDGE)
                label5.place(x=580, y=170)
                entery1 = AutocompleteEntry(frame1, width=15, font=('arial', 15, 'bold'), completevalues=zutatenliste,
                                            textvariable=name)

                tip = Hovertip(label5, 'auf dem Katagorie drucken und ziehen ')
                entery1.place(x=5, y=120)
                entery2 = Entry(frame1, width=5, font=('arial', 15, 'bold'))
                entery2.place(x=210, y=120)
                buttn1 = Button(frame1, text='speichern', font=('arial', 12, 'bold'), bg='green', command=speichern)
                buttn1.place(x=320, y=120)
                buttn2 = Button(frame1, text='Löschen', font=('arial', 12, 'bold'), bg='red', command=loschen)
                buttn2.place(x=200, y=210)
                buttn3 = Button(frame4, text='Speichern', font=('arial', 12, 'bold'), bg='green', command=katspeichern)
                buttn3.place(x=150, y=360)

                #################### insert into Treeview
                show()


            def Zweitepc():
                global kundenframe2
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                kundenframe2.config(
                    width=980, bg='DarkSlateGray4',height=780)
                global Aktivier, Einstellung, Kundendaten, speisekarte1, ZutatenListe, ZutatenPreise, OrderData
                from tkinter import filedialog as fd

                frame1 = Frame(kundenframe2, width=420, bd=2, bg='#E3DCD0', height=500)
                frame1.place(x=20, y=10)
                frame15 = Frame(kundenframe2, width=460, bd=2, bg='#839C9C', height=500)
                frame15.place(x=500, y=10)
                frame2 = Frame(frame15, width=420, bd=2, bg='#839C9C', height=500)
                frame2.place(x=5, y=40)
                label1 = Label(frame1, text='Als Zweite PC Aktiviern', font=('arial', 16, 'bold'), bg='#5D6F72')
                label1.place(x=40, y=10)
                label2 = Label(frame15, text='Aktuell', font=('arial', 16, 'bold'), bg='#EFEBE4')
                label2.place(x=200, y=10)
                Einstellungl = Label(frame1, text='Einstellung', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10, y=100)
                Kundendatenl = Label(frame1, text='Kundendaten', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10, y=150)
                speisekarte1l = Label(frame1, text='speisekarte1', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10, y=200)
                ZutatenListel = Label(frame1, text='ZutatenListe', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10, y=250)
                ZutatenPreisel = Label(frame1, text='ZutatenPreise', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10,
                                                                                                                     y=300)
                OrderDatal = Label(frame1, text='OrderData', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10, y=350)
                DataAnalysisl = Label(frame1, text='DataAnalysis', font=('arial', 12, 'bold'), bg='#E3DCD0').place(x=10,
                                                                                                             y=400)
                Aktivierl22 = Label(frame2, text='Aktiviern:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1, row=1,
                                                                                                              pady=10)
                Einstellungl2 = Label(frame2, text='Einstellung:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1,
                                                                                                                  row=2,
                                                                                                                  pady=10)
                Kundendatenl2 = Label(frame2, text='Kundendaten:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1,
                                                                                                                  row=3,
                                                                                                                  pady=10)
                speisekarte1l2 = Label(frame2, text='speisekarte1:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1,
                                                                                                                    row=4,
                                                                                                                    pady=10)
                ZutatenListel2 = Label(frame2, text='ZutatenListe:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1,
                                                                                                                    row=5,
                                                                                                                    pady=10)
                ZutatenPreisel2 = Label(frame2, text='ZutatenPreise:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1,
                                                                                                                      row=6,
                                                                                                                      pady=10)
                OrderDatal2 = Label(frame2, text='OrderData:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1, row=7,
                                                                                                              pady=10)
                DataAnalysisl2 = Label(frame2, text='DatatAnalysis:', font=('arial', 12, 'bold'), bg='#839C9C').grid(column=1,
                                                                                                              row=8,
                                                                                                              pady=10)
                var1 = IntVar()
                check1 = Checkbutton(frame1, text="Aktiviern", variable=var1)
                check1.place(x=10, y=50)
                conn = sqlite3.connect(resource_path('Data\Local.db'))
                cur = conn.cursor()
                # cur.execute("""create table local (Aktivier text,
                #  Einstellung text,
                #  Kundendaten text,
                #  speisekarte1 text,
                #  ZutatenListe text,
                #  ZutatenPreise text,
                #  OrderData text
                Aktivier = 'Lokal'
                Einstellung = 'Lokal'
                Kundendaten = 'Lokal'
                speisekarte1 = 'Lokal'
                ZutatenListe = 'Lokal'
                ZutatenPreise = 'Lokal'
                OrderData = 'Lokal'
                DataAnalysis='Lokal'
                # )""")
                cur.execute('select* from local')
                fetch = cur.fetchall()
                for izo in fetch:
                    AktivierA = izo[0]
                    EinstellungA = izo[1]
                    KundendatenA = izo[2]
                    speisekarte1A = izo[3]
                    ZutatenListeA = izo[4]
                    ZutatenPreiseA = izo[5]
                    OrderDataA = izo[6]
                    DataAnalysisA = izo[7]
                if AktivierA == 'AKTIVE':
                    colour = 'green'
                else:
                    colour = 'red'
                AktivierlAL = Label(frame2, text=AktivierA, font=('arial', 12, 'bold'), bg=colour).grid(column=2, row=1,
                                                                                                        pady=10)
                EinstellunglAL = Label(frame2, text=EinstellungA, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2,
                                                                                                                 row=2,
                                                                                                                 pady=10)
                KundendatenlAL = Label(frame2, text=KundendatenA, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2,
                                                                                                                 row=3,
                                                                                                                 pady=10)
                speisekarte1lAL = Label(frame2, text=speisekarte1A, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2,
                                                                                                                   row=4,
                                                                                                                   pady=10)
                ZutatenListelAL = Label(frame2, text=ZutatenListeA, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2,
                                                                                                                   row=5,
                                                                                                                   pady=10)
                ZutatenPreiselAL = Label(frame2, text=ZutatenPreiseA, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2,
                                                                                                                     row=6,
                                                                                                                     pady=10)
                OrderDatalAL = Label(frame2, text=OrderDataA, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2, row=7,
                                                                                                             pady=10)
                DataAnalysislAL = Label(frame2, text=DataAnalysisA, font=('arial', 12, 'bold'), bg='#839C9C').grid(column=2,
                                                                                                             row=8,
                                                                                                             pady=10)

                # def regedit():
                #
                #     import winreg
                #     import tkinter as tk
                #     from tkinter import filedialog as fd
                #     from tkinter import Label
                #     import tkinter.messagebox as MessageBox
                #     root = tk.Tk()
                #     root.title('Data Ordner')
                #     root.resizable(False, False)
                #     root.geometry('400x200')
                #     MessageBox.showinfo('Data', 'Data Ordner auswählen ')
                #
                #     def callback1():
                #         global name
                #         name = fd.askdirectory()
                #         lab = Label(root, text=name)
                #         lab.pack(fill=tk.X)
                #
                #     def speichern():
                #         root.destroy()
                #
                #     errmsg = 'Error!'
                #     tk.Button(root, text='Click to Open File',
                #               command=callback).pack()
                #     tk.Button(root, text='speichern',
                #               command=speichern, bg='green').place(x=80, y=160)
                #     tk.mainloop()
                #
                #     path = winreg.HKEY_CURRENT_USER
                #     software = winreg.OpenKeyEx(path, r"SOFTWARE\\")
                #
                #     new_key = winreg.CreateKey(software, "M2BESTELLSYSTEM")
                #     winreg.SetValueEx(new_key, "Einstellung", 0, winreg.REG_SZ, name + '\Einstellung.db')
                #     winreg.SetValueEx(new_key, "BestellungData", 0, winreg.REG_SZ, name + '\BestellungData.db')
                #     winreg.SetValueEx(new_key, "DataListe", 0, winreg.REG_SZ, name + '\DataListe.db')
                #     winreg.SetValueEx(new_key, "Kundendaten", 0, winreg.REG_SZ, name + '\Kundendaten.db')
                #     winreg.SetValueEx(new_key, "OrderData", 0, winreg.REG_SZ, name + '\OrderData.db')
                #     winreg.SetValueEx(new_key, "speisekarte1", 0, winreg.REG_SZ, name + '\speisekarte1.db')
                #     winreg.SetValueEx(new_key, "ZutatenListe", 0, winreg.REG_SZ, name + '\ZutatenListe.db')
                #     winreg.SetValueEx(new_key, "ZutatenPreise", 0, winreg.REG_SZ, name + '\ZutatenPreise.db')
                #     winreg.SetValueEx(new_key, "KundendatenB", 0, winreg.REG_SZ, name + '\Kundendatenb.db')
                #     if new_key:
                #         winreg.CloseKey(new_key)
                #
                #     SARA = winreg.OpenKeyEx(path, r"SOFTWARE\\M2BESTELLSYSTEM\\")
                #     Einstellung = winreg.QueryValueEx(SARA, "Einstellung.db")
                #     BestellungData = winreg.QueryValueEx(SARA, "BestellungData.db")
                #     DataListe = winreg.QueryValueEx(SARA, "DataListe.db")
                #     Kundendaten = winreg.QueryValueEx(SARA, "Kundendaten.db")
                #     OrderData = winreg.QueryValueEx(SARA, "OrderData.db")
                #     speisekarte1 = winreg.QueryValueEx(SARA, "speisekarte1.db")
                #     ZutatenListe = winreg.QueryValueEx(SARA, "ZutatenListe.db")
                #     ZutatenPreise = winreg.QueryValueEx(SARA, "ZutatenPreise.db")
                #     KundendatenB = winreg.QueryValueEx(SARA, "KundendatenB.db")
                #     speichern()

                def callAktivier():
                    global Aktivier

                    akt = var1.get()
                    if akt == 1:
                        Aktivierp = 'AKTIVE'
                    else:
                        Aktivierp = 'Nicht AKTIVE'
                    cur.execute('update local set Aktivier=(?)  ', (Aktivierp,))
                    conn.commit()
                    Zweitepc()

                def callEinstellung():
                    global Einstellung
                    Einstellung = fd.askdirectory()
                    Einstellungp = Einstellung + '/Einstellung.db'
                    cur.execute('update local set Einstellung=(?)  ', (Einstellungp,))
                    conn.commit()
                    Zweitepc()

                def lokalEinstellung():
                    cur.execute('update local set Einstellung=(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()

                def callKundendaten():
                    global Kundendaten
                    Kundendaten = fd.askdirectory()
                    Kundendatenp = Kundendaten + '/Kundendaten.db'
                    cur.execute('update local set Kundendaten =(?)  ', (Kundendatenp,))
                    conn.commit()
                    Zweitepc()

                def lokalkundendaten():
                    cur.execute('update local set Kundendaten=(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()

                def callspeisekarte1():
                    global speisekarte1
                    speisekarte1 = fd.askdirectory()
                    speisekarte1p = speisekarte1 + '/speisekarte1.db'
                    cur.execute('update local set speisekarte1 =(?)  ', (speisekarte1p,))
                    conn.commit()
                    Zweitepc()

                def lokalspeisekarte1():
                    cur.execute('update local set speisekarte1=(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()

                def callZutatenListe():
                    global ZutatenListe
                    ZutatenListe = fd.askdirectory()
                    ZutatenListep = ZutatenListe + '/ZutatenListe.db'
                    cur.execute('update local set ZutatenListe =(?)  ', (ZutatenListep,))
                    conn.commit()
                    Zweitepc()

                def lokalZutatenListe():
                    cur.execute('update local set ZutatenListe=(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()

                def callZutatenPreise():
                    global ZutatenPreise
                    ZutatenPreise = fd.askdirectory()
                    ZutatenPreisep = ZutatenPreise + '/ZutatenPreise.db'
                    cur.execute('update local set ZutatenPreise =(?)  ', (ZutatenPreisep,))
                    conn.commit()
                    Zweitepc()

                def lokalZutatenPreise():
                    cur.execute('update local set ZutatenPreise =(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()

                def callOrderData():
                    global OrderData
                    OrderData = fd.askdirectory()
                    OrderDatap = OrderData + '/OrderData.db'
                    cur.execute('update local set OrderData =(?)  ', (OrderDatap,))
                    conn.commit()
                    Zweitepc()

                def lokalOrderData():
                    cur.execute('update local set OrderData =(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()
                def callDataAnalysis():
                    global DataAnalysis
                    DataAnalysis= fd.askdirectory()
                    DataAnalysisp = DataAnalysis + '/DataAnalysis.db'
                    cur.execute('update local set DataAnalysis =(?)  ', (DataAnalysisp,))
                    conn.commit()
                    Zweitepc()

                def lokalDataAnalysis():
                    cur.execute('update local set DataAnalysis =(?)', ('lokal',))
                    conn.commit()
                    Zweitepc()

                Einstellungb = Button(frame1, text='Einstellung importiern',
                                      command=callEinstellung).place(x=120, y=100)
                Einstellungl = Button(frame1, text='Lokal',
                                      command=lokalEinstellung, bg='red').place(x=300, y=100)
                Kundendatenb = Button(frame1, text='Kundendaten importiern',
                                      command=callKundendaten).place(x=120, y=150)
                Kundendatebl = Button(frame1, text='Lokal',
                                      command=lokalkundendaten, bg='red').place(x=300, y=150)
                speisekarte1b = Button(frame1, text='speisekarte1 importiern',
                                       command=callspeisekarte1).place(x=120, y=200)
                speisekartel = Button(frame1, text='Lokal',
                                      command=lokalspeisekarte1, bg='red').place(x=300, y=200)
                ZutatenListeb = Button(frame1, text='ZutatenListe importiern',
                                       command=callZutatenListe).place(x=120, y=250)
                ZutatenListel = Button(frame1, text='Lokal',
                                       command=lokalZutatenListe, bg='red').place(x=300, y=250)
                ZutatenPreiseb = Button(frame1, text='ZutatenPreise importiern',
                                        command=callZutatenPreise).place(x=120, y=300)
                ZutatenPreisel = Button(frame1, text='Lokal',
                                        command=lokalZutatenPreise, bg='red').place(x=300, y=300)
                OrderDatab = Button(frame1, text='OrderData importiern',
                                    command=callOrderData).place(x=120, y=350)
                OrderDatal = Button(frame1, text='Lokal',
                                    command=lokalOrderData, bg='red').place(x=300, y=350)
                DataAnalysisl = Button(frame1, text='Lokal',
                                    command=lokalDataAnalysis, bg='red').place(x=300, y=400)
                DataAnalysisb = Button(frame1, text='OrderData importiern',
                                    command=callDataAnalysis).place(x=120, y=400)

                btn1 = Button(frame1, text='Speichern', font=('arial', 10, 'bold'), bg='green', command=callAktivier)
                btn1.place(x=120, y=50)


            def Bestellliste():
                global kundenframe2, text1
                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                font_size = ('arial', 12, 'bold')
                kundenframe2.config(width=1010, bg='DarkSlateGray4',height=780)

                Frame12 = Frame(kundenframe2, width=520, bg='white', bd=4, height=1100)
                Frame12.place(x=0, y=0)
                Frame22 = Frame(kundenframe2, width=800, bd=4, height=600)
                Frame22.place(x=435, y=510)
                Frame32 = Frame(kundenframe2, width=510, bg='white', bd=4, height=600)
                Frame32.place(x=0, y=510)
                text2 = Text(kundenframe2, height=40, bg='white', bd=0, font=('arial', 10, 'bold'))
                # bestell_liste: Listbox=Listbox(Frame1,bd=6,width=20,height=25,bg='white')
                # bestell_liste.place(x=35,y=65)
                # bestell_liste.config(font=('arial', 16, 'bold'))
                # nr_liste = Listbox(Frame1, bd=6, width=2, height=25, bg='white')
                # nr_liste.place(x=0, y=65)
                # nr_liste.config(font=('arial', 16, 'bold'))
                Update = Label(Frame12, text=' Heute Bestellung  ', font=font_size, bd=4, bg='SlateGray3', height=1, width=48)
                Update.place(x=0, y=0)
                # nr = Label(Frame1, text='Nr ', font=('arial', 16, 'bold'), bd=4, bg='white', height=1, width=2)
                # nr.place(x=0, y=35)
                # name = Label(Frame1, text='Name ', font=('arial', 16, 'bold'), bd=4, bg='white', height=1 )
                # name.place(x=60, y=35)
                KundenidE = Entry(kundenframe2, width=6, bd=5, bg='white smoke')
                KundenidE.place(x=510, y=35)
                KundenidE.configure(font=font_size)
                NameE = Entry(kundenframe2, bd=5, width=25, bg='white smoke')
                NameE.place(x=510, y=94)
                NameE.configure(font=font_size)
                AdresseE = Entry(kundenframe2, width=30, bd=5, font=("Helvetica", 12, 'bold'))
                AdresseE.place(x=510, y=162)
                HauesnrE = Entry(kundenframe2, bd=5, width=4, bg='white smoke')
                HauesnrE.place(x=810, y=162)
                HauesnrE.configure(font=font_size)
                PLZE = Entry(kundenframe2, bd=5, width=7, bg='white smoke')
                PLZE.place(x=510, y=230)
                PLZE.configure(font=font_size)
                ORTE = Entry(kundenframe2, bd=5, width=15, bg='white smoke')
                ORTE.place(x=680, y=230)
                ORTE.configure(font=font_size)
                TelefonnummerE = Entry(kundenframe2, bd=5, width=20, bg='white smoke')
                TelefonnummerE.place(x=630, y=35)
                TelefonnummerE.configure(font=font_size)
                BestellnummerE = Entry(kundenframe2, bd=5, width=5, bg='white smoke')
                BestellnummerE.place(x=820, y=35)
                BestellnummerE.configure(font=font_size)
                EmailE = Entry(kundenframe2, bd=5, width=30, bg='white smoke')
                EmailE.place(x=510, y=300)
                EmailE.configure(font=font_size)
                BestellzeitE = Entry(kundenframe2, width=20, bd=5)
                BestellzeitE.place(x=510, y=370)
                BestellzeitE.configure(font=font_size)
                bedinerE = Entry(kundenframe2, bd=5, width=10, bg='white smoke')
                bedinerE.place(x=610, y=410)
                bedinerE.configure(font=font_size)
                lieferzeitE = Entry(kundenframe2, bd=5, width=10, bg='white smoke',)
                lieferzeitE.place(x=610, y=453)
                lieferzeitE.configure(font=font_size)
                GpreisE = Entry(kundenframe2, bd=5, width=10, bg='white smoke')
                GpreisE.place(x=800, y=453)
                GpreisE.configure(font=font_size)
                sucheE = Entry(kundenframe2, bd=5, width=15, bg='white smoke')
                sucheE.place(x=170, y=320)
                sucheE.configure(font=font_size)
                # -------------------------------------------- Labels ------------------------------------------------------------#
                Kundenid = Label(kundenframe2, bd=0, text='Kunden/ID', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=10)
                Name = Label(kundenframe2, bd=0, text='Name:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=70)
                Adresse = Label(kundenframe2, bd=0, text='Straße:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=135)
                Hausnr = Label(kundenframe2, bd=0, text='Nr:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=810, y=135)
                PLZ = Label(kundenframe2, bd=0, text='PLZ:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=205)
                Ort = Label(kundenframe2, bd=0, text='ORT:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=675, y=205)
                Telefonnummer = Label(kundenframe2, bd=0, text='Telefonnummer:', font=('arial', 14, 'bold'),
                                      bg='DarkSlateGray4') \
                    .place(x=630, y=10)
                Bestellnummer = Label(kundenframe2, bd=0, text='B/nr:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=810, y=10)
                Email = Label(kundenframe2, bd=0, text='Email:', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=272)
                Bestellzeit = Label(kundenframe2, bd=0, text='Bestellzeit:', font=('arial', 16, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=342)
                bediner = Label(kundenframe2, bd=0, text='Bediener:', font=('arial', 16, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=415)
                LIEFERZEIT = Label(kundenframe2, bd=0, text='Lieferzet:', font=('arial', 16, 'bold'), bg='DarkSlateGray4') \
                    .place(x=510, y=455)
                GesamtPreis = Label(kundenframe2, bd=0, text='Gesamte Preis:', font=('arial', 16, 'bold'), bg='DarkSlateGray4') \
                    .place(x=750, y=415)
                suchenl = Label(kundenframe2, bd=0, text='Name eingeben :', font=('arial', 14, 'bold'), bg='DarkSlateGray4') \
                    .place(x=0, y=320)
                Lieferungl = Label(kundenframe2, bd=0, text='Lieferung ', font=('arial', 20, 'bold'), bg='DarkSlateGray4') \
                    .place(x=345, y=250)
                Abholungl = Label(kundenframe2, bd=0, text='Abholung ', font=('arial', 20, 'bold'), bg='DarkSlateGray4') \
                    .place(x=5, y=470)
                text1 = Text(kundenframe2, height=40, bg='white', bd=0, font=('arial', 10, 'bold'))
                # ---------------------------------------------Bbestelltree---------------------------------------------------------#
                style = ttk.Style()
                style.theme_use("clam")
                style.configure("Tree", background='white', foreground="black", rowheight=50, font=("ARIEL", 14, 'bold'),
                                fieldbackground="silver", height=60)
                style.map('Tree', background=[('selected', 'red')])
                style.configure("Tree.Heading", font=('bold', 12))
                style.configure("Tree.column", font=('bold', 6))
                style.configure('Tree', rowheight=90)
                style.configure('Tree.row', font=('bold', 20))
                myTree = ttk.Treeview(Frame22)
                myTree['columns'] = ("pos", "Grosse", "Anzahl", "Nr", "Speise", "Mit", "Ohne", "Katagorie", "Preis",)
                myTree.column("#0", width=0, stretch=NO)
                myTree.column("pos", anchor=CENTER, width=30, stretch=TRUE, )
                myTree.column("Grosse", anchor=CENTER, width=50)
                myTree.column("Anzahl", anchor=CENTER, width=30)
                myTree.column("Nr", anchor=CENTER, width=35)
                myTree.column("Speise", anchor=W, width=85)
                myTree.column("Mit", anchor=W, width=110, stretch=TRUE)
                myTree.column("Ohne", anchor=W, width=110, stretch=TRUE)
                myTree.column("Katagorie", anchor=CENTER, width=40)
                myTree.column("Preis", anchor=CENTER, width=40)
                myTree.tag_configure('pos', background='gray')
                myTree.heading("#0", text="", anchor=W)
                myTree.heading("pos", text="pos", anchor=W)
                myTree.heading("Grosse", text="Grosse", anchor=W)
                myTree.heading("Anzahl", text="Anzahl", anchor=W)
                myTree.heading("Nr", text="Nr", anchor=W)
                myTree.heading("Speise", text="Speise", anchor=CENTER)
                myTree.heading("Mit", text="Mit", anchor=CENTER)
                myTree.heading("Ohne", text="Ohne", anchor=CENTER)
                myTree.heading("Katagorie", text="Katagorie", anchor=CENTER)
                myTree.heading("Preis", text="Preis", anchor=W)
                myTree.pack()
                # ------------------------------------------------------------------------------------------------------------------#
                besteltree = ttk.Treeview(Frame12)
                besteltree['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr")
                besteltree.column("#0", width=0, stretch=NO)
                besteltree.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                besteltree.column("Name", anchor=CENTER, width=120)
                besteltree.column("Datum/Uhrzeit", anchor=W, width=120)
                besteltree.column("Straße", anchor=W, width=90)
                besteltree.column("Haus/nr", anchor=W, width=40)
                besteltree.tag_configure('pos', background='white')
                besteltree.heading("#0", text="", anchor=W)
                besteltree.heading("Nr", text="Nr", anchor=W)
                besteltree.heading("Name", text="Name", anchor=W)
                besteltree.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                besteltree.heading("Straße", text="Straße", anchor=W)
                besteltree.heading("Haus/nr", text="nr", anchor=W)
                besteltree.pack()
                # ------------------------------------------------AbholTree---------------------------------------------------------#
                besteltree1 = ttk.Treeview(Frame32)
                besteltree1['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr")
                besteltree1.column("#0", width=0, stretch=NO)
                besteltree1.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                besteltree1.column("Name", anchor=CENTER, width=120)
                besteltree1.column("Datum/Uhrzeit", anchor=W, width=120)
                besteltree1.column("Straße", anchor=W, width=100)
                besteltree1.column("Haus/nr", anchor=W, width=40)
                besteltree1.tag_configure('pos', background='white')
                besteltree1.heading("#0", text="", anchor=W)
                besteltree1.heading("Nr", text="Nr", anchor=W)
                besteltree1.heading("Name", text="Name", anchor=W)
                besteltree1.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                besteltree1.heading("Straße", text="Straße", anchor=W)
                besteltree1.heading("Haus/nr", text="nr", anchor=W)
                besteltree1.pack()
                # ----------------------------------------------------------get bestelldata und insert in Tree----------------------#
                abhol = 'ABHOLUNG'
                imhaus = 'IM HAUS'
                cur = connO.cursor()
                cur23 = connO.cursor()

                # Execute the first query
                cur.execute('SELECT * FROM kundeinfo WHERE kstrasse NOT IN (?, ?) ORDER BY ID DESC', (abhol, imhaus))
                firo = cur.fetchall()

                # Execute the second query
                cur23.execute('SELECT * FROM kundeinfo WHERE kstrasse IN (?, ?) ORDER BY ID DESC', (abhol, imhaus))
                firo1 = cur23.fetchall()
                # ------------------------------------------------------------------------------------------------------------------#
                global count
                count = 0
                for ziko in firo:
                    besteltree.insert(parent='', open=True, index='end', iid=count, text='',
                                      values=(ziko[0], ziko[3], ziko[10], ziko[4], ziko[5]))
                    count += 1
                global counts
                counts = 0
                for ziko1 in firo1:
                    besteltree1.insert(parent='', open=True, index='end', iid=counts, text='',
                                       values=(ziko1[0], ziko1[3], ziko1[10], ziko1[4], ziko1[5]))
                    counts += 1

                # --------------------------------------------  Functions  ---------------------------------------------------------#
                # ------------------------------------------------------  Bestellung aus der Data base Löschen ----------------------#
                def clear():
                    an=1
                    cur = connO.cursor()
                    sick100 = besteltree.focus()
                    bestel = (besteltree.item(sick100)['values'][0])
                    da=(besteltree.item(sick100)['values'][2])

                    siko='Storno'
                    print(bestel)
                    conn = sqlite3.connect(resource_path('Data\Pending.db'))
                    cur1 = conn.cursor()
                    cur1.execute('Insert into pending (Anzahl,Datum,Gesamt,Fahrer) values(?,?,?,?)', (an, da, bestel, siko))
                    # conn.commit()

                    cur.execute('delete from kundeinfo where ID = (?)', (bestel,))
                    connO.commit()
                    conn.commit()


                    Bestellliste()

                def clear1():

                    cur = connO.cursor()
                    sick100 = besteltree1.focus()
                    bestel = (besteltree1.item(sick100)['values'][0])

                    cur.execute('delete from kundeinfo where ID = (?)', (bestel,))
                    connO.commit()



                    Bestellliste()

                # ------------------------------------------------------------------------------------------------------------------#
                def bestat():
                    sick100 = besteltree.focus()
                    times = (besteltree.item(sick100)['values'][2])

                    cur = connO.cursor()
                    cur.execute('delete from kundeinfo where zeit=(?)', (times,))
                    cur1 = connO.cursor()
                    cur1.execute('delete from speiseinfo where zeit=(?)', (times,))
                    connO.commit()

                    KundenidE.delete(0, END)
                    for record in myTree.get_children():
                        myTree.delete(record)
                    TelefonnummerE.delete(0, END)
                    NameE.delete(0, END)
                    AdresseE.delete(0, END)
                    HauesnrE.delete(0, END)
                    PLZE.delete(0, END)
                    ORTE.delete(0, END)
                    EmailE.delete(0, END)
                    lieferzeitE.delete(0, END)
                    bedinerE.delete(0, END)
                    BestellzeitE.delete(0, END)
                    GpreisE.delete(0, END)
                    BestellnummerE.delete(0, END)


                # ------------------------------------------------------------------------------------------------------------------#
                # ---------------------------------------------- Bestellung Abrufen (Lieferung) ------------------------------------#
                def clicking(e=NONE):
                    global count
                    KundenidE.config(state=NORMAL)
                    TelefonnummerE.config(state=NORMAL)
                    NameE.config(state=NORMAL)
                    AdresseE.config(state=NORMAL)
                    ORTE.config(state=NORMAL)
                    PLZE.config(state=NORMAL)
                    EmailE.config(state=NORMAL)
                    BestellzeitE.config(state=NORMAL)
                    bedinerE.config(state=NORMAL)
                    lieferzeitE.config(state=NORMAL)
                    HauesnrE.config(state=NORMAL)
                    GpreisE.config(state=NORMAL)
                    BestellnummerE.config(state=NORMAL)
                    count = 0
                    sick100 = besteltree.focus()
                    numos = (besteltree.item(sick100)['values'][0])
                    cur = connO.cursor()
                    cur.execute('select * from kundeinfo where ID =(?) ', (numos,))
                    tor = cur.fetchall()
                    cur1 = connO.cursor()
                    cur1.execute('select zeit from kundeinfo where ID =(?) ', (numos,))
                    tor1 = str(cur1.fetchall())
                    tor1.strip()
                    tor2 = tor1.replace('[', '').replace(']', '').replace('(', '').replace(')', '').replace("'", '').replace(
                        ',',
                        '')
                    for na in tor:
                        KundenidE.delete(0, END)
                        for record in myTree.get_children():
                            myTree.delete(record)

                        TelefonnummerE.delete(0, END)
                        NameE.delete(0, END)
                        AdresseE.delete(0, END)
                        HauesnrE.delete(0, END)
                        PLZE.delete(0, END)
                        ORTE.delete(0, END)
                        EmailE.delete(0, END)
                        lieferzeitE.delete(0, END)
                        bedinerE.delete(0, END)
                        BestellzeitE.delete(0, END)
                        GpreisE.delete(0, END)
                        BestellnummerE.delete(0, END)
                        BestellnummerE.insert(0, na[0])
                        KundenidE.insert(0, na[1])
                        TelefonnummerE.insert(0, na[2])
                        NameE.insert(0, na[3])
                        AdresseE.insert(0, na[4])
                        HauesnrE.insert(0, na[5])
                        PLZE.insert(0, na[6])
                        ORTE.insert(0, na[7])
                        EmailE.insert(0, na[8])
                        BestellzeitE.insert(0, na[10])
                        bedinerE.insert(0, na[9])
                        GpreisE.insert(0, na[13])
                        KundenidE.config(state=DISABLED)
                        TelefonnummerE.config(state=DISABLED)
                        NameE.config(state=DISABLED)
                        AdresseE.config(state=DISABLED)
                        ORTE.config(state=DISABLED)
                        PLZE.config(state=DISABLED)
                        EmailE.config(state=DISABLED)
                        BestellzeitE.config(state=DISABLED)
                        bedinerE.config(state=DISABLED)
                        HauesnrE.config(state=DISABLED)
                        GpreisE.config(state=DISABLED)
                        BestellnummerE.config(state=DISABLED)
                        if na[11] != '':
                            lieferzeitE.insert(0, na[11])
                            lieferzeitE.config(state=DISABLED)
                        cur = connO.cursor()
                        cur.execute('select * from speiseinfo where zeit =(?) ', (tor2,))
                        cos = cur.fetchall()
                        for zol in cos:
                            myTree.insert(parent='', open=True, index='end', iid=count, text='',
                                          values=(zol[1], zol[2], zol[3],
                                                  zol[4], zol[5],
                                                  zol[6], zol[7], zol[8], zol[9]))
                            count += 1

                # ------------------------------------------------------------------------------------------------------------------#
                # --------------------------------------------Bestellunh Abrufen (Abholung)-----------------------------------------#
                def clicking1(e=NONE):
                    global count
                    BestellnummerE.config(state=NORMAL)
                    KundenidE.config(state=NORMAL)
                    TelefonnummerE.config(state=NORMAL)
                    NameE.config(state=NORMAL)
                    AdresseE.config(state=NORMAL)
                    ORTE.config(state=NORMAL)
                    PLZE.config(state=NORMAL)
                    EmailE.config(state=NORMAL)
                    BestellzeitE.config(state=NORMAL)
                    bedinerE.config(state=NORMAL)
                    lieferzeitE.config(state=NORMAL)
                    HauesnrE.config(state=NORMAL)
                    GpreisE.config(state=NORMAL)
                    count = 0
                    sick100 = besteltree1.focus()
                    numos = (besteltree1.item(sick100)['values'][0])

                    cur = connO.cursor()
                    cur.execute('select * from kundeinfo where ID =(?) ', (numos,))
                    tor = cur.fetchall()
                    cur1 = connO.cursor()
                    abhol = 'ABHOLUNG'
                    cur1.execute('select zeit from kundeinfo where ID =(?)  ', (numos,))
                    tor1 = str(cur1.fetchall())
                    tor1.strip()
                    tor2 = tor1.replace('[', '').replace(']', '').replace('(', '').replace(')', '').replace("'", '').replace(
                        ',',
                        '')
                    for na in tor:
                        KundenidE.delete(0, END)
                        for record in myTree.get_children():
                            myTree.delete(record)
                        TelefonnummerE.delete(0, END)
                        NameE.delete(0, END)
                        AdresseE.delete(0, END)
                        HauesnrE.delete(0, END)
                        PLZE.delete(0, END)
                        ORTE.delete(0, END)
                        EmailE.delete(0, END)
                        lieferzeitE.delete(0, END)
                        bedinerE.delete(0, END)
                        BestellzeitE.delete(0, END)
                        GpreisE.delete(0, END)
                        BestellnummerE.delete(0, END)
                        BestellnummerE.insert(0, na[0])
                        KundenidE.insert(0, na[1])
                        TelefonnummerE.insert(0, na[2])
                        NameE.insert(0, na[3])
                        AdresseE.insert(0, na[4])
                        HauesnrE.insert(0, na[5])
                        PLZE.insert(0, na[6])
                        ORTE.insert(0, na[7])
                        EmailE.insert(0, na[8])
                        BestellzeitE.insert(0, na[10])
                        bedinerE.insert(0, na[9])
                        GpreisE.insert(0, na[13])
                        KundenidE.config(state=DISABLED)
                        TelefonnummerE.config(state=DISABLED)
                        NameE.config(state=DISABLED)
                        AdresseE.config(state=DISABLED)
                        ORTE.config(state=DISABLED)
                        PLZE.config(state=DISABLED)
                        EmailE.config(state=DISABLED)
                        BestellzeitE.config(state=DISABLED)
                        bedinerE.config(state=DISABLED)
                        HauesnrE.config(state=DISABLED)
                        GpreisE.config(state=DISABLED)
                        BestellnummerE.config(state=DISABLED)
                        if na[11] != '':
                            lieferzeitE.insert(0, na[11])
                            lieferzeitE.config(state=DISABLED)
                        cur = connO.cursor()
                        cur.execute('select * from speiseinfo where zeit =(?) ', (tor2,))
                        cos = cur.fetchall()

                        for zol in cos:
                            myTree.insert(parent='', open=True, index='end', iid=count, text='',
                                          values=(zol[1], zol[2], zol[3],
                                                  zol[4], zol[5],
                                                  zol[6], zol[7], zol[8], zol[9]))
                            count += 1

                # ------------------------------------------------------------------------------------------------------------------#
                def Reset():
                    KundenidE.config(state=NORMAL)
                    TelefonnummerE.config(state=NORMAL)
                    NameE.config(state=NORMAL)
                    AdresseE.config(state=NORMAL)
                    ORTE.config(state=NORMAL)
                    PLZE.config(state=NORMAL)
                    EmailE.config(state=NORMAL)
                    BestellzeitE.config(state=NORMAL)
                    bedinerE.config(state=NORMAL)
                    lieferzeitE.config(state=NORMAL)
                    HauesnrE.config(state=NORMAL)
                    GpreisE.config(state=NORMAL)
                    BestellnummerE.config(state=NORMAL)
                    KundenidE.delete(0, END)
                    for record in myTree.get_children():
                        myTree.delete(record)
                    for tom in besteltree.get_children():
                        besteltree.delete(tom)
                    TelefonnummerE.delete(0, END)
                    BestellnummerE.delete(0, END)
                    NameE.delete(0, END)
                    AdresseE.delete(0, END)
                    HauesnrE.delete(0, END)
                    PLZE.delete(0, END)
                    ORTE.delete(0, END)
                    EmailE.delete(0, END)
                    lieferzeitE.delete(0, END)
                    bedinerE.delete(0, END)
                    BestellzeitE.delete(0, END)
                    GpreisE.delete(0, END)
                    abhol = 'ABHOLUNG'
                    cur = connO.cursor()
                    cur.execute('select * from kundeinfo where kstrasse!=(?) order by ID DESC    ', (abhol,))
                    firo = cur.fetchall()
                    count = 0
                    for ziko in firo:
                        besteltree.insert(parent='', open=True, index='end', iid=count, text='',
                                          values=(ziko[0], ziko[3], ziko[10], ziko[4]))
                        count += 1

                # ------------------------------------------------------------------------------------------------------------------#
                def search_data():

                    country = 0
                    get_record = sucheE.get()
                    for record in besteltree.get_children():
                        besteltree.delete(record)
                    sucheE.delete(0, END)
                    cur = connO.cursor()
                    cur.execute('select * from kundeinfo where kname like (?) ', (get_record,))
                    toz = cur.fetchall()
                    for item in toz:
                        besteltree.insert(parent='', open=True, index='end', iid=country, text='',
                                          values=(item[0], item[3], item[10], item[4]))
                        country += 1

                def printing101():
                    global logo1, coto, text1, mok, imag12, imag13, iamg14
                    mit = []
                    alle = []
                    mab = []
                    gesamt = []
                    lon = []

                    cur = connE.cursor()
                    cur.execute('select* from Katagorie ')
                    fur = cur.fetchall()

                    cur12 = connE.cursor()
                    cur12.execute('select * from Liefergeld')
                    sol12 = cur12.fetchall()
                    for line in myTree.get_children():
                        sick108 = (myTree.item(line[0])['values'][8])
                    Bz = BestellzeitE.get()
                    Bn = BestellnummerE.get()
                    Bk = KundenidE.get()
                    Bt = TelefonnummerE.get()
                    Bna = NameE.get()
                    Ba = AdresseE.get()
                    Bh = HauesnrE.get()
                    Bp = PLZE.get()
                    Bo = ORTE.get()
                    Bl = lieferzeitE.get()

                    for izo12 in sol12:
                        sazo12 = str(izo12).replace("'", "").replace(',', '').replace(')', '').replace('(', '')
                    for kat in fur:
                        top = str(kat)
                        replace = {'[': '',
                                   ']': '',
                                   '(': '',
                                   ')': '',
                                   ',': '',
                                   "'": ''}
                        tops = top.translate(str.maketrans(replace))
                        alle.insert(0, tops)
                    lenalle = len(alle)
                    sicked = myTree.get_children()
                    for tom in myTree.get_children():
                        sick100 = (myTree.item(tom[0])['values'][0])
                        sick101 = (myTree.item(tom[0])['values'][1])
                        sick102 = (myTree.item(tom[0])['values'][2])
                        sick103 = (myTree.item(tom[0])['values'][3])
                        sick104 = (myTree.item(tom[0])['values'][4])
                        sick105 = (myTree.item(tom[0])['values'][5])
                        sick106 = (myTree.item(tom[0])['values'][6])
                        sick107 = (myTree.item(tom[0])['values'][7])
                        sick108 = (myTree.item(tom[0])['values'][8])
                        mab.insert(0, sick107)
                    text1.delete(1.0, END)
                    text1.insert(END, f'\t    \n')
                    text1.insert(END, f'\n\t--NACHDRUCK--   \n')

                    text1.insert(END, f'\n {restName}   \n')
                    text1.insert(END, f'\n===================')
                    text1.insert(END, f'\n{Bz}\t\t\tBes_nr:{Bn}')
                    text1.insert(END, f'\n\n\n kd:{Bk}\t\ttel:{Bt}')
                    text1.insert(END, f'\n{Bna}')
                    text1.insert(END, f'\n{Ba}\t\t{Bh}')
                    text1.insert(END, f'\n{Bp}\t\t{Bo}')
                    text1.insert(END, f'\n===================')
                    if lieferzeitE.get() != '':
                        text1.insert(END, f'\n=====ACHTUNG=====')
                        text1.insert(END, f'\n==ZU {Bl} UHR==')
                        text1.insert(END, f'\n====================')

                    for i in range(lenalle):
                        if alle[i] in mab:
                            sob = mab.count(alle[i])
                            text1.insert(END, f'\n\t------------{alle[i]}------{sob}--')
                        for tom in myTree.get_children():
                            sick100 = (myTree.item(tom[0])['values'][0])
                            sick101 = (myTree.item(tom[0])['values'][1])
                            sick102 = (myTree.item(tom[0])['values'][2])
                            sick103 = (myTree.item(tom[0])['values'][3])
                            sick104 = (myTree.item(tom[0])['values'][4])
                            sick105 = (myTree.item(tom[0])['values'][5])
                            sick106 = (myTree.item(tom[0])['values'][6])
                            sick107 = (myTree.item(tom[0])['values'][7])
                            sick108 = (myTree.item(tom[0])['values'][8])
                            if sick107 == alle[i]:
                                text1.insert(END, f'\n{sick102}x) {sick104} ({sick101}) {sick108}€')
                                if sick105 != "":
                                    text1.config(font=('arial', 10, 'bold'))
                                    for izo in [sick105]:
                                        hozo = str(izo).split(' ')
                                        mit.append(hozo)
                                        for azo in mit:
                                            for tizo in azo:
                                                text1.insert(END, f'\n\t{tizo}')
                                                mit.clear()
                                    text1.insert(END, f'{sick106}\n')
                    Bg = GpreisE.get()

                    text1.insert(END, f"==================")
                    text1.insert(END, f"\n\t\t Gesamtpreis: {Bg}€")

                    q = text1.get("1.0", "end-1c")

                    filename = tempfile.mktemp(".txt")
                    open(filename, 'w').write(q)

                    os.startfile(filename, "print")

                # ---------------------------------------------- Buttons -----------------------------------------------------------#
                show = Button(kundenframe2, bd=4, text='Abrufen', width=10, height=2, bg='green', command=clicking)
                show.place(x=100, y=250)
                show1 = Button(kundenframe2, bd=4, text='Abrufen', width=10, height=2, bg='green', command=clicking1)
                show1.place(x=150, y=460)
                besteltree.bind('<Double-Button-1>', clicking)
                delete_ = Button(kundenframe2, bd=4, text='STORNIERN', width=10, height=2, bg='red', command=clear)
                delete_.place(x=0, y=250)
                delete_AB = Button(kundenframe2, bd=4, text='STORNIERN', width=10, height=2, bg='red', command=clear1)
                delete_AB.place(x=300, y=460)

                # buchen = Button(new_window5, bd=4, text='Büchen', width=10, height=2, bg='cadetblue', command=bestat)
                # buchen.place(x=1200, y=50)
                reset_ = Button(kundenframe2, bd=4, text='Reset', width=10, height=2, bg='Blue', command=Reset)
                reset_.place(x=200, y=250)
                printbtn = Button(kundenframe2, bd=4, text='Print', width=10, height=2, bg='Green', command=printing101)
                printbtn.place(x=860, y=360)
                sucheb_ = Button(kundenframe2, bd=4, text='Suche', width=8, height=1, bg='CADET BLUE', command=search_data)
                sucheb_.place(x=320, y=320)


            def backup():
                import os
                import shutil
                import time
                import tkinter as tk
                from tkinter import filedialog
                import sqlite3

                # get the directory of the executable file


                # create a backup folder if it doesn't exist
                backup_folder = 'C:/Database Backups'
                if not os.path.exists(backup_folder):
                    os.mkdir(backup_folder)

                # define the database names
                db_names = ['Data\Einstellung.db', 'Data\Kundendaten.db', 'Data\Local.db', 'Data\OrderData.db', 'Data\Pending.db', 'Data\ZutatenListe.db',
                            'Data\ZutatenPreise.db', 'Data\streets.txt', 'Data\printer2.xlsx', 'Data\clock.jpg']

                # define a function to create a backup of the databases
                def backup_databases():
                    # get the current date and time

                    # iterate over the databases and backup each one
                    for db_name in db_names:
                        src_path = db_name
                        dest_path = os.path.join(backup_folder, f'{db_name}')
                        shutil.copy2(src_path, dest_path)
                    # display a message box to confirm the backup
                    tk.messagebox.showinfo('Backup', 'Databases backed up successfully.')
                    window.destroy()

                # define a function to import a backup to the original file
                def import_backup():
                    exe_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Data'))
                    # ask the user to select a backup file to import
                    backup_file = filedialog.askopenfilename(initialdir=backup_folder)
                    # get the original file name from the backup file name
                    original_file = backup_file.split('-')[0]
                    # ask the user to select the file to import the backup into
                    import_file = filedialog.asksaveasfilename(initialdir=exe_dir, initialfile=original_file)
                    # copy the backup file to the import file location
                    shutil.copy2(backup_file, import_file)
                    # display a message box to confirm the import
                    tk.messagebox.showinfo('Import', 'Backup imported successfully.')
                    window.destroy()

                # create the Tkinter window and buttons
                window = tk.Tk()
                window.geometry('350x200')
                window.config(bg='#1A3A3A')
                backup_button = tk.Button(window, text='Databases sichern',bg='red', command=backup_databases,font=('Helvetica', 15, "italic bold"))
                backup_button.place(x=65,y=20)
                import_button = tk.Button(window, text='Databases importiern',bg='green', command=import_backup,font=('Helvetica', 15, "italic bold"))
                import_button.place(x=50,y=120)

                # run the Tkinter event loop
                window.mainloop()
            def Einstellung2():
                global kundenframe2

                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                font_size = ('arial', 12, 'bold')
                kundenframe2.config(width=930, bg='DarkSlateGray4',height=780)
                Frame1 = Frame(kundenframe2, width=500, bg='DarkSlateGray4', bd=4, height=500, relief=RIDGE)
                Frame1.place(x=0, y=0)
                Frame2 = Frame(kundenframe2, width=400, bg='light grey', bd=4, height=500, relief=RIDGE)
                Frame2.place(x=500, y=0)
                Update = Label(kundenframe2, text='Update ', font=font_size, bd=4, bg='SlateGray4', height=1, width=48)
                Update.place(x=0, y=5)
                Aktuell = Label(kundenframe2, text='Aktuell ', font=font_size, bd=4, bg='dark grey', height=1, width=35)
                Aktuell.place(x=500, y=5)



                # --------------------------------------------Style ----------------------------------------------------------##
                def update_style():
                    messagebox.showinfo('Update','um den Update zu aktiviern den program bitte neustarten')
                    style = combobox101.get()  # get the selected value from the combobox
                    conn = connE
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM fast WHERE id=1")
                    result = cursor.fetchone()[0]
                    if result == 0:
                        cursor.execute("INSERT INTO fast (id, name) VALUES (?, ?)", (1, style))
                    else:
                        cursor.execute("UPDATE fast SET name=? WHERE id=1", (style,))
                    conn.commit()



                    Einstellung2()

                    # --------------------------------------------Style ----------------------------------------------------------##
                def update_anzahl():
                    style = anzahlE.get()  # get the selected value from the combobox
                    if style.isdigit():
                        conn = connE
                        cursor = conn.cursor()
                        cursor.execute("SELECT COUNT(*) FROM DruckAnzahl WHERE id=1")
                        result = cursor.fetchone()[0]
                        if result == 0:
                            cursor.execute("INSERT INTO DruckAnzahl (id, Anzahl) VALUES (?, ?)", (1, style))
                        else:
                            cursor.execute("UPDATE DruckAnzahl SET Anzahl=? WHERE id=1", (style,))
                        conn.commit()



                        Einstellung2()
                    else:
                        messagebox.showwarning('Warnung','Bitte ihre angabe korrigieren')
                # -------------------------------------grab style ---------------------------------------------------#

                cur = connE.cursor()
                cur.execute('select* from fast' )
                nado = cur.fetchall()
                frei = ''

                for t in nado:
                    frei = t[1]
                # -------------------------------------grab anzahl ---------------------------------------------------#
                cur2 = connE.cursor()
                cur2.execute('select* from DruckAnzahl')
                nado2 = cur2.fetchall()
                anzahl = ''

                for t2 in nado2:
                    anzahl = t2[1]

                #-------------------------------------- Fahrer Speichern-----------------------------------------------#

                def Fahrer_speich():
                    cur = connE.cursor()
                    siko = str(FahrerE.get())
                    cur.execute("INSERT INTO Lieferanten VALUES (NULL, '" + siko + "')")
                    connE.commit()
                    FahrerE.delete(0, END)
                    Einstellung2()
                # -------------------------------------- Fahrer zeigen -----------------------------------------------#
                Fahreliste = Listbox(kundenframe2, width=20, height=10, font=font_size, bd=0)
                Fahreliste.place(x=550, y=210)
                def fahrer_show():
                    cur = connE.cursor()
                    cur.execute('select Name from Lieferanten ')
                    connE.commit()
                    miko = cur.fetchall()
                    Fahreliste.delete(0, END)
                    for zabir in miko:
                        name=zabir[0]

                        Fahreliste.insert(0,name)
                fahrer_show()

                # -------------------------------------- Fahrer zeigen -----------------------------------------------#
                def fahrer_lösch():

                    cur = connE.cursor()
                    siko = clicked.get()
                    soko = siko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"', '')
                    cur.execute('delete from Lieferanten where name =(?)', (soko,))
                    connE.commit()
                    Einstellung2()
                #-------------------------------------- fahrer drop liste ---------------------------------------------#
                options = []
                cur = connE.cursor()
                cur.execute('select Name from Lieferanten ')
                siko = cur.fetchall()
                for ido in siko:
                    list_name=ido[0]
                    options.insert(0, list_name)
                clicked = StringVar()
                drop = OptionMenu(Frame1, clicked, *options)
                clicked.set(ido[0])
                # --------------------------------------Labels and eteries-------------------------------------------------#
                # --------------------------------------Bestellung seite style-------------------------------------------------#
                Familien_ = Label(Frame1, text='Bestellung/Style ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                  width=48)
                Familien_.place(x=0, y=35)

                Familien_Frei = Label(Frame1, text='Style Auswählen :', font=font_size, bd=4, bg='DarkSlateGray4',
                                      height=1)
                Familien_Frei.place(x=0, y=65)
                Familien_Button = Button(Frame1, text='Update', width=8, height=1, command=update_style, bg='#1fc5a8')
                Familien_Button.place(x=310, y=68)
                Familien_A = Label(kundenframe2, text='Bestellung/Style ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                   width=35)
                Familien_A.place(x=500, y=39)
                Familien_Frei = Label(kundenframe2, text=frei, font=font_size, bd=4, bg='light grey',
                                      height=1)
                Familien_Frei.place(x=550, y=70)
                options = ["90s style", "Original style"]
                combobox101 = ttk.Combobox(Frame1, values=options)
                combobox101.place(x=146, y=68)
                # --------------------------------------Lieferschein anzahl-------------------------------------------------#
                anzahl_ = Label(Frame1, text='Lieferschin Anzahl ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                  width=48)
                anzahl_.place(x=0, y=98)

                anzahl_Druck= Label(Frame1, text='Anzahl Druck Lieferschein :', font=font_size, bd=4, bg='DarkSlateGray4',
                                      height=1)
                anzahl_Druck.place(x=0, y=140)
                anzahlE=Entry(Frame1,width=5,bd=4)
                anzahlE.place(x=220, y=140)
                anzahl_Button = Button(Frame1, text='Update', width=8, height=1, command=update_anzahl, bg='#1fc5a8')
                anzahl_Button.place(x=280, y=139)
                anzahl_A = Label(kundenframe2, text='Lieferschin Anzahl ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                   width=35)
                anzahl_A.place(x=500, y=101)
                anzahl_Frei = Label(kundenframe2, text=anzahl, font=font_size, bd=4, bg='light grey',
                                    height=1)
                anzahl_Frei.place(x=550, y=145)


                # -------------------------------------grab style ---------------------------------------------------#
                Fahrerl_ = Label(Frame1, text='Fahrer erstellen ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                  width=48)
                Fahrerl_.place(x=0, y=175)

                Namel_Fahrer= Label(Frame1, text='Name :', font=font_size, bd=4, bg='DarkSlateGray4',
                                      height=1)
                Namel_Fahrer.place(x=0, y=210)
                FahrerE=Entry(Frame1,width=15,bd=4)
                FahrerE.place(x=80, y=210)

                drop.configure(width=15, bd=0, bg='grey', font=font_size)
                drop.place(x=0, y=260)
                Fahrer_Button = Button(Frame1, text='speichern', width=8, height=1, command=Fahrer_speich, bg='#1fc5a8')
                Fahrer_Button.place(x=200, y=210)
                anzahl_A = Label(kundenframe2, text='Fahrer Liste ', font=font_size, bd=4, bg='dark slate grey', height=1,
                                   width=35)
                anzahl_A.place(x=500, y=178)
                loschenB = Button(Frame1, text='Löschen', bd=4, bg='red', command=fahrer_lösch)
                loschenB.place(x=200, y=260)


                # -------------------------------------grab style ---------------------------------------------------#
            def Data():
                global kundenframe2

                for widget in kundenframe2.winfo_children():
                    widget.destroy()
                font_size = ('arial', 12, 'bold')
                kundenframe2.config(width=930, bg='DarkSlateGray4',height=780)
                #Frames
                leftframe=Frame(kundenframe2,width=300,height=750,bg='SlateGray4')
                leftframe.place(x=0,y=5)
                # conn1 = sqlite3.connect(resource_path('Data\DataAnalysis.db'))
                # cur1 = conn1.cursor()
                # cur1.execute("""Create Table kundeinfo(
                #
                #         ID integer primary key AUTOINCREMENT UNIQUE,
                #         kid INTEGER,
                #         ktelefonnummer INTEGER,
                #         kname TEXT,
                #         kstrasse TEXT,
                #         khausnr INTEGER,
                #         kplz INTEGER,
                #         kort TEXT,
                #         kemail TEXT,
                #         bediener TEXT,
                #         zeit INTEGER,
                #         bestellzeit INTEGER,
                #         externinfo TEXT,
                #         gesamtepreis INTEGER)""")
                # cur1.execute("""Create Table speiseinfo(
                #             zeit INTEGER,
                #             pos INTEGER,
                #             grosse TEXT,
                #             anzahl INTEGER,
                #             nr INTEGER,
                #             speise TEXT,
                #             mit TEXT,
                #             ohne TEXT,
                #             katagorie TEXT,
                #             preis INTEGER,
                #             name TEXT )""")
                # cur1.execute("CREATE TABLE lieferpara(geld text, zeit Integer)")
                # conn1.commit()
                #labels
                Diversl=Label(leftframe,text='Diverse Daten',font=('arial',16,'bold'),bg='SlateGray4')
                Diversl.place(x=60,y=10)
                Häufigstel=Label(leftframe,text='Häufigste bestellet:',font=('arial',12,'bold'),bg='SlateGray4')
                Häufigstel.place(x=0,y=60)






            ############################################seite 2 chef einsetllung##################################################################################################################################
            def openwindow1():


                global kundenframe2, Bestellliste,openchef
                font_size = ("ARIEL", 10, "bold")
                global new_window
                root.wm_state('iconic')
                global new_window
                new_window = Toplevel(root)
                screen_width = root.winfo_screenwidth()
                screen_height = root.winfo_screenheight()
                new_window.geometry("1920x1080")
                openchef+=1

                def on_close():
                    global openchef
                    openchef = 0
                    new_window.destroy()

                new_window.protocol("WM_DELETE_WINDOW", on_close)



                new_window.config(bg='#1A3A3A')
                ########################################speisekarte#################################################################

                ##########################################KUNDENDAETN/EINSTELLUNG##################################################

                #################### Mianframes
                kundenframe2 = Frame(new_window, bd=8, width=1000, height=870, bg='#102b37')
                kundenframe2.place(x=160, y=30)
                kundenframe3 = Frame(new_window, bd=8, width=140, height=700, bg='#051821')
                kundenframe3.place(x=5, y=30)
                kundenframe4 = Frame(new_window, bd=8, width=140, height=700, bg='#051821')
                kundenframe4.place(x=1200, y=30)
                ################################################## PLOT BESTELLUNG ###################################################

                Menulabel = Label(new_window, text='Menüleiste', font=('Helvetica', 14, "italic bold"), bg='#1A3A3A',
                                  fg='silver')
                Menulabel.place(x=10, y=10)
                Speisekartel = Label(new_window, text='Speisekarte', font=('Helvetica', 12, "italic bold"), bg='#1A3A3A',
                                     fg='silver')
                Speisekartel.place(x=17, y=55)
                speiseverbl = Label(new_window, text='--\n\n\n\n-\n\n\n\n-\n\n\n\n-', font=('Helvetica', 10, "italic bold"),
                                    bg='#071317', fg='silver')
                speiseverbl.place(x=5, y=55)
                Einstellungl = Label(new_window, text='Einstellung', font=('Helvetica', 12, "italic bold"), bg='#1A3A3A',
                                     fg='silver')
                Einstellungl.place(x=17, y=300)
                Einstelverbl = Label(new_window, text='--\n\n\n-\n\n\n\n-\n\n\n\n-\n\n\n\n-',
                                     font=('Helvetica', 10, "italic bold"),
                                     bg='#071317', fg='silver')
                Einstelverbl.place(x=5, y=300)
                Kundenl = Label(new_window, text='Kunde', font=('Helvetica', 12, "italic bold"), bg='#1A3A3A',
                                fg='silver')
                Kundenl.place(x=1255, y=55)
                kundenverbl = Label(new_window, text='--\n\n\n-', font=('Helvetica', 10, "italic bold"),
                                    bg='#071317', fg='silver')
                kundenverbl.place(x=1325, y=55)
                Bestellungl = Label(new_window, text='Bestellung', font=('Helvetica', 12, "italic bold"), bg='#1A3A3A',
                                    fg='silver')
                Bestellungl.place(x=1236, y=170)
                Bestellverbl = Label(new_window, text='--\n\n\n-\n\n\n\n-', font=('Helvetica', 10, "italic bold"),
                                     bg='#071317', fg='silver')
                Bestellverbl.place(x=1325, y=170)

                Displabel = Label(new_window, text='Display', font=('Helvetica', 14, "italic bold"), bg='#1A3A3A', fg='silver')
                Displabel.place(x=460, y=5)
                backupverbl = Label(new_window, text='--\n\n\n-', font=('Helvetica', 10, "italic bold"),
                                    bg='#071317', fg='silver')
                backupverbl.place(x=1325, y=340)
                dival = Label(new_window, text='Diverse', font=('Helvetica', 12, "italic bold"), bg='#1A3A3A',
                                    fg='silver')
                dival.place(x=1250, y=340)
                # onlabel = Label(kundenframe2, text='M.B.P.S', font=('Helvetica', 75, "italic bold"), bg='#1A3A3A', fg='silver')
                # onlabel.place(x=190, y=400)

                import numpy as np
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
                from collections import Counter

                cur303 = connO.cursor()
                cur303.execute('select*  from kundeinfo')
                timep = cur303.fetchall()

                timing = [info[10][9:11] for info in timep]

                day = [tag[10][:1] for tag in timep]

                hour_counts = Counter(timing)
                day_counts = Counter(day)

                ytimes = np.arange(10, 24 + 1, 1)
                xlist = [hour_counts[str(i)] for i in ytimes]

                fig = plt.figure()

                ax = fig.add_subplot(211)
                plt.xlabel('Hour')
                plt.ylabel('Number of Orders')
                plt.title('Anzahl Bestellung per Stunde')
                ax.plot(ytimes, xlist)
                canvas = FigureCanvasTkAgg(fig, master=kundenframe2)
                canvas.get_tk_widget().place(x=5, y=20)

                fig2 = plt.figure()
                ax2 = fig2.add_subplot(211)

                months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                counts = [0] * len(months)

                for info in timep:
                    month = int(info[10][3:5])

                    counts[month - 1] += 1

                ax2.bar(months, counts)
                ax2.set_xlabel('Month')
                ax2.set_ylabel('Number of Orders')
                ax2.set_title('Anzahl Bestellung pro Monat')

                canvas2 = FigureCanvasTkAgg(fig2, master=kundenframe2)
                canvas2.draw()
                canvas2.get_tk_widget().place(x=5, y=360)

                # Kundentitle = Frame(new_window, bd=8, width=370, height=65, bg='BLACK')
                # Kundentitle.place(x=384, y=320)
                # Kundentitle1 = Label(new_window, text='Kunden/Einstellung', font=("Ariel", 30, "bold"), bg='#1A3A3A', bd=8)
                # Kundentitle1.place(x=360, y=328)
                Zutaten_Preise = Button(kundenframe3, text=('Zutaten preise\n 💰🥕💰'), bd=4, bg='#0D7377', width=12, height=2,
                                        font=("Ariel", 10, "bold"),
                                        relief=GROOVE, activebackground='#367381', command=zutatenpreise)
                Zutaten_Preise.place(x=5, y=60)
                Bestellungb = Button(kundenframe3, text=('Speisen\n 🍕🍕🍕'), bd=4, bg='#0D7377', width=12, height=2,
                                     command=Bestellung,
                                     fg="BLACK",
                                     relief=GROOVE, activebackground='red', font=("Ariel", 10, "bold"))
                Bestellungb.place(x=5, y=180)

                Zutaten_ = Button(kundenframe3, text=('Zutaten \n 🥕🥦🍅🥒'), bd=4, bg='#0D7377', width=12, height=2,
                                  fg="black", font=("Arial", 10, "bold"),
                                  relief=GROOVE, activebackground='red', command=zutaten)

                Zutaten_.place(x=5, y=120)

                Kunden = Button(kundenframe4, text=('Kunden\n 🧑 👩 👨 '), bd=4, bg='#0D7377', width=12, height=2,
                                fg="black", font=("Ariel", 10, "bold"),
                                relief=GROOVE, activebackground='red', command=kunden)
                Kunden.place(x=5, y=60)
                Diverse_Einstellung = Button(kundenframe3, text=('Diverse\n 👷👷👷 '), bd=4, bg='#0D7377', width=12,
                                             height=2,
                                             fg="black", font=("Ariel", 10, "bold"),
                                             relief=GROOVE, activebackground='red', command=Einstellung)
                Diverse_Einstellung.place(x=5, y=310)
                Bestellung_liste = Button(kundenframe4, text=('Bestellung Liste\n 📝📊📋' ), bd=4, bg='#0D7377', width=12, height=2,
                                          fg="black", font=("Ariel", 10, "bold"),
                                          relief=GROOVE, activebackground='red', command=Bestellliste)
                Bestellung_liste.place(x=5, y=165)

                # Liefrante = Button(kundenframe4, text=('Liefranten\n 🚕 🚙 🛵'), bd=4, bg='#0D7377', width=12, height=2,
                #                    fg="black", font=("Ariel", 10, "bold"),
                #                    relief=GROOVE, activebackground='red', command=openwindow7)
                # Liefrante.place(x=5, y=230)
                Druckerb = Button(kundenframe3, text=('Drucker\n 🖨️ '), bd=4, bg='#0D7377', width=12, height=2,
                                  fg="black", font=("Ariel", 10, "bold"),
                                  relief=GROOVE, activebackground='red', command=Drucker)
                Druckerb.place(x=5, y=370)
                Freiezutaten = Button(kundenframe3, text=('Freiezutataen\n 🎁💸'), bd=4, bg='#0D7377', width=12, height=2,
                                      fg="black", font=("Ariel", 10, "bold"),
                                      relief=GROOVE, activebackground='red', command=Freizutaten)
                Freiezutaten.place(x=5, y=430)
                secondpc = Button(kundenframe3, text=('Second PC\n 💻💻💻'), bd=4, bg='#0D7377', width=12,
                                  fg="black", font=("Ariel", 10, "bold"), height=2,
                                  relief=GROOVE, activebackground='green', command=Zweitepc)
                secondpc.place(x=5, y=490)
                btn3 = Button(kundenframe4, text=("Backup",), bd=1, bg='#475B37', height=1, width=10, fg="black",
                              command=backup,
                              font=('Helvetica', 9,
                                    "italic bold"),
                              relief=GROOVE, activebackground='blue')

                btn3.place(x=36, y=410)
                Diverse_Einstellung2 = Button(kundenframe4, text=('Diverse2\n 👷👷👷 '), bd=4, bg='#0D7377', width=12,
                                             height=2,
                                             fg="black", font=("Ariel", 10, "bold"),
                                             relief=GROOVE, activebackground='red', command=Einstellung2)
                Diverse_Einstellung2.place(x=8, y=350)
                Data_Analyze = Button(kundenframe4, text=('Data Analyze\n 📊📈📉  '), bd=4, bg='#0D7377', width=12,
                                              height=2,
                                              fg="black", font=("Ariel", 10, "bold"),
                                              relief=GROOVE, activebackground='red', command=Data)
                Data_Analyze.place(x=5, y=230)


            #########################################  Passwort zu Chef mit functions  ############################################
            def closepopup():
                top.destroy()


            # ------------------------------------------------------------------------------------------------------------------#
            def passwortget(event=None):


                cur = connE.cursor()
                cur.execute('select* from Passwort ')
                Passwort = cur.fetchall()
                Passwort1 = str(Passwort)
                Passwort2 = Passwort1.replace("(", "").replace(")", "").replace(",", "").replace("'", "").replace("[",
                                                                                                                  "").replace(
                    "]", "")
                connE.commit()
                Password = Passwort2
                passwordE = entery.get()
                if passwordE == Password:
                    openwindow1()
                    closepopup()

                elif passwordE == ('773468810598'):
                    closepopup()
                    Error_report()

                else:
                    closepopup()


            # ----------------------------------------------------------------------------------------------------------------------#
            def popupwin(event=None):
                global top

                if openchef==0:


                    top = Toplevel(root)
                    w13 = 350
                    h13 = 100
                    # Get the width and height of the screen
                    screen_width = root.winfo_screenwidth()
                    screen_height = root.winfo_screenheight()
                    # Calculate the x and y coordinates for the top-left corner of the Toplevel window
                    x = (screen_width - w13) // 2
                    y = (screen_height - h13) // 2
                    top.geometry(f"{w13}x{h13}+{x}+{y}")
                    top.config(bg=colour4)
                    top.title('CHEF Passwort')
                    global entery
                    entery = Entry(top, show="*", width=25)
                    label1 = Label(top, text="Passwort eingeben", font=("Ariel", 12), bg=colour4)
                    entery.focus_force()
                    label1.pack()
                    entery.pack()

                    def password_enter(e):
                        passwortget()

                    # ----------------------------------------------------------------------------------------------------------------------#

                    Button1 = Button(top, text="Ok", width=10, command=passwortget)
                    Button2 = Button(top, text="Quit", width=10, command=closepopup)
                    Button1.place(x=40, y=60)
                    Button2.place(x=185, y=60)
                    top.bind("<Return>", passwortget)
                else:
                    pass

                # ----------------------------------------------------------------------------------------------------------------------#



            ###################################CHEF SEITE ##########################################################################

            # ######################################  seite 2 button  ################################################################

            frame3=Frame(bd=4,bg=colour4,width=500, height=300)
            frame3.place(x=800,y=350)
            label1=Label(frame3,bg=colour4,font=('Helvetica', 20,"italic bold"),text='MindMeshLab')
            label1.place(x=10,y=10)
            label2=Label(frame3,bg=colour4,font=('Helvetica', 12,"italic bold"),text='A Creative Software Development Company')
            label2.place(x=10,y=50)
            label3 = Label(frame3, bg=colour4, font=('Helvetica', 12, "italic bold"),text='www.mindmeshlab.de')
            label3.place(x=10,y=110)


            btn2 = Button(root, text=("Chef👨‍🏭",), bg=colour4, width=12,  fg="black", command=popupwin,height=2,
                          font=('Helvetica', 36,
                                "italic bold"),
                           activebackground='blue')
            btn2.pack(padx=10, pady=10)

            btn2.place(x=80, y=460)


            ###################################################### Backup#######################################################

            # -------------------------------------------------- Liefranet _ Rechunung----------------------------------------------
            def Error_report():
                class Pipe:
                    """mock stdin stdout or stderr"""

                    def __init__(self):
                        self.buffer = queue.Queue()
                        self.reading = False

                    def write(self, data):
                        self.buffer.put(data)

                    def flush(self):
                        pass

                    def readline(self):
                        self.reading = True
                        line = self.buffer.get()
                        self.reading = False
                        return line

                class Console(tk.Frame):
                    """A tkinter widget which behaves like an interpreter"""

                    def __init__(self, parent, _locals, exit_callback):
                        super().__init__(parent)

                        self.text = ConsoleText(self, wrap=tk.WORD)
                        self.text.pack(fill=tk.BOTH, expand=True)

                        self.shell = code.InteractiveConsole(_locals)

                        # make the enter key call the self.enter function
                        self.text.bind("<Return>", self.enter)
                        self.prompt_flag = True
                        self.command_running = False
                        self.exit_callback = exit_callback

                        # replace all input and output
                        sys.stdout = Pipe()
                        sys.stderr = Pipe()
                        sys.stdin = Pipe()

                        def loop():
                            self.read_from_pipe(sys.stdout, "stdout")
                            self.read_from_pipe(sys.stderr, "stderr", foreground='red')

                            self.after(50, loop)

                        self.after(50, loop)

                    def prompt(self):
                        """Add a '>>> ' to the console"""
                        self.prompt_flag = True

                    def read_from_pipe(self, pipe: Pipe, tag_name, **kwargs):
                        """Method for writing data from the replaced stdout and stderr to the console widget"""

                        # write the >>>
                        if self.prompt_flag and not self.command_running:
                            self.text.prompt()
                            self.prompt_flag = False

                        # get data from buffer
                        string_parts = []
                        while not pipe.buffer.empty():
                            part = pipe.buffer.get()
                            string_parts.append(part)

                        # write to console
                        str_data = ''.join(string_parts)
                        if str_data:
                            if self.command_running:
                                insert_position = "end-1c"
                            else:
                                insert_position = "prompt_end"

                            self.text.write(str_data, tag_name, insert_position, **kwargs)

                    def enter(self, e):
                        """The <Return> key press handler"""

                        if sys.stdin.reading:
                            # if stdin requested, then put data in stdin instead of running a new command
                            line = self.text.consume_last_line()
                            line = line + '\n'
                            sys.stdin.buffer.put(line)
                            return

                        # don't run multiple commands simultaneously
                        if self.command_running:
                            return

                        # get the command text
                        command = self.text.read_last_line()
                        try:
                            # compile it
                            compiled = code.compile_command(command)
                            is_complete_command = compiled is not None
                        except (SyntaxError, OverflowError, ValueError):
                            # if there is an error compiling the command, print it to the console
                            self.text.consume_last_line()
                            self.prompt()
                            traceback.print_exc()
                            return

                        # if it is a complete command
                        if is_complete_command:
                            # consume the line and run the command
                            self.text.consume_last_line()

                            self.prompt()
                            self.command_running = True

                            def run_command():
                                try:
                                    self.shell.runcode(compiled)
                                except SystemExit:
                                    self.after(0, self.exit_callback)

                                self.command_running = False

                            threading.Thread(target=run_command).start()

                class ConsoleText(ScrolledText):
                    """
                    A Text widget which handles some application logic,
                    e.g. having a line of input at the end with everything else being uneditable
                    """

                    def __init__(self, *args, **kwargs):
                        super().__init__(*args, **kwargs)

                        # make edits that occur during on_text_change not cause it to trigger again
                        def on_modified(event):
                            flag = self.edit_modified()
                            if flag:
                                self.after(10, self.on_text_change(event))
                            self.edit_modified(False)

                        self.bind("<<Modified>>", on_modified)

                        # store info about what parts of the text have what colour
                        # used when colour info is lost and needs to be re-applied
                        self.console_tags = []

                        # the position just before the prompt (>>>)
                        # used when inserting command output and errors
                        self.mark_set("prompt_end", 1.0)

                        # keep track of where user input/commands start and the committed text ends
                        self.committed_hash = None
                        self.committed_text_backup = ""
                        self.commit_all()

                    def prompt(self):
                        """Insert a prompt"""
                        self.mark_set("prompt_end", 'end-1c')
                        self.mark_gravity("prompt_end", tk.LEFT)
                        self.write(">>> ", "prompt", foreground="blue")
                        self.mark_gravity("prompt_end", tk.RIGHT)

                    def commit_all(self):
                        """Mark all text as committed"""
                        self.commit_to('end-1c')

                    def commit_to(self, pos):
                        """Mark all text up to a certain position as committed"""
                        if self.index(pos) in (self.index("end-1c"), self.index("end")):
                            # don't let text become un-committed
                            self.mark_set("committed_text", "end-1c")
                            self.mark_gravity("committed_text", tk.LEFT)
                        else:
                            # if text is added before the last prompt (">>> "), update the stored position of the tag
                            for i, (tag_name, _, _) in reversed(list(enumerate(self.console_tags))):
                                if tag_name == "prompt":
                                    tag_ranges = self.tag_ranges("prompt")
                                    self.console_tags[i] = ("prompt", tag_ranges[-2], tag_ranges[-1])
                                    break

                        # update the hash and backup
                        self.committed_hash = self.get_committed_text_hash()
                        self.committed_text_backup = self.get_committed_text()

                    def get_committed_text_hash(self):
                        """Get the hash of the committed area - used for detecting an attempt to edit it"""
                        return hashlib.md5(self.get_committed_text().encode()).digest()

                    def get_committed_text(self):
                        """Get all text marked as committed"""
                        return self.get(1.0, "committed_text")

                    def write(self, string, tag_name, pos='end-1c', **kwargs):
                        """Write some text to the console"""

                        # get position of the start of the text being added
                        start = self.index(pos)

                        # insert the text
                        self.insert(pos, string)
                        self.see(tk.END)

                        # commit text
                        self.commit_to(pos)

                        # color text
                        self.tag_add(tag_name, start, pos)
                        self.tag_config(tag_name, **kwargs)

                        # save color in case it needs to be re-colured
                        self.console_tags.append((tag_name, start, self.index(pos)))

                    def on_text_change(self, event):
                        """If the text is changed, check if the change is part of the committed text, and if it is revert the change"""
                        if self.get_committed_text_hash() != self.committed_hash:
                            # revert change
                            self.mark_gravity("committed_text", tk.RIGHT)
                            self.replace(1.0, "committed_text", self.committed_text_backup)
                            self.mark_gravity("committed_text", tk.LEFT)

                            # re-apply colours
                            for tag_name, start, end in self.console_tags:
                                self.tag_add(tag_name, start, end)

                    def read_last_line(self):
                        """Read the user input, i.e. everything written after the committed text"""
                        return self.get("committed_text", "end-1c")

                    def consume_last_line(self):
                        """Read the user input as in read_last_line, and mark it is committed"""
                        line = self.read_last_line()
                        self.commit_all()
                        return line

                if __name__ == '__main__':
                    root = tk.Tk()
                    root.config(background="red")
                    main_window = Console(root, locals(), root.destroy)
                    main_window.pack(fill=tk.BOTH, expand=True)
                    root.mainloop()


            def openwindow7():
                global new_window, new_window2,counter
                font_size = ("ARIEL", 12, "bold")
                new_window6 = Toplevel(root)
                screen_width = 1200
                screen_height = 1100

                new_window6.geometry("1920x1080")

                new_window6.configure(bg='cadet blue')
                new_window6.title(220 * titlespace + "Liste")





                # new_window6.attributes('-fullscreen', True)
                Frame1 = Frame(new_window6, width=300, bg='SlateGray4', bd=4, height=459, relief=RIDGE)
                Frame1.place(x=0, y=0)
                Frame6 = Frame(new_window6, width=300, bg='SlateGray4', bd=4, height=140, relief=RIDGE)
                Frame6.place(x=0, y=500)
                Frame2 = Frame(new_window6, width=450, bg='grey', bd=4, height=400, relief=RIDGE)
                Frame2.place(x=300, y=60)
                Frame3 = Frame(new_window6, width=560, bg='grey', bd=4, height=750, relief=RIDGE)
                Frame3.place(x=860, y=60)
                Frame4 = Frame(new_window6, width=450, bg='grey', bd=4, height=400, relief=RIDGE)
                Frame4.place(x=300, y=500)
                # Frame6 = Frame(new_window6, width=600, bg='white', bd=4, height=260, relief=RIDGE)
                # Frame6.place(x=0, y=650)
                Frame5 = Frame(new_window6, width=600, bg='SlateGray4', bd=4, height=140, relief=RIDGE)
                Frame5.place(x=5, y=650)

                # conn=sqlite3.connect('Einstellung.db')
                # cur=conn.cursor()
                # cur.execute("""Create Table Lieferanten ( ID integer primary key AUTOINCREMENT, Name TEXT ) """)
                # conn.commit()
                # ------------------------------------------------ functions--------------------------------------------------------#
                def name_speich():

                    cur = connE.cursor()
                    siko = NameE.get()
                    cur.execute('insert into Lieferanten values(Null,?)', (siko,))
                    connE.commit()
                    NameE.delete(0, END)
                    new_window6.destroy()
                    openwindow7()

                # ------------------------------------------------Labels------------------------------------------------------------#
                fahrerl = Label(Frame1, text='Fahrer erstellen ', bg='cadet blue', bd=0, relief=RIDGE, font=font_size, width=28)
                fahrerl.place(x=0, y=5)
                Namel = Label(Frame1, text='Name:', bg='SlateGray4', bd=0, font=font_size)
                Namel.place(x=0, y=35)
                fahrerlo = Label(Frame1, text='Fahrer Löschen ', bg='cadet blue', bd=0, relief=RIDGE, font=font_size, width=28)
                fahrerlo.place(x=0, y=65)
                fahrerlis = Label(Frame1, text='Fahrer Liste', bg='cadet blue', bd=0, relief=RIDGE, font=font_size, width=28)
                fahrerlis.place(x=0, y=140)
                namel = Label(Frame1, text='Namen', bg='SlateGray4', bd=0, relief=RIDGE, font=font_size, width=10)
                namel.place(x=0, y=165)
                namef = Label(Frame6, text='Fahrer', bg='SlateGray4', bd=0, relief=RIDGE, font=('arial', 16, 'bold'))
                namef.place(x=15, y=5)
                nummer = Label(new_window6, text=' BestellungsNr:', bg='cadet blue', bd=0, font=('arial', 16, 'bold'))
                nummer.place(x=300, y=25)
                ofennebes = Label(new_window6, text=' Offene Bestellung:', bg='cadet blue', bd=0, font=('arial', 16, 'bold'))
                ofennebes.place(x=870, y=25)
                pendingl = Label(new_window6, text=' Pending', bg='cadet blue', bd=0, font=('arial', 16, 'bold'))
                pendingl.place(x=470, y=470)
                Anzahl = Label(Frame5, text=' Anzahl/B', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                Anzahl.place(x=175, y=10)
                Datum = Label(Frame5, text=' Datum', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                Datum.place(x=290, y=10)
                GPreis = Label(Frame5, text=' Gesamt preis', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                GPreis.place(x=390, y=10)
                ffahrer = Label(Frame5, text=' Fahrer', bg='SlateGray4', bd=0, font=('arial', 16, 'bold'))
                ffahrer.place(x=20, y=10)
                # -----------------------------------------------------------Enetry_Liste----------------------------------------------#
                NameE = Entry(Frame1, width=15, font=font_size, bd=3, )
                NameE.place(x=60, y=35)
                Fahreliste = Listbox(Frame1, width=20, height=10, font=font_size, bd=0)
                Fahreliste.place(x=10, y=195)
                nummerL = Listbox(new_window6, width=3, height=0, font=font_size)
                nummerL.place(x=1066, y=25)
                nummerE = Entry(new_window6, width=3, font=font_size, bd=3)
                nummerE.place(x=460, y=25)
                AnzE = Entry(Frame5, width=3, font=font_size, bd=3)
                AnzE.place(x=195, y=40)
                DatE = Entry(Frame5, width=9, font=font_size, bd=3)
                DatE.place(x=290, y=40)
                GesE = Entry(Frame5, width=9, font=font_size, bd=3)
                GesE.place(x=410, y=40)
                # -----------------------------------------------------------DropBox----------------------------------------------------#
                options = []
                options2=[]
                cur = connE.cursor()
                cur.execute('select Name from Lieferanten ')
                siko = cur.fetchall()
                for ido in siko:
                    options.insert(0, ido)
                conn12 = sqlite3.connect(resource_path('Data\Pending.db'))
                cur12 = conn12.cursor()
                cur12.execute('select Fahrer from pending ')
                open_fahrer=cur12.fetchall()
                if open_fahrer:
                    for fahrers in open_fahrer:
                        options2.insert(0,fahrers)
                else:
                    options2.insert(0,' ')
                clicked = StringVar()

                clicked1 = StringVar()
                clicked2 = StringVar()
                drop = OptionMenu(Frame1, clicked, *options)
                clicked.set(ido[0])
                drop.configure(width=15, bd=0, bg='grey', font=font_size)
                drop.place(x=0, y=95)
                drop1 = OptionMenu(Frame6, clicked1, *options)
                drop1.configure(width=10, bd=0, bg='grey', font=('arial', 16, 'bold'))
                drop1.place(x=0, y=30)

                # ------------------------------------------------------------------------------------------------------------------#
                def name_lösch():

                    cur = connE.cursor()
                    siko = clicked.get()
                    soko = siko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"', '')
                    cur.execute('delete from Lieferanten where name =(?)', (soko,))
                    connE.commit()
                    new_window6.destroy()
                    openwindow7()

                # ----------------------------------------------------------#

                cur = connE.cursor()
                cur.execute('select Name from Lieferanten ')
                connE.commit()
                miko = cur.fetchall()
                for zabir in miko:
                    Fahreliste.insert(0, zabir)
                # ---------------------------------------------------------------Buttons------------------------------------------------#
                NameB = Button(Frame1, text='speichern', bd=4, bg='green', command=name_speich).place(x=210, y=33)
                loschenB = Button(Frame1, text='Löschen', bd=4, bg='red', command=name_lösch)
                loschenB.place(x=200, y=93)
                # conn = sqlite3.connect('Pending.db')
                # cur = conn.cursor()
                # cur.execute("""Create Table Rechnung ( Nr INTEGER, Name TEXT,Datum TEXT, Stasse TEXT, Hnr TEXT,Preis INTEGER ,Fahrer TEXT ) """)
                # conn.commit()
                # ------------------------------------------------------Tree1---------------------------------------------------------#
                # style = ttk.Style()
                # style.theme_use("clam")
                # style.configure("Tree1", background='white', foreground="black", rowheight=20, font=("ARIEL", 9, 'bold'),
                #                 fieldbackground="silver")
                # style.map('Tree1', background=[('selected', 'red')])
                # style.configure("Tree1.Heading", font=('bold', 12))
                # style.configure("Tree1.column", font=('bold', 6))
                # style.configure('Tree1', rowheight=20)
                # style.configure('Tree1.row', font=('bold', 20))
                besteltree = ttk.Treeview(Frame2, height=18)
                besteltree['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                besteltree.column("#0", width=0, stretch=NO)
                besteltree.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                besteltree.column("Name", anchor=CENTER, width=120)
                besteltree.column("Datum/Uhrzeit", anchor=W, width=160)
                besteltree.column("Straße", anchor=W, width=120)
                besteltree.column("Haus/nr", anchor=W, width=40)
                besteltree.column("Preis", anchor=W, width=60)
                besteltree.tag_configure('pos', background='white')
                besteltree.heading("#0", text="", anchor=W)
                besteltree.heading("Nr", text="Nr", anchor=W)
                besteltree.heading("Name", text="Name", anchor=W)
                besteltree.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                besteltree.heading("Straße", text="Straße", anchor=W)
                besteltree.heading("Haus/nr", text="nr", anchor=W)
                besteltree.heading("Preis", text="Preis", anchor=W)
                besteltree.pack()
                # ------------------------------------------------------- Tree2---------------------------------------------------------#
                besteltree1 = ttk.Treeview(Frame3, height=27)
                besteltree1['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                besteltree1.column("#0", width=0, stretch=NO)
                besteltree1.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                besteltree1.column("Name", anchor=CENTER, width=120)
                besteltree1.column("Datum/Uhrzeit", anchor=W, width=160)
                besteltree1.column("Straße", anchor=W, width=120)
                besteltree1.column("Haus/nr", anchor=W, width=40)
                besteltree1.column("Preis", anchor=W, width=60)
                besteltree1.tag_configure('pos', background='white')
                besteltree1.heading("#0", text="", anchor=W)
                besteltree1.heading("Nr", text="Nr", anchor=W)
                besteltree1.heading("Name", text="Name", anchor=W)
                besteltree1.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                besteltree1.heading("Straße", text="Straße", anchor=W)
                besteltree1.heading("Haus/nr", text="nr", anchor=W)
                besteltree1.heading("Preis", text="Preis", anchor=W)
                besteltree1.pack()
                # ----------------------------------------------------- Tree 3 ---------------------------------------------------------#
                besteltree2 = ttk.Treeview(Frame4, height=5)
                besteltree2['columns'] = ("Nr", "Name", "Datum/Uhrzeit", "Straße", "Haus/nr", "Preis")
                besteltree2.column("#0", width=0, stretch=NO)
                besteltree2.column("Nr", anchor=CENTER, width=40, stretch=TRUE, )
                besteltree2.column("Name", anchor=CENTER, width=120)
                besteltree2.column("Datum/Uhrzeit", anchor=W, width=160)
                besteltree2.column("Straße", anchor=W, width=120)
                besteltree2.column("Haus/nr", anchor=W, width=40)
                besteltree2.column("Preis", anchor=W, width=60)
                besteltree2.tag_configure('pos', background='white')
                besteltree2.heading("#0", text="", anchor=W)
                besteltree2.heading("Nr", text="Nr", anchor=W)
                besteltree2.heading("Name", text="Name", anchor=W)
                besteltree2.heading("Datum/Uhrzeit", text="Datum/Uhrzeit", anchor=W)
                besteltree2.heading("Straße", text="Straße", anchor=W)
                besteltree2.heading("Haus/nr", text="nr", anchor=W)
                besteltree2.heading("Preis", text="Preis", anchor=W)
                besteltree2.pack()

                # -------------------------------------------- Tress Functions ---------------------------------------------------------#
                def besteliste():
                    for record in besteltree1.get_children():
                        besteltree1.delete(record)
                    count = 0
                    conn1 = sqlite3.connect(resource_path('Data\OrderData.db'))
                    cur1 = conn1.cursor()
                    cur1.execute('select * from kundeinfo ')
                    firo = cur1.fetchall()
                    for ziko in firo:
                        besteltree1.insert(parent='', open=True, index='0', iid=count, text='',
                                           values=(ziko[0], ziko[3], ziko[10], ziko[4], ziko[5], ziko[13]))
                        count += 1
                    zuzu = besteltree1.get_children()
                    moso = len(zuzu)
                    nummerL.delete(0, END)
                    nummerL.insert(0, moso)

                besteliste()
                global zaro
                zaro = 0

                # ----------------------------------------------------------------------------------------------------------------------#
                def add_fun():
                    global zaro
                    zopl = nummerE.get()
                    cur = connO.cursor()
                    cur.execute('select* from kundeinfo where ID=(?)', (zopl,))
                    bol = cur.fetchall()
                    for opo in bol:
                        besteltree.insert(parent='', open=True, index='end', iid=zaro, text='',
                                          values=(opo[0], opo[3], opo[10], opo[4], opo[5], opo[13]))
                        zaro += 1
                    cur.execute('select* from speiseninfo where zeit=(?)', (opo[10],))
                    samo=cur.fetchall()
                    print(opo10)
                    for sam in samo:
                        print(sam)
                    # curd=connD.cursor()
                    #
                    # cur1.execute("""INSERT INTO speiseinfo (zeit, pos, grosse, anzahl, nr, speise, mit, ohne, katagorie, preis, name)
                    #                 VALUES (:zeit, :pos, :grosse, :anzahl, :nr, :speise, :mit, :ohne, :katagorie, :preis, :name)""",
                    #              {'zeit': timed,
                    #               'pos': sick100,
                    #               'grosse': sick101,
                    #               'anzahl': sick102,
                    #               'nr': sick103,
                    #               'speise': sick104,
                    #               'mit': sick105 + ' ',
                    #               'ohne': sick106,
                    #               'katagorie': sick107,
                    #               'preis': sick108,
                    #               'name': sick3
                    #               })
                    # connO.commit()
                    # cur.execute('delete from kundeinfo where ID=(?)', (zopl,))
                    connO.commit()
                    nummerE.delete(0, END)
                    besteliste()

                # ----------------------------------------------------------------------------------------------------------------------#
                def add_abhol():
                    global zaro
                    zopl = 'ABHOLUNG'
                    cur = connO.cursor()
                    cur.execute('select* from kundeinfo where kstrasse =(?)', (zopl,))
                    bol = cur.fetchall()
                    for opo in bol:
                        besteltree.insert(parent='', open=True, index='end', iid=zaro, text='',
                                          values=(opo[0], opo[3], opo[10], opo[4], opo[5], opo[13]))
                        zaro += 1
                    cur.execute('delete from kundeinfo WHERE kstrasse =(?)', (zopl,))
                    connO.commit()
                    nummerE.delete(0, END)
                    besteliste()

                # ----------------------------------------------------------------------------------------------------------------------#
                def pending():
                    global zazo
                    selected = besteltree.selection()[0]
                    values = besteltree.item(selected, 'values')
                    besteltree.delete(selected)
                    countss = 0
                    besteltree2.insert(parent='', open=True, index='end', text='',
                                       values=(values[0], values[1], values[2], values[3], values[4], values[5]))
                    countss += 1

                # ------------------------------------------------- Tree Buttons -------------------------------------------------------#
                add = Button(new_window6, text='Add', bg='green', width=4, bd=4, command=add_fun, state=DISABLED)
                add.place(x=510, y=25)
                pend = Button(new_window6, text='Pend', bg='orange', width=7, bd=4, command=pending, state=DISABLED)
                pend.place(x=560, y=25)
                addabh = Button(new_window6, text='Alle Abholung', bg='green', width=11, bd=4, command=add_abhol,
                                state=DISABLED)
                addabh.place(x=1100, y=20)

                # ----------------------------------------------------------------------------------------------------------------------#
                def Speichern():
                    siko = str(clicked1.get())
                    soko = siko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"', '')

                    conn = sqlite3.connect(resource_path('Data\Pending.db'))
                    cur = conn.cursor()
                    cur.execute("SELECT * FROM Rechnung WHERE Fahrer = ?", (soko,))
                    rows = cur.fetchall()

                    if len(rows) > 0:
                        for line in besteltree.get_children():
                            sick105 = (besteltree.item(line)['values'][5])
                            cur.execute("UPDATE Rechnung SET Preis = Preis + ? WHERE Fahrer = ?", (sick105, soko))
                    else:
                        for line in besteltree.get_children():
                            sick100 = (besteltree.item(line)['values'][0])
                            sick101 = (besteltree.item(line)['values'][1])
                            sick102 = (besteltree.item(line)['values'][2])
                            sick103 = (besteltree.item(line)['values'][3])
                            sick104 = (besteltree.item(line)['values'][4])
                            sick105 = (besteltree.item(line)['values'][5])
                            conn = sqlite3.connect(resource_path('Data\Pending.db'))
                            cur = conn.cursor()
                            cur.execute(
                                """ insert into Rechnung (Nr,Name,Datum,Strasse,Hnr,Preis,Fahrer)values(:Nr,:Name,:Datum,:Strasse,:Hnr,:Preis,:Fahrer)""",
                                {'Nr': sick100,
                                 'Name': sick101,
                                 'Datum': sick102,
                                 'Strasse': sick103,
                                 'Hnr': sick104,
                                 'Preis': sick105,
                                 'Fahrer': soko
                                 })
                            conn.commit()

                    conn.commit()


                    for record in besteltree.get_children():
                        besteltree.delete(record)
                    new_window6.overrideredirect(False)
                    drop1.config(state=ACTIVE)
                    clicked1.set('')
                    Strat.config(state=ACTIVE)
                    add.config(state=DISABLED)
                    pend.config(state=DISABLED)
                    addabh.config(state=DISABLED)
                    Done.config(state=DISABLED)


                Done = Button(new_window6, text='Fertig', font=font_size, bg='red', width=5, height=1, bd=6,
                              command=Speichern, state=DISABLED)
                Done.place(x=650, y=20)

                # ---------------------------------------- Straten Abrechnung ------------------------------------------------------#
                def abrechnung():
                    sok = clicked1.get()
                    sol = len(sok)
                    if sol < 3:
                        messagebox.showerror('Error', 'Fahrer aussuchen')
                    else:
                        drop1.config(state=DISABLED)
                        add.config(state=ACTIVE)
                        pend.config(state=ACTIVE)
                        addabh.config(state=ACTIVE)
                        new_window6.overrideredirect(True)
                        Strat.config(state=DISABLED)
                        Done.config(state=ACTIVE)

                Strat = Button(Frame6, text='Abrechnung', font=font_size, bg='blue', bd=6, command=abrechnung)
                Strat.place(x=170, y=84)

                def reset():
                    AnzE.config(state=NORMAL)
                    DatE.config(state=NORMAL)
                    GesE.config(state=NORMAL)
                    AnzE.delete(0, END)
                    DatE.delete(0, END)
                    GesE.delete(0, END)

                    clicked2.set(' ')

                rest = Button(Frame5, text='Reset', bg='red', font=font_size, bd=2, command=reset)
                rest.place(x=20, y=90)

                # --------------------------------- jeder Fahrer Rechnung(summe) --------------------------------------------------#
                def fahrer(e=None):
                    count = 0
                    fahrer_list=[]
                    AnzE.config(state=NORMAL)
                    DatE.config(state=NORMAL)
                    GesE.config(state=NORMAL)

                    soko = str(clicked2.get())

                    siko = soko.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"', '')
                    conn = sqlite3.connect(resource_path('Data\Pending.db'))
                    cur = conn.cursor()
                    cur1 = conn.cursor()
                    cur2 = conn.cursor()
                    cur3 = conn.cursor()
                    cur.execute('select * from Rechnung ')
                    cur3.execute('select Fahrer from Rechnung')
                    solo=cur3.fetchall()
                    for sol in solo:
                        fahrer_list.append(str(sol).replace('(','').replace(')','').replace(',','').replace("'",""))
                    if siko in fahrer_list:
                        cur2.execute('select ROUND (SUM(Preis),0) from Rechnung Where Fahrer =(?)', (siko,))
                        cur1.execute('select count(*) from Rechnung Where Fahrer =(?)', (siko,))
                        tol = cur1.fetchall()
                        zoro = cur2.fetchall()
                        fob = cur.fetchall()
                        connE.commit()
                        datum = datetime.now().strftime("%d.%m.%Y")
                        for ziko in fob:
                            anzahl = ziko[0]
                            AnzE.delete(0, END)
                            DatE.delete(0, END)
                            GesE.delete(0, END)
                            AnzE.insert(0, tol)
                            GesE.insert(0, zoro)
                            DatE.insert(0, datum)
                        AnzE.config(state=DISABLED)
                        DatE.config(state=DISABLED)
                        GesE.config(state=DISABLED)
                    else:
                        AnzE.delete(0, END)
                        DatE.delete(0, END)
                        GesE.delete(0, END)
                    conn.close()
                drop2 = OptionMenu(Frame5, clicked2, *options, command=fahrer)
                drop2.configure(width=10, bd=0, bg='grey', font=('arial', 16, 'bold'))
                drop2.place(x=0, y=45)
                # ------------------------------------------------------------------------------------------------------------------#
                def unpending():
                    selected = besteltree2.selection()[0]
                    values = besteltree2.item(selected, 'values')
                    besteltree2.delete(selected)
                    zazo = 0
                    besteltree.insert(parent='', open=True, index='end', text='',
                                      values=(values[0], values[1], values[2], values[3], values[4], values[5]))
                unpend = Button(new_window6, text='Unpend', font=font_size, bg='green', bd=6, command=unpending)
                unpend.place(x=770, y=460)
                conn = sqlite3.connect(resource_path('Data\Pending.db'))
                cur = conn.cursor()
                conn.commit()
                conn.close()
                ############################################# Alle Rechnung von Alle Fahrer##########################################
                def fahreren():
                    ge = GesE.get().strip()
                    if ge !=  None and ge !='':
                        fa = clicked2.get()
                        an = AnzE.get()
                        da = DatE.get()
                        gm = GesE.get()
                        siko = fa.replace("'", "").replace('(', '').replace(')', '').replace(',', '').replace('"', '')
                        an = AnzE.get()
                        da = DatE.get()
                        ge = GesE.get()
                        text2 = Text(new_window6, height=40, bg='white', bd=0, font=('arial', 10, 'bold'))
                        text2.insert(END, f'\n{restName}   \n')
                        text2.insert(END, f'\n=============')
                        text2.insert(END, f'\n Fahrer:{siko}')
                        text2.insert(END, f'\n Anzahl:{an}')
                        text2.insert(END, f'\n Datum:{da}')
                        text2.insert(END, f'\n Preis:{gm} €')
                        text2.insert(END, f'\n')
                        text2.insert(END, f'\n')
                        lp = text2.get("1.0", "end-1c")
                        printfile = tempfile.mktemp(".txt")
                        open(printfile, 'w').write(lp)
                        os.startfile(printfile, "print")
                        conn = sqlite3.connect(resource_path('Data\Pending.db'))
                        cur = conn.cursor()
                        cur.execute('Insert into pending (Anzahl,Datum,Gesamt,Fahrer) values(?,?,?,?)', (an, da, ge, siko))
                        conn.commit()
                        conn1 = sqlite3.connect(resource_path('Data\Pending.db'))
                        cur1 = conn1.cursor()
                        cur1.execute('delete  from Rechnung where fahrer =(?)', (siko,))
                        conn1.commit()
                        # new_window6.destroy()
                        # openwindow7()
                        conn.close()
                        conn1.close()
                        AnzE.config(state=NORMAL)
                        DatE.config(state=NORMAL)
                        GesE.config(state=NORMAL)
                        AnzE.delete(0,END)
                        DatE.delete(0,END)
                        GesE.delete(0,END)
                        clicked.set('Fahrer')
                    elif len(GesE.get()) == 0:

                        messagebox.showerror('error', 'diese fahrer wurde schon gedruckt ')

                dru = Button(Frame5, text='Drucken', bg='green', font=font_size, bd=2, command=fahreren)
                dru.place(x=100, y=90)

                # liefer = Label(Frame5, text='Mindestpreise :', bg='white', font=font_size, bd=2)
                # liefer.place(x=50, y=195)
                # liefer = Label(Frame5, text='Mindestpreise :', bg='white', font=font_size, bd=2)
                # liefer.place(x=50, y=195)

                ####--------------------------------------     FIERAMT      -------------------------------------------------------#
                def fieramt():
                    global new_window2
                    storno='Storno'
                    conn = sqlite3.connect(resource_path('Data\Pending.db'))
                    cur = conn.cursor()
                    cur.execute('select Fahrer FROM pending where Fahrer !=(?) ',(storno,))
                    fahrer=cur.fetchall()
                    if len(fahrer) > 0:
                        zuzu = besteltree1.get_children()
                        moso = int(nummerL.get(0))
                        if moso > 1:
                            messagebox.showerror('Warning', 'da sind noch offene Bestellung bitte zu ordnen')
                        else:
                            ask = messagebox.askyesno('Fieramt', 'Sind Sie sicher')
                            if ask == 1:
                                conn = sqlite3.connect(resource_path('Data\Pending.db'))
                                cur = conn.cursor()
                                cur.execute('select * from pending where Fahrer !=(?) ',(storno,))
                                all = cur.fetchall()
                                conn1 = sqlite3.connect(resource_path('Data\Pending.db'))
                                cur1 = conn.cursor()
                                cur1.execute('select * FROM pending where Fahrer =(?) ',(storno, ))
                                them=cur1.fetchall()
                                text3 = Text(new_window6, height=40, bg='white', bd=0, font=('arial', 10, 'bold'))
                                text3.insert(END, f'\n{restName}  \n')
                                text3.insert(END, f'\n=============')
                                for alles in all:
                                    text3.insert(END, f'\n Fahrer:{alles[3]}')
                                    text3.insert(END, f'\n Anzahl:{alles[0]}')
                                    text3.insert(END, f'\n Datum:{alles[1]}')
                                    text3.insert(END, f'\n Preis:{alles[2]} €')
                                    text3.insert(END, f'\n')
                                    text3.insert(END, f'\n')
                                text3.insert(END, f'\n STORNO:')
                                for the in them:
                                    text3.insert(END, f'\n Datum:{the[1]}')
                                    text3.insert(END, f'\n Bestellung Nummer:{the[2]} ')
                                    # Select the columns "Preis" and "Anzahl" from the "pending" table
                                cur.execute('SELECT Gesamt, Anzahl FROM pending')
                                # Fetch all rows and calculate the sum of "Preis" and "Anzahl"
                                rows = cur.fetchall()
                                preis_sum = sum(float(row[0]) for row in rows)
                                anzahl_sum = sum(row[1] for row in rows)
                                text3.insert(END, f'\n=============')
                                text3.insert(END, f'\n Gesamt')
                                text3.insert(END, f'\n Anzahl:{anzahl_sum}')
                                text3.insert(END, f'\n Preis:{preis_sum} €')
                                lp = text3.get("1.0", "end-1c")
                                printfile = tempfile.mktemp(".txt")
                                open(printfile, 'w').write(lp)
                                os.startfile(printfile, "print")
                                text3.delete("1.0", "end-1c")
                                time.sleep(3)
                                conn.close()
                                messagebox.showinfo('FIERAMT',
                                                'DAS PROGRAM BITTE NICHT SCHLIESEN DER SCHLIEST VON ALLEINE EIN SCHÖNES FIERAMT')
                                connK.close()
                                connZu.close()
                                connO.close()
                                connS.close()
                                connZ.close()
                                conn1.close()
                                try:
                                    new_window2.destroy()
                                except:
                                    pass
                                new_window.destroy()
                                new_window6.destroy()

                                root.destroy()
                                os.remove(resource_path('Data\Pending.db'))
                                os.remove(resource_path('Data\OrderData.db'))
                                time.sleep(3)

                                conn = sqlite3.connect(resource_path('Data\Pending.db'))
                                cur = conn.cursor()
                                cur.execute('Create Table pending (Anzahl INTEGER , Datum INTEGER, Gesamt INTEGER ,Fahrer TEXT)')

                                cur.execute(
                                """Create Table Rechnung ( Nr INTEGER, Datum TEXT,  Preis INTEGER ,Fahrer TEXT ) """)
                                conn.commit()
                                conn1 = sqlite3.connect(resource_path('Data\OrderData.db'))
                                cur1 = conn1.cursor()
                                cur1.execute("""Create Table kundeinfo(
        
                                        ID integer primary key AUTOINCREMENT UNIQUE,
                                        kid INTEGER,
                                        ktelefonnummer INTEGER,
                                        kname TEXT,
                                        kstrasse TEXT,
                                        khausnr INTEGER,
                                        kplz INTEGER,
                                        kort TEXT,
                                        kemail TEXT,
                                        bediener TEXT,
                                        zeit INTEGER,
                                        bestellzeit INTEGER,
                                        externinfo TEXT,
                                        gesamtepreis INTEGER)""")
                                cur1.execute("""Create Table speiseinfo(
                                            zeit INTEGER,
                                            pos INTEGER,
                                            grosse TEXT,
                                            anzahl INTEGER,
                                            nr INTEGER,
                                            speise TEXT,
                                            mit TEXT,
                                            ohne TEXT,
                                            katagorie TEXT,
                                            preis INTEGER,
                                            name TEXT )""")
                                cur1.execute("CREATE TABLE lieferpara(geld text, zeit Integer)")
                                conn1.commit()
                                conn.commit()
                                conn1.close()
                                conn.close()
                                import sys
                                    # terminate the program
                                sys.exit()
                                # register the cleanup function to be called when the program is closing
                    else:
                        messagebox.showerror('Warning', 'Bitte alle fahrer Rechnung ausdrucken ausdrucken')
                fierB = Button(new_window6, text='Feierabend', bd=4, font=font_size, height=3, bg='red', command=fieramt)
                fierB.place(x=1340, y=715)
            root.mainloop()
        else:
            messagebox.showwarning('warning', 'Das program hat keine license Mehr Bitte kontaktieren Sie uns Unetr www.MindMeshLab.com')
create_root()